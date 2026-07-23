"""Modular Action Layer between the chat endpoint and OpenRouter.

Each action is a plugin that registers itself with the ActionRegistry.
When a user message arrives, ActionLayer checks every registered action.
If one matches, it executes the action and returns the result.
If none match, the message passes through to the existing AI flow.
"""

import logging
import re
from datetime import date, timedelta
from urllib.parse import quote

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Plugin base
# ---------------------------------------------------------------------------

class BaseAction:
    """Subclass to create a new action plugin.

    Must set:
        action_type (str) — unique id, e.g. ``"create_task"``
        keywords (frozenset) — trigger words for pre-filtering
        patterns (list of compiled regex) — matched against message text

    May override:
        execute(text, user) -> str
            Execute the action and return a human-readable result string.
            Default returns a placeholder.
    """

    action_type = ''
    keywords = frozenset()
    patterns = []

    def detect(self, text):
        """Return ``True`` if *text* looks like this action."""
        text_lower = text.lower()
        if not any(kw in text_lower for kw in self.keywords):
            return False
        return any(p.search(text_lower) for p in self.patterns)

    def execute(self, text, user):
        """Execute this action against the database.

        *text* — original user message
        *user* — authenticated Django user
        Returns a human-readable result string, or raises on failure.
        """
        return self._placeholder(text)

    def _placeholder(self, text):
        return f"CRM action detected: {self.action_type}"


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------

class ActionRegistry:
    """Holds all registered action plugins."""

    def __init__(self):
        self._actions = {}

    def register(self, action_cls):
        action = action_cls() if isinstance(action_cls, type) else action_cls
        self._actions[action.action_type] = action
        logger.debug('Action registered: %s', action.action_type)
        return action

    def get(self, action_type):
        return self._actions.get(action_type)

    @property
    def all(self):
        return list(self._actions.values())


# ---------------------------------------------------------------------------
# Global registry
# ---------------------------------------------------------------------------

registry = ActionRegistry()


def register(action_cls):
    """Decorator to register an action plugin."""
    registry.register(action_cls)
    return action_cls


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

class ActionLayer:
    """Entry point. Call ``handle(text, user)`` from the chat view.

    Returns a result string if a CRM action is detected and executed, or
    ``None`` to let the message pass through to OpenRouter.
    """

    # Tiebreaker: when multiple actions match, prefer the one whose entity
    # keyword appears as a whole word in the text.
    # Entity-only keywords (no generic field keywords) for position-based routing.
    _ENTITY_ROUTING_KEYWORDS = {
        'create_company': r'\bcompan(y|ies)\b',
        'delete_company': r'\bcompan(y|ies)\b',
        'update_company': r'\bcompan(y|ies)\b',
        'view_company': r'\bcompan(y|ies)\b',
        'create_deal': r'\bdeal\b',
        'delete_deal': r'\bdeals?\b',
        'update_deal': r'\bdeals?\b',
        'view_deal': r'\bdeals?\b|\bpipeline\b',
        'create_lead': r'\blead\b|\bprospect\b',
        'create_contact': r'\bcontact\b',
        'create_task': r'\btask\b|\bto-?do\b',
        'create_event': r'\bevent\b|\bmeeting\b|\bappointment\b',
        'create_campaign': r'\bcampaign\b',
        'create_notification': r'\bnotification\b|\balert\b',
        'create_workflow': r'\bworkflow\b|\bautomation\b|\brule\b',
        'delete_contact': r'\bcontact\b|\bperson\b',
        'delete_lead': r'\bleads?\b|\bprospects?\b',
        'delete_task': r'\btasks?\b',
        'delete_event': r'\bevents?\b|\bmeetings?\b|\bappointments?\b',
        'delete_campaign': r'\bcampaigns?\b',
        'delete_workflow': r'\bworkflows?\b|\bautomations?\b|\brules?\b',
        'delete_notification': r'\bnotifications?\b|\balerts?\b',
        'update_contact': r'\bcontact\b|\bperson\b',
        'update_lead': r'\bleads?\b|\bprospects?\b',
        'update_task': r'\btasks?\b|\bto-?do\b',
        'update_event': r'\bevents?\b|\bmeetings?\b|\bappointments?\b',
        'update_campaign': r'\bcampaigns?\b',
        'update_workflow': r'\bworkflows?\b|\bautomations?\b|\brules?\b',
        'update_notification': r'\bnotifications?\b|\balerts?\b',
        'view_contact': r'\bcontacts?\b|\bpersons?\b',
        'view_lead': r'\bleads?\b|\bprospects?\b',
        'view_task': r'\btasks?\b|\bto-?dos?\b',
        'view_event': r'\bevents?\b|\bmeetings?\b|\bappointments?\b',
        'view_campaign': r'\bcampaigns?\b',
        'view_workflow': r'\bworkflows?\b|\bautomations?\b|\brules?\b',
        'view_notification': r'\bnotifications?\b|\balerts?\b',
        'compose_email': r'\bemail\b|\bmail\b',
        'send_email': r'\bemail\b|\bmail\b',
        'analytics': r'\bhow\s+many\b|\bcount\b|\btotal\b|\bnumber\s+of\b|\bhow\s+much\b|\brevenue\b|\bstatistics?\b|\bstats?\b|\bsummary\b',
        'view_lead': r'\bleads?\b|\bprospects?\b',
        'dashboard': r'\bdashboard\b|\boverview\b|\bhow\s+is\b|\bcrm\s+(?:summary|overview|doing)\b|\b(?:show|view|get)\s+(?:my\s+)?dashboard\b',
        'recommendations': r'\brecommend|today\'?s?\s+(?:recommend|priorit|action)|show\s+my\s+priorit|'
                           r'what\s+(?:should|needs\s+my\s+attention|meetings\s+are)|'
                           r'any\s+overdue|which\s+(?:leads\s+require|campaigns\s+need)|'
                           r'give\s+me\s+(?:today\'?s?\s+)?(?:recommend|priorit)|'
                           r'what\s+should\s+i\s+focus',
    }
    # Full keyword dict including generic field keywords (used as fallback).
    _ENTITY_KEYWORDS = {
        'create_company': r'\bcompan(y|ies)\b',
        'delete_company': r'\bcompan(y|ies)\b',
        'update_company': r'\bcompan(y|ies)\b',
        'view_company': r'\bcompan(y|ies)\b',
        'create_deal': r'\bdeal\b',
        'delete_deal': r'\bdeals?\b',
        'update_deal': r'\bdeals?\b',
        'view_deal': r'\bdeals?\b|\bpipeline\b',
        'create_lead': r'\blead\b|\bprospect\b',
        'create_contact': r'\bcontact\b',
        'create_task': r'\btask\b|\bto-?do\b',
        'create_event': r'\bevent\b|\bmeeting\b|\bappointment\b|\bschedule\b|\bcalendar\b',
        'create_campaign': r'\bcampaign\b',
        'create_notification': r'\bnotification\b|\balert\b',
        'create_workflow': r'\bworkflow\b|\bautomation\b|\brule\b',
        'delete_contact': r'\bcontact\b|\bperson\b',
        'delete_lead': r'\bleads?\b|\bprospects?\b',
        'delete_task': r'\btasks?\b',
        'delete_event': r'\bevents?\b|\bmeetings?\b|\bappointments?\b',
        'delete_campaign': r'\bcampaigns?\b',
        'update_contact': r'\bcontact\b|\bperson\b',
        'update_lead': r'\bleads?\b|\bprospects?\b',
        'update_task': r'\btasks?\b|\bto-?do\b',
        'update_event': r'\bevents?\b|\bmeetings?\b|\bappointments?\b|' + 
                        r'\blocation\b|\btime\b|\bdate\b|\bstatus\b|\btype\b|'
                        r'\b(?:am|pm)\b',
        'compose_email': r'\bemail\b|\bmail\b|\bcompose\b|\bdraft\b',
        'send_email': r'\bemail\b|\bmail\b',
        'view_contact': r'\bcontacts?\b|\bpersons?\b',
        'view_lead': r'\bleads?\b|\bprospects?\b',
        'view_deal': r'\bdeals?\b|\bpipeline\b',
        'view_task': r'\btasks?\b|\bto-?dos?\b',
        'view_event': r'\bevents?\b|\bmeetings?\b|\bappointments?\b|\bcalendar\b|\bschedule\b',
        'view_campaign': r'\bcampaigns?\b',
        'update_campaign': r'\bcampaigns?\b|' +
                           r'\bname\b|\bsubject\b|\bstatus\b|\bbody\b|'
                           r'\bcontent\b|\bschedule\b',
        'view_workflow': r'\bworkflows?\b|\bautomations?\b|\brules?\b',
        'update_workflow': r'\bworkflows?\b|\bautomations?\b|\brules?\b|' +
                          r'\bname\b|\bstatus\b|\bactive\b|\btrigger\b|'
                          r'\bdescription\b',
        'delete_workflow': r'\bworkflows?\b|\bautomations?\b|\brules?\b',
        'view_notification': r'\bnotifications?\b|\balerts?\b',
        'update_notification': r'\bnotifications?\b|\balerts?\b|' +
                              r'\bmark\b|\bread\b|\bunread\b|\bpriority\b',
        'delete_notification': r'\bnotifications?\b|\balerts?\b',
        'analytics': r'\bhow\s+many\b|\bcount\b|\btotal\b|\bnumber\s+of\b|\bhow\s+much\b|\brevenue\b',
        'dashboard': r'\bdashboard\b|\boverview\b|\bhow\s+is\b|\bcrm\s+(?:summary|overview|doing)\b|\b(?:show|view|get)\s+(?:my\s+)?dashboard\b',
        'recommendations': r'\brecommend|today\'?s?\s+(?:recommend|priorit|action)|show\s+my\s+priorit|'
                           r'what\s+(?:should|needs\s+my\s+attention|meetings\s+are)|'
                           r'any\s+overdue|which\s+(?:leads\s+require|campaigns\s+need)|'
                           r'give\s+me\s+(?:today\'?s?\s+)?(?:recommend|priorit)|'
                           r'what\s+should\s+i\s+focus',
        'crm_insights': r'\bsummary\b|\binsight\b|\bpriorit|\brecommend|\burgen|\boverdue\b|\bfocus\b|\bwhat\s+should\b|\bhow\s+is\b',
        'smart_actions': r'\b(?:meeting|meet|follow.?up|prepare|launch|notify|when|remind|and\s+(?:then|also|create|schedule|notify))\b',
        'crm_reports': r'\b(?:export|report|download|generate|print|summary\s+report|sales\s+report)\b',
    }

    @staticmethod
    def handle(text, user=None, request=None):
        if not text or not text.strip():
            return None
        text_lower = text.lower().strip()
        matched = []
        for action in registry.all:
            if action.detect(text_lower):
                matched.append(action)

        if not matched:
            return ActionLayer._fallback_search(text, user)

        # Attach request to the chosen action so it can build absolute URLs
        def _dispatch(action):
            if request is not None:
                action._current_request = request
            return action.execute(text, user)

        if len(matched) > 1:
            best = None
            # Verbs like "create/add/new" are typically followed directly
            # by the entity type ("create workflow X").  For other verbs
            # the entity type tends to appear at the end ("show X workflow").
            verb_first = bool(re.match(
                r'(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
                r'(?:create|add|new|send|compose|write|draft|how\s+(?:many|much)|'
                r'what\s+(?:should|needs|meetings)|which\s+(?:leads|campaigns)|'
                r'any\s+overdue|show\s+(?:my|me|today)|give\s+me)\b',
                text_lower,
            ))

            if verb_first:
                best_pos = len(text_lower) + 1
                for action in matched:
                    kw = ActionLayer._ENTITY_ROUTING_KEYWORDS.get(action.action_type)
                    if kw:
                        m = re.search(kw, text_lower)
                        if m and m.start() < best_pos:
                            best = action
                            best_pos = m.start()
            else:
                best_pos = -1
                for action in matched:
                    kw = ActionLayer._ENTITY_ROUTING_KEYWORDS.get(action.action_type)
                    if kw:
                        for m in re.finditer(kw, text_lower):
                            if m.start() > best_pos:
                                best = action
                                best_pos = m.start()

            # Fallback: no entity routing keyword matched → use earliest
            # full-keyword match (original behaviour)
            if best is None:
                best_pos = len(text_lower) + 1
                for action in matched:
                    kw = ActionLayer._ENTITY_KEYWORDS.get(action.action_type)
                    if kw:
                        m = re.search(kw, text_lower)
                        if m and m.start() < best_pos:
                            best = action
                            best_pos = m.start()

            # Tiebreaker: when both compose_email and send_email match
            # at the same position, prefer send_email if a send-like
            # verb is present (send, email-as-verb, dispatch, transmit).
            if best is not None and best.action_type == 'compose_email':
                has_send_intent = bool(re.search(
                    r'\b(?:send|dispatch|transmit)\b', text_lower,
                )) or bool(re.match(
                    r'(?:please\s+)?email\s+\w+', text_lower,
                ))
                if has_send_intent:
                    for action in matched:
                        if action.action_type == 'send_email':
                            best = action
                            break

            if best:
                logger.info(
                    'Action detected: %s | user=%s msg=%s',
                    best.action_type, user, text[:80],
                )
                return _dispatch(best)

        # Check if compose_email and send_email both matched but no
        # entity keyword matched either — prefer send_email for send intent.
        send_action = None
        compose_action = None
        for a in matched:
            if a.action_type == 'send_email':
                send_action = a
            elif a.action_type == 'compose_email':
                compose_action = a
        if send_action and compose_action:
            has_send_intent = bool(re.search(
                r'\b(?:send|dispatch|transmit)\b', text_lower,
            )) or bool(re.match(
                r'(?:please\s+)?email\s+\w+', text_lower,
            ))
            if has_send_intent:
                logger.info(
                    'Action detected: %s | user=%s msg=%s',
                    send_action.action_type, user, text[:80],
                )
                return _dispatch(send_action)

        action = matched[0]
        logger.info(
            'Action detected: %s | user=%s msg=%s',
            action.action_type, user, text[:80],
        )
        return _dispatch(action)

    @staticmethod
    def _fallback_search(text, user):
        """When no action matches, try searching all entities for *text*.

        If exactly one object matches → show its details.
        If multiple objects match → show grouped results.
        If none → return ``None`` so the message passes to the AI.
        """
        if not user or not user.is_authenticated:
            return None

        query = text.strip()
        if not query or len(query) < 2:
            return None

        from django.db.models import Q

        results = []  # (type_label, obj_list)

        # Contacts
        contacts = list(user.contacts.filter(
            Q(full_name__icontains=query)
        )[:5])
        if contacts:
            results.append(('Contact', contacts))

        # Leads
        leads = list(user.leads.filter(
            Q(lead_name__icontains=query) | Q(contact_person__icontains=query)
        )[:5])
        if leads:
            results.append(('Lead', leads))

        # Tasks
        tasks = list(user.tasks.filter(
            Q(title__icontains=query)
        )[:5])
        if tasks:
            results.append(('Task', tasks))

        # Events
        events = list(user.events.filter(
            Q(title__icontains=query)
        )[:5])
        if events:
            results.append(('Event', events))

        # Campaigns
        campaigns = list(user.campaigns.filter(
            Q(name__icontains=query)
        )[:5])
        if campaigns:
            results.append(('Campaign', campaigns))

        # Workflows
        workflows = list(user.workflows.filter(
            Q(name__icontains=query)
        )[:5])
        if workflows:
            results.append(('Workflow', workflows))

        # Notifications
        notifications = list(user.notifications.filter(
            Q(title__icontains=query)
        )[:5])
        if notifications:
            results.append(('Notification', notifications))

        if not results:
            return None

        total = sum(len(items) for _, items in results)

        if total == 1:
            # Single match — open it directly
            type_label, items = results[0]
            obj = items[0]
            return SearchAction._format_single_result(type_label, obj)

        # Multiple matches — grouped results
        lines = [f'I found **{total}** matches for "{query}":']
        for type_label, items in results:
            lines.append(f'\n**{type_label}**')
            for obj in items:
                name = SearchAction._obj_name(type_label, obj)
                lines.append(f'- {name}')
        lines.append(
            '\nWhich one would you like to open? '
            'You can say the number or the name.'
        )
        return '\n'.join(lines)


def _is_entity_list_ref(name, ref_set):
    if not name:
        return False
    cleaned = re.sub(r'^(?:my|the|a|an|all)\s+', '', name.lower().strip()).strip()
    return cleaned in ref_set


def _build_entity_list(qs, name_field, status_field=None, status_fn=None):
    items = list(qs.order_by('-created_at'))
    if not items:
        return 'You have no items yet.'
    lines = []
    for i, obj in enumerate(items, 1):
        label = getattr(obj, name_field, str(obj))
        if status_fn:
            extra = status_fn(obj)
        elif status_field:
            extra = getattr(obj, status_field, '')
        else:
            extra = ''
        if extra:
            lines.append(f'{i}. **{label}** - {extra}')
        else:
            lines.append(f'{i}. **{label}**')
    return '\n'.join(lines)


_ENTITY_LIST_REFS = {
    'view_task': frozenset({'task', 'tasks', 'to-do', 'to-dos', 'todo', 'todos'}),
    'view_contact': frozenset({'contact', 'contacts', 'person', 'persons', 'people'}),
    'view_company': frozenset({'company', 'companies', 'organization', 'organizations', 'org', 'orgs'}),
    'view_deal': frozenset({'deal', 'deals', 'pipeline'}),
    'view_lead': frozenset({'lead', 'leads', 'prospect', 'prospects'}),
    'view_event': frozenset({'event', 'events', 'meeting', 'meetings', 'appointment', 'appointments', 'calendar'}),
    'view_campaign': frozenset({'campaign', 'campaigns'}),
    'view_workflow': frozenset({'workflow', 'workflows', 'automation', 'automations', 'rule', 'rules'}),
    'view_notification': frozenset({'notification', 'notifications', 'alert', 'alerts'}),
}


# ── View filter helpers for "show all pending tasks" type queries ──

_VIEW_FILTER_CONFIGS = {
    'view_task': {
        'entity_words': frozenset({'task', 'tasks', 'todo', 'todos', 'to-do', 'to-dos'}),
        'label': 'Tasks',
        'format_item': lambda t: f'{t.title} - {t.get_status_display()}',
        'filters': {
            'status': {
                'field': 'status',
                'map': {
                    'pending': 'pending', 'todo': 'pending', 'not started': 'pending', 'open': 'pending',
                    'in progress': 'in_progress', 'in-progress': 'in_progress', 'doing': 'in_progress', 'started': 'in_progress',
                    'completed': 'completed', 'done': 'completed', 'finished': 'completed',
                }
            },
            'priority': {
                'field': 'priority',
                'map': {
                    'low': 'low', 'minor': 'low',
                    'medium': 'medium', 'normal': 'medium', 'moderate': 'medium',
                    'high': 'high', 'important': 'high',
                }
            },
        },
    },
    'view_company': {
        'entity_words': frozenset({'company', 'companies', 'organization', 'organizations'}),
        'label': 'Companies',
        'format_item': lambda c: f'{c.name} - {c.industry or "—"} ({c.status})',
        'filters': {
            'status': {
                'field': 'status',
                'map': {
                    'active': 'Active', 'inactive': 'Inactive',
                    'lead': 'Lead', 'prospect': 'Prospect',
                    'customer': 'Customer', 'partner': 'Partner',
                    'former': 'Former',
                }
            },
            'industry': {
                'field': 'industry',
                'map': {
                    'technology': 'Technology', 'tech': 'Technology',
                    'healthcare': 'Healthcare', 'health': 'Healthcare',
                    'finance': 'Finance', 'banking': 'Finance',
                    'education': 'Education',
                    'manufacturing': 'Manufacturing',
                    'retail': 'Retail',
                    'real estate': 'Real Estate',
                    'consulting': 'Consulting',
                    'media': 'Media',
                    'telecommunications': 'Telecommunications', 'telecom': 'Telecommunications',
                    'transportation': 'Transportation',
                    'energy': 'Energy',
                    'hospitality': 'Hospitality',
                    'agriculture': 'Agriculture',
                }
            },
        },
    },
    'view_deal': {
        'entity_words': frozenset({'deal', 'deals', 'pipeline'}),
        'label': 'Deals',
        'format_item': lambda d: f'{d.deal_name} - {d.stage} (${d.value:,.2f})' if d.value else f'{d.deal_name} - {d.stage}',
        'filters': {
            'stage': {
                'field': 'stage',
                'map': {
                    'new': 'New', 'qualified': 'Qualified',
                    'proposal sent': 'Proposal Sent', 'proposal': 'Proposal Sent',
                    'negotiation': 'Negotiation',
                    'contract review': 'Contract Review', 'contract': 'Contract Review',
                    'won': 'Won', 'lost': 'Lost',
                }
            },
            'priority': {
                'field': 'priority',
                'map': {
                    'low': 'Low', 'minor': 'Low',
                    'medium': 'Medium', 'normal': 'Medium', 'moderate': 'Medium',
                    'high': 'High', 'important': 'High',
                    'urgent': 'Urgent', 'critical': 'Urgent',
                }
            },
        },
    },
    'view_lead': {
        'entity_words': frozenset({'lead', 'leads', 'prospect', 'prospects'}),
        'label': 'Leads',
        'format_item': lambda l: f'{l.lead_name} - {l.get_status_display()}',
        'filters': {
            'status': {
                'field': 'status',
                'map': {
                    'new': 'New', 'contacted': 'Contacted', 'qualified': 'Qualified',
                    'proposal sent': 'Proposal Sent', 'proposal': 'Proposal Sent',
                    'negotiation': 'Negotiation', 'won': 'Won', 'lost': 'Lost', 'closed': 'Closed',
                }
            },
            'priority': {
                'field': 'priority',
                'map': {
                    'low': 'Low', 'minor': 'Low',
                    'medium': 'Medium', 'normal': 'Medium', 'moderate': 'Medium',
                    'high': 'High', 'important': 'High',
                    'urgent': 'Urgent', 'critical': 'Urgent',
                }
            },
        },
    },
    'view_event': {
        'entity_words': frozenset({'event', 'events', 'meeting', 'meetings', 'appointment', 'appointments'}),
        'label': 'Events',
        'format_item': lambda e: f"{e.title} - {e.start_date.strftime('%b %d') if e.start_date else '—'} ({e.get_status_display()})",
        'filters': {
            'status': {
                'field': 'status',
                'map': {
                    'scheduled': 'scheduled', 'planned': 'scheduled', 'upcoming': 'scheduled',
                    'completed': 'completed', 'done': 'completed', 'finished': 'completed', 'past': 'completed',
                    'cancelled': 'cancelled', 'canceled': 'cancelled',
                }
            },
            'type': {
                'field': 'event_type',
                'map': {
                    'meeting': 'meeting', 'call': 'call', 'phone': 'call',
                    'reminder': 'reminder', 'personal': 'personal',
                }
            },
        },
    },
    'view_campaign': {
        'entity_words': frozenset({'campaign', 'campaigns'}),
        'label': 'Campaigns',
        'format_item': lambda c: f'{c.name} - {c.get_status_display()}',
        'filters': {
            'status': {
                'field': 'status',
                'map': {
                    'draft': 'Draft',
                    'scheduled': 'Scheduled',
                    'sent': 'Sent', 'running': 'Sent', 'active': 'Sent',
                }
            },
        },
    },
    'view_workflow': {
        'entity_words': frozenset({'workflow', 'workflows', 'automation', 'automations', 'rule', 'rules'}),
        'label': 'Workflows',
        'format_item': lambda w: f'{w.name} - {"Active" if w.is_active else "Inactive"}',
        'filters': {
            'active': {
                'field': 'is_active',
                'map': {
                    'active': True, 'enabled': True, 'on': True,
                    'inactive': False, 'disabled': False, 'off': False,
                }
            },
        },
    },
    'view_notification': {
        'entity_words': frozenset({'notification', 'notifications', 'alert', 'alerts'}),
        'label': 'Notifications',
        'format_item': lambda n: f'{n.title} - {"Read" if n.is_read else "Unread"} ({n.get_priority_display()})',
        'filters': {
            'read': {
                'field': 'is_read',
                'map': {
                    'read': True, 'seen': True,
                    'unread': False, 'new': False,
                }
            },
            'priority': {
                'field': 'priority',
                'map': {
                    'low': 'low', 'minor': 'low',
                    'medium': 'medium', 'normal': 'medium', 'moderate': 'medium',
                    'high': 'high', 'important': 'high',
                    'urgent': 'high', 'critical': 'high',
                }
            },
        },
    },
}


def _apply_entity_filter(raw_name, qs, config_key):
    """
    Detect view filter patterns (e.g. "pending tasks", "tasks with status pending")
    and return a formatted filtered list, or *None* if no filter matches.
    """
    if not raw_name:
        return None
    text = raw_name.lower().strip()
    config = _VIEW_FILTER_CONFIGS.get(config_key)
    if not config:
        return None

    entity_words = config['entity_words']
    filters = config['filters']
    format_item = config['format_item']

    def _match_filter_against(filter_value, fcfg):
        fv = re.sub(r'^(?:my|your|our|the|a|an|all)\s+', '', filter_value.strip().lower()).strip()
        if not fv:
            return None
        fmap = fcfg['map']
        if fv in fmap:
            return fmap[fv]
        for key, mapped in fmap.items():
            key_lower = key.lower() if isinstance(key, str) else key
            if isinstance(key_lower, str) and len(fv) > 1:
                if fv == key_lower or key_lower.startswith(fv) or fv.startswith(key_lower):
                    return mapped
        return None

    def _match_filter(filter_value):
        fv = re.sub(r'^(?:my|your|our|the|a|an|all)\s+', '', filter_value.strip().lower()).strip()
        if not fv:
            return None
        for fk, fcfg in filters.items():
            res = _match_filter_against(fv, fcfg)
            if res is not None:
                return (fk, fcfg, res)
        return None

    # Pattern: "entity with (value) (type)" / "entity with (type) (value)"
    m = re.match(r'^(.+?)\s+with\s+(.+)$', text)
    if m:
        ep = re.sub(r'^(?:my|your|our|the|a|an|all)\s+', '', m.group(1).strip()).strip()
        ep_singular = ep.rstrip('s')
        if ep in entity_words or ep_singular in entity_words:
            after_with = m.group(2).strip()
            # Try "value type" — "high priority", "won status"
            m2 = re.match(r'^(.+?)\s+(status|priority|type|active|read)$', after_with, re.IGNORECASE)
            if m2:
                ftype = m2.group(2).lower()
                fval = m2.group(1).strip()
                if ftype in filters:
                    res = _match_filter_against(fval, filters[ftype])
                    if res:
                        return _format_filtered_list(qs, format_item, (ftype, filters[ftype], res), config)
            # Try "type value" — "status pending", "priority high"
            m3 = re.match(r'^(status|priority|type|active|read)\s+(.+)$', after_with, re.IGNORECASE)
            if m3:
                ftype = m3.group(1).lower()
                fval = m3.group(2).strip()
                if ftype in filters:
                    res = _match_filter_against(fval, filters[ftype])
                    if res:
                        return _format_filtered_list(qs, format_item, (ftype, filters[ftype], res), config)
            # Plain value
            res = _match_filter(after_with)
            if res:
                return _format_filtered_list(qs, format_item, res, config)

    # Pattern: split words, locate entity word
    words = text.split()
    for i, w in enumerate(words):
        if w in entity_words:
            val_before = ' '.join(words[:i])
            if val_before:
                res = _match_filter(val_before)
                if res:
                    return _format_filtered_list(qs, format_item, res, config)
            val_after = ' '.join(words[i + 1:])
            if val_after:
                val_after = re.sub(r'\s+(?:with|of|in|for|by)$', '', val_after).strip()
                if val_after:
                    res = _match_filter(val_after)
                    if res:
                        return _format_filtered_list(qs, format_item, res, config)
            break

    # Standalone filter word (entity implied by action context)
    res = _match_filter(text)
    if res:
        return _format_filtered_list(qs, format_item, res, config)

    return None


def _format_filtered_list(qs, format_item, filter_result, config):
    """Apply a matched filter, query, and return a formatted list."""
    filter_kind, fcfg, mapped_value = filter_result
    field = fcfg['field']
    qs = qs.filter(**{field: mapped_value})
    items = list(qs.order_by('-created_at'))
    if not items:
        return f'No {config["label"].lower()} found matching **{filter_kind}**.'
    lines = [f'**{config["label"]}** (filtered by {filter_kind}):']
    for i, obj in enumerate(items, 1):
        lines.append(f'{i}. {format_item(obj)}')
    return '\n'.join(lines)


# ===========================================================================
# Built-in action plugins
# ===========================================================================

# ── Helpers shared across actions ──

_WEEKDAYS = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday',
             'saturday', 'sunday']

_MONTH_MAP = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12,
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
    'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9,
    'oct': 10, 'nov': 11, 'dec': 12,
}


def _parse_date_from_text(text):
    """Extract a ``date`` from *text* and return ``(date_obj, cleaned_text)``.

    Handles:
      - today, tomorrow
      - next week, next <weekday>, this <weekday>
      - in N days / in N weeks
      - Month Day, Day Month, YYYY-MM-DD, DD/MM/YYYY
      - due <date>, by <date>, on <date>
    Returns ``(None, text)`` if nothing found.
    """
    today = date.today()
    original = text
    text_lower = text.lower()

    LD = r'(?:due|by|on|before)\s+'  # optional leader word prefix

    # ── Relative dates ──

    for pattern, compute in [
        (rf'{LD}tomorrow', lambda m: today + timedelta(days=1)),
        (r'\btomorrow\b', lambda m: today + timedelta(days=1)),
        (rf'{LD}today', lambda m: today),
        (r'\btoday\b', lambda m: today),
        (rf'{LD}next\s+week', lambda m: today + timedelta(days=7)),
        (r'\bnext\s+week\b', lambda m: today + timedelta(days=7)),
    ]:
        m = re.search(pattern, text_lower)
        if m:
            return (compute(m), _remove_match(text, m))

    # next Monday/Tuesday/... (always the following week)
    for leader in [rf'{LD}next', r'\bnext']:
        m = re.search(
            leader + r'\s+(' + '|'.join(_WEEKDAYS) + r')\b', text_lower,
        )
        if m:
            target = _WEEKDAYS.index(m.group(1))
            delta = (target - today.weekday() + 7) % 7
            delta += 7  # always next week, not the upcoming occurrence
            return (today + timedelta(days=delta),
                    _remove_match(text, m))

    # this Monday/Tuesday/...
    for leader in [rf'{LD}this', r'\bthis']:
        m = re.search(
            leader + r'\s+(' + '|'.join(_WEEKDAYS) + r')\b', text_lower,
        )
        if m:
            target = _WEEKDAYS.index(m.group(1))
            delta = target - today.weekday()
            if delta <= 0:
                delta += 7
            return (today + timedelta(days=delta),
                    _remove_match(text, m))

    # Standalone weekday (e.g. "Monday" or "on Monday") — upcoming occurrence
    for leader in [rf'{LD}', r'\b']:
        m = re.search(
            leader + r'(' + '|'.join(_WEEKDAYS) + r')\b', text_lower,
        )
        if m:
            target = _WEEKDAYS.index(m.group(1))
            delta = target - today.weekday()
            if delta <= 0:
                delta += 7
            return (today + timedelta(days=delta),
                    _remove_match(text, m))

    # in N days / in N weeks
    m = re.search(r'\bin\s+(\d+)\s+(day|days|week|weeks)\b', text_lower)
    if m:
        num = int(m.group(1))
        unit = m.group(2)
        delta = num * 7 if unit in ('week', 'weeks') else num
        return (today + timedelta(days=delta),
                _remove_match(text, m))

    # ── Absolute dates ──

    # Month Day (e.g. "march 5", "mar 5th")
    for name, num in _MONTH_MAP.items():
        for prefix in [rf'{LD}', r'\b']:
            m = re.search(
                prefix + name + r'\s+(\d{1,2})(?:st|nd|rd|th)?\b', text_lower,
            )
            if m:
                try:
                    d = date(today.year, num, int(m.group(1)))
                    if d < today:
                        d = date(today.year + 1, num, int(m.group(1)))
                    return (d, _remove_match(text, m))
                except ValueError:
                    continue

    # Day Month (e.g. "5 march", "5th mar")
    for name, num in _MONTH_MAP.items():
        for prefix in [rf'{LD}', r'\b']:
            m = re.search(
                prefix + r'(\d{1,2})(?:st|nd|rd|th)?\s+' + name + r'\b', text_lower,
            )
            if m:
                try:
                    d = date(today.year, num, int(m.group(1)))
                    if d < today:
                        d = date(today.year + 1, num, int(m.group(1)))
                    return (d, _remove_match(text, m))
                except ValueError:
                    continue

    # YYYY-MM-DD
    m = re.search(r'\b(\d{4})-(\d{1,2})-(\d{1,2})\b', text)
    if m:
        try:
            return (date(int(m.group(1)), int(m.group(2)), int(m.group(3))),
                    _remove_match(text, m))
        except ValueError:
            pass

    # DD/MM/YYYY
    m = re.search(r'\b(\d{1,2})/(\d{1,2})/(\d{4})\b', text)
    if m:
        try:
            return (date(int(m.group(3)), int(m.group(2)), int(m.group(1))),
                    _remove_match(text, m))
        except ValueError:
            try:
                return (date(int(m.group(3)), int(m.group(1)), int(m.group(2))),
                        _remove_match(text, m))
            except ValueError:
                pass

    return (None, text)


def _parse_time_from_text(text):
    """Extract a ``time`` from *text*. Returns ``(time_obj, cleaned_text)``
    or ``(None, text)``.

    Handles:
      - at 3pm, at 3:30 PM
      - at 15:00 (24-hour)
    """
    from datetime import time as time_class

    # 12-hour with am/pm
    m = re.search(
        r'\b(?:at\s+|by\s+)?'
        r'(\d{1,2})(?::(\d{2}))?\s*(am|pm)\b',
        text, re.IGNORECASE,
    )
    if m:
        hour = int(m.group(1))
        minute = int(m.group(2)) if m.group(2) else 0
        ampm = m.group(3).lower()
        if ampm == 'pm' and hour < 12:
            hour += 12
        elif ampm == 'am' and hour == 12:
            hour = 0
        try:
            return (time_class(hour, minute), _remove_match(text, m))
        except ValueError:
            pass

    # 24-hour HH:MM
    m = re.search(r'\b(\d{2}):(\d{2})\b', text)
    if m:
        try:
            t = time_class(int(m.group(1)), int(m.group(2)))
            return (t, _remove_match(text, m))
        except ValueError:
            pass

    return (None, text)


def _parse_datetime_from_text(text):
    """Extract a ``datetime`` from *text*.  Returns ``(datetime_obj, cleaned_text)``
    or ``(None, text)``.

    Combines date + time parsing in a single pass so that both date and
    time words are removed from the returned text.  The returned datetime
    is timezone-aware (uses Django's current timezone).
    """
    from datetime import datetime as dt_class
    from django.utils import timezone

    d, body = _parse_date_from_text(text)
    if not d:
        return (None, text)
    t, body = _parse_time_from_text(body)
    if t:
        dt = dt_class.combine(d, t)
    else:
        dt = dt_class.combine(d, dt_class.min.time())
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt)
    return (dt, body)


def _remove_match(text, match):
    """Remove a regex match region from *text*, preserving original case."""
    return (text[:match.start()] + text[match.end():]).strip()


def _extract_priority(text):
    """Extract priority from *text*.  Returns ``(priority, cleaned_text)``.

    Priority is one of ``'high'``, ``'medium'``, ``'low'``, or ``None``.
    """
    text_lower = text.lower()
    rules = [
        (r'\bhigh\s+priority\b', 'high'),
        (r'\bpriority\s+high\b', 'high'),
        (r'\burgent\b', 'high'),
        (r'\bcritical\b', 'high'),
        (r'\bmedium\s+priority\b', 'medium'),
        (r'\bpriority\s+medium\b', 'medium'),
        (r'\bnormal\s+priority\b', 'medium'),
        (r'\blow\s+priority\b', 'low'),
        (r'\bpriority\s+low\b', 'low'),
        (r'\bminor\b', 'low'),
    ]
    for pattern, value in rules:
        m = re.search(pattern, text_lower)
        if m:
            return (value, _remove_match(text, m))
    return (None, text)


def _extract_title(text):
    """Extract a meaningful title from the remaining message text."""
    text = re.sub(r'\s+', ' ', text).strip()
    # Strip leading/trailing punctuation (including quotes)
    text = re.sub(r'^[\'\"`:\-;.,\s]+', '', text).strip()
    text = re.sub(r'[\'\"`:\-;.,\s]+$', '', text).strip()
    # Remove trailing prepositions / noise
    text = re.sub(r'\b(due|by|on|at|for|to|with|about)\s*$', '', text, flags=re.IGNORECASE).strip()
    return text


def _strip_leading_noise(text):
    """Strip leading prepositions/articles/possessives/demonstratives that may appear between verb and entity.
    
    Handles sequences, e.g. ``"from the campaign"`` → ``"campaign"``.
    """
    return re.sub(
        r'^(?:(?:from|for|about|regarding|the|a|an|my|your|our|this|that|these|those)\s+)+',
        '', text, flags=re.IGNORECASE,
    ).strip()


# ---------------------------------------------------------------------------
# Update-action helpers
# ---------------------------------------------------------------------------

# Maps natural-language field phrases → model field names for each entity

_CONTACT_UPDATE_FIELDS = {
    'phone': 'phone', 'phone number': 'phone', 'mobile': 'phone',
    'telephone': 'phone', 'tel': 'phone',
    'email': 'email', 'e-mail': 'email', 'mail': 'email',
    'company': 'company', 'organization': 'company', 'org': 'company',
    'job title': 'job_title', 'designation': 'job_title',
    'notes': 'notes', 'note': 'notes',
    'tags': 'tags', 'tag': 'tags',
    'name': 'full_name', 'full name': 'full_name',
}

_LEAD_UPDATE_FIELDS = {
    'status': 'status',
    'priority': 'priority',
    'source': 'source',
    'lead name': 'lead_name', 'name': 'lead_name',
    'company': 'company',
    'contact person': 'contact_person',
    'expected revenue': 'expected_revenue', 'revenue': 'expected_revenue',
    'email': 'email', 'phone': 'phone',
    'notes': 'notes', 'note': 'notes',
}

_TASK_UPDATE_FIELDS = {
    'status': 'status',
    'priority': 'priority',
    'due date': 'due_date', 'deadline': 'due_date',
    'due time': 'due_time',
    'title': 'title', 'name': 'title',
    'description': 'description', 'desc': 'description',
    'notes': 'description',
}

_EVENT_UPDATE_FIELDS = {
    'title': 'title', 'name': 'title',
    'date': 'start_date', 'start date': 'start_date',
    'time': 'start_time', 'start time': 'start_time',
    'end time': 'end_time',
    'location': 'location', 'place': 'location',
    'status': 'status',
    'type': 'event_type', 'event type': 'event_type',
    'description': 'description',
    'meeting link': 'meeting_link', 'link': 'meeting_link', 'url': 'meeting_link',
}

_CAMPAIGN_UPDATE_FIELDS = {
    'name': 'name',
    'subject': 'subject',
    'body': 'body', 'content': 'body', 'description': 'body',
    'status': 'status',
    'budget': 'budget',
    'scheduled at': 'scheduled_at', 'scheduled_at': 'scheduled_at',
    'scheduled date': 'scheduled_at', 'scheduled time': 'scheduled_at',
    'schedule': 'scheduled_at', 'date': 'scheduled_at', 'time': 'scheduled_at',
}

_NOTIFICATION_UPDATE_FIELDS = {
    'title': 'title',
    'message': 'message', 'msg': 'message', 'content': 'message',
    'link': 'link',
    'read': 'is_read', 'status': 'is_read',
    'priority': 'priority',
}

_WORKFLOW_UPDATE_FIELDS = {
    'name': 'name',
    'description': 'description', 'desc': 'description',
    'status': 'is_active', 'active': 'is_active', 'enabled': 'is_active',
    'trigger': 'trigger_type', 'trigger type': 'trigger_type',
}


def _parse_update_field_value(text, field_map):
    """Extract (model_field, raw_value, remaining) from an update message.

    Tries:
      1.  ``<field> to <value>`` / ``<field> as <value>`` / ``<field>: <value>``
      2.  ``as <value>`` at end (status-only updates)
      3.  ``'s <field> to <value>`` (possessive)
    Returns ``(None, None, text)`` when nothing is found.
    """
    text_stripped = text.strip()
    text_lower = text_stripped.lower()

    # Build alternation sorted longest-first so longer phrases match before
    # substrings (e.g. "phone number" before "phone").
    keys = sorted(field_map.keys(), key=len, reverse=True)
    field_alt = '|'.join(re.escape(k) for k in keys)

    # 0.  "<field> of <name> to/as <value>"  → "company of Rahul Sharma to Google"
    m = re.search(
        r'\b(' + field_alt + r')\s+of\s+(.+?)\s+(?:to|as)\s+(.+?)$',
        text_stripped, flags=re.IGNORECASE,
    )
    if m:
        raw_field = m.group(1).lower()
        value = m.group(3).strip()
        remaining = m.group(2).strip()
        return (field_map[raw_field], value, remaining)

    # 1.  "<field> to/as/is/= <value>"  or  "<field>: <value>" at end
    m = re.search(
        r'\b(' + field_alt + r')(?:\s+(?:to|as|is|=)|:\s*)\s*(.+?)$',
        text_stripped, flags=re.IGNORECASE,
    )
    if m:
        raw_field = m.group(1).lower()
        value = m.group(2).strip()
        remaining = text_stripped[:m.start()].strip()
        remaining = re.sub(r"'s\s*$", '', remaining, flags=re.IGNORECASE).strip()
        remaining = re.sub(
            r'\s+(?:to|for|with|about|regarding)\s*$', '', remaining,
            flags=re.IGNORECASE,
        ).strip()
        return (field_map[raw_field], value, remaining)

    # 2.  possessive:  "<name>'s <field> to/as/is/= <value>"  or  "<name>'s <field>: <value>"
    m = re.search(
        r"'s\s+(" + field_alt + r")(?:\s+(?:to|as|is|=)|:\s*)\s*(.+?)$",
        text_stripped, flags=re.IGNORECASE,
    )
    if m:
        raw_field = m.group(1).lower()
        value = m.group(2).strip()
        remaining = text_stripped[:m.start()].strip()
        remaining = re.sub(r"\s*'?\s*$", '', remaining)
        return (field_map[raw_field], value, remaining)

    # 3.  "as <value>" at end  (status-only)
    m = re.search(r'\bas\s+(.+?)$', text_stripped, flags=re.IGNORECASE)
    if m:
        value = m.group(1).strip()
        remaining = text_stripped[:m.start()].strip()
        remaining = re.sub(r'\s+(?:for|with|about)\s*$', '', remaining, flags=re.IGNORECASE).strip()
        return ('status', value, remaining)

    return (None, None, text)


def _normalise_contact_field_value(field, raw, user):
    """Validate + normalise a contact field value.  Returns (field, value)
    or raises ``ValueError`` explaining the problem."""
    if field == 'full_name':
        val = re.sub(r'\s+', ' ', raw).strip()
        if not val:
            raise ValueError('Name cannot be empty.')
        return (field, val)

    if field == 'email':
        val = raw.lower().strip()
        if not re.match(r'[^@\s]+@[^@\s]+\.[^@\s]+$', val):
            raise ValueError(f'"{raw}" is not a valid email address.')
        return (field, val)

    if field in ('phone', 'company', 'job_title', 'notes', 'tags'):
        return (field, raw.strip())

    raise ValueError(f'Unknown contact field: "{field}".')


def _normalise_lead_field_value(field, raw, user):
    """Validate + normalise a lead field value.  Returns (field, value)
    or raises ``ValueError``."""
    from leads.models import Lead

    if field == 'lead_name':
        val = re.sub(r'\s+', ' ', raw).strip()
        if not val:
            raise ValueError('Lead name cannot be empty.')
        return (field, val)

    if field == 'status':
        val = raw.strip().lower()
        status_map = {
            'new': 'New', 'contacted': 'Contacted', 'qualified': 'Qualified',
            'proposal sent': 'Proposal Sent', 'proposal': 'Proposal Sent',
            'negotiation': 'Negotiation',
            'won': 'Won', 'closed won': 'Won',
            'lost': 'Lost', 'closed lost': 'Lost',
        }
        if val in status_map:
            return (field, status_map[val])
        # Try partial match
        for k, mapped in status_map.items():
            if k.startswith(val) or val.startswith(k):
                return (field, mapped)
        raise ValueError(f'Unknown lead status: "{raw}". '
                         f'Valid: {", ".join(dict(Lead.STATUS_CHOICES).values())}.')

    if field == 'priority':
        val = raw.strip().lower()
        priority_map = {'low': 'Low', 'medium': 'Medium', 'high': 'High', 'urgent': 'Urgent'}
        if val in priority_map:
            return (field, priority_map[val])
        for k, mapped in priority_map.items():
            if k.startswith(val) or val.startswith(k):
                return (field, mapped)
        raise ValueError(f'Unknown priority: "{raw}". Use Low, Medium, High, or Urgent.')

    if field == 'source':
        from leads.models import Lead
        val = raw.strip().lower()
        source_map = {s.lower(): s for s in dict(Lead.SOURCE_CHOICES)}
        if val in source_map:
            return (field, source_map[val])
        for k, mapped in source_map.items():
            if k.startswith(val) or val.startswith(k):
                return (field, mapped)
        return (field, raw.strip().title())

    if field == 'expected_revenue':
        cleaned = re.sub(r'[$,]', '', raw).strip()
        try:
            return (field, float(cleaned))
        except ValueError:
            raise ValueError(f'Could not parse revenue from "{raw}".')

    if field in ('company', 'contact_person', 'email', 'phone', 'notes'):
        return (field, raw.strip())

    raise ValueError(f'Unknown lead field: "{field}".')


def _normalise_task_field_value(field, raw, user):
    """Validate + normalise a task field value.  Returns (field, value)
    or raises ``ValueError``."""
    from tasks.models import Task

    if field in ('title', 'description', 'notes'):
        return (field, raw.strip())

    if field == 'priority':
        val = raw.strip().lower()
        for choice_val, _ in Task.PRIORITY_CHOICES:
            if val == choice_val or val.startswith(choice_val) or choice_val.startswith(val):
                return (field, choice_val)
        raise ValueError(f'Unknown priority: "{raw}". '
                         f'Valid: {", ".join(dict(Task.PRIORITY_CHOICES).values())}.')

    if field == 'status':
        val = raw.strip().lower()
        status_map = {
            'pending': 'pending', 'todo': 'pending', 'not started': 'pending',
            'in progress': 'in_progress', 'in-progress': 'in_progress',
            'inprogress': 'in_progress', 'doing': 'in_progress',
            'completed': 'completed', 'done': 'completed', 'finished': 'completed',
        }
        if val in status_map:
            return (field, status_map[val])
        for k, mapped in status_map.items():
            if k.startswith(val) or val.startswith(k):
                return (field, mapped)
        raise ValueError(f'Unknown task status: "{raw}". '
                         f'Valid: {", ".join(dict(Task.STATUS_CHOICES).values())}.')

    if field == 'due_date':
        d, _ = _parse_date_from_text(raw)
        if d:
            return (field, d)
        raise ValueError(f'Could not parse a date from "{raw}".')

    if field == 'due_time':
        t, _ = _parse_time_from_text(raw)
        if t:
            return (field, t)
        raise ValueError(f'Could not parse a time from "{raw}".')

    raise ValueError(f'Unknown task field: "{field}".')


def _normalise_event_field_value(field, raw, user):
    """Validate + normalise an event field value.  Returns (field, value)
    or raises ``ValueError``."""
    from calendars.models import Event

    if field in ('title', 'description', 'location', 'meeting_link'):
        return (field, raw.strip())

    if field == 'start_date':
        d, _ = _parse_date_from_text(raw)
        if d:
            return (field, d)
        # Could also be a relative reference like "tomorrow" → _parse_date_from_text handles that
        raise ValueError(f'Could not parse a date from "{raw}".')

    if field == 'start_time':
        t, _ = _parse_time_from_text(raw)
        if t:
            return (field, t)
        raise ValueError(f'Could not parse a time from "{raw}".')

    if field == 'end_time':
        t, _ = _parse_time_from_text(raw)
        if t:
            return (field, t)
        raise ValueError(f'Could not parse a time from "{raw}".')

    if field == 'status':
        val = raw.strip().lower()
        status_map = {
            'scheduled': 'scheduled', 'planned': 'scheduled',
            'completed': 'completed', 'done': 'completed', 'finished': 'completed',
            'cancelled': 'cancelled', 'canceled': 'cancelled', 'cancel': 'cancelled',
        }
        if val in status_map:
            return (field, status_map[val])
        for k, mapped in status_map.items():
            if k.startswith(val) or val.startswith(k):
                return (field, mapped)
        raise ValueError(f'Unknown event status: "{raw}". '
                         f'Valid: {", ".join(dict(Event.STATUS_CHOICES).values())}.')

    if field == 'event_type':
        val = raw.strip().lower()
        type_map = {
            'meeting': 'meeting', 'call': 'call', 'phone': 'call',
            'reminder': 'reminder', 'personal': 'personal',
        }
        if val in type_map:
            return (field, type_map[val])
        for k, mapped in type_map.items():
            if k.startswith(val) or val.startswith(k):
                return (field, mapped)
        raise ValueError(f'Unknown event type: "{raw}". '
                         f'Valid: {", ".join(dict(Event.EVENT_TYPE_CHOICES).values())}.')

    raise ValueError(f'Unknown event field: "{field}".')


# ═══════════════════════════════════════════════════════════════════════════
#  search
# ═══════════════════════════════════════════════════════════════════════════

@register
class SearchAction(BaseAction):
    """Search all CRM entities and return grouped results or open directly."""
    action_type = 'search'
    keywords = frozenset({'search', 'find', 'lookup', 'look', 'results'})
    patterns = [
        re.compile(r'(search|find|look\s?up)\s+(?:for\s+)?'),
    ]

    @staticmethod
    def _obj_name(type_label, obj):
        if type_label == 'Contact':
            return f'{obj.full_name}  ({obj.email or obj.phone or "—"})'
        if type_label == 'Lead':
            return f'{obj.lead_name}  ({obj.get_status_display()})'
        if type_label == 'Task':
            return f'{obj.title}  ({obj.get_status_display()})'
        if type_label == 'Event':
            return f'{obj.title}  ({obj.start_date})'
        if type_label == 'Campaign':
            return f'{obj.name}  ({obj.get_status_display()})'
        if type_label == 'Workflow':
            return f'{obj.name}  ({"Active" if obj.is_active else "Inactive"})'
        if type_label == 'Notification':
            return f'{obj.title}  ({"Read" if obj.is_read else "Unread"})'
        return str(obj)

    @staticmethod
    def _format_single_result(type_label, obj):
        if type_label == 'Contact':
            created = obj.created_at.strftime('%b %d, %Y') if obj.created_at else '—'
            updated = obj.updated_at.strftime('%b %d, %Y') if obj.updated_at else '—'
            rows = [
                f'| **Name**     | {obj.full_name}',
                f'| **Email**    | {obj.email or "—"}',
                f'| **Phone**    | {obj.phone or "—"}',
                f'| **Company**  | {obj.company or "—"}',
                f'| **Position** | {obj.job_title or "—"}',
                f'| **Notes**    | {obj.notes or "—"}',
                f'| **Created**  | {created}',
                f'| **Updated**  | {updated}',
            ]
            return '### Contact Details\n' + '\n'.join(rows)
        if type_label == 'Lead':
            revenue = f'${obj.expected_revenue:,.2f}' if obj.expected_revenue else '—'
            assigned = obj.assigned_user.get_full_name() or str(obj.assigned_user) if obj.assigned_user else '—'
            rows = [
                f'| **Name**            | {obj.lead_name}',
                f'| **Email**           | {obj.email or "—"}',
                f'| **Phone**           | {obj.phone or "—"}',
                f'| **Status**          | {obj.get_status_display()}',
                f'| **Priority**        | {obj.get_priority_display()}',
                f'| **Expected Revenue**| {revenue}',
                f'| **Source**          | {obj.get_source_display()}',
                f'| **Notes**           | {obj.notes or "—"}',
                f'| **Owner**           | {assigned}',
            ]
            return '### Lead Details\n' + '\n'.join(rows)
        if type_label == 'Task':
            due_date = obj.due_date.strftime('%b %d, %Y') if obj.due_date else '—'
            due_time = obj.due_time.strftime('%I:%M %p').lstrip('0') if obj.due_time else ''
            due_str = f'{due_date} {due_time}'.strip() if due_time else due_date
            contact_name = obj.contact.full_name if obj.contact else '—'
            rows = [
                f'| **Title**       | {obj.title}',
                f'| **Description** | {obj.description or "—"}',
                f'| **Due**         | {due_str}',
                f'| **Priority**    | {obj.get_priority_display()}',
                f'| **Status**      | {obj.get_status_display()}',
                f'| **Contact**     | {contact_name}',
            ]
            return '### Task Details\n' + '\n'.join(rows)
        if type_label == 'Event':
            start_date = obj.start_date.strftime('%b %d, %Y') if obj.start_date else '—'
            start_time = obj.start_time.strftime('%I:%M %p').lstrip('0') if obj.start_time else '—'
            end_time = obj.end_time.strftime('%I:%M %p').lstrip('0') if obj.end_time else '—'
            time_str = f'{start_time} – {end_time}' if obj.start_time else '—'
            rows = [
                f'| **Title**  | {obj.title}',
                f'| **Date**   | {start_date}',
                f'| **Time**   | {time_str}',
                f'| **Location**| {obj.location or "—"}',
                f'| **Type**   | {obj.get_event_type_display()}',
                f'| **Status** | {obj.get_status_display()}',
            ]
            return '### Event Details\n' + '\n'.join(rows)
        if type_label == 'Campaign':
            scheduled = obj.scheduled_at.strftime('%b %d, %Y at %I:%M %p') if obj.scheduled_at else '—'
            rows = [
                f'| **Name**    | {obj.name}',
                f'| **Status**  | {obj.get_status_display()}',
                f'| **Subject** | {obj.subject or "—"}',
                f'| **Body**    | {obj.body[:500] if obj.body else "—"}',
                f'| **Scheduled**| {scheduled}',
                f'| **Created** | {obj.created_at.strftime("%b %d, %Y") if obj.created_at else "—"}',
            ]
            return '### Campaign Details\n' + '\n'.join(rows)
        if type_label == 'Workflow':
            actions_list = ', '.join(a.get_action_type_display() for a in obj.actions.all()) if obj.actions.exists() else '—'
            rows = [
                f'| **Name**     | {obj.name}',
                f'| **Status**   | {"Active" if obj.is_active else "Inactive"}',
                f'| **Trigger**  | {obj.get_trigger_type_display()}',
                f'| **Description**| {obj.description[:500] if obj.description else "—"}',
                f'| **Actions**  | {actions_list}',
                f'| **Created**  | {obj.created_at.strftime("%b %d, %Y") if obj.created_at else "—"}',
            ]
            return '### Workflow Details\n' + '\n'.join(rows)
        if type_label == 'Notification':
            rows = [
                f'| **Title**   | {obj.title}',
                f'| **Status**  | {"Read" if obj.is_read else "Unread"}',
                f'| **Message** | {obj.message[:500] if obj.message else "—"}',
                f'| **Link**    | {obj.link or "—"}',
                f'| **Created** | {obj.created_at.strftime("%b %d, %Y at %I:%M %p") if obj.created_at else "—"}',
            ]
            return '### Notification Details\n' + '\n'.join(rows)
        return str(obj)

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:search|find|look\s?up)\s+(?:for\s+)?',
            '', body, flags=re.IGNORECASE,
        ).strip()
        return _extract_title(body)

    def execute(self, text, user):
        query = self._parse(text)
        if not query:
            return 'What would you like to search for? I can search contacts, leads, tasks, events, campaigns, workflows, and notifications.'

        from django.db.models import Q

        results = []

        contacts = list(user.contacts.filter(
            Q(full_name__icontains=query)
        )[:5])
        if contacts:
            results.append(('Contact', contacts))

        leads = list(user.leads.filter(
            Q(lead_name__icontains=query) | Q(contact_person__icontains=query)
        )[:5])
        if leads:
            results.append(('Lead', leads))

        tasks = list(user.tasks.filter(
            Q(title__icontains=query)
        )[:5])
        if tasks:
            results.append(('Task', tasks))

        events = list(user.events.filter(
            Q(title__icontains=query)
        )[:5])
        if events:
            results.append(('Event', events))

        campaigns = list(user.campaigns.filter(
            Q(name__icontains=query)
        )[:5])
        if campaigns:
            results.append(('Campaign', campaigns))

        workflows = list(user.workflows.filter(
            Q(name__icontains=query)
        )[:5])
        if workflows:
            results.append(('Workflow', workflows))

        notifications = list(user.notifications.filter(
            Q(title__icontains=query)
        )[:5])
        if notifications:
            results.append(('Notification', notifications))

        if not results:
            return f'I could not find any contacts, leads, tasks, events, campaigns, workflows, or notifications matching "{query}".'

        total = sum(len(items) for _, items in results)

        if total == 1:
            type_label, items = results[0]
            return self._format_single_result(type_label, items[0])

        lines = [f'I found **{total}** matches for "{query}":']
        for type_label, items in results:
            lines.append(f'\n**{type_label}**')
            for obj in items:
                lines.append(f'- {self._obj_name(type_label, obj)}')
        lines.append(
            '\nWhich one would you like to open? '
            'You can type its name or number.'
        )
        return '\n'.join(lines)


# ═══════════════════════════════════════════════════════════════════════════
#  create_task
# ═══════════════════════════════════════════════════════════════════════════

@register
class CreateTaskAction(BaseAction):
    action_type = 'create_task'
    keywords = frozenset({'create', 'add', 'new', 'make', 'task', 'to-do', 'todo'})
    patterns = [re.compile(r'(create|add|new|make).*(task|to-?do)')]

    def execute(self, text, user):
        params = self._parse(text)
        title = params.get('title', '').strip()
        if not title:
            return (
                'I need a title to create a task. '
                'What should the task be called?'
            )

        priority = params.get('priority', 'medium')
        if priority not in ('low', 'medium', 'high'):
            priority = 'medium'

        contact = None
        if '_contact_raw' in params:
            contact = self._resolve_contact(params['_contact_raw'], user)
        if not contact:
            contact = self._find_contact_in_text(text, user)

        # Clean contact name from title if present
        if contact and contact.full_name:
            for name_part in [contact.full_name, contact.full_name.split()[0]]:
                if name_part.lower() in title.lower():
                    title = re.sub(
                        re.escape(name_part), '', title, flags=re.IGNORECASE,
                    ).strip()
                    title = re.sub(r'^(?:to|for)\s+', '', title, flags=re.IGNORECASE).strip()
                    title = re.sub(r'\s+', ' ', title).strip()
                    break

        if not title:
            return (
                'I need a title to create a task. '
                'What should the task be called?'
            )

        from tasks.models import Task
        from django.db import transaction

        try:
            with transaction.atomic():
                task = Task(
                    owner=user,
                    title=title,
                    description=params.get('description', ''),
                    priority=priority,
                    due_date=params.get('due_date'),
                )
                if contact:
                    task.contact = contact
                task.save()
                task.refresh_from_db()
        except Exception as e:
            logger.exception('Failed to create task for user=%s', user)
            return f'Failed to create task: {e}'

        lines = [f'Task created successfully: **{task.title}**']
        if task.due_date:
            lines.append(f'Due: {task.due_date.strftime("%b %d, %Y")}')
        lines.append(f'Priority: {task.priority.title()}')
        if task.contact:
            lines.append(f'Contact: {task.contact.full_name}')
        if task.description:
            lines.append(f'Description: {task.description[:200]}')
        return '  \n'.join(lines)

    def _parse(self, text):
        """Parse *text* into a params dict."""
        params = {}
        body = text

        # 1. Extract priority FIRST (before prefix is stripped)
        priority, body = _extract_priority(body)
        if priority:
            params['priority'] = priority

        # 2. Strip action prefix — find "task" or "to-do" and take everything after it
        m = re.search(
            r'\b(?:create|add|new|make)\s'
            r'(?:.*?\s+)?'
            r'(?:task|to-?do)\s*',
            body, flags=re.IGNORECASE,
        )
        if m:
            body = body[m.end():].strip()
        body = re.sub(r'^to\s+', '', body, flags=re.IGNORECASE).strip()

        # 3. Extract raw contact name from @mention only
        m = re.search(r'@(\w[\w\s]{0,30}?)', body)
        if m:
            params['_contact_raw'] = m.group(1).strip()
            body = _remove_match(body, m)

        # 4. Extract due date
        due_date, body = _parse_date_from_text(body)
        if due_date:
            params['due_date'] = due_date

        # 5. Extract description after colon / "description:" / "note:"
        for sep in [r'\bdescription\s*:\s*', r'\bdesc\s*:\s*',
                     r'\bnote\s*:\s*', r'\bdetails?\s*:\s*']:
            m = re.search(sep, body, flags=re.IGNORECASE)
            if m:
                after = body[m.end():].strip()
                if after:
                    params['description'] = after
                    body = body[:m.start()].strip()
                    break

        # 6. Strip leading "to " / "for " from body
        body = re.sub(r'^(?:to|for)\s+', '', body, flags=re.IGNORECASE).strip()

        # 7. Remaining body → title
        title = _extract_title(body)
        if title:
            params['title'] = title

        return params

    def _resolve_contact(self, raw_name, user):
        if not raw_name or not user:
            return None
        try:
            from contacts.models import Contact
            return Contact.objects.filter(
                owner=user, full_name__icontains=raw_name,
            ).first()
        except Exception:
            return None

    def _find_contact_in_text(self, text, user):
        """Scan *text* for words that match contact names in the DB."""
        if not text or not user:
            return None
        try:
            from contacts.models import Contact
            contacts = Contact.objects.filter(owner=user)
            text_lower = text.lower()
            for c in contacts:
                if c.full_name and c.full_name.lower() in text_lower:
                    return c
                first = c.full_name.split()[0] if c.full_name else ''
                if first and first.lower() in text_lower:
                    return c
        except Exception:
            pass
        return None


@register
class UpdateTaskAction(BaseAction):
    action_type = 'update_task'
    keywords = frozenset({
        'update', 'edit', 'change', 'modify', 'rename', 'mark', 'set',
        'task', 'done', 'complete', 'finish',
    })
    patterns = [
        re.compile(r'\b(update|edit|change|modify|rename|mark|set)\b.*(task)'),
        re.compile(r'(task).*(as|to)\s+(completed|done|finished|pending|in.progress|in_progress)'),
    ]

    def _parse(self, text):
        """Parse *text* into params dict: keys ``title``, ``field``, ``value``."""
        params = {}
        body = text

        # 1. Strip action verb
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:update|edit|change|modify|rename|mark|set)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # 1b. Strip leading prepositions/articles that may appear before entity
        body = _strip_leading_noise(body)

        # 2. Strip entity type "task"
        body = re.sub(r'^tasks?\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'\s+from\s+(?:the\s+)?tasks?[.,;:]*\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        # Insert ":" between bare field keywords and values ("status done" → "status: done")
        _task_fk = '|'.join(re.escape(k) for k in _TASK_UPDATE_FIELDS.keys())
        body = re.sub(
            r'\b(' + _task_fk + r')\s+(?!(?:to|as|is|=|of)\s)(?=\S)',
            r'\1: ', body, flags=re.IGNORECASE,
        ).strip()

        # 3. Try field-value extraction
        field, value, remaining = _parse_update_field_value(body, _TASK_UPDATE_FIELDS)
        if field:
            params['field'] = field
            params['value'] = value
            remaining = re.sub(r'\s+from\s+(?:the\s+)?tasks?[.,;:]*\s*$', '', remaining, flags=re.IGNORECASE).strip()
            remaining = re.sub(r'\s+tasks?[.,;:]*\s*$', '', remaining, flags=re.IGNORECASE).strip()
            params['title'] = _extract_title(remaining)
            return params

        # 4. Fallback: extract date / time / priority from body
        d, body = _parse_date_from_text(body)
        if d:
            body = re.sub(r'\s+from\s+(?:the\s+)?tasks?[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
            body = re.sub(r'\s+tasks?[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
            params['field'] = 'due_date'
            params['value'] = d
            params['title'] = _extract_title(body)
            return params

        priority, body = _extract_priority(body)
        if priority:
            body = re.sub(r'\s+from\s+(?:the\s+)?tasks?[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
            body = re.sub(r'\s+tasks?[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
            params['field'] = 'priority'
            params['value'] = priority
            params['title'] = _extract_title(body)
            return params

        # 5. Remaining body is the title (for "mark task X as done", etc.)
        #    But first check for trailing status word
        m = re.search(r'\b(as\s+)?(completed|done|finished|pending|in.progress|in_progress)\s*$', body, flags=re.IGNORECASE)
        if m:
            title_text = re.sub(r'\s+from\s+(?:the\s+)?tasks?[.,;:]*\s*$', '', body[:m.start()], flags=re.IGNORECASE).strip()
            title_text = re.sub(r'\s+tasks?[.,;:]*\s*$', '', title_text, flags=re.IGNORECASE).strip()
            params['field'] = 'status'
            params['value'] = m.group(2)
            params['title'] = _extract_title(title_text)
        else:
            body = re.sub(r'\s+from\s+(?:the\s+)?tasks?[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
            body = re.sub(r'\s+tasks?[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
            params['title'] = _extract_title(body)

        return params

    def execute(self, text, user):
        params = self._parse(text)
        title = params.get('title', '').strip()
        if not title:
            return (
                'I need a task title to update. '
                'Which task should I update?'
            )

        field = params.get('field', '')
        value = params.get('value')

        from tasks.models import Task
        from django.db import transaction
        from django.core.exceptions import ValidationError

        # Find the task
        qs = user.tasks.all()
        task = qs.filter(title__icontains=title).first()
        if not task:
            # Try exact match
            task = qs.filter(title__iexact=title).first()
        if not task:
            return f'I could not find a task matching "{title}".'

        # Normalise the field + value
        try:
            norm_field, norm_value = _normalise_task_field_value(field, str(value), user)
        except ValueError as e:
            return str(e)

        old_value = getattr(task, norm_field)

        # Update
        try:
            with transaction.atomic():
                setattr(task, norm_field, norm_value)
                task.full_clean()
                task.save()
                task.refresh_from_db()
        except ValidationError as e:
            error_details = []
            for fld, errors in e.message_dict.items():
                for err in errors:
                    error_details.append(f'{fld}: {err}')
            return (
                'Failed to update task — validation error(s):\n'
                + '\n'.join(error_details)
            )
        except Exception as e:
            logger.exception('Failed to update task for user=%s', user)
            return f'Failed to update task: {e}'

        lines = [f'Task **"{task.title}"** updated.']
        if norm_field == 'status':
            lines.append(f'Status: {task.get_status_display()}')
        elif norm_field == 'priority':
            lines.append(f'Priority: {task.get_priority_display()}')
        elif norm_field == 'due_date':
            lines.append(f'Due: {task.due_date.strftime("%b %d, %Y") if task.due_date else "None"}')
        else:
            lines.append(f'{norm_field.replace("_", " ").title()}: {norm_value}')
        return '  \n'.join(lines)


@register
class DeleteTaskAction(BaseAction):
    action_type = 'delete_task'
    keywords = frozenset({'delete', 'remove', 'erase', 'cancel', 'task'})
    patterns = [
        re.compile(r'(delete|remove|erase|cancel)\s+task\b'),
        re.compile(r'(delete|remove|erase|cancel).*\btask\b'),
    ]

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:delete|remove|erase|cancel)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = _strip_leading_noise(body)
        body = re.sub(r'^tasks?\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'\s+from\s+(?:the\s+)?tasks?[.,;:]*\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'\s+tasks?[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
        return _extract_title(body)

    def execute(self, text, user):
        title = self._parse(text)
        if not title:
            return 'I need a task title to delete. Which task should I delete?'

        from tasks.models import Task
        from django.db import transaction

        qs = user.tasks.all()
        task = qs.filter(title__icontains=title).first()
        if not task:
            task = qs.filter(title__iexact=title).first()
        if not task:
            return f'I could not find a task matching "{title}".'

        try:
            with transaction.atomic():
                task.delete()
        except Exception as e:
            logger.exception('Failed to delete task for user=%s', user)
            return f'Failed to delete task: {e}'

        return f'Task **"{title}"** has been deleted.'


@register
class ViewTaskAction(BaseAction):
    action_type = 'view_task'
    keywords = frozenset({'task', 'show', 'view', 'open', 'details', 'tell', 'about'})
    patterns = [
        re.compile(r'(show|view|open).*(task)'),
        re.compile(r'(task).*(details|info|information)'),
        re.compile(r'(tell|show).*(about).*(task)'),
    ]

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:show|view|open)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'^(?:tell\s+me\s+about|tell\s+about|tell\s+me)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'^task\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'^(?:details|info|information)\s+(?:for|about|on)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'\s+(?:details|info|information)\s*$', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'\s+task\s*$', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'^about\s+', '', body, flags=re.IGNORECASE).strip()
        return _extract_title(body)

    def execute(self, text, user):
        title = self._parse(text)
        if not title:
            return 'I need a task title. Which task would you like to view?'

        from tasks.models import Task

        qs = user.tasks.all()
        task = qs.filter(title__icontains=title).first()
        if not task:
            task = qs.filter(title__iexact=title).first()
        if not task:
            if _is_entity_list_ref(title, _ENTITY_LIST_REFS['view_task']):
                all_items = qs.order_by('-created_at')
                if not all_items:
                    return 'You have no tasks yet.'
                lines = ['**Your Tasks:**']
                for i, t in enumerate(all_items, 1):
                    lines.append(f'{i}. **{t.title}** - {t.get_status_display()}')
                return '\n'.join(lines)
            filtered = _apply_entity_filter(title, qs, 'view_task')
            if filtered:
                return filtered
            return f'I could not find a task matching "{title}".'

        due_date = task.due_date.strftime('%b %d, %Y') if task.due_date else '—'
        due_time = task.due_time.strftime('%I:%M %p').lstrip('0') if task.due_time else ''
        due_str = f'{due_date} {due_time}'.strip() if due_time else due_date

        contact_name = task.contact.full_name if task.contact else '—'

        rows = [
            f'| **Title**       | {task.title}',
            f'| **Description** | {task.description or "—"}',
            f'| **Due**         | {due_str}',
            f'| **Priority**    | {task.get_priority_display()}',
            f'| **Status**      | {task.get_status_display()}',
            f'| **Contact**     | {contact_name}',
        ]
        table = '\n'.join(rows)
        return f'### Task Details\n{table}'


@register
class CompleteTaskAction(BaseAction):
    action_type = 'complete_task'
    keywords = frozenset({'mark', 'complete', 'finish', 'done', 'task'})
    patterns = [
        re.compile(r'(mark|complete|finish|done).*(task)'),
        re.compile(r'(task).*(complete|done|finished)'),
    ]

    def extract_params(self, text):
        return {}


@register
class ReopenTaskAction(BaseAction):
    action_type = 'reopen_task'
    keywords = frozenset({'reopen', 'unarchive', 'task'})
    patterns = [re.compile(r'(reopen|unarchive).*(task)')]

    def extract_params(self, text):
        return {}


@register
class CreateContactAction(BaseAction):
    action_type = 'create_contact'
    keywords = frozenset({'create', 'add', 'new', 'contact', 'person'})
    patterns = [re.compile(r'(create|add|new).*(contact|person)')]

    def execute(self, text, user):
        params = self._parse(text)
        full_name = params.get('full_name', '').strip()
        if not full_name:
            return (
                'I need a contact name to create a contact. '
                'What should the contact be called?'
            )

        from contacts.models import Contact
        from django.db import transaction
        from django.core.exceptions import ValidationError

        sid = None
        try:
            with transaction.atomic():
                sid = transaction.savepoint()
                contact = Contact(
                    owner=user,
                    full_name=full_name,
                    company=params.get('company', ''),
                    job_title=params.get('job_title', ''),
                    email=params.get('email', ''),
                    phone=params.get('phone', ''),
                    tags=params.get('tags', ''),
                    notes=params.get('notes', ''),
                )
                contact.full_clean()
                contact.save()
                contact.refresh_from_db()

                saved_pk = contact.pk
                try:
                    reloaded = Contact.objects.get(pk=saved_pk)
                except Contact.DoesNotExist:
                    transaction.savepoint_rollback(sid)
                    logger.error(
                        'Contact pk=%s was NOT committed -- DB get() returned None. '
                        'user=%s msg=%s',
                        saved_pk, user, text[:120],
                    )
                    return (
                        'Failed to create contact: the record was not persisted '
                        'to the database. Please try again.'
                    )

        except ValidationError as e:
            logger.error(
                'Contact validation failed for user=%s msg=%s error=%s',
                user, text[:120], e.message_dict,
            )
            error_details = []
            for field, errors in e.message_dict.items():
                for err in errors:
                    error_details.append(f'{field}: {err}')
            return (
                'Failed to create contact -- validation error(s):\n'
                + '\n'.join(error_details)
            )
        except Exception as e:
            logger.exception(
                'Failed to create contact for user=%s msg=%s', user, text[:120],
            )
            if sid is not None:
                logger.warning('Transaction rolled back at savepoint %s', sid)
            return f'Failed to create contact: {e}'

        lines = [
            'Contact created successfully (ID: {}): **{}**'.format(
                contact.pk, contact.full_name,
            ),
        ]
        if contact.job_title:
            lines.append(f'Title: {contact.job_title}')
        if contact.company:
            lines.append(f'Company: {contact.company}')
        if contact.email:
            lines.append(f'Email: {contact.email}')
        if contact.phone:
            lines.append(f'Phone: {contact.phone}')
        if contact.tags:
            lines.append(f'Tags: {contact.tags}')
        if contact.notes:
            lines.append(f'Notes: {contact.notes[:200]}')
        return '  \n'.join(lines)

    def _parse(self, text):
        """Parse *text* into a params dict."""
        params = {}
        body = text

        # 1. Strip action prefix -- find "contact" or "person" after verb
        m = re.search(
            r'\b(?:create|add|new)\s+(?:a\s+|an\s+)?(?:new\s+)?'
            r'(?:contact|person)\s*',
            body, flags=re.IGNORECASE,
        )
        if m:
            body = body[m.end():].strip()
        body = re.sub(
            r'^(?:for|named|called|about)\s+', '', body, flags=re.IGNORECASE,
        ).strip()

        # 2. Extract structured fields via keyword:value patterns
        _end = r'(?=\s+(?:company|title|tags?|notes?|phone|email)|$)'
        field_patterns = [
            ('email', r'\bemail\s*[:=]\s*(\S+@\S+\.\S+)'),
            ('email', r'\b([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})\b'),
            ('phone', r'\bphone\s*[:=]\s*(.+?)' + _end),
            ('phone', r'\b(?:tel|mobile|cell|phone)\s*[:=]\s*(.+?)' + _end),
            ('company', r'\bcompany\s*[:=]\s*(.+?)' + _end),
            ('job_title', r'\btitle\s*[:=]\s*(.+?)' + _end),
            ('tags', r'\btags?\s*[:=]\s*(.+?)' + _end),
            ('notes', r'\bnotes?\s*[:=]\s*(.+?)$'),
        ]
        for key, pattern in field_patterns:
            if key in params and params[key]:
                continue
            m = re.search(pattern, body, flags=re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if key == 'email':
                    val = val.lower()
                if val:
                    params[key] = val
                    body = _remove_match(body, m)

        # 3. Extract company from natural language ("at X", "from X")
        if 'company' not in params:
            m = re.search(
                r'\b(?:at|from)\s+'
                r'([A-Z][A-Za-z0-9\s]{1,40}?)'
                r'(?=\s+(?:tags?|notes?|phone|email|title)|$)',
                body,
            )
            if m:
                cand = m.group(1).strip()
                if len(cand) > 1:
                    params['company'] = cand
                    body = _remove_match(body, m)

        # 4. Extract standalone phone number
        if 'phone' not in params:
            m = re.search(
                r'\b(\+?\d[\d\s\-().]{6,20}\d)\b', body,
            )
            if m:
                params['phone'] = m.group(1).strip()
                body = _remove_match(body, m)

        # 5. Remaining body -> full_name
        name = _extract_title(body)
        if name:
            params['full_name'] = name

        return params


@register
class UpdateContactAction(BaseAction):
    """Update a contact's fields from natural language."""
    action_type = 'update_contact'
    keywords = frozenset({
        'update', 'edit', 'change', 'modify', 'set', 'rename',
        'contact', 'person', 'phone', 'email', 'company',
        'notes', 'tags', 'mobile', 'number',
    })
    patterns = [
        re.compile(r'(update|edit|change|modify|rename|set)\b'),
        re.compile(r"(contact|person).*(?:'s\s+)?(?:phone|email|company|number|mobile|tag)"),
        re.compile(r"(?:phone|email|mobile|telephone|company|organization|notes?|tags?|job.?title|designation)\s+(?:number\s+)?(?:to|as)"),
        re.compile(r"(?:phone|email|mobile|telephone|company|organization|notes?|tags?|job.?title|designation)\s*:"),
    ]

    def _parse(self, text):
        """Parse text → ``{name, updates: [(field, value), ...]}``."""
        params = {'updates': []}
        body = text

        # 1. Strip action verb
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:update|edit|change|modify|rename|set)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # 1b. Strip leading prepositions/articles that may appear before entity
        body = _strip_leading_noise(body)

        # 2. Strip entity type
        body = re.sub(r'^(?:contacts?|people|persons?)\s+', '', body, flags=re.IGNORECASE).strip()

        keys = sorted(_CONTACT_UPDATE_FIELDS.keys(), key=len, reverse=True)
        field_alt = '|'.join(re.escape(k) for k in keys)

        # 3a. Format "field of name to value" (single update)
        m = re.search(
            r'\b(' + field_alt + r')\s+of\s+(.+?)\s+(?:to|as)\s+(.+?)$',
            body, flags=re.IGNORECASE,
        )
        if m:
            raw_field = m.group(1).lower()
            params['updates'].append((_CONTACT_UPDATE_FIELDS[raw_field], m.group(3).strip()))
            name = re.sub(r"'s\s*$", '', m.group(2).strip()).strip()
            params['name'] = _extract_title(name)
            return params

        # 3b. Extract all "field: value" / "field to value" pairs
        pair_re = r'\b(' + field_alt + r')(?:\s+(?:to|as|is|=)|:\s*)\s*'
        matches = list(re.finditer(pair_re, body, flags=re.IGNORECASE))
        if matches:
            name_text = re.sub(r"'s\s*$", '', body[:matches[0].start()].strip()).strip()
            params['name'] = _extract_title(name_text)
            for i, m in enumerate(matches):
                raw_field = m.group(1).lower()
                val_start = m.end()
                val_end = matches[i + 1].start() if i + 1 < len(matches) else len(body)
                value = body[val_start:val_end].strip().rstrip('. ')
                if value:
                    params['updates'].append((_CONTACT_UPDATE_FIELDS[raw_field], value))
            return params

        # 4. Fallback: extract phone/email from body
        m = re.search(r'\b(\+?\d[\d\s\-().]{6,20}\d)\b', body)
        if m:
            params['updates'].append(('phone', m.group(1).strip()))
            name_text = re.sub(r"'s\s*$", '', body[:m.start()].strip()).strip()
            params['name'] = _extract_title(name_text)
            return params

        m = re.search(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})\b', body)
        if m:
            params['updates'].append(('email', m.group(1).lower()))
            name_text = re.sub(r"'s\s*$", '', body[:m.start()].strip()).strip()
            params['name'] = _extract_title(name_text)
            return params

        # 5. No field-value identified — remaining body is just a name
        name = re.sub(r"'s\s*$", '', body).strip()
        params['name'] = _extract_title(name)
        return params

    def execute(self, text, user):
        params = self._parse(text)
        name = params.get('name', '').strip()
        if not name:
            return 'I need a contact name to update. Which contact should I update?'

        updates = params.get('updates', [])
        if not updates:
            return (
                f'I found contact "{name}", but what would you like to change? '
                f'You can update phone, email, company, job title, tags, or notes.'
            )

        from contacts.models import Contact
        from django.db import transaction
        from django.core.exceptions import ValidationError

        # Find the contact
        qs = user.contacts.all()
        contact = qs.filter(full_name__icontains=name).first()
        if not contact:
            contact = qs.filter(full_name__iexact=name).first()
        if not contact:
            return f'I could not find a contact matching "{name}".'

        # Normalise and set all field values
        changed = []
        for field, raw_value in updates:
            try:
                norm_field, norm_value = _normalise_contact_field_value(
                    field, str(raw_value), user,
                )
            except ValueError as e:
                return str(e)
            setattr(contact, norm_field, norm_value)
            changed.append((norm_field, norm_value))

        # Save once with all changes
        try:
            with transaction.atomic():
                contact.full_clean()
                contact.save()
                contact.refresh_from_db()
        except ValidationError as e:
            error_details = []
            for fld, errors in e.message_dict.items():
                for err in errors:
                    error_details.append(f'{fld}: {err}')
            return (
                'Failed to update contact — validation error(s):\n'
                + '\n'.join(error_details)
            )
        except Exception as e:
            logger.exception('Failed to update contact for user=%s', user)
            return f'Failed to update contact: {e}'

        lines = [f'Contact **"{contact.full_name}"** updated.']
        for norm_field, norm_value in changed:
            readable = norm_field.replace('_', ' ').title()
            lines.append(f'{readable}: {norm_value}')
        return '  \n'.join(lines)


@register
class DeleteContactAction(BaseAction):
    action_type = 'delete_contact'
    keywords = frozenset({'delete', 'remove', 'erase', 'contact', 'person'})
    patterns = [
        re.compile(r'(delete|remove|erase)\b'),
    ]

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:delete|remove|erase)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = _strip_leading_noise(body)
        body = re.sub(r'^(?:contact|person)\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'\s+(?:contact|person)\s*$', '', body, flags=re.IGNORECASE).strip()
        return _extract_title(body)

    def execute(self, text, user):
        name = self._parse(text)
        if not name:
            return 'Contact not found.'

        from contacts.models import Contact

        qs = user.contacts.all()
        contact = qs.filter(full_name__icontains=name).first()
        if not contact:
            contact = qs.filter(full_name__iexact=name).first()
        if not contact:
            return 'Contact not found.'

        from django.db import transaction
        try:
            with transaction.atomic():
                contact.delete()
        except Exception as e:
            logger.exception('Failed to delete contact for user=%s', user)
            return f'Failed to delete contact: {e}'

        # Verify deletion
        if Contact.objects.filter(pk=contact.pk).exists():
            return 'Failed to delete contact.'

        return 'Contact deleted successfully.'


@register
class ViewContactAction(BaseAction):
    action_type = 'view_contact'
    keywords = frozenset({
        'contact', 'person', 'show', 'view', 'open', 'details',
        'tell', 'about', 'who', 'give', 'information',
    })
    patterns = [
        re.compile(r'(tell|show|view).*(about)'),
        re.compile(r'^who\s+is\b'),
        re.compile(r'(show|view|open).*(details?|info|information)'),
        re.compile(r'^(?:contact|give)\b'),
        re.compile(r'(details?|info|information)\s+(about|on|for|of)'),
        re.compile(r'^(?:show|view|open|see|display)\s+(?!me\b)'),
    ]

    def _parse(self, text):
        body = text

        # Strip any leading intent prefix
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:'
            r'tell\s+(?:me\s+)?(?:about|everything\s+about)\s+|'
            r'show\s+(?:me\s+)?(?:details?\s+(?:of|for|about|on)\s+|'
            r'(?:everything|full|complete)\s+about\s+|about\s+)|'
            r'who\s+is\s+|'
            r'contact\s+|'
            r'give\s+(?:me\s+)?(?:information\s+(?:about|on)\s+|'
            r'details?\s+(?:about|on)\s+)?|'
            r'information\s+(?:about|on)\s+|'
            r'details?\s+(?:about|on|for|of)\s+|'
            r'about\s+|'
            r'(?:view|open|see|display)\s+'
            r')',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # Strip remaining noise words anywhere
        body = re.sub(
            r'\b(?:about|details?|info|information|everything|full|'
            r'complete|my|the|a|an|contact|person|entry|record|profile)\b\s*',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # Strip trailing context clauses
        body = re.sub(
            r"\s+(?:in|from|of|for|under)\s+(?:my|the|your)?\s*"
            r"(?:contacts?|crm|database|system|directory|roster|list|address\s*book)?\s*$",
            '', body, flags=re.IGNORECASE,
        ).strip()

        # Handle bare verb + name (e.g. "Show Rahul Sharma")
        body = re.sub(
            r'^(?:show|view|open|see|display)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # Strip leading prepositions that survive intent stripping
        # e.g. "Show contact details of Rahul" -> "contact" stripped, "details" noise-stripped,
        # then "of Rahul" remains — the "of" must be removed.
        body = re.sub(
            r'^(?:'
            r'details?\s+(?:of|for|about|on)\s+|'
            r'info\s+(?:about|on)\s+|'
            r'information\s+(?:about|on)\s+|'
            r'(?:of|for|about|on|in|from)\s+'
            r')',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # Trim punctuation and normalize whitespace
        body = re.sub(r'^[\s"\'.,;:!?\-]+|[\s"\'.,;:!?\-]+$', '', body).strip()
        body = re.sub(r'\s+', ' ', body).strip()

        return body

    def _related_leads(self, contact, user):
        from leads.models import Lead
        from django.db.models import Q
        q = Q()
        if contact.email:
            q |= Q(email__iexact=contact.email)
        if contact.phone:
            q |= Q(phone=contact.phone)
        if contact.full_name:
            parts = contact.full_name.split()
            for part in parts:
                if len(part) > 2:
                    q |= Q(contact_person__icontains=part)
        return user.leads.filter(q).distinct()[:10]

    def _related_tasks(self, contact, user):
        return contact.tasks.all()[:10]

    def _related_events(self, contact, user):
        return contact.events.all()[:10]

    def _related_notifications(self, contact, user):
        from django.db.models import Q
        q = Q()
        if contact.full_name:
            parts = contact.full_name.split()
            for part in parts:
                if len(part) > 2:
                    q |= Q(title__icontains=part) | Q(message__icontains=part)
        if contact.email:
            q |= Q(title__icontains=contact.email) | Q(message__icontains=contact.email)
        return user.notifications.filter(q).distinct()[:10] if q else []

    def _related_campaigns(self, contact, user):
        from django.db.models import Q
        q = Q()
        if contact.full_name:
            parts = contact.full_name.split()
            for part in parts:
                if len(part) > 2:
                    q |= Q(name__icontains=part) | Q(subject__icontains=part)
        if contact.email:
            q |= Q(name__icontains=contact.email) | Q(subject__icontains=contact.email)
        return user.campaigns.filter(q).distinct()[:10] if q else []

    def execute(self, text, user):
        name = self._parse(text)
        if not name:
            return 'I need a contact name. Which contact would you like to view?'

        from django.db.models import Q
        from contacts.models import Contact

        qs = user.contacts.all()
        contact = qs.filter(
            Q(full_name__icontains=name) |
            Q(email__icontains=name) |
            Q(phone__icontains=name)
        ).first()
        if not contact:
            if _is_entity_list_ref(name, _ENTITY_LIST_REFS['view_contact']):
                all_items = qs.order_by('full_name')
                if not all_items:
                    return 'You have no contacts yet.'
                lines = ['**Your Contacts:**']
                for i, c in enumerate(all_items, 1):
                    lines.append(f'{i}. **{c.full_name}**{f" ({c.email})" if c.email else ""}')
                return '\n'.join(lines)
            return f'I could not find a contact matching "{name}".'

        created = contact.created_at.strftime('%b %d, %Y') if contact.created_at else '—'
        updated = contact.updated_at.strftime('%b %d, %Y') if contact.updated_at else '—'

        rows = [
            f'| **Name**     | {contact.full_name}',
            f'| **Email**    | {contact.email or "—"}',
            f'| **Phone**    | {contact.phone or "—"}',
            f'| **Company**  | {contact.company or "—"}',
            f'| **Position** | {contact.job_title or "—"}',
            f'| **Tags**     | {contact.tags or "—"}',
            f'| **Notes**    | {contact.notes or "—"}',
            f'| **Created**  | {created}',
            f'| **Updated**  | {updated}',
        ]
        parts = [f'### Contact Details\n' + '\n'.join(rows)]

        # ── Related Leads ──
        related_leads = self._related_leads(contact, user)
        if related_leads:
            lead_rows = []
            for l in related_leads:
                status = l.get_status_display()
                priority = l.get_priority_display()
                rev = f'${l.expected_revenue:,.2f}' if l.expected_revenue else '—'
                lead_rows.append(
                    f'| **{l.lead_name}** | {status} | {priority} | {rev}'
                )
            header = '| Lead | Status | Priority | Expected Revenue |'
            sep = '|---|---|---|---|'
            parts.append(
                '#### Related Leads\n' + header + '\n' + sep + '\n' + '\n'.join(lead_rows)
            )

        # ── Related Tasks ──
        related_tasks = self._related_tasks(contact, user)
        if related_tasks:
            task_rows = []
            for t in related_tasks:
                due = t.due_date.strftime('%b %d') if t.due_date else '—'
                task_rows.append(
                    f'| **{t.title}** | {t.get_status_display()} | {t.get_priority_display()} | {due}'
                )
            header = '| Task | Status | Priority | Due |'
            sep = '|---|---|---|---|'
            parts.append(
                '#### Related Tasks\n' + header + '\n' + sep + '\n' + '\n'.join(task_rows)
            )

        # ── Related Events ──
        related_events = self._related_events(contact, user)
        if related_events:
            event_rows = []
            for e in related_events:
                d = e.start_date.strftime('%b %d') if e.start_date else '—'
                t = e.start_time.strftime('%I:%M %p').lstrip('0') if e.start_time else '—'
                event_rows.append(
                    f'| **{e.title}** | {d} | {t} | {e.get_event_type_display()} | {e.get_status_display()}'
                )
            header = '| Event | Date | Time | Type | Status |'
            sep = '|---|---|---|---|---|'
            parts.append(
                '#### Related Events\n' + header + '\n' + sep + '\n' + '\n'.join(event_rows)
            )

        # ── Related Notifications ──
        related_notifications = self._related_notifications(contact, user)
        if related_notifications:
            notif_rows = []
            for n in related_notifications:
                read = '✓' if n.is_read else '○'
                notif_rows.append(
                    f'| {read} | **{n.title}** | {n.created_at.strftime("%b %d")}'
                )
            header = '| | Notification | Created |'
            sep = '|---|---|---|'
            parts.append(
                '#### Related Notifications\n' + header + '\n' + sep + '\n' + '\n'.join(notif_rows)
            )

        # ── Related Campaigns ──
        related_campaigns = self._related_campaigns(contact, user)
        if related_campaigns:
            camp_rows = []
            for c in related_campaigns:
                camp_rows.append(
                    f'| **{c.name}** | {c.get_status_display()} | {c.created_at.strftime("%b %d")}'
                )
            header = '| Campaign | Status | Created |'
            sep = '|---|---|---|'
            parts.append(
                '#### Related Campaigns\n' + header + '\n' + sep + '\n' + '\n'.join(camp_rows)
            )

        return '\n\n'.join(parts)


@register
class CreateLeadAction(BaseAction):
    action_type = 'create_lead'
    keywords = frozenset({'create', 'add', 'new', 'lead', 'prospect'})
    patterns = [re.compile(r'(create|add|new).*(lead|prospect)')]

    _STATUS_MAP = {
        'new': 'New', 'contacted': 'Contacted', 'qualified': 'Qualified',
        'proposal sent': 'Proposal Sent', 'proposal': 'Proposal Sent',
        'negotiation': 'Negotiation', 'won': 'Won', 'lost': 'Lost',
        'closed won': 'Won', 'closed lost': 'Lost',
    }

    _SOURCE_MAP = {
        'website': 'Website', 'referral': 'Referral',
        'linkedin': 'LinkedIn', 'facebook': 'Facebook',
        'instagram': 'Instagram', 'cold email': 'Cold Email',
        'email': 'Cold Email', 'event': 'Event',
        'other': 'Other', 'call': 'Other', 'phone': 'Other',
    }

    def execute(self, text, user):
        params = self._parse(text)
        lead_name = params.get('lead_name', '').strip()
        if not lead_name:
            return (
                'I need a lead name to create a lead. '
                'What should the lead be called?'
            )

        from leads.models import Lead
        from django.db import transaction
        from django.core.exceptions import ValidationError

        sid = None
        try:
            with transaction.atomic():
                sid = transaction.savepoint()
                lead = Lead(
                    owner=user,
                    lead_name=lead_name,
                    company=params.get('company', ''),
                    contact_person=params.get('contact_person', ''),
                    email=params.get('email', ''),
                    phone=params.get('phone', ''),
                    status=params.get('status', 'New'),
                    priority=params.get('priority', 'Medium'),
                    source=params.get('source', 'Website'),
                    expected_revenue=params.get('revenue'),
                    notes=params.get('notes', ''),
                )
                lead.full_clean()
                lead.save()
                lead.refresh_from_db()

                # ── Post-save existence verification ──
                saved_pk = lead.pk
                try:
                    reloaded = Lead.objects.get(pk=saved_pk)
                except Lead.DoesNotExist:
                    transaction.savepoint_rollback(sid)
                    logger.error(
                        'Lead pk=%s was NOT committed — DB get() returned None. '
                        'user=%s msg=%s',
                        saved_pk, user, text[:120],
                    )
                    return (
                        'Failed to create lead: the record was not persisted to '
                        'the database. Please try again.'
                    )

        except ValidationError as e:
            logger.error(
                'Lead validation failed for user=%s msg=%s error=%s',
                user, text[:120], e.message_dict,
            )
            error_details = []
            for field, errors in e.message_dict.items():
                for err in errors:
                    error_details.append(f'{field}: {err}')
            return (
                'Failed to create lead — validation error(s):\n'
                + '\n'.join(error_details)
            )
        except Exception as e:
            logger.exception(
                'Failed to create lead for user=%s msg=%s', user, text[:120],
            )
            if sid is not None:
                logger.warning('Transaction rolled back at savepoint %s', sid)
            return f'Failed to create lead: {e}'

        lines = [f'Lead created successfully: **{lead.lead_name}**']
        if lead.company:
            lines.append(f'Company: {lead.company}')
        if lead.contact_person:
            lines.append(f'Contact: {lead.contact_person}')
        if lead.email:
            lines.append(f'Email: {lead.email}')
        if lead.phone:
            lines.append(f'Phone: {lead.phone}')
        lines.append(f'Status: {lead.status}')
        lines.append(f'Priority: {lead.priority}')
        if lead.source:
            lines.append(f'Source: {lead.source}')
        if lead.expected_revenue is not None:
            lines.append(f'Expected Revenue: ${lead.expected_revenue}')
        if lead.notes:
            lines.append(f'Notes: {lead.notes[:200]}')
        return '  \n'.join(lines)

    def _parse(self, text):
        """Parse *text* into a params dict."""
        params = {}
        body = text

        # 1. Extract lead-specific priority (has 'Urgent' as separate value)
        priority, body = self._extract_lead_priority(body)
        if priority:
            params['priority'] = priority

        # 2. Strip action prefix — find "lead", "prospect", or "deal" after verb
        m = re.search(
            r'\b(?:create|add|new)\s+(?:a\s+|an\s+)?(?:new\s+)?'
            r'(?:lead|prospect|deal)\s*',
            body, flags=re.IGNORECASE,
        )
        if m:
            body = body[m.end():].strip()
        # Also strip leading "for " / "named " / "called " after prefix
        body = re.sub(
            r'^(?:for|named|called|about)\s+', '', body, flags=re.IGNORECASE,
        ).strip()

        # 3. Extract structured fields via keyword:value patterns
        #    Using lookahead (?=...) so trailing keywords aren't consumed.
        _end = r'(?=\s+(?:status|source|priority|revenue|company|contact|notes?|phone|email)|$)'
        field_patterns = [
            ('email', r'\bemail\s*[:=]\s*(\S+@\S+\.\S+)'),
            ('email', r'\b([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})\b'),
            ('phone', r'\bphone\s*[:=]\s*(.+?)' + _end),
            ('phone', r'\b(?:tel|mobile|cell|phone)\s*[:=]\s*(.+?)' + _end),
            ('source', r'\bsource\s*[:=]\s*(.+?)' + _end),
            ('status', r'\bstatus\s*[:=]\s*(.+?)' + _end),
            ('revenue', r'\b(?:expected\s+)?revenue\s*[:=]\s*\$?([\d,]+(?:\.\d{1,2})?)'),
            ('revenue', r'\b(?:expected\s+)?revenue\s*(?:of|:)?\s*\$?([\d,]+(?:\.\d{1,2})?)'),
            ('company', r'\bcompany\s*[:=]\s*(.+?)' + _end),
            ('contact_person', r'\bcontact\s*(?:person\s+)?[:=]\s*(.+?)' + _end),
            ('notes', r'\bnotes?\s*[:=]\s*(.+?)$'),
        ]
        for key, pattern in field_patterns:
            if key in params and params[key]:
                continue
            m = re.search(pattern, body, flags=re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if key == 'source':
                    val = self._normalise_source(val)
                elif key == 'status':
                    val = self._normalise_status(val)
                elif key == 'revenue':
                    val = self._parse_revenue(val)
                elif key == 'email':
                    val = val.lower()
                if val:
                    params[key] = val
                    body = _remove_match(body, m)

        # 4. Extract company/contact from natural language (not keyworded)
        if 'company' not in params:
            m = re.search(r'\b(?:at|from|of)\s+([A-Z][A-Za-z0-9\s]{1,40}?)(?:\s+(?:contact|status|source|priority|revenue|notes|phone|email)|$)', body)
            if m:
                cand = m.group(1).strip()
                if len(cand) > 1:
                    params['company'] = cand
                    body = _remove_match(body, m)

        # 5. Extract phone pattern (standalone phone number)
        if 'phone' not in params:
            m = re.search(
                r'\b(\+?\d[\d\s\-().]{6,20}\d)\b', body,
            )
            if m:
                params['phone'] = m.group(1).strip()
                body = _remove_match(body, m)

        # 6. Source from context words
        if 'source' not in params:
            src = self._detect_source_in_text(body)
            if src:
                params['source'] = src
                # Don't remove from body — the source word might be part of the name

        # 7. Remaining body → lead name
        name = _extract_title(body)
        if name:
            params['lead_name'] = name

        return params

    @staticmethod
    def _extract_lead_priority(text):
        """Lead-specific priority extraction (includes 'Urgent' as separate value)."""
        text_lower = text.lower()
        rules = [
            (r'\burgent\b', 'Urgent'),
            (r'\bhigh\s+priority\b', 'High'),
            (r'\bpriority\s+high\b', 'High'),
            (r'\bcritical\b', 'High'),
            (r'\bmedium\s+priority\b', 'Medium'),
            (r'\bpriority\s+medium\b', 'Medium'),
            (r'\bnormal\s+priority\b', 'Medium'),
            (r'\blow\s+priority\b', 'Low'),
            (r'\bpriority\s+low\b', 'Low'),
            (r'\bminor\b', 'Low'),
        ]
        for pattern, value in rules:
            m = re.search(pattern, text_lower)
            if m:
                return (value, _remove_match(text, m))
        return (None, text)

    @staticmethod
    def _normalise_source(val):
        val = val.strip().lower()
        for key, mapped in CreateLeadAction._SOURCE_MAP.items():
            if val == key or val.startswith(key):
                return mapped
        return val.title()

    @staticmethod
    def _normalise_status(val):
        val = val.strip().lower()
        for key, mapped in CreateLeadAction._STATUS_MAP.items():
            if val == key or val.startswith(key):
                return mapped
        return val.title()

    @staticmethod
    def _parse_revenue(val):
        val = val.strip().replace(',', '')
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _detect_source_in_text(text):
        text_lower = text.lower()
        for key, mapped in CreateLeadAction._SOURCE_MAP.items():
            if key in text_lower:
                return mapped
        return None


@register
class UpdateLeadAction(BaseAction):
    """Update a lead's fields from natural language."""
    action_type = 'update_lead'
    keywords = frozenset({
        'update', 'edit', 'change', 'modify', 'set', 'move', 'rename',
        'lead', 'prospect',
        'qualified', 'won', 'lost', 'contacted',
    })
    patterns = [
        re.compile(r'\b(update|edit|change|modify|set|move|rename)\b.*(lead|prospect)'),
        re.compile(r'(lead|prospect).*(?:status|priority|source)\s+(?:to|as|is)'),
        re.compile(r'(lead|prospect).*(qualified|won|lost|contacted)'),
        re.compile(r'\b(update|edit|change|modify|set|move|rename)\b.*(status|name|priority|source|company|email|phone)'),
        re.compile(r'\b(update|edit|change|modify|set|move|rename)\b.*\b(qualified|won|lost|contacted)\b'),
    ]

    _STATUS_MAP = {
        'new': 'New', 'contacted': 'Contacted', 'qualified': 'Qualified',
        'proposal sent': 'Proposal Sent', 'proposal': 'Proposal Sent',
        'negotiation': 'Negotiation', 'won': 'Won', 'lost': 'Lost',
        'closed won': 'Won', 'closed lost': 'Lost',
    }

    def _parse(self, text):
        """Parse text → ``{name, updates: [(field, value), ...]}``."""
        params = {'updates': []}
        body = text

        # Handle rename before entity type stripping (name may contain "lead")
        if re.search(r'\brename\b', text, re.IGNORECASE) and ' to ' in body:
            stripped = re.sub(
                r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
                r'(?:rename)\s+',
                '', body, flags=re.IGNORECASE,
            ).strip()
            m = re.match(r'^(.+?)\s+to\s+(.+)$', stripped, flags=re.IGNORECASE)
            if m:
                old_name = m.group(1).strip()
                new_name = m.group(2).strip()
                new_name = re.sub(
                    r'\s+(?:leads?|prospects?|deals?)[.,;:]*\s*$',
                    '', new_name, flags=re.IGNORECASE,
                ).strip()
                params['name'] = _extract_title(old_name)
                params['updates'].append(('lead_name', new_name))
                return params

        # 1. Strip action verb
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:update|edit|change|modify|rename|set|move)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # 1b. Strip leading prepositions/articles that may appear before entity
        body = _strip_leading_noise(body)

        # 2. Strip entity type
        body = re.sub(r'^(?:leads?|prospects?|deals?)\s+', '', body, flags=re.IGNORECASE).strip()

        keys = sorted(_LEAD_UPDATE_FIELDS.keys(), key=len, reverse=True)
        field_alt = '|'.join(re.escape(k) for k in keys)

        # 3a. Format "field of name to value" (single update)
        m = re.search(
            r'\b(' + field_alt + r')\s+of\s+(.+?)\s+(?:to|as)\s+(.+?)$',
            body, flags=re.IGNORECASE,
        )
        if m:
            raw_field = m.group(1).lower()
            params['updates'].append((_LEAD_UPDATE_FIELDS[raw_field], m.group(3).strip()))
            name = re.sub(r"'s\s*$", '', m.group(2).strip()).strip()
            name = re.sub(r'\s+(?:leads?|prospects?|deals?)[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
            params['name'] = _extract_title(name)
            return params

        # 3b. Extract all "field value" / "field to value" / "field: value" pairs
        pair_re = r'\b(' + field_alt + r')(?:\s+(?:to|as|is=)\s*|:\s*|\s+)(?=\S)'
        matches = list(re.finditer(pair_re, body, flags=re.IGNORECASE))
        if matches:
            name_text = re.sub(r"'s\s*$", '', body[:matches[0].start()].strip()).strip()
            name_text = re.sub(r'\s+(?:leads?|prospects?|deals?)[.,;:]*\s*$', '', name_text, flags=re.IGNORECASE).strip()
            params['name'] = _extract_title(name_text)
            for i, m in enumerate(matches):
                raw_field = m.group(1).lower()
                val_start = m.end()
                val_end = matches[i + 1].start() if i + 1 < len(matches) else len(body)
                value = body[val_start:val_end].strip().rstrip('. ')
                if value:
                    params['updates'].append((_LEAD_UPDATE_FIELDS[raw_field], value))
            return params

        # 4. Fallback: check for status-only: "<name> qualified/won/lost/contacted"
        m = re.search(r'\b(qualified|won|lost|contacted)\b', body, flags=re.IGNORECASE)
        if m:
            params['updates'].append(('status', m.group(1)))
            name_text = re.sub(r"'s\s*$", '', body[:m.start()].strip()).strip()
            name_text = re.sub(r'\s+(?:leads?|prospects?|deals?)[.,;:]*\s*$', '', name_text, flags=re.IGNORECASE).strip()
            params['name'] = _extract_title(name_text)
            return params

        # 5. No field-value identified — remaining body is just the name
        name = re.sub(r"'s\s*$", '', body).strip()
        name = re.sub(r'\s+(?:leads?|prospects?|deals?)[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
        params['name'] = _extract_title(name)
        return params

    def execute(self, text, user):
        params = self._parse(text)
        name = params.get('name', '').strip()
        if not name:
            return 'I need a lead name to update. Which lead should I update?'

        updates = params.get('updates', [])
        if not updates:
            return (
                f'I found lead "{name}", but what would you like to change? '
                f'You can update status, priority, source, company, email, phone, '
                f'contact person, revenue, or notes.'
            )

        from leads.models import Lead
        from django.db import transaction
        from django.core.exceptions import ValidationError

        # Find the lead
        qs = user.leads.all()
        lead = qs.filter(lead_name__icontains=name).first()
        if not lead:
            lead = qs.filter(lead_name__iexact=name).first()
        if not lead:
            # Cross-entity fallback: maybe this is a workflow or campaign rename
            if re.search(r'\brename\b', text, re.IGNORECASE):
                from workflows.models import Workflow
                wf = user.workflows.filter(name__icontains=name).first()
                if wf:
                    uw = UpdateWorkflowAction()
                    return uw.execute(text, user)
                from campaigns.models import Campaign
                camp = user.campaigns.filter(name__icontains=name).first()
                if camp:
                    uc = UpdateCampaignAction()
                    return uc.execute(text, user)
            return f'I could not find a lead matching "{name}".'

        # Normalise and set all field values
        changed = []
        for field, raw_value in updates:
            try:
                norm_field, norm_value = _normalise_lead_field_value(
                    field, str(raw_value), user,
                )
            except ValueError as e:
                return str(e)
            setattr(lead, norm_field, norm_value)
            changed.append((norm_field, norm_value))

        # Save once with all changes
        try:
            with transaction.atomic():
                lead.full_clean()
                lead.save()
                lead.refresh_from_db()
        except ValidationError as e:
            error_details = []
            for fld, errors in e.message_dict.items():
                for err in errors:
                    error_details.append(f'{fld}: {err}')
            return (
                'Failed to update lead — validation error(s):\n'
                + '\n'.join(error_details)
            )
        except Exception as e:
            logger.exception('Failed to update lead for user=%s', user)
            return f'Failed to update lead: {e}'

        lines = [f'Lead **"{lead.lead_name}"** updated.']
        for norm_field, norm_value in changed:
            readable = norm_field.replace('_', ' ').title()
            lines.append(f'{readable}: {norm_value}')
        return '  \n'.join(lines)

    @staticmethod
    def _normalise_status(val):
        """Map raw status string to Lead STATUS_CHOICES value."""
        val = val.strip().lower()
        for key, mapped in UpdateLeadAction._STATUS_MAP.items():
            if val == key or val.startswith(key) or key.startswith(val):
                return mapped
        return val.title()


@register
class DeleteLeadAction(BaseAction):
    action_type = 'delete_lead'
    keywords = frozenset({'delete', 'remove', 'erase', 'cancel', 'lead', 'prospect'})
    patterns = [
        re.compile(r'(delete|remove|erase|cancel)\b'),
    ]

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:delete|remove|erase|cancel)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = _strip_leading_noise(body)
        body = re.sub(r'^(?:leads?|prospects?|deals?)\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'\s+from\s+(?:the\s+)?(?:leads?|prospects?|deals?)[.,;:]*\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'\s+(?:leads?|prospects?|deals?)[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'\s+from\s+(?:the\s+)?(?:crm|database|system|records)\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        name = re.sub(r"'s\s*$", '', body).strip()
        return _extract_title(name)

    def execute(self, text, user):
        lead_name = self._parse(text)
        if not lead_name:
            return 'Lead not found.'

        from leads.models import Lead

        qs = user.leads.all()
        lead = qs.filter(lead_name__icontains=lead_name).first()
        if not lead:
            lead = qs.filter(lead_name__iexact=lead_name).first()
        if not lead:
            return 'Lead not found.'

        from django.db import transaction
        try:
            with transaction.atomic():
                lead.delete()
        except Exception as e:
            logger.exception('Failed to delete lead for user=%s', user)
            return f'Failed to delete lead: {e}'

        # Verify deletion
        if Lead.objects.filter(pk=lead.pk).exists():
            return 'Failed to delete lead.'

        return 'Lead deleted successfully.'


@register
class ViewLeadAction(BaseAction):
    action_type = 'view_lead'
    keywords = frozenset({
        'lead', 'prospect', 'show', 'view', 'open', 'details',
        'tell', 'about', 'who',
    })
    patterns = [
        re.compile(r'(show|view|open).*(lead|prospect)'),
        re.compile(r'(lead|prospect).*(details|info|information|about)'),
        re.compile(r'(tell|show).*(about).*(lead|prospect)'),
        re.compile(r'^(lead|prospect)\b'),
        re.compile(r'^who\s+is\b'),
    ]

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:show|view|open)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'^(?:tell\s+me\s+about|tell\s+about|tell\s+me)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'^about\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'^who\s+is\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'^(?:lead|prospect|deal)\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'^(?:details?|info|information)(?:\s+(?:for|about|on)\s+|\s+)',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'\s+(?:lead|prospect|deal)\s*$', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'\s+(?:details?|info|information)\s*$', '', body, flags=re.IGNORECASE).strip()
        # Strip remaining standalone noise words
        body = re.sub(
            r'^(?:the|a|an|about|details?|info|information|everything|full|'
            r'complete|my|this|that)\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        name = re.sub(r"'s\s*$", '', body).strip()
        return _extract_title(name)

    def execute(self, text, user):
        lead_name = self._parse(text)
        if not lead_name:
            return 'I need a lead name. Which lead would you like to view?'

        from leads.models import Lead

        qs = user.leads.all()
        lead = qs.filter(lead_name__icontains=lead_name).first()
        if not lead:
            lead = qs.filter(lead_name__iexact=lead_name).first()
        if not lead:
            if _is_entity_list_ref(lead_name, _ENTITY_LIST_REFS['view_lead']):
                all_items = qs.order_by('-created_at')
                if not all_items:
                    return 'You have no leads yet.'
                lines = ['**Your Leads:**']
                for i, l in enumerate(all_items, 1):
                    lines.append(f'{i}. **{l.lead_name}** - {l.get_status_display()}')
                return '\n'.join(lines)
            filtered = _apply_entity_filter(lead_name, qs, 'view_lead')
            if filtered:
                return filtered
            return f'I could not find a lead matching "{lead_name}".'

        revenue = f'${lead.expected_revenue:,.2f}' if lead.expected_revenue else '—'
        assigned_to = lead.assigned_user.get_full_name() or str(lead.assigned_user) if lead.assigned_user else '—'
        created = lead.created_at.strftime('%b %d, %Y') if lead.created_at else '—'
        updated = lead.updated_at.strftime('%b %d, %Y') if lead.updated_at else '—'

        rows = [
            f'| **Name**            | {lead.lead_name}',
            f'| **Email**           | {lead.email or "—"}',
            f'| **Phone**           | {lead.phone or "—"}',
            f'| **Status**          | {lead.get_status_display()}',
            f'| **Priority**        | {lead.get_priority_display()}',
            f'| **Source**          | {lead.get_source_display()}',
            f'| **Expected Revenue**| {revenue}',
            f'| **Assigned To**     | {assigned_to}',
            f'| **Notes**           | {lead.notes or "—"}',
            f'| **Created**         | {created}',
            f'| **Updated**         | {updated}',
        ]
        table = '\n'.join(rows)
        return f'### Lead Details\n{table}'


@register
class AssignLeadAction(BaseAction):
    action_type = 'assign_lead'
    keywords = frozenset({'assign', 'reassign', 'transfer', 'lead'})
    patterns = [re.compile(r'(assign|reassign|transfer).*(lead)')]

    def extract_params(self, text):
        return {}


# ═══════════════════════════════════════════════════════════════════════════
#  Deal Actions
# ═══════════════════════════════════════════════════════════════════════════

@register
class CreateDealAction(BaseAction):
    action_type = 'create_deal'
    keywords = frozenset({
        'create', 'add', 'new', 'make', 'start', 'open', 'launch',
        'deal', 'pipeline',
    })
    patterns = [
        re.compile(r'(?:create|add|make|new|start|open|launch)\s+(?:a\s+)?(?:new\s+)?deal\b'),
        re.compile(r'deal\s+(?:called|named|for|titled)\b'),
        re.compile(r'(?:i\s+)?(?:want\s+to\s+)?(?:create|add|make)\s+.*\bdeal\b'),
        re.compile(r'\bnew\s+deal\b'),
    ]

    _STAGE_ALIASES = {
        'new': 'New', 'qualified': 'Qualified',
        'proposal': 'Proposal Sent', 'proposal sent': 'Proposal Sent',
        'negotiation': 'Negotiation',
        'contract': 'Contract Review', 'contract review': 'Contract Review',
        'won': 'Won', 'closed won': 'Won',
        'lost': 'Lost', 'closed lost': 'Lost',
    }
    _SOURCE_ALIASES = {
        'website': 'Website', 'referral': 'Referral', 'linkedin': 'LinkedIn',
        'facebook': 'Facebook', 'instagram': 'Instagram',
        'cold email': 'Cold Email', 'cold outreach': 'Cold Email',
        'event': 'Event', 'conference': 'Event', 'webinar': 'Event',
        'other': 'Other',
    }

    def _parse(self, text):
        params = {}
        body = text

        # 1. Strip action prefix
        m = re.search(
            r'\b(?:create|add|make|new|start|open|launch)\s+'
            r'(?:a\s+)?(?:new\s+)?(?:deal\s+)'
            r'(?:(?:called|named|for|titled)\s+)?',
            body, flags=re.IGNORECASE,
        )
        if m:
            body = body[m.end():].strip()
        body = re.sub(
            r'^(?:for|named|called|about|titled)\s+', '', body, flags=re.IGNORECASE,
        ).strip()
        body = _strip_leading_noise(body)
        body = re.sub(r"'s\s*$", '', body).strip()

        # 2. Extract value: "$50,000", "50k", "worth $50,000", "value of 50000", "$50,000 deal"
        val_match = re.search(
            r'(?:worth|value\s+(?:of\s+)?|priced?\s+(?:at\s+)?)?'
            r'\$?(\d[\d,.]*)\s*(?:k\b|K\b|,000\b)?',
            body, flags=re.IGNORECASE,
        )
        if val_match:
            raw_val = val_match.group(1).replace(',', '')
            multiplier = 1
            if val_match.group(0).lower().endswith('k'):
                multiplier = 1000
            elif val_match.group(0).lower().endswith(',000'):
                multiplier = 1000
            params['value'] = str(int(float(raw_val) * multiplier))
            body = _remove_match(body, val_match)

        # Also match "50k" pattern without $ sign
        if 'value' not in params:
            k_match = re.search(r'\b(\d+)\s*k\b', body, flags=re.IGNORECASE)
            if k_match:
                params['value'] = str(int(k_match.group(1)) * 1000)
                body = _remove_match(body, k_match)

        # 3. Extract stage: "in negotiation", "at proposal stage", "stage qualified"
        stage_pat = '|'.join(reversed(sorted(self._STAGE_ALIASES.keys(), key=len)))
        stage_match = re.search(
            rf'(?:in|at|stage\s+(?:is\s+)?)\s+({stage_pat})\b',
            body, flags=re.IGNORECASE,
        )
        if stage_match:
            raw = stage_match.group(1).lower()
            params['stage'] = self._STAGE_ALIASES.get(raw, raw.title())
            body = _remove_match(body, stage_match)

        # 4. Extract priority: "high priority", "urgent", "priority high"
        priority, body = _extract_priority(body)
        if priority:
            params['priority'] = priority.capitalize()

        # 5. Extract expected close date: "closing next month", "close date July 30", "due in 2 weeks"
        date_match = re.search(
            r'(?:clos(?:e|ing)|due|expected|target|by|before)\s+(?:date\s+)?(?:is\s+)?',
            body, flags=re.IGNORECASE,
        )
        if date_match:
            date_body = body[date_match.end():]
            d, _ = _parse_date_from_text(date_body)
            if d:
                params['expected_close_date'] = d
                body = body[:date_match.start()] + ' ' + body[date_match.end():]
        else:
            # Try bare date reference
            d, cleaned = _parse_date_from_text(body)
            if d:
                params['expected_close_date'] = d
                body = cleaned

        # 6. Extract company: "at Acme Corp", "for Acme", "company: Acme"
        company_match = re.search(
            r'\b(?:at|for|with|company\s*[:=]\s*|client\s*[:=]\s*)'
            r'\s*([A-Z][A-Za-z0-9\s&.,]{1,50}?)'
            r'(?=\s+(?:worth|value|stage|priority| clos| due| source| probab)|$)',
            body,
        )
        if company_match:
            params['company'] = company_match.group(1).strip()
            body = _remove_match(body, company_match)

        # Also try "company: X" pattern
        if 'company' not in params:
            cm = re.search(r'\bcompany\s*[:=]\s*(.+?)(?:\s+(?:worth|value|stage|priority| clos| due| source)|$)', body, flags=re.IGNORECASE)
            if cm:
                params['company'] = cm.group(1).strip()
                body = _remove_match(body, cm)

        # 7. Extract source: "from LinkedIn", "source: referral"
        source_match = re.search(
            r'\b(?:from|source\s*[:=]\s*)\s*(' + '|'.join(self._SOURCE_ALIASES.keys()) + r')\b',
            body, flags=re.IGNORECASE,
        )
        if source_match:
            raw_src = source_match.group(1).lower()
            params['source'] = self._SOURCE_ALIASES.get(raw_src, raw_src.title())
            body = _remove_match(body, source_match)

        # 8. Extract probability: "70% probability", "probability 70", "70 percent"
        prob_match = re.search(
            r'(\d{1,3})\s*%\s*(?:probab|chance|likely|prob)?',
            body, flags=re.IGNORECASE,
        )
        if not prob_match:
            prob_match = re.search(
                r'probab(?:ility)?\s*(?:of\s+)?(\d{1,3})',
                body, flags=re.IGNORECASE,
            )
        if prob_match:
            prob_val = int(prob_match.group(1))
            if 0 <= prob_val <= 100:
                params['probability'] = prob_val
                body = _remove_match(body, prob_match)

        # 9. Extract description/notes: "description: X", "notes: X"
        notes_match = re.search(r'\bnotes?\s*[:=]\s*(.+?)$', body, flags=re.IGNORECASE)
        if notes_match:
            params['notes'] = notes_match.group(1).strip()
            body = body[:notes_match.start()].strip()

        desc_match = re.search(r'\bdesc(?:ription)?\s*[:=]\s*(.+?)$', body, flags=re.IGNORECASE)
        if desc_match:
            params['description'] = desc_match.group(1).strip()
            body = body[:desc_match.start()].strip()

        # 10. Remaining body → deal name
        name = _extract_title(body)
        if name:
            _DEAL_TRIGGER_WORDS = frozenset({
                'create', 'add', 'make', 'new', 'start', 'open', 'launch',
                'deal', 'pipeline', 'a', 'an', 'the',
            })
            name_words = set(name.lower().split())
            if not name_words.issubset(_DEAL_TRIGGER_WORDS):
                params['deal_name'] = name

        return params

    def execute(self, text, user):
        params = self._parse(text)
        name = params.get('deal_name', '').strip()
        if not name:
            return (
                'I need a deal name to create a deal. '
                'What should the deal be called?'
            )

        from deals.models import Deal
        from django.db import transaction
        from django.core.exceptions import ValidationError

        if Deal.objects.filter(owner=user, deal_name__iexact=name).exists():
            return f'A deal named "{name}" already exists.'

        sid = None
        try:
            with transaction.atomic():
                sid = transaction.savepoint()
                deal = Deal(
                    owner=user,
                    deal_name=name,
                    company=params.get('company', ''),
                    value=params.get('value'),
                    stage=params.get('stage', 'New'),
                    priority=params.get('priority', 'Medium'),
                    source=params.get('source', 'Website'),
                    probability=params.get('probability', 0),
                    expected_close_date=params.get('expected_close_date'),
                    description=params.get('description', ''),
                    notes=params.get('notes', ''),
                )
                deal.full_clean()
                deal.save()
                deal.refresh_from_db()

                saved_pk = deal.pk
                try:
                    reloaded = Deal.objects.get(pk=saved_pk)
                except Deal.DoesNotExist:
                    transaction.savepoint_rollback(sid)
                    logger.error(
                        'Deal pk=%s was NOT committed -- DB get() returned None. '
                        'user=%s msg=%s',
                        saved_pk, user, text[:120],
                    )
                    return (
                        'Failed to create deal: the record was not persisted '
                        'to the database. Please try again.'
                    )

        except ValidationError as e:
            logger.error(
                'Deal validation failed for user=%s msg=%s error=%s',
                user, text[:120], e.message_dict,
            )
            error_details = []
            for field, errors in e.message_dict.items():
                for err in errors:
                    error_details.append(f'{field}: {err}')
            return (
                'Failed to create deal -- validation error(s):\n'
                + '\n'.join(error_details)
            )
        except Exception as e:
            logger.exception(
                'Failed to create deal for user=%s msg=%s', user, text[:120],
            )
            if sid is not None:
                logger.warning('Transaction rolled back at savepoint %s', sid)
            return f'Failed to create deal: {e}'

        from workflows.services.engine import fire_trigger
        fire_trigger('deal_created', deal)

        value_str = f'${deal.value:,.2f}' if deal.value else 'Not set'
        lines = [
            f'Deal created successfully (ID: {deal.pk}): **{deal.deal_name}**',
        ]
        lines.append(f'Stage: {deal.stage}')
        lines.append(f'Value: {deal.currency} {value_str}')
        lines.append(f'Priority: {deal.priority}')
        if deal.company:
            lines.append(f'Company: {deal.company}')
        if deal.source != 'Website':
            lines.append(f'Source: {deal.source}')
        if deal.probability:
            lines.append(f'Probability: {deal.probability}%')
        if deal.expected_close_date:
            lines.append(f'Expected Close: {deal.expected_close_date}')
        if deal.description:
            lines.append(f'Description: {deal.description[:200]}')
        if deal.notes:
            lines.append(f'Notes: {deal.notes[:200]}')
        return '\n'.join(lines)


@register
class UpdateDealAction(BaseAction):
    action_type = 'update_deal'
    keywords = frozenset({
        'update', 'edit', 'change', 'modify', 'move', 'set', 'rename',
        'deal', 'pipeline', 'stage', 'value', 'company', 'source',
        'probability', 'priority', 'close', 'date',
    })
    patterns = [
        re.compile(r'(update|edit|change|modify|move|set|rename)\b.*\bdeal\b'),
        re.compile(r'\bdeal\b.*(?:update|edit|change|modify|move|set)'),
        re.compile(r'(move|change)\s+.*\b(stage|pipeline)\b'),
        re.compile(r'(?:set|change|update)\s+.*\b(?:value|company|source|priority|probability|close\s*date)\b'),
        re.compile(r'\bdeal\b.*(?:to|value|company|source|priority|probability|close)'),
        re.compile(r'(?:value|company|source|priority|probability)\s+(?:to|of|as)\b'),
    ]

    _STAGE_ALIASES = {
        'qualified': 'Qualified', 'proposal': 'Proposal Sent',
        'proposal sent': 'Proposal Sent', 'proposal-sent': 'Proposal Sent',
        'negotiation': 'Negotiation',
        'contract': 'Contract Review', 'contract review': 'Contract Review',
        'won': 'Won', 'closed won': 'Won', 'winner': 'Won',
        'lost': 'Lost', 'closed lost': 'Lost', 'loser': 'Lost',
        'new': 'New',
    }
    _PRIORITY_MAP = {
        'low': 'Low', 'medium': 'Medium', 'high': 'High', 'urgent': 'Urgent',
    }
    _SOURCE_ALIASES = {
        'website': 'Website', 'referral': 'Referral', 'linkedin': 'LinkedIn',
        'facebook': 'Facebook', 'instagram': 'Instagram',
        'cold email': 'Cold Email', 'cold outreach': 'Cold Email',
        'event': 'Event', 'conference': 'Event', 'webinar': 'Event',
        'other': 'Other',
    }

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:update|edit|change|modify|move|set|rename)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = _strip_leading_noise(body)
        return body

    def execute(self, text, user):
        body = self._parse(text)
        from deals.models import Deal

        deal_name = ''
        stage = None
        priority = None
        value = None
        company = None
        source = None
        probability = None
        close_date = None
        description = None
        notes = None

        for_match = re.search(
            r'\bfor\s+(.+?)$', body, flags=re.IGNORECASE,
        )
        if for_match:
            deal_name = for_match.group(1).strip()
            body = body[:for_match.start()].strip()

        stage_match = re.search(
            r"(?:to|stage\s+to|stage\s+is)\s+(" + '|'.join(self._STAGE_ALIASES.keys()) + r")\b",
            body, flags=re.IGNORECASE,
        )
        if stage_match:
            raw = stage_match.group(1).lower()
            stage = self._STAGE_ALIASES.get(raw, raw.title())

        pri_match = re.search(
            r"(?:priority\s+(?:to|is)\s+)(" + '|'.join(self._PRIORITY_MAP.keys()) + r")\b",
            body, flags=re.IGNORECASE,
        )
        if not pri_match:
            pri_match = re.search(
                r"priority\s+(" + '|'.join(self._PRIORITY_MAP.keys()) + r")\b",
                body, flags=re.IGNORECASE,
            )
        if pri_match:
            raw = pri_match.group(1).lower()
            priority = self._PRIORITY_MAP.get(raw, raw.title())

        val_match = re.search(r"value\s+(?:to|is)\s+\$?(\d[\d,.]*)", body, flags=re.IGNORECASE)
        if not val_match:
            val_match = re.search(r"(?:set|change|update)\s+(?:the\s+)?value\s+(?:to|as)\s+\$?(\d[\d,.]*)", body, flags=re.IGNORECASE)
        if not val_match:
            val_match = re.search(r"\$?(\d[\d,.]*)\s*(?:k\b|K\b)", body, flags=re.IGNORECASE)
        if val_match:
            raw_val = val_match.group(1).replace(',', '')
            multiplier = 1
            if val_match.group(0).lower().endswith('k'):
                multiplier = 1000
            value = str(int(float(raw_val) * multiplier))

        company_match = re.search(
            r"(?:company|client)\s+(?:to|is)\s+(.+?)(?:\s+(?:to|value|stage|priority|source|probab|close|for)\b|$)",
            body, flags=re.IGNORECASE,
        )
        if not company_match:
            company_match = re.search(
                r"(?:set|change|update)\s+(?:the\s+)?(?:company|client)\s+(?:to|as)\s+(.+?)(?:\s+(?:to|value|stage|priority|source|probab|close|for)\b|$)",
                body, flags=re.IGNORECASE,
            )
        if company_match:
            company = company_match.group(1).strip()

        source_match = re.search(
            r"(?:source|from)\s+(?:to|is)\s+(" + '|'.join(self._SOURCE_ALIASES.keys()) + r")\b",
            body, flags=re.IGNORECASE,
        )
        if not source_match:
            source_match = re.search(
                r"(?:set|change|update)\s+(?:the\s+)?(?:source|from)\s+(?:to|as)\s+(" + '|'.join(self._SOURCE_ALIASES.keys()) + r")\b",
                body, flags=re.IGNORECASE,
            )
        if source_match:
            raw_src = source_match.group(1).lower()
            source = self._SOURCE_ALIASES.get(raw_src, raw_src.title())

        prob_match = re.search(r"probab(?:ility)?\s+(?:to|is)\s+(\d{1,3})", body, flags=re.IGNORECASE)
        if not prob_match:
            prob_match = re.search(r"probab(?:ility)?\s+(\d{1,3})", body, flags=re.IGNORECASE)
        if not prob_match:
            prob_match = re.search(r"(?:set|change|update)\s+(?:the\s+)?probab(?:ility)?\s+(?:to|as)\s+(\d{1,3})", body, flags=re.IGNORECASE)
        if prob_match:
            prob_val = int(prob_match.group(1))
            if 0 <= prob_val <= 100:
                probability = prob_val

        close_match = re.search(
            r"(?:clos(?:e|ing)|expected)\s+(?:date\s+)?(?:to|is)\s+",
            body, flags=re.IGNORECASE,
        )
        if close_match:
            date_body = body[close_match.end():]
            d, _ = _parse_date_from_text(date_body)
            if d:
                close_date = d

        desc_match = re.search(r"\bdesc(?:ription)?\s+(?:to|is|:)\s+(.+?)(?:\s+(?:to|value|stage|priority|source|probab|close|notes|for)\b|$)", body, flags=re.IGNORECASE)
        if desc_match:
            description = desc_match.group(1).strip()

        notes_match = re.search(r"\bnotes?\s+(?:to|is|:)\s+(.+?)$", body, flags=re.IGNORECASE)
        if notes_match:
            notes = notes_match.group(1).strip()

        if not deal_name:
            first_field_pos = len(body)
            for m in [stage_match, pri_match, val_match, company_match, source_match, prob_match, close_match, desc_match, notes_match]:
                if m and m.start() < first_field_pos:
                    first_field_pos = m.start()
            deal_name = body[:first_field_pos].strip()
            deal_name = re.sub(r"'s\s*$", '', deal_name).strip()
            deal_name = re.sub(r'\s+deal\s*$', '', deal_name, flags=re.IGNORECASE).strip()
            deal_name = re.sub(r'\s+(?:stage|priority|value|company|source|probability|close|date|description|notes)\s*$', '', deal_name, flags=re.IGNORECASE).strip()
            deal_name = _extract_title(deal_name)

        if not deal_name:
            return 'Could not identify the deal. Please specify a deal name.'

        deal = user.deals.filter(deal_name__icontains=deal_name).first()
        if not deal:
            deal = user.deals.filter(deal_name__iexact=deal_name).first()
        if not deal:
            return f'Deal "{deal_name}" not found.'

        changes = []
        if stage:
            old_stage = deal.stage
            deal.stage = stage
            deal.update_status_from_stage()
            changes.append(f'stage changed from "{old_stage}" to "{stage}"')
        if priority:
            deal.priority = priority
            changes.append(f'priority set to "{priority}"')
        if value:
            deal.value = value
            changes.append(f'value updated to ${value}')
        if company:
            deal.company = company
            changes.append(f'company set to "{company}"')
        if source:
            deal.source = source
            changes.append(f'source set to "{source}"')
        if probability is not None:
            deal.probability = probability
            changes.append(f'probability set to {probability}%')
        if close_date:
            deal.expected_close_date = close_date
            changes.append(f'expected close date set to {close_date}')
        if description:
            deal.description = description
            changes.append(f'description updated')
        if notes:
            deal.notes = notes
            changes.append(f'notes updated')

        if not changes:
            return f'No changes detected for deal "{deal.deal_name}".'

        deal.save()
        from workflows.services.engine import fire_trigger
        if stage:
            fire_trigger('deal_stage_changed', deal)
            if stage == 'Won':
                fire_trigger('deal_won', deal)
            elif stage == 'Lost':
                fire_trigger('deal_lost', deal)
        else:
            fire_trigger('deal_updated', deal)
        return f'Deal "{deal.deal_name}" updated: {", ".join(changes)}.'


@register
class DeleteDealAction(BaseAction):
    action_type = 'delete_deal'
    keywords = frozenset({
        'delete', 'remove', 'erase', 'cancel', 'destroy', 'drop',
        'deal', 'pipeline',
    })
    patterns = [
        re.compile(r'(?:delete|remove|erase|cancel|destroy|drop)\b.*\bdeal\b'),
        re.compile(r'\bdeal\b.*(?:delete|remove|erase|cancel|destroy|drop)'),
        re.compile(r'(?:delete|remove|erase|cancel|destroy|drop)\s+(?:the\s+)?deal\b'),
        re.compile(r'\bdelete\s+all\s+deals?\b'),
    ]

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:delete|remove|erase|cancel)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = _strip_leading_noise(body)
        body = re.sub(r'^(?:deals?)\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'\s+(?:deals?)[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r"\s+from\s+(?:the\s+)?(?:crm|database|system)\s*$", '', body, flags=re.IGNORECASE).strip()
        name = re.sub(r"'s\s*$", '', body).strip()
        return _extract_title(name)

    def execute(self, text, user):
        deal_name = self._parse(text)
        if not deal_name:
            return 'Deal not found.'
        from deals.models import Deal
        deal = user.deals.filter(deal_name__icontains=deal_name).first()
        if not deal:
            deal = user.deals.filter(deal_name__iexact=deal_name).first()
        if not deal:
            return 'Deal not found.'
        from django.db import transaction
        try:
            with transaction.atomic():
                deal.delete()
        except Exception as e:
            logger.exception('Failed to delete deal for user=%s', user)
            return f'Failed to delete deal: {e}'
        if Deal.objects.filter(pk=deal.pk).exists():
            return 'Failed to delete deal.'
        return 'Deal deleted successfully.'


@register
class ViewDealAction(BaseAction):
    action_type = 'view_deal'
    keywords = frozenset({
        'deal', 'deals', 'pipeline', 'show', 'view', 'open', 'details',
        'worth', 'value', 'closing', 'won', 'lost', 'priority', 'high',
    })
    patterns = [
        re.compile(r'(show|view|open).*(deal|pipeline)'),
        re.compile(r'(deal|pipeline).*(details|info|information|about|summary)'),
        re.compile(r'(worth|value|closing|won|lost|high|priority).*(deal)'),
        re.compile(r'(deal).*(worth|value|closing|priority)'),
        re.compile(r'show\s+(?:all\s+)?(?:won|lost|open|high\s+priority)\s+deals?'),
        re.compile(r'(?:which|what)\s+deals?\s+.*(?:closing|worth|value|priority)'),
        re.compile(r'pipeline\s+summary'),
    ]

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:show|view|open)\s+(?:all\s+)?(?:the\s+)?',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'^(?:deals?)\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'\s+(?:deals?)[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
        return _extract_title(body)

    def _pipeline_summary(self, user):
        from deals.models import Deal
        qs = Deal.objects.filter(owner=user)
        total = qs.count()
        won = qs.filter(stage='Won').count()
        lost = qs.filter(stage='Lost').count()
        open_deals = total - won - lost
        revenue = qs.filter(stage='Won').aggregate(Sum('value'))['value__sum'] or 0
        potential = qs.exclude(stage__in=['Won', 'Lost']).aggregate(Sum('value'))['value__sum'] or 0
        avg = revenue / won if won else 0
        win_rate = round((won / total * 100), 1) if total else 0

        return (
            f"**Pipeline Summary**\n\n"
            f"Total Deals: {total}\n"
            f"Open Deals: {open_deals}\n"
            f"Won: {won} | Lost: {lost}\n"
            f"Win Rate: {win_rate}%\n"
            f"Revenue: ${revenue:,.2f}\n"
            f"Potential: ${potential:,.2f}\n"
            f"Avg Deal Size: ${avg:,.2f}"
        )

    def execute(self, text, user):
        from django.db.models import Sum, Q
        from deals.models import Deal

        text_lower = text.lower()

        # Pipeline summary
        if 'pipeline summary' in text_lower or 'show pipeline' in text_lower:
            return self._pipeline_summary(user)

        # Show high priority deals
        if re.search(r'show\s+(?:all\s+)?high\s+priority\s+deals?', text_lower):
            deals = Deal.objects.filter(owner=user, priority='High').order_by('-value')[:10]
            if not deals:
                return 'No high priority deals found.'
            lines = ['**High Priority Deals**', '']
            for d in deals:
                lines.append(f'\u2022 {d.deal_name} — {d.currency} {d.value:,.2f} (Stage: {d.stage})' if d.value else f'\u2022 {d.deal_name} (Stage: {d.stage})')
            return '\n'.join(lines)

        # Show won/lost/open deals
        if re.search(r'show\s+(all\s+)?won\s+deals?', text_lower):
            deals = Deal.objects.filter(owner=user, stage='Won').order_by('-value')[:10]
            if not deals:
                return 'No won deals found.'
            lines = ['**Won Deals**', '']
            for d in deals:
                lines.append(f'\u2022 {d.deal_name} — {d.currency} {d.value:,.2f}' if d.value else f'\u2022 {d.deal_name}')
            return '\n'.join(lines)

        if re.search(r'show\s+(all\s+)?lost\s+deals?', text_lower):
            deals = Deal.objects.filter(owner=user, stage='Lost').order_by('-value')[:10]
            if not deals:
                return 'No lost deals found.'
            lines = ['**Lost Deals**', '']
            for d in deals:
                lines.append(f'\u2022 {d.deal_name} — {d.currency} {d.value:,.2f}' if d.value else f'\u2022 {d.deal_name}')
            return '\n'.join(lines)

        # Deals worth more than $X
        val_match = re.search(r'worth\s+more\s+than\s+\$?(\d[\d,.]*)', text_lower)
        if val_match:
            threshold = float(val_match.group(1).replace(',', ''))
            deals = Deal.objects.filter(owner=user, value__gte=threshold).order_by('-value')[:10]
            if not deals:
                return f'No deals worth more than ${threshold:,.2f} found.'
            lines = [f'**Deals worth more than ${threshold:,.2f}**', '']
            for d in deals:
                lines.append(f'\u2022 {d.deal_name} — {d.currency} {d.value:,.2f} (Stage: {d.stage})')
            return '\n'.join(lines)

        # Deals closing this week
        if re.search(r'closing\s+this\s+week', text_lower):
            today = date.today()
            end_of_week = today + timedelta(days=(6 - today.weekday()))
            deals = Deal.objects.filter(
                owner=user,
                expected_close_date__gte=today,
                expected_close_date__lte=end_of_week,
            ).exclude(stage__in=['Won', 'Lost']).order_by('expected_close_date')[:10]
            if not deals:
                return 'No deals closing this week.'
            lines = ['**Deals closing this week**', '']
            for d in deals:
                lines.append(
                    f'\u2022 {d.deal_name} — {d.currency} {d.value:,.2f} '
                    f'(closes {d.expected_close_date}, Stage: {d.stage})'
                )
            return '\n'.join(lines)

        # General view — search for specific deal
        query = self._parse(text)
        if not query:
            return 'Please specify a deal name.'
        deal = user.deals.filter(deal_name__icontains=query).first()
        if not deal:
            deal = user.deals.filter(deal_name__iexact=query).first()
        if not deal:
            return f'Deal "{query}" not found.'
        value_str = f'${deal.value:,.2f}' if deal.value else 'Not set'
        return (
            f'**{deal.deal_name}**\n'
            f'Company: {deal.company or "—"}\n'
            f'Stage: {deal.stage} | Probability: {deal.probability}%\n'
            f'Value: {deal.currency} {value_str}\n'
            f'Priority: {deal.priority}\n'
            f'Close Date: {deal.expected_close_date or "—"}\n'
            f'Status: {deal.status}'
        )


def _extract_event_times(text):
    """Extract start and end times from *text*.

    Returns ``(start_time, end_time, cleaned_text)``.
    Handles ``from 3pm to 4pm``, ``3pm - 4pm``, ``at 3pm``.
    """
    from datetime import time as time_class

    body = text
    text_lower = text.lower()

    # "from 3pm to 4pm", "from 3:00 PM until 4:00 PM"
    m = re.search(
        r'\bfrom\s+'
        r'(\d{1,2})(?::(\d{2}))?\s*(am|pm)\s*'
        r'(?:to|until|-)\s*'
        r'(\d{1,2})(?::(\d{2}))?\s*(am|pm)',
        text_lower,
    )
    if m:
        hour1 = int(m.group(1))
        min1 = int(m.group(2)) if m.group(2) else 0
        ampm1 = m.group(3).lower() if m.group(3) else None
        hour2 = int(m.group(4))
        min2 = int(m.group(5)) if m.group(5) else 0
        ampm2 = m.group(6).lower()
        if ampm1 == 'pm' and hour1 < 12:
            hour1 += 12
        elif ampm1 == 'am' and hour1 == 12:
            hour1 = 0
        if ampm2 == 'pm' and hour2 < 12:
            hour2 += 12
        elif ampm2 == 'am' and hour2 == 12:
            hour2 = 0
        try:
            start = time_class(hour1, min1)
            end = time_class(hour2, min2)
            return (start, end, _remove_match(text, m))
        except ValueError:
            pass

    # "3pm - 4:30pm", "3:00 PM - 4:00 PM"
    m = re.search(
        r'(\d{1,2})(?::(\d{2}))?\s*(am|pm)\s*-\s*'
        r'(\d{1,2})(?::(\d{2}))?\s*(am|pm)',
        text_lower,
    )
    if m:
        hour1 = int(m.group(1))
        min1 = int(m.group(2)) if m.group(2) else 0
        ampm1 = m.group(3).lower() if m.group(3) else None
        hour2 = int(m.group(4))
        min2 = int(m.group(5)) if m.group(5) else 0
        ampm2 = m.group(6).lower()
        if ampm1 == 'pm' and hour1 < 12:
            hour1 += 12
        elif ampm1 == 'am' and hour1 == 12:
            hour1 = 0
        if ampm2 == 'pm' and hour2 < 12:
            hour2 += 12
        elif ampm2 == 'am' and hour2 == 12:
            hour2 = 0
        try:
            start = time_class(hour1, min1)
            end = time_class(hour2, min2)
            return (start, end, _remove_match(text, m))
        except ValueError:
            pass

    # Single time (start only)
    t, body = _parse_time_from_text(text)
    return (t, None, body)


@register
class CreateEventAction(BaseAction):
    action_type = 'create_event'
    keywords = frozenset({
        'create', 'add', 'schedule', 'new', 'event', 'meeting',
        'appointment', 'call', 'reminder',
    })
    patterns = [
        re.compile(r'(create|add|schedule|new).*(event|meeting|appointment|call|reminder)'),
        re.compile(r'schedule\s+'),
    ]

    _STATUS_MAP = {
        'scheduled': 'scheduled', 'confirmed': 'scheduled',
        'completed': 'completed', 'done': 'completed',
        'cancelled': 'cancelled', 'canceled': 'cancelled',
        'cancel': 'cancelled',
    }

    _EVENT_TYPE_MAP = {
        'meeting': 'meeting',
        'call': 'call',
        'phone': 'call',
        'phone call': 'call',
        'reminder': 'reminder',
        'personal': 'personal',
    }

    def execute(self, text, user):
        params = self._parse(text)
        title = params.get('title', '').strip()
        if not title:
            return (
                'I need a title to create an event. '
                'What should the event be called?'
            )

        start_date = params.get('start_date') or date.today()

        from calendars.models import Event
        from django.db import transaction
        from django.core.exceptions import ValidationError

        sid = None
        try:
            with transaction.atomic():
                sid = transaction.savepoint()
                event = Event(
                    owner=user,
                    title=title,
                    description=params.get('description', ''),
                    start_date=start_date,
                    start_time=params.get('start_time'),
                    end_time=params.get('end_time'),
                    location=params.get('location', ''),
                    status=params.get('status', 'scheduled'),
                    event_type=params.get('event_type', 'meeting'),
                )
                event.full_clean()
                event.save()
                event.refresh_from_db()

                saved_pk = event.pk
                try:
                    reloaded = Event.objects.get(pk=saved_pk)
                except Event.DoesNotExist:
                    transaction.savepoint_rollback(sid)
                    logger.error(
                        'Event pk=%s was NOT committed -- DB get() returned None. '
                        'user=%s msg=%s',
                        saved_pk, user, text[:120],
                    )
                    return (
                        'Failed to create event: the record was not persisted '
                        'to the database. Please try again.'
                    )

        except ValidationError as e:
            logger.error(
                'Event validation failed for user=%s msg=%s error=%s',
                user, text[:120], e.message_dict,
            )
            error_details = []
            for field, errors in e.message_dict.items():
                for err in errors:
                    error_details.append(f'{field}: {err}')
            return (
                'Failed to create event -- validation error(s):\n'
                + '\n'.join(error_details)
            )
        except Exception as e:
            logger.exception(
                'Failed to create event for user=%s msg=%s', user, text[:120],
            )
            if sid is not None:
                logger.warning('Transaction rolled back at savepoint %s', sid)
            return f'Failed to create event: {e}'

        lines = [
            'Event created successfully (ID: {}): **{}**'.format(
                event.pk, event.title,
            ),
        ]
        if event.start_date:
            label = event.start_date.strftime('%b %d, %Y')
            if event.start_time:
                label += event.start_time.strftime(' at %I:%M %p').lstrip('0')
            lines.append(f'When: {label}')
        if event.end_time:
            lines.append(
                'End: {}'.format(event.end_time.strftime('%I:%M %p').lstrip('0')),
            )
        if event.location:
            lines.append(f'Location: {event.location}')
        if event.description:
            lines.append(f'Description: {event.description[:200]}')
        lines.append(f'Status: {event.status.title()}')
        return '  \n'.join(lines)

    def _parse(self, text):
        """Parse *text* into a params dict."""
        params = {}
        body = text

        # 1. Strip action prefix
        m = re.search(
            r'\b(?:create|add|schedule|new)\s+(?:a\s+|an\s+)?(?:new\s+)?'
            r'(?:event|meeting|appointment|call|reminder)\s*',
            body, flags=re.IGNORECASE,
        )
        if m:
            body = body[m.end():].strip()
        # Fallback: strip bare "schedule" verb without entity keyword
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?schedule\s+(?:a\s+|an\s+)?',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'^(?:for|named|called|about)\s+', '', body, flags=re.IGNORECASE,
        ).strip()

        # 2. Extract structured fields via keyword:value patterns
        _end = r'(?=\s+(?:location|status|type|notes?|description)|$)'
        field_patterns = [
            ('location', r'\blocation\s*[:=]\s*(.+?)' + _end),
            ('status', r'\bstatus\s*[:=]\s*(.+?)' + _end),
            ('event_type', r'\b(?:event\s+)?type\s*[:=]\s*(.+?)' + _end),
            ('description', r'\bdescription\s*[:=]\s*(.+?)$'),
            ('notes', r'\bnotes?\s*[:=]\s*(.+?)$'),
        ]
        for key, pattern in field_patterns:
            if key in params and params[key]:
                continue
            m = re.search(pattern, body, flags=re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if key == 'status':
                    val = self._normalise_event_status(val)
                if key == 'event_type':
                    val = self._normalise_event_type(val)
                if val:
                    params[key] = val
                    body = _remove_match(body, m)

        # 3. Extract event type from natural language
        if 'event_type' not in params:
            text_lower = body.lower()
            if re.search(r'\b(phone|call)\b', text_lower):
                params['event_type'] = 'call'
            elif re.search(r'\breminder\b', text_lower):
                params['event_type'] = 'reminder'
            elif re.search(r'\bpersonal\b', text_lower):
                params['event_type'] = 'personal'
        if 'event_type' in params:
            params['event_type'] = self._normalise_event_type(params['event_type'])

        # 4. Extract date
        start_date, body = _parse_date_from_text(body)
        if start_date:
            params['start_date'] = start_date

        # 5. Extract start / end times
        start_time, end_time, body = _extract_event_times(body)
        if start_time:
            params['start_time'] = start_time
        if end_time:
            params['end_time'] = end_time

        # 6. Strip leading prepositions that may have been exposed after
        #    date/time removal (e.g. "for project review" -> "project review")
        body = re.sub(
            r'^(?:for|with|about|regarding|at)\s+', '',
            body, flags=re.IGNORECASE,
        ).strip()

        # 7. Remaining body -> title
        title = _extract_title(body)
        if title:
            params['title'] = title

        return params

    @staticmethod
    def _normalise_event_status(val):
        val = val.strip().lower()
        for key, mapped in CreateEventAction._STATUS_MAP.items():
            if val == key or val.startswith(key):
                return mapped
        return val

    @staticmethod
    def _normalise_event_type(val):
        val_lower = val.strip().lower()
        return CreateEventAction._EVENT_TYPE_MAP.get(val_lower, val)


@register
class UpdateEventAction(BaseAction):
    """Update an event's fields from natural language."""
    action_type = 'update_event'
    keywords = frozenset({
        'update', 'edit', 'change', 'modify', 'reschedule', 'move', 'rename',
        'event', 'meeting', 'appointment', 'call', 'reminder',
    })
    patterns = [
        re.compile(r'\b(update|edit|change|modify|reschedule|move|rename)\b.*\b(event|meeting|appointment|call|reminder)\b'),
        re.compile(r'(event|meeting|appointment).*(?:to|as)\s+(?:\d|\w+)'),
        re.compile(r'\b(move|reschedule)\s+'),
        re.compile(r'(change|update|set|modify|rename)\s+(?:the\s+)?'
                   r'(?:time|date|location|status|type|description|title|name|link|'
                   r'end\s*time|start\s*time)\b'),
        re.compile(r'(?:update|change|edit|modify|rename)\s+\w+(?:\s+\w+)*\s+'
                   r'(?:to|at)\s+\d{1,2}\s*(?:am|pm)\b'),
    ]

    def _parse(self, text):
        """Parse text → ``{identifier, field, value}``."""
        params = {}
        body = text

        # 1. Strip action verb
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:update|edit|change|modify|reschedule|move|rename)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # 1b. Strip leading prepositions/articles that may appear before entity
        body = _strip_leading_noise(body)

        # 2. Strip entity type (with optional plural)
        body = re.sub(
            r'^(?:events?|meetings?|appointments?|calls?|reminders?)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'\s+from\s+(?:the\s+)?(?:events?|meetings?|appointments?|calls?|reminders?|calendar|schedule)[.,;:]*\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        # Insert ":" between bare field keywords and values ("status done" → "status: done")
        _evt_fk = '|'.join(re.escape(k) for k in _EVENT_UPDATE_FIELDS.keys())
        body = re.sub(
            r'\b(' + _evt_fk + r')\s+(?!(?:to|as|is|=|of)\s)(?=\S)',
            r'\1: ', body, flags=re.IGNORECASE,
        ).strip()

        # 3. Try direct field-value extraction (supports plain "field value")
        keys = sorted(_EVENT_UPDATE_FIELDS.keys(), key=len, reverse=True)
        field_alt = '|'.join(re.escape(k) for k in keys)
        m = re.search(
            r'\b(' + field_alt + r')(?:\s+(?:to|as|is|=)|:\s*)\s*(.+?)$',
            body, flags=re.IGNORECASE,
        )
        if m:
            raw_field = m.group(1).lower()
            params['field'] = _EVENT_UPDATE_FIELDS[raw_field]
            params['value'] = m.group(2).strip()
            raw_ident = body[:m.start()]
            raw_ident = re.sub(r"'s\s*$", '', raw_ident.strip(), flags=re.IGNORECASE)
            raw_ident = re.sub(r'^(?:the|a|an)\b\s*', '', raw_ident, flags=re.IGNORECASE).strip()
            raw_ident = re.sub(
                r'\s+from\s+(?:the\s+)?(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$',
                '', raw_ident, flags=re.IGNORECASE,
            ).strip()
            raw_ident = re.sub(
                r'\s+(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$',
                '', raw_ident, flags=re.IGNORECASE,
            ).strip()
            params['identifier'] = _extract_title(raw_ident)
            return params

        # 4. Try _parse_update_field_value for possessive / "as" formats
        field, value, remaining = _parse_update_field_value(body, _EVENT_UPDATE_FIELDS)
        if field:
            params['field'] = field
            params['value'] = value
            params['identifier'] = re.sub(r"'s\s*$", '', remaining.strip())
            params['identifier'] = re.sub(
                r'\s+from\s+(?:the\s+)?(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$',
                '', params['identifier'], flags=re.IGNORECASE,
            ).strip()
            params['identifier'] = re.sub(
                r'\s+(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$',
                '', params['identifier'], flags=re.IGNORECASE,
            ).strip()
            return params

        # 5. For "move/reschedule <event> to <time/date>":
        #    Try time first
        t, time_body = _parse_time_from_text(body)
        if t:
            params['field'] = 'start_time'
            params['value'] = t
            cleaned = re.sub(r'\b(?:to|at)\b\s*', '', time_body, flags=re.IGNORECASE).strip()
            cleaned = re.sub(r"'s\b", '', cleaned, flags=re.IGNORECASE).strip()
            cleaned = re.sub(r'^(?:events?|meetings?|appointments?|calls?|reminders?)\s+', '', cleaned, flags=re.IGNORECASE).strip()
            cleaned = re.sub(r'\s+from\s+(?:the\s+)?(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$', '', cleaned, flags=re.IGNORECASE).strip()
            cleaned = re.sub(r'\s+(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$', '', cleaned, flags=re.IGNORECASE).strip()
            params['identifier'] = _extract_title(cleaned)
            return params

        #    Then try date — try to extract a second date for "move <event> to <date>"
        d, date_body = _parse_date_from_text(body)
        if d:
            # Try extracting a second date (the target)
            d2, date_body2 = _parse_date_from_text(date_body)
            if d2:
                params['field'] = 'start_date'
                params['value'] = d2
                cleaned = re.sub(r'\b(?:to|on)\b\s*', '', date_body2, flags=re.IGNORECASE).strip()
                cleaned = re.sub(r"'s\b", '', cleaned, flags=re.IGNORECASE).strip()
                cleaned = re.sub(r'^(?:events?|meetings?|appointments?|calls?|reminders?)\s+', '', cleaned, flags=re.IGNORECASE).strip()
                cleaned = re.sub(r'\s+from\s+(?:the\s+)?(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$', '', cleaned, flags=re.IGNORECASE).strip()
                cleaned = re.sub(r'\s+(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$', '', cleaned, flags=re.IGNORECASE).strip()
                params['identifier'] = _extract_title(cleaned)
                return params
            # Single date — use it as the value
            params['field'] = 'start_date'
            params['value'] = d
            cleaned = re.sub(r'\b(?:to|on)\b\s*', '', date_body, flags=re.IGNORECASE).strip()
            cleaned = re.sub(r"'s\b", '', cleaned, flags=re.IGNORECASE).strip()
            cleaned = re.sub(r'^(?:events?|meetings?|appointments?|calls?|reminders?)\s+', '', cleaned, flags=re.IGNORECASE).strip()
            cleaned = re.sub(r'\s+from\s+(?:the\s+)?(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$', '', cleaned, flags=re.IGNORECASE).strip()
            cleaned = re.sub(r'\s+(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$', '', cleaned, flags=re.IGNORECASE).strip()
            params['identifier'] = _extract_title(cleaned)
            return params

        # 6. Check for status word at end
        m = re.search(
            r'\b(as\s+)?(completed|done|finished|cancelled|scheduled)\s*$',
            body, flags=re.IGNORECASE,
        )
        if m:
            params['field'] = 'status'
            params['value'] = m.group(2)
            raw_ident = body[:m.start()]
            raw_ident = re.sub(r"'s\s*$", '', raw_ident.strip(), flags=re.IGNORECASE)
            raw_ident = re.sub(r'^(?:the|a|an)\b\s*', '', raw_ident, flags=re.IGNORECASE).strip()
            raw_ident = re.sub(r'^(?:events?|meetings?|appointments?|calls?|reminders?)\s+', '', raw_ident, flags=re.IGNORECASE).strip()
            raw_ident = re.sub(r'\s+from\s+(?:the\s+)?(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$', '', raw_ident, flags=re.IGNORECASE).strip()
            raw_ident = re.sub(r'\s+(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$', '', raw_ident, flags=re.IGNORECASE).strip()
            params['identifier'] = _extract_title(raw_ident)
            return params

        # 7. Check for "to <time>" pattern (without explicit field)
        m = re.search(
            r'\b(?:to|at)\s+(\d{1,2})(?::(\d{2}))?\s*(am|pm)\b',
            body, flags=re.IGNORECASE,
        )
        if m:
            from datetime import time as time_class
            hour = int(m.group(1))
            minute = int(m.group(2)) if m.group(2) else 0
            ampm = m.group(3).lower()
            if ampm == 'pm' and hour < 12:
                hour += 12
            elif ampm == 'am' and hour == 12:
                hour = 0
            try:
                t = time_class(hour, minute)
                params['field'] = 'start_time'
                params['value'] = t
                raw_ident = body[:m.start()]
                raw_ident = re.sub(r"'s\s*$", '', raw_ident.strip(), flags=re.IGNORECASE)
                raw_ident = re.sub(r'^(?:the|a|an)\b\s*', '', raw_ident, flags=re.IGNORECASE).strip()
                raw_ident = re.sub(r'^(?:events?|meetings?|appointments?|calls?|reminders?)\s+', '', raw_ident, flags=re.IGNORECASE).strip()
                raw_ident = re.sub(r'\s+from\s+(?:the\s+)?(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$', '', raw_ident, flags=re.IGNORECASE).strip()
                raw_ident = re.sub(r'\s+(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$', '', raw_ident, flags=re.IGNORECASE).strip()
                params['identifier'] = _extract_title(raw_ident)
                return params
            except ValueError:
                pass

        # 8. No field-value — body is the identifier
        body_ident = re.sub(r"'s\s*$", '', body.strip())
        body_ident = re.sub(
            r'\s+from\s+(?:the\s+)?(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$',
            '', body_ident, flags=re.IGNORECASE,
        ).strip()
        body_ident = re.sub(
            r'\s+(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$',
            '', body_ident, flags=re.IGNORECASE,
        ).strip()
        params['identifier'] = _extract_title(body_ident)
        return params

    def _find_event(self, identifier, user):
        """Find an event matching *identifier* (title or date)."""
        from calendars.models import Event
        qs = user.events.all()

        # Try title match first
        if identifier:
            event = qs.filter(title__icontains=identifier).first()
            if event:
                return event

        # Try date match (today/this week)
        d, _ = _parse_date_from_text(identifier)
        if d:
            event = qs.filter(start_date=d).first()
            if event:
                return event

        # Try exact title
        if identifier:
            event = qs.filter(title__iexact=identifier).first()
            if event:
                return event

        return None

    def execute(self, text, user):
        params = self._parse(text)
        identifier = params.get('identifier', '').strip()
        field = params.get('field', '')
        value = params.get('value')

        # If we have a field+value but no explicit identifier, try to find
        # the most relevant event (e.g. today's events for status updates)
        if not identifier and field and value is not None:
            from calendars.models import Event
            from django.utils import timezone
            today = timezone.now().date()
            recent = user.events.filter(start_date__gte=today).order_by('start_date').first()
            if recent:
                identifier = recent.title

        if not identifier:
            return (
                'I need to know which event to update. '
                'Please provide the event title or date.'
            )

        if not field or value is None:
            return (
                f'I found event matching "{identifier}", but what would you like to change? '
                f'You can update the title, date, time, end time, location, status, type, '
                f'description, or link.'
            )

        from calendars.models import Event
        from django.db import transaction
        from django.core.exceptions import ValidationError

        event = self._find_event(identifier, user)
        if not event:
            return f'I could not find an event matching "{identifier}".'

        try:
            norm_field, norm_value = _normalise_event_field_value(field, str(value), user)
        except ValueError as e:
            return str(e)

        try:
            with transaction.atomic():
                setattr(event, norm_field, norm_value)
                event.full_clean()
                event.save()
                event.refresh_from_db()
        except ValidationError as e:
            error_details = []
            for fld, errors in e.message_dict.items():
                for err in errors:
                    error_details.append(f'{fld}: {err}')
            return (
                'Failed to update event — validation error(s):\n'
                + '\n'.join(error_details)
            )
        except Exception as e:
            logger.exception('Failed to update event for user=%s', user)
            return f'Failed to update event: {e}'

        readable_field = norm_field.replace('_', ' ').title()
        if norm_field == 'start_date':
            display = event.start_date.strftime('%b %d, %Y')
        elif norm_field == 'start_time':
            display = event.start_time.strftime('%I:%M %p').lstrip('0') if event.start_time else 'None'
        elif norm_field == 'end_time':
            display = event.end_time.strftime('%I:%M %p').lstrip('0') if event.end_time else 'None'
        elif norm_field == 'status':
            display = event.get_status_display()
        elif norm_field == 'event_type':
            display = event.get_event_type_display()
        else:
            display = str(norm_value)

        lines = [
            f'Event **"{event.title}"** updated.',
            f'{readable_field}: {display}',
        ]
        return '  \n'.join(lines)


@register
class DeleteEventAction(BaseAction):
    action_type = 'delete_event'
    keywords = frozenset({'delete', 'remove', 'cancel', 'erase', 'event', 'meeting', 'appointment'})
    patterns = [
        re.compile(r'(delete|remove|cancel|erase)\s+(event|meeting|appointment)'),
        re.compile(r'(delete|remove|cancel|erase).*\b(event|meeting|appointment)\b'),
        re.compile(r'(?:delete|remove|cancel|erase)\s+(?=[a-zA-Z])'),
    ]

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:delete|remove|erase|cancel)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = _strip_leading_noise(body)
        body = re.sub(r'^(?:events?|meetings?|appointments?|calls?)\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'\s+from\s+(?:the\s+)?(?:events?|meetings?|appointments?|calls?|reminders?|calendar|schedule)'
            r'[.,;:]*\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        # Strip entity-type words from end (e.g. "demo meeting" -> "demo")
        body = re.sub(
            r'\s+(?:events?|meetings?|appointments?|calls?|reminders?)[.,;:]*\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        # Strip entity-type word + "with <name>" suffix (e.g. "demo meeting with Kunal" -> "demo")
        body = re.sub(
            r'\s+(?:events?|meetings?|appointments?|calls?|reminders?)\s+with\s+.+$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r"'s\b", '', body, flags=re.IGNORECASE).strip()
        return _extract_title(body)

    def _find_event(self, identifier, user):
        """Find an event matching *identifier* (title or date)."""
        from calendars.models import Event
        qs = user.events.all()

        if identifier:
            event = qs.filter(title__icontains=identifier).first()
            if event:
                return event

        d, _ = _parse_date_from_text(identifier)
        if d:
            event = qs.filter(start_date=d).first()
            if event:
                return event

        if identifier:
            event = qs.filter(title__iexact=identifier).first()
            if event:
                return event

        return None

    def execute(self, text, user):
        identifier = self._parse(text)
        if not identifier:
            return 'I need to know which event to delete. Please provide the event title or date.'

        from calendars.models import Event

        # Try to find the event — first by title, then by date
        event = self._find_event(identifier, user)

        if not event:
            # Try using the original text to extract a date
            d, _ = _parse_date_from_text(text)
            if d:
                matches = list(user.events.filter(start_date=d))
                if len(matches) == 1:
                    event = matches[0]
                elif len(matches) > 1:
                    names = '\n'.join(f'  {i+1}. **{e.title}** ({e.start_date})' for i, e in enumerate(matches))
                    return (
                        f'I found multiple events on that date:\n'
                        f'{names}\n'
                        f'Please specify which one you want to delete.'
                    )
                else:
                    return f'I could not find any events on {d.strftime("%b %d, %Y")}.'

        if not event:
            return f'I could not find an event matching "{identifier}".'

        from django.db import transaction
        try:
            with transaction.atomic():
                event.delete()
        except Exception as e:
            logger.exception('Failed to delete event for user=%s', user)
            return f'Failed to delete event: {e}'

        # Verify deletion
        if Event.objects.filter(pk=event.pk).exists():
            return 'Failed to delete event.'

        return f'Event **"{event.title}"** has been deleted.'


@register
class ViewEventAction(BaseAction):
    action_type = 'view_event'
    keywords = frozenset({
        'event', 'meeting', 'appointment', 'show', 'view', 'open', 'details',
        'tell', 'about', 'agenda', 'schedule', 'calendar',
    })
    patterns = [
        re.compile(r'(show|view|open).*(event|meeting|appointment)'),
        re.compile(r'(event|meeting|appointment).*(details|info|information)'),
        re.compile(r'(tell|show).*(about).*(event|meeting|appointment)'),
        re.compile(r'(?:do\s+(?:i|we)\s+have|is\s+there|are\s+there|what\s+'
                   r'(?:meetings?|events?|appointments?)|when\s+(?:is|are))\b'),
        re.compile(r'(?:agenda|schedule)\b'),
    ]

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:show\s+(?:me\s+)?|view\s+|open\s+|see\s+|display\s+)',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'^(?:tell\s+me\s+about|tell\s+about|tell\s+me)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'^about\s+', '', body, flags=re.IGNORECASE).strip()
        # Strip question patterns
        body = re.sub(
            r'^(?:do\s+(?:i|we)\s+have\s+(?:any\s+)?|is\s+there\s+(?:a|an|any)\s+|'
            r'are\s+there\s+(?:any\s+)?|what\s+(?:meetings?|events?|appointments?)\s+'
            r'(?:do\s+(?:i|we)\s+have|are\s+)?|when\s+(?:is|are)\s+)',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'^(?:event|meeting|appointment)\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'^(?:details?|info|information)(?:\s+(?:for|about|on)\s+|\s+)',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'\s+(?:event|meeting|appointment)\s*$', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'\s+(?:details?|info|information)\s*$', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'^(?:details?|info|information)\s*$', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'^(?:(?:my|the|a|an)\s+)?(?:event|meeting|appointment)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'^(?:in|from|of|for|on)\s+(?:my|the|your)?\s*'
            r'(?:calendar|schedule|event|meeting|appointment)\s*',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r"\s+(?:in|from|of|for|on)\s+(?:my|the|your)?\s*"
            r"(?:calendars?|schedule|events?|meetings?|appointments?)?\s*$",
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'^[\s"\'.,;:!?\-]+|[\s"\'.,;:!?\-]+$', '', body).strip()
        body = re.sub(r"'s\b", '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'\s+', ' ', body).strip()
        return body

    def _find_event(self, identifier, user):
        from calendars.models import Event
        qs = user.events.all()

        if identifier:
            event = qs.filter(title__icontains=identifier).first()
            if event:
                return event

        d, _ = _parse_date_from_text(identifier)
        if d:
            event = qs.filter(start_date=d).first()
            if event:
                return event

        if identifier:
            event = qs.filter(title__iexact=identifier).first()
            if event:
                return event

        return None

    def execute(self, text, user):
        identifier = self._parse(text)
        if not identifier:
            return 'I need to know which event to show. Please provide the event title or date.'

        event = self._find_event(identifier, user)

        if not event:
            d, _ = _parse_date_from_text(text)
            if d:
                event = user.events.filter(start_date=d).first()

        if not event:
            if _is_entity_list_ref(identifier, _ENTITY_LIST_REFS['view_event']):
                all_items = user.events.all().order_by('start_date', 'start_time')
                if not all_items:
                    return 'You have no events yet.'
                lines = ['**Your Events:**']
                for i, e in enumerate(all_items, 1):
                    d = e.start_date.strftime('%b %d') if e.start_date else '—'
                    lines.append(f'{i}. **{e.title}** - {d} ({e.get_status_display()})')
                return '\n'.join(lines)
            filtered = _apply_entity_filter(identifier, user.events.all(), 'view_event')
            if filtered:
                return filtered
            return f'I could not find an event matching "{identifier}".'

        start_date = event.start_date.strftime('%b %d, %Y') if event.start_date else '—'
        start_time = event.start_time.strftime('%I:%M %p').lstrip('0') if event.start_time else '—'
        end_time = event.end_time.strftime('%I:%M %p').lstrip('0') if event.end_time else '—'
        time_str = f'{start_time} – {end_time}' if event.start_time else '—'
        created = event.created_at.strftime('%b %d, %Y') if event.created_at else '—'
        updated = event.updated_at.strftime('%b %d, %Y') if event.updated_at else '—'

        rows = [
            f'| **Title**      | {event.title}',
            f'| **Date**       | {start_date}',
            f'| **Time**       | {time_str}',
            f'| **Location**   | {event.location or "—"}',
            f'| **Type**       | {event.get_event_type_display()}',
            f'| **Status**     | {event.get_status_display()}',
            f'| **Description**| {event.description or "—"}',
            f'| **Created**    | {created}',
            f'| **Updated**    | {updated}',
        ]
        table = '\n'.join(rows)
        return f'### Event Details\n{table}'


@register
class CreateCampaignAction(BaseAction):
    action_type = 'create_campaign'
    keywords = frozenset({'create', 'add', 'new', 'campaign', 'email'})
    patterns = [re.compile(r'(create|add|new).*(campaign)')]

    _STATUS_MAP = {
        'draft': 'Draft',
        'scheduled': 'Scheduled',
        'sent': 'Sent',
        'completed': 'Sent',
        'done': 'Sent',
        'finished': 'Sent',
    }

    def execute(self, text, user):
        params = self._parse(text)
        name = params.get('name', '').strip()
        if not name:
            return (
                'I need a campaign name to create a campaign. '
                'What should the campaign be called?'
            )

        from campaigns.models import Campaign
        from django.db import transaction
        from django.core.exceptions import ValidationError

        sid = None
        try:
            with transaction.atomic():
                sid = transaction.savepoint()
                campaign = Campaign(
                    owner=user,
                    name=name,
                    subject=params.get('subject', name),
                    body=params.get('body', ''),
                    status=params.get('status', 'Draft'),
                    scheduled_at=params.get('scheduled_at'),
                )
                campaign.full_clean()
                campaign.save()
                campaign.refresh_from_db()

                saved_pk = campaign.pk
                try:
                    reloaded = Campaign.objects.get(pk=saved_pk)
                except Campaign.DoesNotExist:
                    transaction.savepoint_rollback(sid)
                    logger.error(
                        'Campaign pk=%s was NOT committed -- DB get() returned None. '
                        'user=%s msg=%s',
                        saved_pk, user, text[:120],
                    )
                    return (
                        'Failed to create campaign: the record was not persisted '
                        'to the database. Please try again.'
                    )

        except ValidationError as e:
            logger.error(
                'Campaign validation failed for user=%s msg=%s error=%s',
                user, text[:120], e.message_dict,
            )
            error_details = []
            for field, errors in e.message_dict.items():
                for err in errors:
                    error_details.append(f'{field}: {err}')
            return (
                'Failed to create campaign -- validation error(s):\n'
                + '\n'.join(error_details)
            )
        except Exception as e:
            logger.exception(
                'Failed to create campaign for user=%s msg=%s',
                user, text[:120],
            )
            if sid is not None:
                logger.warning('Transaction rolled back at savepoint %s', sid)
            return f'Failed to create campaign: {e}'

        lines = [
            'Campaign created successfully (ID: {}): **{}**'.format(
                campaign.pk, campaign.name,
            ),
        ]
        if campaign.subject and campaign.subject != campaign.name:
            lines.append(f'Subject: {campaign.subject}')
        if campaign.scheduled_at:
            lines.append(
                'Scheduled: {}'.format(
                    campaign.scheduled_at.strftime('%b %d, %Y at %I:%M %p'),
                ),
            )
        lines.append(f'Status: {campaign.status}')
        if campaign.body:
            lines.append(f'Body: {campaign.body[:200]}')
        return '  \n'.join(lines)

    def _parse(self, text):
        """Parse *text* into a params dict."""
        params = {}
        body = text

        # 1. Strip action prefix
        m = re.search(
            r'\b(?:create|add|new)\s+(?:a\s+|an\s+)?(?:new\s+)?'
            r'(?:email\s+)?campaign\s*',
            body, flags=re.IGNORECASE,
        )
        if m:
            body = body[m.end():].strip()
        body = re.sub(
            r'^(?:for|named|called|about)\s+', '', body, flags=re.IGNORECASE,
        ).strip()

        # 2. Extract structured fields via keyword:value patterns
        _end = r'(?=\s+(?:name|subject|body|description|status)|$)'
        field_patterns = [
            ('name', r'\bname\s*[:=]\s*(.+?)' + _end),
            ('subject', r'\bsubject\s*[:=]\s*(.+?)' + _end),
            ('body', r'\b(?:body|description|content)\s*[:=]\s*(.+?)' + _end),
            ('status', r'\bstatus\s*[:=]\s*(.+?)' + _end),
        ]
        for key, pattern in field_patterns:
            if key in params and params[key]:
                continue
            m = re.search(pattern, body, flags=re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if key == 'status':
                    val = self._normalise_campaign_status(val)
                if val:
                    params[key] = val
                    body = _remove_match(body, m)

        # 3. Parse scheduled_at from natural language date/time
        dt, body = _parse_datetime_from_text(body)
        if dt:
            params['scheduled_at'] = dt

        # 4. Strip schedule/target noise words left after date removal
        body = re.sub(
            r'\bschedule(?:d)?\s+(?:for|on|at)?\s*',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'^(?:for|about|regarding)\s+', '', body, flags=re.IGNORECASE,
        ).strip()

        # 5. Remaining body -> campaign name (if not already extracted via name: field)
        if 'name' not in params:
            name = _extract_title(body)
            if name:
                params['name'] = name

        return params

    @staticmethod
    def _normalise_campaign_status(val):
        val = val.strip().lower()
        for key, mapped in CreateCampaignAction._STATUS_MAP.items():
            if val == key or val.startswith(key):
                return mapped
        return val.title()


@register
class PauseCampaignAction(BaseAction):
    action_type = 'pause_campaign'
    keywords = frozenset({'pause', 'stop', 'halt', 'campaign'})
    patterns = [re.compile(r'(pause|stop|halt).*(campaign)')]

    def extract_params(self, text):
        return {}


@register
class ResumeCampaignAction(BaseAction):
    action_type = 'resume_campaign'
    keywords = frozenset({'resume', 'unpause', 'continue', 'campaign'})
    patterns = [re.compile(r'(resume|unpause|continue).*(campaign)')]

    def extract_params(self, text):
        return {}


@register
class DeleteCampaignAction(BaseAction):
    action_type = 'delete_campaign'
    keywords = frozenset({'delete', 'remove', 'erase', 'cancel', 'campaign'})
    patterns = [
        re.compile(r'(delete|remove|erase|cancel)\s+campaign\b'),
        re.compile(r'(delete|remove|erase|cancel).*\bcampaign\b'),
        re.compile(r'(?:delete|remove|erase|cancel)\s+(?=[a-zA-Z])'),
    ]

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:delete|remove|erase|cancel)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = _strip_leading_noise(body)
        body = re.sub(r'^campaigns?\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'\s+(?:from\s+)?campaigns?[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
        return _extract_title(body)

    def execute(self, text, user):
        name = self._parse(text)
        if not name:
            return 'I need a campaign name to delete. Which campaign should I delete?'

        from campaigns.models import Campaign

        qs = user.campaigns.all()
        campaign = qs.filter(name__iexact=name).first()
        if not campaign:
            campaign = qs.filter(name__icontains=name).first()
        if not campaign:
            return f'I could not find a campaign matching "{name}".'

        try:
            from django.db import transaction
            with transaction.atomic():
                campaign.delete()
        except Exception as e:
            logger.exception('Failed to delete campaign for user=%s', user)
            return f'Failed to delete campaign: {e}'

        return f'Campaign **"{name}"** has been deleted.'


@register
class ViewCampaignAction(BaseAction):
    action_type = 'view_campaign'
    keywords = frozenset({
        'campaign', 'show', 'view', 'open', 'details',
        'tell', 'about', 'give', 'information',
    })
    patterns = [
        re.compile(r'(tell|show|view).*(about)'),
        re.compile(r'(show|view|open).*(details?|info|information)'),
        re.compile(r'(details?|info|information)\s+(about|on|for|of)'),
        re.compile(r'^(?:show|view|open|see|display)\s+'),
        re.compile(r'(show|view|open).*(campaign)'),
    ]

    def _parse(self, text):
        body = text

        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:'
            r'tell\s+(?:me\s+)?(?:about|everything\s+about)\s+|'
            r'show\s+(?:me\s+)?(?:details?\s+(?:of|for|about|on)\s+|'
            r'(?:everything|full|complete)\s+about\s+|about\s+)|'
            r'who\s+is\s+|'
            r'give\s+(?:me\s+)?(?:information\s+(?:about|on)\s+|'
            r'details?\s+(?:about|on)\s+)?|'
            r'information\s+(?:about|on)\s+|'
            r'details?\s+(?:about|on|for|of)\s+|'
            r'about\s+|'
            r'(?:view|open|see|display)\s+'
            r')',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(
            r'\b(?:about|details?|info|information|everything|full|'
            r'complete|my|the|a|an|campaign|entry|record|profile)\b\s*',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(
            r'^campaign\s+', '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(
            r"\s+(?:in|from|of|for|under)\s+(?:my|the|your)?\s*"
            r"(?:campaigns?|crm|database|system|list)?\s*$",
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(
            r'^(?:show|view|open|see|display)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(
            r'^(?:details?\s+(?:of|for|about|on)\s+|'
            r'info\s+(?:about|on)\s+|'
            r'information\s+(?:about|on)\s+|'
            r'(?:of|for|about|on|in|from)\s+)',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(r'^[\s"\'.,;:!?\-]+|[\s"\'.,;:!?\-]+$', '', body).strip()
        body = re.sub(r'\s+', ' ', body).strip()

        return body

    def execute(self, text, user):
        name = self._parse(text)
        if not name:
            return 'I need a campaign name to show. Which campaign would you like to view?'

        from campaigns.models import Campaign

        qs = user.campaigns.all()
        matches = list(qs.filter(name__icontains=name))
        if len(matches) == 0:
            if _is_entity_list_ref(name, _ENTITY_LIST_REFS['view_campaign']):
                all_items = qs.order_by('-created_at')
                if not all_items:
                    return 'You have no campaigns yet.'
                lines = ['**Your Campaigns:**']
                for i, c in enumerate(all_items, 1):
                    lines.append(f'{i}. **{c.name}** - {c.get_status_display()}')
                return '\n'.join(lines)
            filtered = _apply_entity_filter(name, qs, 'view_campaign')
            if filtered:
                return filtered
            return f'I could not find a campaign matching "{name}".'
        if len(matches) > 1:
            names_list = '\n'.join(f'  {i+1}. **{c.name}**' for i, c in enumerate(matches))
            return (
                f'I found multiple campaigns matching "{name}":\n'
                f'{names_list}\n'
                f'Please specify which one you want to view.'
            )

        campaign = matches[0]
        lines = [f'**{campaign.name}**']
        lines.append(f'Status: {campaign.status}')
        if campaign.subject:
            lines.append(f'Subject: {campaign.subject}')
        if campaign.body:
            lines.append(f'Body: {campaign.body[:500]}')
        if campaign.scheduled_at:
            from django.utils import timezone
            local = timezone.localtime(campaign.scheduled_at)
            lines.append(f'Scheduled: {local.strftime("%B %d, %Y at %I:%M %p")}')
        lines.append(f'Created: {campaign.created_at.strftime("%B %d, %Y")}')
        return '\n'.join(lines)


@register
class UpdateCampaignAction(BaseAction):
    action_type = 'update_campaign'
    keywords = frozenset({
        'update', 'edit', 'change', 'modify', 'set', 'rename', 'mark',
        'campaign', 'name', 'subject', 'status', 'body', 'content',
        'schedule', 'date', 'time',
    })
    patterns = [
        re.compile(r'(update|edit|change|modify|rename|set)\b'),
        re.compile(r'\bmark\b'),
        re.compile(r"(campaign).*(?:'s\s+)?(?:name|subject|status|body|content|schedule)"),
        re.compile(r"(name|subject|status|body|content|schedule)\s+(?:to|as)"),
        re.compile(r"(name|subject|status|body|content|schedule)\s*:"),
    ]

    def _parse(self, text):
        params = {'updates': []}
        body = text

        # Handle "mark (name) as (status)" pattern
        mark_m = re.match(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'mark\s+(.+?)\s+as\s+(.+)$',
            body, flags=re.IGNORECASE,
        )
        if mark_m:
            raw_name = mark_m.group(1).strip()
            raw_name = re.sub(r'\s+campaign$', '', raw_name, flags=re.IGNORECASE).strip()
            raw_name = re.sub(r'^campaign\s+', '', raw_name, flags=re.IGNORECASE).strip()
            if raw_name and raw_name.lower() not in ('campaign', 'the campaign', 'my campaign'):
                params['name'] = raw_name
            params['updates'].append(('status', mark_m.group(2).strip()))
            return params

        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:update|edit|change|modify|rename|set|mark)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = _strip_leading_noise(body)

        body = re.sub(r'^campaigns?\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'\s+from\s+(?:the\s+)?campaigns?[.,;:]*\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()

        keys = sorted(_CAMPAIGN_UPDATE_FIELDS.keys(), key=len, reverse=True)
        field_alt = '|'.join(re.escape(k) for k in keys)

        m = re.search(
            r'\b(' + field_alt + r')\s+of\s+(.+?)\s+(?:to|as)\s+(.+?)$',
            body, flags=re.IGNORECASE,
        )
        if m:
            raw_field = m.group(1).lower()
            params['updates'].append((_CAMPAIGN_UPDATE_FIELDS[raw_field], m.group(3).strip()))
            name = re.sub(r"'s\s*$", '', m.group(2).strip()).strip()
            name = re.sub(r'\s+from\s+(?:the\s+)?campaigns?[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
            name = re.sub(r'\s+campaigns?[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
            params['name'] = _extract_title(name)
            return params

        pair_re = r'\b(' + field_alt + r')(?:\s+(?:to|as|is|=)\s*|:\s*|\s+)(?=\S)'
        matches = list(re.finditer(pair_re, body, flags=re.IGNORECASE))
        if matches:
            name_text = re.sub(r"'s\s*$", '', body[:matches[0].start()].strip()).strip()
            name_text = re.sub(r'\s+from\s+(?:the\s+)?campaigns?[.,;:]*\s*$', '', name_text, flags=re.IGNORECASE).strip()
            name_text = re.sub(r'\s+campaigns?[.,;:]*\s*$', '', name_text, flags=re.IGNORECASE).strip()
            params['name'] = _extract_title(name_text)
            for i, m in enumerate(matches):
                raw_field = m.group(1).lower()
                val_start = m.end()
                val_end = matches[i + 1].start() if i + 1 < len(matches) else len(body)
                value = body[val_start:val_end].strip().rstrip('. ')
                if value:
                    params['updates'].append((_CAMPAIGN_UPDATE_FIELDS[raw_field], value))
            return params

        name = re.sub(r"'s\s*$", '', body).strip()
        name = re.sub(r'\s+from\s+(?:the\s+)?campaigns?[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
        name = re.sub(r'\s+campaigns?[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
        params['name'] = _extract_title(name)
        return params

    def execute(self, text, user):
        params = self._parse(text)
        name = params.get('name', '').strip()
        if not name:
            from campaigns.models import Campaign
            latest = user.campaigns.order_by('-created_at').first()
            if latest:
                name = latest.name
            else:
                return 'I could not find any campaigns to update. Create one first.'

        from campaigns.models import Campaign
        from django.db import transaction

        qs = user.campaigns.all()
        campaign = qs.filter(name__iexact=name).first()
        if not campaign:
            campaign = qs.filter(name__icontains=name).first()
        if not campaign:
            return f'I could not find a campaign matching "{name}".'

        updates = params.get('updates', [])
        if not updates:
            return (
                f'I found campaign **"{campaign.name}"**, but what would you like to change? '
                f'You can update name, subject, body, status, or schedule.'
            )

        applied = []
        for field, value in updates:
            normalized = value.strip()
            if field == 'status':
                mapped = CreateCampaignAction._normalise_campaign_status(normalized)
                if mapped:
                    normalized = mapped
                else:
                    normalized = normalized.title()
                valid = [c[0] for c in Campaign.STATUS_CHOICES]
                if normalized not in valid:
                    return (
                        f'Invalid status "{value}". '
                        f'Valid options are: {", ".join(valid)}.'
                    )
            elif field == 'scheduled_at':
                dt, _ = _parse_datetime_from_text(normalized)
                if dt:
                    setattr(campaign, field, dt)
                    applied.append(f'{field} set to {dt.strftime("%B %d, %Y at %I:%M %p")}')
                    continue
                else:
                    from dateutil import parser as dateparser
                    try:
                        dt = dateparser.parse(normalized)
                        from django.utils import timezone
                        if timezone.is_naive(dt):
                            dt = timezone.make_aware(dt)
                        setattr(campaign, field, dt)
                        applied.append(f'{field} set to {dt.strftime("%B %d, %Y at %I:%M %p")}')
                        continue
                    except Exception:
                        return (
                            f'I could not understand the schedule "{value}". '
                            f'Please provide a date and time like '
                            f'"tomorrow at 3pm" or "2026-01-15 14:00".'
                        )
            setattr(campaign, field, normalized)
            applied.append(f'{field} set to "{normalized}"')

        if not applied:
            return f'No changes were made to campaign **"{campaign.name}"**.'

        try:
            with transaction.atomic():
                campaign.save()
        except Exception as e:
            logger.exception('Failed to update campaign for user=%s', user)
            return f'Failed to update campaign: {e}'

        changed = '\n'.join(f'  \u2022 {a}' for a in applied)
        return f'Campaign **"{campaign.name}"** has been updated:\n{changed}'


@register
class CreateNotificationAction(BaseAction):
    action_type = 'create_notification'
    keywords = frozenset({
        'create', 'add', 'send', 'new', 'notification', 'alert', 'notify',
    })
    patterns = [
        re.compile(r'(create|add|send|new).*(notification|alert|notify)'),
    ]

    def execute(self, text, user):
        params = self._parse(text)
        title = params.get('title', '').strip()
        if not title:
            return (
                'I need a title to create a notification. '
                'What should the notification be about?'
            )

        from workflows.models import Notification
        from django.db import transaction
        from django.core.exceptions import ValidationError

        sid = None
        try:
            with transaction.atomic():
                sid = transaction.savepoint()
                notification = Notification(
                    owner=user,
                    title=title,
                    message=params.get('message', ''),
                    link=params.get('link', ''),
                )
                notification.full_clean()
                notification.save()
                notification.refresh_from_db()

                saved_pk = notification.pk
                try:
                    reloaded = Notification.objects.get(pk=saved_pk)
                except Notification.DoesNotExist:
                    transaction.savepoint_rollback(sid)
                    logger.error(
                        'Notification pk=%s was NOT committed -- DB get() '
                        'returned None. user=%s msg=%s',
                        saved_pk, user, text[:120],
                    )
                    return (
                        'Failed to create notification: the record was not '
                        'persisted to the database. Please try again.'
                    )

        except ValidationError as e:
            logger.error(
                'Notification validation failed for user=%s msg=%s error=%s',
                user, text[:120], e.message_dict,
            )
            error_details = []
            for field, errors in e.message_dict.items():
                for err in errors:
                    error_details.append(f'{field}: {err}')
            return (
                'Failed to create notification -- validation error(s):\n'
                + '\n'.join(error_details)
            )
        except Exception as e:
            logger.exception(
                'Failed to create notification for user=%s msg=%s',
                user, text[:120],
            )
            if sid is not None:
                logger.warning('Transaction rolled back at savepoint %s', sid)
            return f'Failed to create notification: {e}'

        lines = [
            'Notification created successfully (ID: {}): **{}**'.format(
                notification.pk, notification.title,
            ),
        ]
        if notification.message:
            lines.append(f'Message: {notification.message[:200]}')
        if notification.link:
            lines.append(f'Link: {notification.link}')
        return '  \n'.join(lines)

    def _parse(self, text):
        """Parse *text* into a params dict."""
        params = {}
        body = text

        # 1. Strip action prefix
        m = re.search(
            r'\b(?:create|add|send|new)\s+(?:a\s+|an\s+)?(?:new\s+)?'
            r'(?:notification|alert)\s*',
            body, flags=re.IGNORECASE,
        )
        if m:
            body = body[m.end():].strip()
        body = re.sub(
            r'^(?:for|about|regarding)\s+', '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'^(?:titled|called|named)\s+', '', body, flags=re.IGNORECASE,
        ).strip()

        # 2. Extract structured fields via keyword:value patterns
        _end = r'(?=\s+(?:title|message|msg|link|priority)|$)'
        field_patterns = [
            ('title', r'\btitle\s*[:=]\s*(.+?)' + _end),
            ('message', r'\b(?:message|msg)\s*[:=]\s*(.+?)' + _end),
            ('link', r'\blink\s*[:=]\s*(.+?)' + _end),
        ]
        for key, pattern in field_patterns:
            if key in params and params[key]:
                continue
            m = re.search(pattern, body, flags=re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if val:
                    params[key] = val
                    body = _remove_match(body, m)

        # 3. Remaining body -> title (if not already extracted)
        if 'title' not in params:
            t = _extract_title(body)
            if t:
                params['title'] = t

        return params


@register
class ViewNotificationAction(BaseAction):
    action_type = 'view_notification'
    keywords = frozenset({
        'notification', 'alert', 'show', 'view', 'list', 'read',
        'latest', 'recent', 'all', 'unread',
    })
    patterns = [
        re.compile(r'(show|view|list|get|display|read)\s+'
                   r'(?:my\s+)?(?:recent\s+)?(?:notifications?|alerts?)'),
        re.compile(r'(show|view|list|get|display|tell)\s+(?:latest|recent|all|unread)'),
        re.compile(r'^(?:show|view|list|get|display|tell)\s+'),
        re.compile(r'(notifications?|alerts?)\s+(?:list|details?|info)'),
        re.compile(r'(tell|show).*(about).*(notification|alert)'),
    ]

    def _parse(self, text):
        body = text.lower()

        # Detect "latest"/"recent"/"unread" qualifier
        if re.search(r'\b(latest|recent|newest)\b', body):
            return {'qualifier': 'latest'}

        if re.search(r'\b(unread)\b', body):
            return {'qualifier': 'unread'}

        if re.search(r'\b(all)\b', body):
            return {'qualifier': 'all'}

        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:show|view|list|get|display|read)\s+'
            r'(?:my\s+)?(?:recent\s+)?(?:notifications?|alerts?)?\s*',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(
            r'\b(?:my|the|a|an|notification|alert|notifications|alerts|'
            r'recent|latest|all|unread)\b\s*',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(r'^[\s"\'.,;:!?\-]+|[\s"\'.,;:!?\-]+$', '', body).strip()
        body = re.sub(r'\s+', ' ', body).strip()

        if body:
            return {'title': body}
        return {}

    def execute(self, text, user):
        params = self._parse(text)

        from workflows.models import Notification

        qs = user.notifications.all().order_by('-created_at')

        qualifier = params.get('qualifier')
        if qualifier == 'latest':
            notification = qs.first()
            if not notification:
                return 'You have no notifications yet.'
            return self._format_notification(notification)

        if qualifier == 'unread':
            qs = qs.filter(is_read=False)

        title = params.get('title')
        if title:
            if _is_entity_list_ref(title, _ENTITY_LIST_REFS['view_notification']):
                recent = list(qs[:10])
                if not recent:
                    return 'You have no notifications yet.'
                if len(recent) == 1:
                    return self._format_notification(recent[0])
                lines = [f'**Notifications** ({len(recent)} shown):']
                for i, n in enumerate(recent):
                    status = 'Read' if n.is_read else 'Unread'
                    lines.append(f'{i+1}. **{n.title}** - {status}')
                return '\n'.join(lines)
            filtered = _apply_entity_filter(title, qs, 'view_notification')
            if filtered:
                return filtered
            matches = list(qs.filter(title__icontains=title))
            if len(matches) == 0:
                return f'I could not find a notification matching "{title}".'
            if len(matches) > 1:
                names_list = '\n'.join(
                    f'  {i+1}. **{n.title}**' for i, n in enumerate(matches[:10])
                )
                return (
                    f'I found multiple notifications matching "{title}":\n'
                    f'{names_list}\n'
                    f'Please specify which one.'
                )
            return self._format_notification(matches[0])

        # No specific match — show recent list
        recent = list(qs[:10])
        if not recent:
            return 'You have no notifications yet.'

        if len(recent) == 1:
            return self._format_notification(recent[0])

        lines = [f'**Notifications** ({len(recent)} shown):']
        for i, n in enumerate(recent):
            status = '📖' if n.is_read else '🔴'
            lines.append(f'{i+1}. {status} **{n.title}**')
        return '\n'.join(lines)

    @staticmethod
    def _format_notification(n):
        status = 'Read' if n.is_read else 'Unread'
        lines = [f'**{n.title}**']
        lines.append(f'Status: {status}')
        if n.message:
            lines.append(f'Message: {n.message[:500]}')
        if n.link:
            lines.append(f'Link: {n.link}')
        lines.append(f'Created: {n.created_at.strftime("%B %d, %Y at %I:%M %p")}')
        return '\n'.join(lines)


@register
class UpdateNotificationAction(BaseAction):
    action_type = 'update_notification'
    keywords = frozenset({
        'update', 'edit', 'change', 'modify', 'set', 'mark', 'rename',
        'read', 'unread', 'notification', 'alert', 'priority',
        'title', 'message', 'latest', 'recent',
    })
    patterns = [
        re.compile(r'(update|edit|change|modify|set|rename)\b'),
        re.compile(r'mark\s+(?:as\s+)?(read|unread)\b'),
        re.compile(r'\bmark\b'),
        re.compile(r'(notification|alert).*(title|message|priority|read|status)'),
        re.compile(r'(title|message|priority)\s+(?:to|as)'),
        re.compile(r'(title|message|priority)\s*:'),
    ]

    def _parse(self, text):
        params = {'updates': []}
        body = text
        text_lower = text.lower()

        # Handle "mark [as] read/unread [the] [latest] notification"
        # Pattern: "mark [as] read" (bare)
        mark_bare = re.match(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'mark\s+(?:as\s+)?(read|unread)\s*$',
            text, flags=re.IGNORECASE,
        )
        if mark_bare:
            params['updates'].append(('is_read', mark_bare.group(1).lower() == 'read'))
            return params

        # Pattern: "mark [the] latest notification [as] read/unread"
        mark_latest = re.match(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'mark\s+(?:the\s+)?(?:latest|most\s+recent)\s+'
            r'(?:notification|alert)\s+(?:as\s+)?(read|unread)\s*$',
            text, flags=re.IGNORECASE,
        )
        if mark_latest:
            params['updates'].append(('is_read', mark_latest.group(1).lower() == 'read'))
            params['qualifier'] = 'latest'
            return params

        # Pattern: "mark <title> notification [as] read/unread"
        mark_title = re.match(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'mark\s+(.+?)\s+(?:notification|alert)\s+'
            r'(?:as\s+)?(read|unread)\s*$',
            text, flags=re.IGNORECASE,
        )
        if mark_title:
            params['name'] = _extract_title(mark_title.group(1).strip())
            params['updates'].append(('is_read', mark_title.group(2).lower() == 'read'))
            return params

        # Pattern: "mark <name> as read/unread" (without "notification" keyword)
        mark_name_as = re.match(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'mark\s+(.+?)\s+as\s+(read|unread)\s*$',
            text, flags=re.IGNORECASE,
        )
        if mark_name_as:
            params['name'] = _extract_title(mark_name_as.group(1).strip())
            params['updates'].append(('is_read', mark_name_as.group(2).lower() == 'read'))
            return params

        # Strip verb
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:update|edit|change|modify|set|rename)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = _strip_leading_noise(body)

        # Handle "latest" / "most recent"
        if re.search(r'\b(latest|most\s+recent)\b', text_lower) and not re.search(r'\bmark\b', text_lower):
            params['qualifier'] = 'latest'

        body = re.sub(r'^(?:notifications?|alerts?)\s+', '', body, flags=re.IGNORECASE).strip()

        body = re.sub(r'\s+from\s+(?:notifications?|alerts?)[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'\s+(?:notifications?|alerts?)[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()

        keys = sorted(_NOTIFICATION_UPDATE_FIELDS.keys(), key=len, reverse=True)
        field_alt = '|'.join(re.escape(k) for k in keys)

        body = re.sub(
            r'\b(' + field_alt + r')\s+(?!(?:to|as|is|=|of)\s)(?=\S)',
            r'\1: ', body, flags=re.IGNORECASE,
        ).strip()

        m = re.search(
            r'\b(' + field_alt + r')\s+of\s+(.+?)\s+(?:to|as)\s+(.+?)$',
            body, flags=re.IGNORECASE,
        )
        if m:
            raw_field = m.group(1).lower()
            params['updates'].append((_NOTIFICATION_UPDATE_FIELDS[raw_field], m.group(3).strip()))
            name = re.sub(r"'s\s*$", '', m.group(2).strip()).strip()
            name = re.sub(r'\s+from\s+(?:notifications?|alerts?)[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
            name = re.sub(r'\s+(?:notifications?|alerts?)[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
            params['name'] = _extract_title(name)
            return params

        pair_re = r'\b(' + field_alt + r')(?:\s+(?:to|as|is|=)\s*|:\s*|\s+)(?=\S)'
        matches = list(re.finditer(pair_re, body, flags=re.IGNORECASE))
        if matches:
            name_text = re.sub(r"'s\s*$", '', body[:matches[0].start()].strip()).strip()
            name_text = re.sub(r'\s+from\s+(?:notifications?|alerts?)[.,;:]*\s*$', '', name_text, flags=re.IGNORECASE).strip()
            name_text = re.sub(r'\s+(?:notifications?|alerts?)[.,;:]*\s*$', '', name_text, flags=re.IGNORECASE).strip()
            params['name'] = _extract_title(name_text)
            for i, m in enumerate(matches):
                raw_field = m.group(1).lower()
                val_start = m.end()
                val_end = matches[i + 1].start() if i + 1 < len(matches) else len(body)
                value = body[val_start:val_end].strip().rstrip('. ')
                if value:
                    params['updates'].append((_NOTIFICATION_UPDATE_FIELDS[raw_field], value))
            return params

        name = re.sub(r"'s\s*$", '', body).strip()
        name = re.sub(r'\s+from\s+(?:notifications?|alerts?)[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
        name = re.sub(r'\s+(?:notifications?|alerts?)[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
        params['name'] = _extract_title(name)
        return params

    def execute(self, text, user):
        params = self._parse(text)

        from workflows.models import Notification
        from django.db import transaction

        qs = user.notifications.all().order_by('-created_at')

        qualifier = params.get('qualifier')
        name = params.get('name', '').strip()
        updates = params.get('updates', [])

        if qualifier == 'latest':
            notification = qs.first()
            if not notification:
                return 'You have no notifications to update.'
        elif name:
            matches = list(qs.filter(title__icontains=name))
            if len(matches) == 0:
                return f'I could not find a notification matching "{name}".'
            if len(matches) > 1:
                names_list = '\n'.join(
                    f'  {i+1}. **{n.title}**' for i, n in enumerate(matches[:10])
                )
                return (
                    f'I found multiple notifications matching "{name}":\n'
                    f'{names_list}\n'
                    f'Please specify which one.'
                )
            notification = matches[0]
        else:
            if not updates:
                return (
                    'I need to know which notification to update. '
                    'Please specify a title or say "latest".'
                )
            notification = qs.first()
            if not notification:
                return 'You have no notifications to update.'

        if not updates:
            return (
                f'I found notification **"{notification.title}"**, '
                f'but what would you like to change? '
                f'You can mark as read/unread or update title or message.'
            )

        applied = []
        for field, value in updates:
            if field == 'is_read':
                if isinstance(value, bool):
                    normalized = value
                elif isinstance(value, str):
                    normalized = value.lower().strip() in ('read', 'true', 'yes', '1')
                else:
                    normalized = bool(value)
                applied.append(f'marked as {"read" if normalized else "unread"}')
            elif field == 'priority':
                raw = value.strip().lower()
                nmap = {'low': 'low', 'minor': 'low', 'medium': 'medium', 'normal': 'medium', 'high': 'high', 'important': 'high', 'urgent': 'high', 'critical': 'high'}
                normalized = nmap.get(raw, raw)
                applied.append(f'priority set to "{normalized.capitalize()}"')
            else:
                normalized = value.strip()
                applied.append(f'{field} set to "{normalized}"')
            setattr(notification, field, normalized)

        try:
            with transaction.atomic():
                notification.save()
        except Exception as e:
            logger.exception('Failed to update notification for user=%s', user)
            return f'Failed to update notification: {e}'

        changed = '\n'.join(f'  \u2022 {a}' for a in applied)
        return f'Notification **"{notification.title}"** has been updated:\n{changed}'


@register
class DeleteNotificationAction(BaseAction):
    action_type = 'delete_notification'
    keywords = frozenset({
        'delete', 'remove', 'erase', 'cancel', 'notification', 'alert',
        'latest', 'recent',
    })
    patterns = [
        re.compile(r'(delete|remove|erase|cancel)\s+notification\b'),
        re.compile(r'(delete|remove|erase|cancel).*\b(notification|alert)\b'),
        re.compile(r'(?:delete|remove|erase|cancel)\s+(?:latest|recent)\b'),
        re.compile(r'(?:delete|remove|erase|cancel)\s+(?=[a-zA-Z])'),
    ]

    def _parse(self, text):
        body = text
        text_lower = text.lower()

        if re.search(r'\b(latest|most\s+recent)\b', text_lower):
            return {'qualifier': 'latest'}

        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:delete|remove|erase|cancel)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = _strip_leading_noise(body)
        body = re.sub(r'^(?:notifications?|alerts?)\s*', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'\s+from\s+(?:notifications?|alerts?)[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(r'\s+(?:notifications?|alerts?)[.,;:]*\s*$', '', body, flags=re.IGNORECASE).strip()
        name = _extract_title(body)
        if name:
            return {'name': name}
        return {}

    def execute(self, text, user):
        params = self._parse(text)

        from workflows.models import Notification
        from django.db import transaction

        qs = user.notifications.all().order_by('-created_at')

        qualifier = params.get('qualifier')
        name = params.get('name', '').strip()

        if qualifier == 'latest':
            notification = qs.first()
            if not notification:
                return 'You have no notifications to delete.'
        elif name:
            matches = list(qs.filter(title__icontains=name))
            if len(matches) == 0:
                return f'I could not find a notification matching "{name}".'
            if len(matches) > 1:
                names_list = '\n'.join(
                    f'  {i+1}. **{n.title}**' for i, n in enumerate(matches[:10])
                )
                return (
                    f'I found multiple notifications matching "{name}":\n'
                    f'{names_list}\n'
                    f'Please specify which one.'
                )
            notification = matches[0]
        else:
            return 'I need a notification title to delete, or say "latest".'

        try:
            with transaction.atomic():
                notification.delete()
        except Exception as e:
            logger.exception('Failed to delete notification for user=%s', user)
            return f'Failed to delete notification: {e}'

        return f'Notification **"{notification.title}"** has been deleted.'


@register
class CreateWorkflowAction(BaseAction):
    action_type = 'create_workflow'
    keywords = frozenset({
        'create', 'add', 'new', 'workflow', 'automation', 'rule',
    })
    patterns = [
        re.compile(r'(create|add|new).*(workflow|automation|rule)'),
    ]

    _TRIGGER_MAP = {
        'one day before a meeting': 'meeting_reminder',
        'day before a meeting': 'meeting_reminder',
        'before a meeting': 'meeting_reminder',
        'meeting reminder': 'meeting_created',
        'remind me about a meeting': 'meeting_created',
        'remind me before a meeting': 'meeting_created',
        'meeting_reminder': 'meeting_reminder',
        'when a new lead is created': 'lead_created',
        'when a lead is created': 'lead_created',
        'when a new contact is added': 'contact_created',
        'when a contact is added': 'contact_created',
        'when a contact is created': 'contact_created',
        'when a contact is updated': 'contact_updated',
        'contact updated': 'contact_updated',
        'contact update': 'contact_updated',
        'when a task is completed': 'task_completed',
        'when a task is created': 'task_created',
        'when a campaign starts': 'campaign_started',
        'when a campaign is started': 'campaign_started',
        'when a campaign is created': 'campaign_created',
        'when a campaign completes': 'campaign_completed',
        'when a campaign is completed': 'campaign_completed',
        'contact created': 'contact_created',
        'contact creation': 'contact_created',
        'new contact': 'contact_created',
        'lead created': 'lead_created',
        'lead creation': 'lead_created',
        'new lead': 'lead_created',
        'lead updated': 'lead_updated',
        'lead qualified': 'lead_qualified',
        'lead won': 'lead_won',
        'won lead': 'lead_won',
        'won deal': 'lead_won',
        'lead lost': 'lead_lost',
        'lost lead': 'lead_lost',
        'lost deal': 'lead_lost',
        'task created': 'task_created',
        'task creation': 'task_created',
        'new task': 'task_created',
        'task completed': 'task_completed',
        'task completion': 'task_completed',
        'task overdue': 'task_overdue',
        'meeting created': 'meeting_created',
        'meeting creation': 'meeting_created',
        'new meeting': 'meeting_created',
        'meeting finished': 'meeting_finished',
        'campaign started': 'campaign_started',
        'campaign created': 'campaign_created',
        'campaign creation': 'campaign_created',
        'new campaign': 'campaign_created',
        'campaign completed': 'campaign_completed',
        'daily': 'scheduled_daily',
        'weekly': 'scheduled_weekly',
        'monthly': 'scheduled_monthly',
    }

    _ACTION_MAP = {
        'create a task': 'create_task',
        'create task': 'create_task',
        'creates a task': 'create_task',
        'follow up task': 'create_task',
        'followup task': 'create_task',
        'follow-up task': 'create_task',
        'update task': 'update_task',
        'update a task': 'update_task',
        'send an email': 'send_email',
        'send email': 'send_email',
        'sends an email': 'send_email',
        'create a calendar event': 'create_event',
        'create calendar event': 'create_event',
        'create an event': 'create_event',
        'create event': 'create_event',
        'create calendar_event': 'create_calendar_event',
        'create_calendar_event': 'create_calendar_event',
        'assign a lead': 'assign_lead',
        'assign lead': 'assign_lead',
        'change lead status': 'change_lead_status',
        'add a tag': 'add_tag',
        'add tag': 'add_tag',
        'remove a tag': 'remove_tag',
        'remove tag': 'remove_tag',
        'generate ai summary': 'ai_summary',
        'generate summary': 'ai_summary',
        'generate ai email': 'ai_email',
        'generate email': 'ai_email',
        'generate ai followup': 'ai_followup',
        'generate follow up': 'ai_followup',
        'generate followup': 'ai_followup',
        'create a notification': 'create_notification',
        'create notification': 'create_notification',
        'send a notification': 'create_notification',
        'send notification': 'create_notification',
        'notify user': 'create_notification',
        'notify me': 'create_notification',
        'send_notification': 'send_notification',
        'create a contact': 'create_contact',
        'create contact': 'create_contact',
        'creates a contact': 'create_contact',
        'create_contact': 'create_contact',
        'create a lead': 'create_lead',
        'create lead': 'create_lead',
        'creates a lead': 'create_lead',
        'create_lead': 'create_lead',
        'webhook': 'webhook',
    }

    def execute(self, text, user):
        params = self._parse(text)

        name = params.get('name', '').strip()
        if not name:
            return (
                'I need a name to create a workflow. '
                'What should the workflow be called?'
            )

        trigger_type = params.get('trigger_type', '').strip()
        if not trigger_type:
            trigger_type = 'lead_created'

        from workflows.models import Workflow, WorkflowAction
        from django.db import transaction
        from django.core.exceptions import ValidationError

        sid = None
        try:
            with transaction.atomic():
                sid = transaction.savepoint()
                workflow = Workflow(
                    owner=user,
                    name=name,
                    description=params.get('description', ''),
                    trigger_type=trigger_type,
                    is_active=params.get('is_active', True),
                )
                workflow.full_clean()
                workflow.save()
                workflow.refresh_from_db()

                saved_pk = workflow.pk
                try:
                    reloaded = Workflow.objects.get(pk=saved_pk)
                except Workflow.DoesNotExist:
                    transaction.savepoint_rollback(sid)
                    logger.error(
                        'Workflow pk=%s was NOT committed -- DB get() '
                        'returned None. user=%s msg=%s',
                        saved_pk, user, text[:120],
                    )
                    return (
                        'Failed to create workflow: the record was not '
                        'persisted to the database. Please try again.'
                    )

                actions = params.get('actions', [])
                created_actions = []
                for idx, action_type in enumerate(actions):
                    wa = WorkflowAction(
                        workflow=workflow,
                        action_type=action_type,
                        order=idx,
                    )
                    wa.full_clean()
                    wa.save()
                    created_actions.append(wa)

        except ValidationError as e:
            logger.error(
                'Workflow validation failed for user=%s msg=%s error=%s',
                user, text[:120], e.message_dict,
            )
            error_details = []
            for field, errors in e.message_dict.items():
                for err in errors:
                    error_details.append(f'{field}: {err}')
            return (
                'Failed to create workflow -- validation error(s):\n'
                + '\n'.join(error_details)
            )
        except Exception as e:
            logger.exception(
                'Failed to create workflow for user=%s msg=%s',
                user, text[:120],
            )
            if sid is not None:
                logger.warning('Transaction rolled back at savepoint %s', sid)
            return f'Failed to create workflow: {e}'

        lines = [
            'Workflow created successfully (ID: {}): **{}**'.format(
                workflow.pk, workflow.name,
            ),
        ]
        lines.append('Trigger: {}'.format(workflow.get_trigger_type_display()))
        lines.append('Active: {}'.format('Yes' if workflow.is_active else 'No'))
        if workflow.description:
            lines.append('Description: {}'.format(workflow.description[:200]))
        if created_actions:
            acts = ', '.join(
                wa.get_action_type_display() for wa in created_actions
            )
            lines.append('Actions: {}'.format(acts))
        return '  \n'.join(lines)

    def _parse(self, text):
        """Parse *text* into a params dict."""
        params = {}
        body = text

        # 1. Strip action prefix
        m = re.search(
            r'\b(?:create|add|new)\s+(?:a\s+|an\s+)?(?:new\s+)?'
            r'(?:workflow|automation|rule)\s*',
            body, flags=re.IGNORECASE,
        )
        if m:
            body = body[m.end():].strip()
        body = re.sub(
            r'^(?:for|about|regarding)\s+', '', body, flags=re.IGNORECASE,
        ).strip()

        # 2. Extract structured fields
        _end = r'(?=\s+(?:name|trigger|action|actions|description|status|active)|$)'
        field_patterns = [
            ('name', r'\bname\s*[:=]?\s*(.+?)' + _end),
            ('trigger_raw', r'\btrigger\s*[:=]?\s*(.+?)' + _end),
            ('actions_raw', r'\b(?:action|actions)\s*[:=]?\s*(.+?)' + _end),
            ('description', r'\bdescription\s*[:=]?\s*(.+?)' + _end),
            ('status', r'\b(?:status|active)\s*[:=]?\s*(.+?)' + _end),
        ]
        actions_raw = None
        for key, pattern in field_patterns:
            if key in params and params[key]:
                continue
            m = re.search(pattern, body, flags=re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if key == 'trigger_raw':
                    mapped = self._resolve_trigger(val)
                    if mapped:
                        params['trigger_type'] = mapped
                        body = _remove_match(body, m)
                    continue
                if key == 'actions_raw':
                    actions_raw = val
                    body = _remove_match(body, m)
                    continue
                if key == 'status':
                    params['is_active'] = val.lower() in (
                        'active', 'true', 'yes', '1', 'on',
                    )
                    body = _remove_match(body, m)
                    continue
                if val:
                    params[key] = val
                    body = _remove_match(body, m)

        # Resolve action strings to action_type slugs
        if actions_raw:
            parts = [a.strip() for a in re.split(r'[,;]', actions_raw) if a.strip()]
            for part in parts:
                mapped = self._resolve_action(part)
                if mapped:
                    params.setdefault('actions', []).append(mapped)

        # 3. Try to detect trigger from natural language
        if 'trigger_type' not in params:
            body_lower = body.lower()
            for phrase, mapped in sorted(
                self._TRIGGER_MAP.items(),
                key=lambda x: -len(x[0]),
            ):
                if phrase in body_lower:
                    params['trigger_type'] = mapped
                    body = re.sub(
                        r'\b' + re.escape(phrase) + r's?\b', '', body,
                        flags=re.IGNORECASE,
                    ).strip()
                    break
            else:
                for slug in [
                    'contact_created', 'lead_created', 'lead_updated',
                    'lead_qualified', 'lead_won', 'lead_lost',
                    'task_created', 'task_completed', 'task_overdue',
                    'meeting_created', 'meeting_finished',
                    'campaign_started', 'campaign_completed',
                ]:
                    words = slug.split('_')
                    pattern = r'\b' + r'\s.*\b'.join(words) + r'\b'
                    if re.search(pattern, body_lower):
                        params['trigger_type'] = slug
                        for w in words:
                            body = re.sub(
                                r'\b' + re.escape(w) + r'\b', '', body,
                                flags=re.IGNORECASE,
                            ).strip()
                        body = ''
                        break

        # 4. Detect actions from remaining natural language
        if 'actions' not in params:
            body_lower = body.lower()
            for phrase, mapped in sorted(
                self._ACTION_MAP.items(),
                key=lambda x: -len(x[0]),
            ):
                if phrase in body_lower:
                    params.setdefault('actions', []).append(mapped)
                    body = re.sub(
                        r'\b' + re.escape(phrase) + r's?\b', '', body,
                        flags=re.IGNORECASE,
                    ).strip()

        # 5. Remaining body -> name (if not already extracted)
        if 'name' not in params:
            n = _extract_title(body)
            if n:
                params['name'] = n

        # 6. Fallback name from trigger when body is empty
        if 'name' not in params and 'trigger_type' in params:
            params['name'] = params['trigger_type'].replace(
                '_', ' ',
            ).title() + ' Workflow'

        return params

    @staticmethod
    def _resolve_trigger(val):
        val_lower = val.strip().lower()
        for choice_key in [
            'contact_created', 'contact_updated',
            'lead_created', 'lead_updated',
            'lead_qualified', 'lead_won', 'lead_lost',
            'task_created', 'task_completed', 'task_overdue',
            'meeting_created', 'meeting_finished', 'meeting_reminder',
            'campaign_created', 'campaign_started', 'campaign_completed',
            'scheduled_daily', 'scheduled_weekly', 'scheduled_monthly',
        ]:
            if val_lower == choice_key or val_lower == choice_key.replace('_', ' '):
                return choice_key
        for phrase, mapped in CreateWorkflowAction._TRIGGER_MAP.items():
            if phrase in val_lower or val_lower in phrase:
                return mapped
        return None

    @staticmethod
    def _resolve_action(val):
        val_lower = val.strip().lower()
        for choice_key in [
            'create_task', 'update_task', 'send_email', 'create_event',
            'create_calendar_event',
            'assign_lead', 'change_lead_status', 'add_tag', 'remove_tag',
            'ai_summary', 'ai_email', 'ai_followup',
            'create_notification', 'send_notification',
            'create_contact', 'create_lead',
            'webhook',
        ]:
            if val_lower == choice_key or val_lower == choice_key.replace('_', ' '):
                return choice_key
        for phrase, mapped in CreateWorkflowAction._ACTION_MAP.items():
            if phrase in val_lower or val_lower in phrase:
                return mapped
        return None


@register
class ViewWorkflowAction(BaseAction):
    action_type = 'view_workflow'
    keywords = frozenset({
        'workflow', 'automation', 'rule', 'show', 'view', 'open',
        'details', 'tell', 'about', 'give', 'information',
    })
    patterns = [
        re.compile(r'(tell|show|view).*(about)'),
        re.compile(r'(show|view|open).*(details?|info|information)'),
        re.compile(r'(details?|info|information)\s+(about|on|for|of)'),
        re.compile(r'^(?:show|view|open|see|display)\s+'),
        re.compile(r'(show|view|open).*(workflow|automation|rule)'),
    ]

    def _parse(self, text):
        body = text

        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:'
            r'tell\s+(?:me\s+)?(?:about|everything\s+about)\s+|'
            r'show\s+(?:me\s+)?(?:details?\s+(?:of|for|about|on)\s+|'
            r'(?:everything|full|complete)\s+about\s+|about\s+)|'
            r'give\s+(?:me\s+)?(?:information\s+(?:about|on)\s+|'
            r'details?\s+(?:about|on)\s+)?|'
            r'information\s+(?:about|on)\s+|'
            r'details?\s+(?:about|on|for|of)\s+|'
            r'about\s+|'
            r'(?:view|open|see|display)\s+'
            r')',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # Strip bare verb before noise words so "Show workflow" → "workflow"
        body = re.sub(
            r'^(?:show|view|open|see|display)\b\s*',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(
            r'\b(?:about|details?|info|information|everything|full|'
            r'complete|my|the|a|an|workflow|automation|rule|entry|record)\b\s*',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(
            r'^workflow\s+', '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(
            r"\s+(?:in|from|of|for|under)\s+(?:my|the|your)?\s*"
            r"(?:workflows?|crm|database|system|list)?\s*$",
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(
            r'^(?:details?\s+(?:of|for|about|on)\s+|'
            r'info\s+(?:about|on)\s+|'
            r'information\s+(?:about|on)\s+|'
            r'(?:of|for|about|on|in|from)\s+)',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = re.sub(r'^[\s"\'.,;:!?\-]+|[\s"\'.,;:!?\-]+$', '', body).strip()
        body = re.sub(r'\s+', ' ', body).strip()

        return body

    def execute(self, text, user):
        name = self._parse(text)
        if not name:
            return 'I need a workflow name to show. Which workflow would you like to view?'

        from workflows.models import Workflow

        qs = user.workflows.all()
        matches = list(qs.filter(name__icontains=name))
        if len(matches) == 0:
            if _is_entity_list_ref(name, _ENTITY_LIST_REFS['view_workflow']):
                all_items = qs.order_by('-created_at')
                if not all_items:
                    return 'You have no workflows yet.'
                lines = ['**Your Workflows:**']
                for i, w in enumerate(all_items, 1):
                    lines.append(f'{i}. **{w.name}** - {"Active" if w.is_active else "Inactive"}')
                return '\n'.join(lines)
            filtered = _apply_entity_filter(name, qs, 'view_workflow')
            if filtered:
                return filtered
            return f'I could not find a workflow matching "{name}".'
        if len(matches) > 1:
            names_list = '\n'.join(f'  {i+1}. **{w.name}**' for i, w in enumerate(matches))
            return (
                f'I found multiple workflows matching "{name}":\n'
                f'{names_list}\n'
                f'Please specify which one you want to view.'
            )

        workflow = matches[0]
        lines = [f'**{workflow.name}**']
        lines.append(f'Status: {"Active" if workflow.is_active else "Inactive"}')
        lines.append(f'Trigger: {workflow.get_trigger_type_display()}')
        if workflow.description:
            lines.append(f'Description: {workflow.description[:500]}')
        lines.append(f'Created: {workflow.created_at.strftime("%B %d, %Y")}')
        actions = workflow.actions.all()
        if actions:
            acts = ', '.join(a.get_action_type_display() for a in actions)
            lines.append(f'Actions: {acts}')
        return '\n'.join(lines)


@register
class UpdateWorkflowAction(BaseAction):
    action_type = 'update_workflow'
    keywords = frozenset({
        'update', 'edit', 'change', 'modify', 'set', 'rename',
        'workflow', 'automation', 'rule', 'name', 'status',
        'active', 'trigger', 'description', 'enable', 'disable',
    })
    patterns = [
        re.compile(r'(update|edit|change|modify|rename|set)\b'),
        re.compile(r"(workflow|automation|rule).*(?:'s\s+)?"
                   r"(?:name|status|active|trigger|description)"),
        re.compile(r"(name|status|active|trigger|description)\s+(?:to|as)"),
        re.compile(r"(name|status|active|trigger|description)\s*:"),
        re.compile(r'(enable|disable)\b'),
    ]

    def _parse(self, text):
        params = {'updates': []}
        body = text

        text_lower = text.lower()

        # Handle enable/disable before verb strip
        enable_m = re.match(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(enable|disable)\s+(?:the\s+)?(?:workflow|automation|rule)\s+(.+)$',
            text, flags=re.IGNORECASE,
        )
        if enable_m:
            verb = enable_m.group(1).lower()
            params['name'] = _extract_title(enable_m.group(2).strip())
            params['updates'].append(('is_active', verb == 'enable'))
            return params

        # Handle "enable/disable <name> workflow|automation|rule"
        enable_m2 = re.match(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(enable|disable)\s+(.+?)\s+(?:workflow|automation|rule)\s*$',
            text, flags=re.IGNORECASE,
        )
        if enable_m2:
            verb = enable_m2.group(1).lower()
            params['name'] = _extract_title(enable_m2.group(2).strip())
            params['updates'].append(('is_active', verb == 'enable'))
            return params

        # Handle bare enable/disable
        bare_m = re.match(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(enable|disable)\s+(?:the\s+)?(?:workflow|automation|rule)\s*$',
            text, flags=re.IGNORECASE,
        )
        if bare_m:
            return params

        # Strip action verb
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:update|edit|change|modify|rename|set|enable|disable)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()

        body = _strip_leading_noise(body)

        # Handle rename before entity strip: "rename X to Y" → name=X, update name=Y
        rename_m = re.match(
            r'(.+?)\s+to\s+(.+)$', body, flags=re.IGNORECASE,
        )
        if rename_m and re.search(r'\brename\b', text_lower):
            params['name'] = _extract_title(rename_m.group(1).strip())
            params['updates'].append(('name', rename_m.group(2).strip()))
            return params

        body = re.sub(r'^(?:workflows?|automations?|rules?)\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'\s+from\s+(?:the\s+)?(?:workflows?|automations?|rules?)[.,;:]*\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        # Insert ":" between bare field keywords and values ("status active" → "status: active")
        _wf_fk = '|'.join(re.escape(k) for k in _WORKFLOW_UPDATE_FIELDS.keys())
        body = re.sub(
            r'\b(' + _wf_fk + r')\s+(?!(?:to|as|is|=|of)\s)(?=\S)',
            r'\1: ', body, flags=re.IGNORECASE,
        ).strip()

        keys = sorted(_WORKFLOW_UPDATE_FIELDS.keys(), key=len, reverse=True)
        field_alt = '|'.join(re.escape(k) for k in keys)

        m = re.search(
            r'\b(' + field_alt + r')\s+of\s+(.+?)\s+(?:to|as)\s+(.+?)$',
            body, flags=re.IGNORECASE,
        )
        if m:
            raw_field = m.group(1).lower()
            params['updates'].append((_WORKFLOW_UPDATE_FIELDS[raw_field], m.group(3).strip()))
            name = re.sub(r"'s\s*$", '', m.group(2).strip()).strip()
            name = re.sub(r'\s+from\s+(?:the\s+)?(?:workflows?|automations?|rules?)[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
            name = re.sub(r'\s+(?:workflows?|automations?|rules?)[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
            params['name'] = _extract_title(name)
            return params

        pair_re = r'\b(' + field_alt + r')(?:\s+(?:to|as|is|=)\s*|:\s*|\s+)(?=\S)'
        matches = list(re.finditer(pair_re, body, flags=re.IGNORECASE))
        if matches:
            name_text = re.sub(r"'s\s*$", '', body[:matches[0].start()].strip()).strip()
            name_text = re.sub(r'\s+from\s+(?:the\s+)?(?:workflows?|automations?|rules?)[.,;:]*\s*$', '', name_text, flags=re.IGNORECASE).strip()
            name_text = re.sub(r'\s+(?:workflows?|automations?|rules?)[.,;:]*\s*$', '', name_text, flags=re.IGNORECASE).strip()
            params['name'] = _extract_title(name_text)
            for i, m in enumerate(matches):
                raw_field = m.group(1).lower()
                val_start = m.end()
                val_end = matches[i + 1].start() if i + 1 < len(matches) else len(body)
                value = body[val_start:val_end].strip().rstrip('. ')
                if value:
                    params['updates'].append((_WORKFLOW_UPDATE_FIELDS[raw_field], value))
            return params

        name = re.sub(r"'s\s*$", '', body).strip()
        name = re.sub(r'\s+from\s+(?:the\s+)?(?:workflows?|automations?|rules?)[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
        name = re.sub(r'\s+(?:workflows?|automations?|rules?)[.,;:]*\s*$', '', name, flags=re.IGNORECASE).strip()
        params['name'] = _extract_title(name)
        return params

    def execute(self, text, user):
        params = self._parse(text)
        name = params.get('name', '').strip()
        if not name:
            latest = user.workflows.order_by('-created_at').first()
            if latest:
                name = latest.name
            else:
                return 'I could not find any workflows to update. Create one first.'

        from workflows.models import Workflow
        from django.db import transaction

        qs = user.workflows.all()
        workflow = qs.filter(name__icontains=name).first()
        if not workflow:
            workflow = qs.filter(name__iexact=name).first()
        if not workflow:
            return f'I could not find a workflow matching "{name}".'
        updates = params.get('updates', [])
        if not updates:
            return (
                f'I found workflow **"{workflow.name}"**, but what would you like to change? '
                f'You can update name, description, status, or trigger.'
            )

        applied = []
        for field, value in updates:
            normalized = value.strip()
            if field == 'is_active':
                lower_val = normalized.lower()
                normalized = lower_val in ('active', 'true', 'yes', '1', 'on', 'enable', 'enabled')
                applied.append(f'status set to {"Active" if normalized else "Inactive"}')
            elif field == 'trigger_type':
                resolved = CreateWorkflowAction._resolve_trigger(normalized)
                if resolved:
                    normalized = resolved
                    applied.append(f'trigger set to {normalized}')
                else:
                    return (
                        f'I could not understand the trigger "{value}". '
                        f'Try something like "lead_created" or "task_completed".'
                    )
            else:
                applied.append(f'{field} set to "{normalized}"')
            setattr(workflow, field, normalized)

        try:
            with transaction.atomic():
                workflow.save()
        except Exception as e:
            logger.exception('Failed to update workflow for user=%s', user)
            return f'Failed to update workflow: {e}'

        changed = '\n'.join(f'  \u2022 {a}' for a in applied)
        return f'Workflow **"{workflow.name}"** has been updated:\n{changed}'


@register
class DeleteWorkflowAction(BaseAction):
    action_type = 'delete_workflow'
    keywords = frozenset({'delete', 'remove', 'erase', 'cancel', 'workflow', 'automation', 'rule'})
    patterns = [
        re.compile(r'(delete|remove|erase|cancel)\s+workflow\b'),
        re.compile(r'(delete|remove|erase|cancel).*\b(workflow|automation|rule)\b'),
        re.compile(r'(?:delete|remove|erase|cancel)\s+(?=[a-zA-Z])'),
    ]

    def _parse(self, text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:delete|remove|erase|cancel)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = _strip_leading_noise(body)
        body = re.sub(r'^(?:workflows?|automations?|rules?)\s+', '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'\s+from\s+(?:the\s+)?(?:workflows?|automations?|rules?)[.,;:]*\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'\s+(?:workflows?|automations?|rules?)[.,;:]*\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()
        return _extract_title(body)

    def execute(self, text, user):
        name = self._parse(text)
        if not name:
            return 'I need a workflow name to delete. Which workflow should I delete?'

        from workflows.models import Workflow
        from django.db import transaction

        qs = user.workflows.all()
        workflow = qs.filter(name__icontains=name).first()
        if not workflow:
            workflow = qs.filter(name__iexact=name).first()
        if not workflow:
            return f'I could not find a workflow matching "{name}".'

        try:
            with transaction.atomic():
                workflow.delete()
        except Exception as e:
            logger.exception('Failed to delete workflow for user=%s', user)
            return f'Failed to delete workflow: {e}'

        return f'Workflow **"{name}"** has been deleted.'


# ═══════════════════════════════════════════════════════════════════════════
#  modify_anything — catch-all update / delete / cancel
# ═══════════════════════════════════════════════════════════════════════════

@register
class ModifyAnythingAction(BaseAction):
    """Catch-all for update/delete/cancel patterns that specific actions miss
    (e.g. plural entity names like "contacts" / "tasks", indirect references
    like "change company of John").
    """

    action_type = 'modify_anything'
    keywords = frozenset({
        'update', 'edit', 'change', 'modify', 'set', 'rename',
        'delete', 'remove', 'erase', 'cancel',
    })
    patterns = [
        # Any text containing an update verb
        re.compile(r'\b(?:update|edit|change|modify|set|rename|delete|remove|erase|cancel)\b'),
        # "I want to ..."
        re.compile(r'\bi\s+(?:want\s+)?(?:to\s+)?(?:update|edit|change|modify|set|rename|delete|remove|erase|cancel)\b'),
        # Possessive "<name>'s <attr> to <value>"
        re.compile(r"'s\s+(?:phone|email|company|status|priority|name|title|number)\s+(?:to|as)\b"),
    ]

    _ENTITY_MAP = {
        'contact': ('contacts', 'full_name'),
        'person': ('contacts', 'full_name'),
        'lead': ('leads', 'lead_name'),
        'prospect': ('leads', 'lead_name'),
        'deal': ('leads', 'lead_name'),
        'task': ('tasks', 'title'),
        'event': ('events', 'title'),
        'meeting': ('events', 'title'),
        'appointment': ('events', 'title'),
        'reminder': ('events', 'title'),
        'campaign': ('campaigns', 'name'),
        'workflow': ('workflows', 'name'),
        'notification': ('notifications', 'title'),
        'alert': ('notifications', 'title'),
    }

    _QUERYSET_BUILDERS = {
        'contacts': lambda u: u.contacts.all(),
        'leads': lambda u: u.leads.all(),
        'tasks': lambda u: u.tasks.all(),
        'events': lambda u: u.events.all(),
        'campaigns': lambda u: u.campaigns.all(),
        'workflows': lambda u: u.workflows.all() if hasattr(u, 'workflows') else None,
        'notifications': lambda u: u.notifications.all() if hasattr(u, 'notifications') else None,
    }

    _NAME_FIELDS = {
        'contacts': 'full_name',
        'leads': 'lead_name',
        'tasks': 'title',
        'events': 'title',
        'campaigns': 'name',
        'workflows': 'name',
        'notifications': 'title',
    }


    # ── helpers ──

    @staticmethod
    def _strip_prefix(text):
        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:update|edit|change|modify|set|rename|'
            r'delete|remove|erase|cancel)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'^(?:the|my|this|that|a|an)\s+', '', body, flags=re.IGNORECASE).strip()
        return body

    @staticmethod
    def _get_queryset(user, rel):
        builder = ModifyAnythingAction._QUERYSET_BUILDERS.get(rel)
        if builder:
            return builder(user)
        return None

    @staticmethod
    def _get_name_field(rel):
        return ModifyAnythingAction._NAME_FIELDS.get(rel, 'title')

    @staticmethod
    def _is_delete(text_lower):
        return bool(re.search(r'\b(?:delete|remove|erase|cancel)\b', text_lower))

    @staticmethod
    def _is_update(text_lower):
        return bool(re.search(r'\b(?:update|edit|change|modify|set|rename)\b', text_lower))

    @staticmethod
    def _is_question(text_lower):
        """Return True if *text_lower* looks like a question, not a command."""
        return bool(re.search(
            r'\b(?:how\s+(?:do|can|should|would|will|to|does)'
            r'|what\s+(?:is|are|does|do)'
            r'|can\s+you'
            r'|tell\s+(?:me|us)\s+(?:how|what|about))',
            text_lower,
        ))

    @staticmethod
    def _normalise_value(field, value, entity_type):
        """Map raw value strings to model-choice-friendly values."""
        val_lower = value.strip().lower()

        # Status normalisation
        status_map = {
            'contacts': {
                'active': 'active', 'inactive': 'inactive',
            },
            'leads': {
                'new': 'New', 'contacted': 'Contacted',
                'qualified': 'Qualified', 'proposal': 'Proposal Sent',
                'proposal sent': 'Proposal Sent',
                'negotiation': 'Negotiation',
                'won': 'Won', 'lost': 'Lost',
            },
            'tasks': {
                'pending': 'pending', 'in progress': 'in_progress',
                'in progress': 'in_progress', 'completed': 'completed',
                'done': 'completed', 'finished': 'completed',
                'cancelled': 'cancelled',
            },
            'events': {
                'scheduled': 'Scheduled', 'completed': 'Completed',
                'cancelled': 'Cancelled', 'rescheduled': 'Rescheduled',
            },
        }
        if field == 'status':
            mapping = status_map.get(entity_type, {})
            if val_lower in mapping:
                return mapping[val_lower]
            return value

        # Priority normalisation
        priority_map = {
            'high': 'High', 'medium': 'Medium', 'low': 'Low',
        }
        if field == 'priority' and val_lower in priority_map:
            return priority_map[val_lower]

        return value

    @staticmethod
    def _extract_field_value_for_entity(body, entity_type):
        """Try to extract ``(field, value)`` from *body* for *entity_type*."""
        # Phone
        m = re.search(r'(\+?\d[\d\s\-().]{6,20}\d)\b', body)
        if m:
            return ('phone', m.group(1).strip())
        # Email
        m = re.search(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})\b', body)
        if m:
            return ('email', m.group(1).lower())

        # Status / priority keywords
        for pat, field, vals in [
            (r'\b(?:qualified|won|lost|new|contacted)\b', 'status',
             ['new', 'contacted', 'qualified', 'won', 'lost']),
            (r'\b(?:high|medium|low)\b', 'priority',
             ['high', 'medium', 'low']),
            (r'\b(?:completed|done|finished|pending|cancelled|active)\b', 'status',
             ['completed', 'done', 'finished', 'pending', 'cancelled', 'active']),
        ]:
            m = re.search(pat, body, re.IGNORECASE)
            if m:
                val = m.group(0).lower()
                norm = {
                    'done': 'completed', 'finished': 'completed',
                }
                return (field, norm.get(val, val))

        field_map = {
            'contacts': _CONTACT_UPDATE_FIELDS,
            'leads': _LEAD_UPDATE_FIELDS,
            'tasks': _TASK_UPDATE_FIELDS,
            'events': _EVENT_UPDATE_FIELDS,
        }.get(entity_type)

        if field_map:
            keys = sorted(field_map.keys(), key=len, reverse=True)
            field_alt = '|'.join(re.escape(k) for k in keys)

            # Pattern A: "<attr> to <value>" (simple, e.g. "company to MegaCorp")
            m = re.search(
                r'\b(' + field_alt + r')\s+(?:to|as)\s+(.+)', body, re.IGNORECASE,
            )
            if m:
                raw_field = m.group(1).lower()
                value = m.group(2).strip()
                norm_field = field_map.get(raw_field)
                if norm_field:
                    return (norm_field, value)

            # Pattern B: "<attr> of <entity> to <value>" (e.g. "company of John to Acme")
            m = re.search(
                r'\b(' + field_alt + r')\s+(?:of|for)\s+'
                r'.*?\s+(?:to|as)\s+(.+)',
                body, re.IGNORECASE,
            )
            if m:
                raw_field = m.group(1).lower()
                value = m.group(2).strip()
                norm_field = field_map.get(raw_field)
                if norm_field:
                    return (norm_field, value)

        return (None, None)

    # ── Name extraction pipeline ──

    def _clean_name_body(self, text):
        """Strip prefix, attribute phrases, entity words, and trailing
        prepositions from *text* to extract a candidate entity name."""
        body = self._strip_prefix(text)

        # 1) Strip trailing "to <value>" / "as <value>"
        body = re.sub(
            r'\s+(?:to|as)\s+\S+(?:\s+\S+)*\s*$', '', body,
            flags=re.IGNORECASE,
        ).strip()

        # 2) Strip attribute-field phrase (e.g. "company of John" → "John")
        body = re.sub(
            r'\b(?:name|title|status|priority|phone|email|company|'
            r'date|time|location|description|number)\s+(?:of|for)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # 3) Strip possessive "'s" and trailing noise words
        body = re.sub(r"'s\b", '', body, flags=re.IGNORECASE).strip()
        body = re.sub(
            r'\b(?:the|my|this|that|a|an|phone|email|company|status|'
            r'priority|name|title|number)\s*$',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # 4) Strip entity-type words
        body = re.sub(
            r'\b(?:contacts?|leads?|tasks?|events?|meetings?|'
            r'campaigns?|appointments?|reminders?|persons?|'
            r'workflows?|notifications?|alerts?)\b',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # 5) Strip trailing prepositions + remainder
        body = re.sub(
            r'\b(?:from|in|of|to|at|for)\b.*$', '', body,
            flags=re.IGNORECASE,
        ).strip()

        return body

    # ── execute ──

    def execute(self, text, user):
        if not user or not user.is_authenticated:
            return None

        text_lower = text.lower().strip()

        # Don't interfere with questions
        if self._is_question(text_lower):
            return None

        is_delete = self._is_delete(text_lower)
        is_update = self._is_update(text_lower)
        if not is_delete and not is_update:
            return None

        from django.db.models import Q

        # ── Extract candidate name ──
        name_body = self._clean_name_body(text)
        if not name_body:
            return 'I need to know which item to modify.'

        # ── Search every entity type ──
        seen = set()
        candidates = []
        for rel_key, name_fld in self._NAME_FIELDS.items():
            if rel_key in seen:
                continue
            seen.add(rel_key)
            qs = self._get_queryset(user, rel_key)
            if qs is None:
                continue
            for m in qs.filter(Q(**{f'{name_fld}__icontains': name_body}))[:3]:
                candidates.append((rel_key, m))

        if len(candidates) == 0:
            return f'I could not find anything matching **"{name_body}"**.'
        if len(candidates) > 1:
            lines = [f'I found multiple matches for **"{name_body}"**:']
            for rel_key, m in candidates:
                nf = self._get_name_field(rel_key)
                lines.append(
                    f'- **{getattr(m, nf)}** ({rel_key.rstrip("s")})'
                )
            lines.append('Please specify which one you want to modify.')
            return '\n'.join(lines)

        rel, obj = candidates[0]
        name_field = self._get_name_field(rel)
        obj_name = getattr(obj, name_field)
        label = rel.rstrip('s')

        # ── DELETE ──
        if is_delete:
            from django.db import transaction
            try:
                with transaction.atomic():
                    obj.delete()
            except Exception as e:
                logger.exception('Failed to delete %s for user=%s', label, user)
                return f'Failed to delete {label}: {e}'
            return f'{label.title()} **"{obj_name}"** has been deleted.'

        # ── UPDATE ──
        field, value = self._extract_field_value_for_entity(text, rel)
        if field and value is not None:
            from django.db import transaction
            from django.core.exceptions import ValidationError
            value = self._normalise_value(field, value, rel)
            try:
                with transaction.atomic():
                    setattr(obj, field, value)
                    obj.full_clean()
                    obj.save()
            except ValidationError as e:
                errors = []
                for fld, errs in e.message_dict.items():
                    for err in errs:
                        errors.append(f'{fld}: {err}')
                return 'Validation error(s):\n' + '\n'.join(errors)
            except Exception as e:
                logger.exception('Failed to update %s for user=%s', label, user)
                return f'Failed to update {label}: {e}'
            return (
                f'{label.title()} **"{obj_name}"** updated.\n'
                f'{field.replace("_", " ").title()}: {value}'
            )

        return (
            f'I found {label} **"{obj_name}"**. '
            f'What would you like to change?'
        )


# ═══════════════════════════════════════════════════════════════════════════
#  compose_email — AI Email Composer
# ═══════════════════════════════════════════════════════════════════════════

@register
class ComposeEmailAction(BaseAction):
    """Detect email-writing intents, extract recipient/purpose, look up
    Contact model, and use the LLM to generate a professional email draft."""

    action_type = 'compose_email'
    keywords = frozenset({
        'write', 'draft', 'compose', 'create', 'send',
        'email', 'mail', 'follow', 'up',
    })
    patterns = [
        re.compile(r'(?:write|draft|compose|create|send)\s+(?:an?\s+)?(?:email|mail)\b'),
        re.compile(r'(?:write|draft|compose|create|send)\s+(?:this|that|the)\s+(?:email|mail)\b'),
        re.compile(r'(?:write|draft|compose|create|send)\s+(?:a\s+)?\w+\s+(?:email|mail)\b'),
        re.compile(r'(?:write|draft|compose|create|send)\s+(?:a\s+)?follow[-\s]up'),
        re.compile(r'email\s+\w+'),
    ]

    _EMAIL_RE = re.compile(r'[\w.+-]+@[\w-]+(?:\.[\w-]+)+')
    _BULK_RE = re.compile(
        r'(?:everyone|everybody|all\s+(?:contacts?|leads?|tasks?|events?|'
        r'campaigns?|workflows?|notifications?|users?|people|persons?|of\s+them|of\s+you))',
        re.IGNORECASE,
    )

    def _parse(self, text):
        params = {'recipient': '', 'subject': '', 'purpose': '', 'email_address': '', 'bulk_all': False}
        body = text

        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:write|draft|compose|create|send)\s+'
            r'(?:an?\s+)?(?:\w+(?:\s+\w+)*\s+)?(?:follow[-\s]up\s+)?(?:email|mail)\s+(?:to|for)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()

        # If still matched the verb-prefix above, try "email <name>"
        if body == text and re.match(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?email\s+', body, re.IGNORECASE,
        ):
            body = re.sub(
                r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?email\s+',
                '', body, flags=re.IGNORECASE,
            ).strip()

        # Extract purpose after "about / regarding / concerning"
        m = re.search(
            r'\b(?:about|regarding|concerning)\s+(.+?)$',
            body, flags=re.IGNORECASE,
        )
        if m:
            params['purpose'] = m.group(1).strip()
            body = body[:m.start()].strip()
        else:
            # Extract purpose after action verbs (thanking, to discuss, etc.)
            m = re.search(
                r'\b(?:thanking|to\s+(?:thank|discuss|follow\s+up|schedule|confirm|update|share|provide|review|go\s+over))\s+(.+?)$',
                body, flags=re.IGNORECASE,
            )
            if m:
                params['purpose'] = m.group(0).strip()
                body = body[:m.start()].strip()

        # Detect bulk / "everyone" recipient
        bulk_m = self._BULK_RE.search(body)
        if bulk_m:
            params['bulk_all'] = True
            # Remove the bulk phrase from body so remaining text can be purpose
            body = self._BULK_RE.sub('', body).strip()
            # If nothing left, recipient is blank (will be inferred in execute)
            if not body:
                return params
            # Otherwise continue parsing remaining text as purpose
            body = re.sub(r'^', '', body).strip()
            m2 = re.search(
                r'\b(?:about|regarding|concerning)\s+(.+?)$',
                body, flags=re.IGNORECASE,
            )
            if m2:
                params['purpose'] = m2.group(1).strip()
            elif 'purpose' not in params or not params['purpose']:
                params['purpose'] = _extract_title(body)
            return params

        # Detect inline email address
        m = self._EMAIL_RE.search(body)
        if m:
            params['email_address'] = m.group()
            pre = body[:m.start()].strip().rstrip(' <(')
            if pre:
                pre = re.sub(
                    r'\s+(?:at|@)\s*$', '', pre, flags=re.IGNORECASE,
                ).strip()
                params['recipient'] = _extract_title(pre)
        else:
            params['recipient'] = _extract_title(body)

        # Strip trailing context that isn't part of a person's name
        # e.g. "Rittik Pati on kolkata on 15th august" → "Rittik Pati"
        if params['recipient'] and not params.get('email_address'):
            stripped = re.sub(
                r'\s+(?:on|at|in|about|regarding|concerning|for|from)\s+.+$',
                '', params['recipient'], flags=re.IGNORECASE,
            ).strip()
            if stripped:
                params['recipient'] = stripped

        return params

    def _build_prompt(self, recipient_name, email_addr, company, position, purpose,
                      sender_name, sender_email, sender_company, crm_context='',
                      crm_source=''):
        lines = [
            'You are a professional email writer for a CRM platform.',
            'Generate ONLY the subject and email body.',
            'Do NOT include To, From, or signature -- those are added by the system.',
            '',
            '--- RECIPIENT ---',
            f'Name: {recipient_name or "Unknown"}',
        ]
        if email_addr:
            lines.append(f'Email: {email_addr}')
        if company:
            lines.append(f'Company: {company}')
        if position:
            lines.append(f'Position: {position}')
        if crm_source:
            lines.append(f'CRM Source: {crm_source}')
        if purpose:
            lines.append(f'Purpose / Topic: {purpose}')

        lines += [
            '',
            '--- SENDER (YOU) ---',
            f'Name: {sender_name}',
            f'Email: {sender_email}',
        ]
        if sender_company:
            lines.append(f'Company: {sender_company}')

        if crm_context:
            lines += [
                '',
                '--- CRM CONTEXT (use these actual values in the email) ---',
                crm_context,
            ]

        lines += [
            '',
            '--- STRICT RULES (VIOLATION = REJECTED) ---',
            '',
            '1. PLAIN TEXT ONLY. Never use markdown, **bold**, bullets, tables, or code blocks.',
            '',
            '2. SALUTATION: "Dear <First Name>," or "Hello <First Name>,".',
            '   Never "Hi there", "Hey", "Greetings", "To whom it may concern".',
            '   Use the recipient first name from the RECIPIENT section above.',
            '',
            '3. BODY: Introduction. Purpose. Relevant information. Professional closing.',
            '   Use natural language -- no AI-sounding or robotic wording.',
            '   Reference any CRM data (event title, date, time, meeting link, task, campaign) naturally.',
            '',
            '4. SIGNATURE: Do NOT include any signature in the body.',
            '   The system appends "Best regards,\n\nName\nEmail" automatically.',
            '',
            '5. ZERO PLACEHOLDERS. The following are FORBIDDEN:',
            '   [Your Name] [Your Title] [Your Company] [Company Name] [Your Email]',
            '   [Phone] [Phone number] [Your PhoneNumber] [Website] [LinkedIn] [Calendly]',
            '   [Date] [Time] [Meeting Link] [Location] [Topic 1] [Insert ...]',
            '   [Please confirm ...] [Office address ...] [TBD] [Proposal] [Key Topic]',
            '   [Attendee Name] [Product Name] [Calendar Link] [Any Placeholder]',
            '',
            '6. USE ACTUAL DATA from RECIPIENT, SENDER, and CRM CONTEXT sections above.',
            '   - Recipient name is "{0}". Use "Dear {0}," as salutation.'.format(
                (recipient_name or '').split()[0] if recipient_name else 'there'),
            '   - If a meeting title is in CRM CONTEXT, reference it in the body.',
            '   - If a meeting date/time exists, use it naturally (do NOT write "[Date]").',
            '   - If a meeting link exists, include it. Otherwise omit the section entirely.',
            '   - If a task describes the topic, reference it naturally in the body.',
            '',
            '7. WHEN DATA IS UNAVAILABLE, REWRITE NATURALLY:',
            '   BAD:  "Meeting Link: [Meeting Link]"',
            '   GOOD: (omit the section entirely)',
            '   BAD:  "Time: [Time]"',
            '   GOOD: "at the scheduled time"',
            '   BAD:  "Product: [Product Name]"',
            '   GOOD: "our CRM platform"',
            '   BAD:  "Date: [Date]"',
            '   GOOD: "tomorrow" or "next week"',
            '   BAD:  "Dear [Attendee Name]"',
            '   GOOD: "Dear Rahul,"',
            '   BAD:  "[Company Name]"',
            '   GOOD: (leave blank if not available)',
            '',
            '8. Subject: professional, short, specific (8-12 words).',
            '   Examples: "Tomorrow\'s Meeting Confirmation",',
            '   "Follow-up After Yesterday\'s Discussion",',
            '   "Proposal for CRM Automation", "Thank You for Your Time".',
            '',
            '9. Write a detailed, professional email (8-15 sentences minimum).',
            '   Include a proper greeting, context, key details,',
            '   and a clear call to action. Make it feel personal and thoughtful.',
            '   The email must be immediately sendable with zero editing.',
            '',
            'Respond ONLY with valid JSON (no markdown, no code fences):',
            '{',
            '  "subject": "<subject line>",',
            '  "body": "<email body -- salutation + message + closing line only>"',
            '}',
        ]
        return '\n'.join(lines)

    @staticmethod
    def _strip_reasoning(text):
        """Strip internal reasoning / chain-of-thought from LLM output."""
        if not text:
            return text
        # Remove blocks wrapped in  think...  tags (standard CoT)
        text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL)
        # Remove lines that are pure reasoning headers (case-insensitive)
        reasoning_headers = [
            r'think\s*:',
            r'thought\s*:',
            r'reasoning\s*:',
            r'analysis\s*:',
            r'planning\s*:',
            r'internal\s+notes\s*:',
        ]
        pattern = '|'.join(f'(?:{h})' for h in reasoning_headers)
        text = re.sub(
            rf'^(?:{pattern}).*$',
            '', text, flags=re.MULTILINE | re.IGNORECASE,
        ).strip()
        # Collapse repeated blank lines left behind
        text = re.sub(r'\n{3,}', '\n\n', text)
        return text.strip()

    @staticmethod
    def _post_process_body(body, sender_name, sender_email):
        if not body:
            return body
        # 1 — Replace all known placeholders with actual sender info
        body = body.replace('[Your Name]', sender_name)
        body = body.replace('[your name]', sender_name)
        body = body.replace('[YOUR NAME]', sender_name)
        body = body.replace('[Your Email]', sender_email or '')
        body = body.replace('[your email]', sender_email or '')
        body = body.replace('[YOUR EMAIL]', sender_email or '')
        body = body.replace('[email address]', sender_email or '')
        body = body.replace('[Email Address]', sender_email or '')
        if sender_name:
            body = body.replace('Your Name', sender_name)
        if sender_email:
            body = body.replace('Your Email', sender_email)

        # 2 — Strip any trailing LLM-generated signature (closing phrase +
        #     optional name/email lines) so the canonical one always wins.
        body = re.sub(
            r'(?:\n\s*)?'
            r'(?:Best\s+regards|Kind\s+regards|Regards|Sincerely|'
            r'Thanks|Thank\s+you|Warmly|Best|Cheers|'
            r'Yours\s+truly|Yours\s+sincerely|Respectfully)'
            r'[,!.]?[ \t]*'
            r'(?:\n.*)?$',
            '', body, flags=re.IGNORECASE | re.DOTALL,
        ).strip()

        # 3 — Append canonical sender signature
        sig = f'Best regards,\n\n{sender_name}'
        if sender_email:
            sig += f'\n{sender_email}'
        if body:
            return body + '\n\n' + sig
        return sig

    @staticmethod
    def _clean_placeholders(text, sender_company=None):
        """Strip every remaining ``[...]`` placeholder, orphaned label lines,
        placeholder sentences, and divider lines from the LLM-generated text.

        Never removes real content — only purges tokens the model
        hallucinated because it lacked the actual value.
        """
        if not text:
            return text

        # 1 — Replace known sender placeholders with real values (if
        #     available) BEFORE line-level removal so the replacement
        #     takes effect and the line is not discarded entirely.
        text = re.sub(
            r'\[(?:'
            r'Your\s+Title[^\]]*|your\s+title[^\]]*'
            r'|YourTitle[^\]]*|yourtitle[^\]]*'
            r'|Your\s+Phone\s*[Nn]umber[^\]]*'
            r'|your\s+phone\s*[Nn]umber[^\]]*'
            r'|Position[^\]]*|position[^\]]*'
            r'|Company\b(?!\s*[Nn]ame)[^\]]*'
            r'|company\b(?!\s*[Nn]ame)[^\]]*'
            r'|Phone\s+[Nn]umber[^\]]*|phone\s+[Nn]umber[^\]]*'
            r'|Phone\b[^\]]*|phone\b[^\]]*'
            r'|Website\b[^\]]*|website\b[^\]]*'
            r'|LinkedIn\b[^\]]*|linkedin\b[^\]]*'
            r'|Calendly\b[^\]]*|calendly\b[^\]]*'
            r')\]',
            '', text,
        )
        if sender_company:
            text = re.sub(
                r'\[(?:Your\s+Company|your\s+company'
                r'|Company\s+Name|company\s+name)\]',
                sender_company, text,
            )
        else:
            text = re.sub(
                r'\s*\[(?:Your\s+Company|your\s+company'
                r'|Company\s+Name|company\s+name)\]',
                '', text,
            )

        # 2 — Remove entire lines whose only purpose was to be a template
        #     fill-in header, e.g. ``Key Topic 1: budget planning`` or
        #     ``Insert time: tomorrow``.  Must run BEFORE orphaned-label
        #     removal so the full line is still intact.
        text = re.sub(
            r'^[ \t]*(?:'
            r'Key\s+Topic\s+\d'
            r'|Specific\s+Resource'
            r'|Insert\s+\w+'
            r')\s*[:–—-]?.*\n?',
            '', text, flags=re.MULTILINE | re.IGNORECASE,
        )

        # 3 — Remove orphaned label headers that have no value after them
        #     e.g. "Meeting Link:" on its own line.
        text = re.sub(
            r'^[ \t]*(?:'
            r'Meeting\s+Link|Link|Video\s+[Cc]all|[Cc]all\s+[Ll]ink'
            r'|Location|Time|Date|Phone|In[- ]person'
            r'|Your\s+Title|Your\s+Company|Company\s+Name|Position|Designation'
            r'|Add\s+[Dd]etails'
            r'|LinkedIn|Calendly|Budget|Team'
            r'|Decision\s+Makers?'
            r'|Proposed\s+Date|Proposed\s+Time'
            r')\s*:\s*\n?',
            '', text, flags=re.MULTILINE,
        )

        # 4 — Remove entire lines that still contain ``[...]`` (including
        #     sentences with inline placeholders like "Send by [date]").
        text = re.sub(
            r'^[^\n]*\[[^\]]*\][^\n]*\n?',
            '', text, flags=re.MULTILINE,
        )

        # 5 — Remove any ``---`` / ``___`` divider lines the model inserted
        text = re.sub(r'^[ \t]*[-_]{3,}[ \t]*\n?', '', text, flags=re.MULTILINE)

        # 6 — Catch any stray bracket fragments the line-level pass missed
        text = re.sub(r'\s*\[[^\]]*\]\s*', ' ', text)
        text = text.replace('[', '')
        text = text.replace(']', '')

        # 7 — Remove any "Quick recap:" / "Key points:" / "Summary:" or
        #     similar headers that ended up empty after placeholder removal.
        text = re.sub(
            r'^[ \t]*(?:Quick\s+recap|Key\s+points?|Summary|Notes)[ \t]*:?\s*\n?',
            '', text, flags=re.MULTILINE | re.IGNORECASE,
        )

        # 8 — Clean up whitespace artifacts left behind
        text = re.sub(r'  +', ' ', text)
        text = re.sub(r'\n{3,}', '\n\n', text)
        text = re.sub(r'\n[ \t]+', '\n', text)

        return text.strip()

    def _parse_json_response(self, text):
        cleaned = text.strip()
        if cleaned.startswith('```') and '```' in cleaned[3:]:
            start = cleaned.find('\n') + 1
            end = cleaned.rfind('```')
            cleaned = cleaned[start:end].strip()
        import json
        try:
            return json.loads(cleaned)
        except (json.JSONDecodeError, ValueError):
            return None

    def _find_matching_people(self, name, user):
        """Search across every CRM module that could reference a person.
        Returns a list of dicts {source, name, email, company, position}
        ordered by module priority (structured person data first)."""
        seen = set()
        results = []

        def add(source, obj):
            key = (type(obj).__name__, obj.pk)
            if key not in seen:
                seen.add(key)
                pname = (
                    obj.full_name
                    if hasattr(obj, 'full_name')
                    else obj.lead_name
                )
                results.append({
                    'source': source,
                    'name': pname,
                    'email': (obj.email or '').strip(),
                    'company': (obj.company or '').strip(),
                    'position': (obj.job_title or '').strip() if hasattr(obj, 'job_title') else '',
                })

        def add_owner(source, obj):
            """Extract the owner (User FK) as the person for entity modules."""
            key = (type(obj).__name__, obj.pk)
            if key not in seen:
                seen.add(key)
                owner = getattr(obj, 'owner', None)
                if owner:
                    oname = owner.get_full_name() or owner.username
                    oemail = (owner.email or '').strip()
                    results.append({
                        'source': source,
                        'name': oname,
                        'email': oemail,
                        'company': '',
                        'position': '',
                    })
                # If no owner, skip (no person data to extract)

        def add_linked_person(source, obj, field_name):
            """Extract a linked Contact or Lead as the person."""
            related = getattr(obj, field_name, None)
            if related:
                key = (type(related).__name__, related.pk)
                if key not in seen:
                    seen.add(key)
                    pname = (
                        related.full_name
                        if hasattr(related, 'full_name')
                        else getattr(related, 'lead_name', str(related))
                    )
                    results.append({
                        'source': source,
                        'name': pname,
                        'email': (getattr(related, 'email', '') or '').strip(),
                        'company': (getattr(related, 'company', '') or '').strip(),
                        'position': (getattr(related, 'job_title', '') or '').strip(),
                    })

        from contacts.models import Contact
        from leads.models import Lead
        from calendars.models import Event
        from tasks.models import Task
        from campaigns.models import Campaign
        from workflows.models import Workflow, Notification

        # ── 1. Contacts (structured person data) ──
        for c in user.contacts.filter(full_name__icontains=name):
            add('Contact', c)

        # ── 2. Leads (structured person data) ──
        for l in user.leads.filter(lead_name__icontains=name):
            add('Lead', l)
        for l in user.leads.filter(contact_person__icontains=name):
            add('Lead', l)

        # ── 3. Events → linked Contact / Lead / Owner ──
        for e in user.events.filter(
            contact__full_name__icontains=name,
        ).select_related('contact'):
            add_linked_person('Event', e, 'contact')
        for e in user.events.filter(
            lead__lead_name__icontains=name,
        ).select_related('lead'):
            add_linked_person('Event', e, 'lead')
        for e in user.events.filter(title__icontains=name).select_related('owner'):
            add_owner('Event', e)

        # ── 4. Tasks → linked Contact / Lead / Owner ──
        for t in user.tasks.filter(
            contact__full_name__icontains=name,
        ).select_related('contact'):
            add_linked_person('Task', t, 'contact')
        for t in user.tasks.filter(
            lead__lead_name__icontains=name,
        ).select_related('lead'):
            add_linked_person('Task', t, 'lead')
        for t in user.tasks.filter(title__icontains=name).select_related('owner'):
            add_owner('Task', t)

        # ── 5. Campaigns ──
        for c in user.campaigns.filter(name__icontains=name).select_related('owner'):
            add_owner('Campaign', c)
        for c in user.campaigns.filter(subject__icontains=name).select_related('owner'):
            add_owner('Campaign', c)

        # ── 6. Workflows ──
        for w in user.workflows.filter(name__icontains=name).select_related('owner'):
            add_owner('Workflow', w)
        for w in user.workflows.filter(description__icontains=name).select_related('owner'):
            add_owner('Workflow', w)

        # ── 7. Notifications ──
        for n in user.notifications.filter(title__icontains=name).select_related('owner'):
            add_owner('Notification', n)
        for n in user.notifications.filter(message__icontains=name).select_related('owner'):
            add_owner('Notification', n)

        return results

    @staticmethod
    def _find_all_people(user):
        """Gather every distinct person (name + email) from the user's
        Contacts and Leads so the LLM can address a bulk email."""
        if not user or not user.is_authenticated:
            return []
        seen = set()
        results = []
        from contacts.models import Contact
        from leads.models import Lead
        for c in user.contacts.exclude(email='').exclude(email__isnull=True):
            key = c.email.lower().strip()
            if key and key not in seen:
                seen.add(key)
                results.append({
                    'name': c.full_name or c.email.split('@')[0],
                    'email': c.email.strip(),
                    'company': (c.company or '').strip(),
                    'position': (c.job_title or '').strip(),
                    'source': 'Contact',
                })
        for l in user.leads.exclude(email='').exclude(email__isnull=True):
            key = l.email.lower().strip()
            if key and key not in seen:
                seen.add(key)
                results.append({
                    'name': l.lead_name or l.email.split('@')[0],
                    'email': l.email.strip(),
                    'company': (l.company or '').strip(),
                    'position': '',
                    'source': 'Lead',
                })
        if user.email:
            key = user.email.lower().strip()
            if key not in seen:
                seen.add(key)
                results.append({
                    'name': user.get_full_name() or user.username,
                    'email': user.email.strip(),
                    'company': '',
                    'position': '',
                    'source': 'User',
                })
        return results

    def _gather_crm_context(self, recipient_name, email_addr, user):
        """Collect CRM data (events, tasks, campaigns) related to the
        recipient so the LLM can use actual values instead of placeholders."""
        if not user or not user.is_authenticated:
            return ''
        context_parts = []

        from calendars.models import Event as CalendarEvent
        from tasks.models import Task
        from campaigns.models import Campaign
        from contacts.models import Contact
        from leads.models import Lead
        from workflows.models import Workflow, Notification
        from django.db.models import Q

        name_lower = recipient_name.lower().strip() if recipient_name else ''
        # Extract meaningful keywords from the recipient text (skip noise words)
        stopwords = {'the', 'a', 'an', 'to', 'for', 'of', 'in', 'on', 'at', 'by',
                     'with', 'from', 'and', 'or', 'attendee', 'participant',
                     'tomorrow', 'today', 'yesterday', 'next', 'this', 'that',
                     'meeting', 'email', 'mail', 'write', 'draft', 'compose'}
        keywords = [w for w in name_lower.split() if w not in stopwords and len(w) > 2]

        def _match_q(field, term):
            """Build a case-insensitive contains query for a single term."""
            return Q(**{f'{field}__icontains': term})

        def _entity_matches(model, fields, filters=None):
            """Search a model for entities matching any keyword across fields."""
            if not keywords:
                return model.objects.none()
            q = Q()
            for kw in keywords:
                for f in fields:
                    q |= _match_q(f, kw)
            base = model.objects.filter(q)
            if filters:
                base = base.filter(**filters)
            return base.distinct()

        # ── 1. Calendar events ──
        events = _entity_matches(
            CalendarEvent,
            ['title', 'description', 'location'],
            {'owner': user, 'status': 'scheduled'},
        ).order_by('start_date', 'start_time')[:3]
        for ev in events:
            parts = [f'Event: {ev.title}']
            if ev.start_date:
                parts.append(f'Date: {ev.start_date}')
            if ev.start_time:
                parts.append(f'Time: {ev.start_time}'
                             + (f' - {ev.end_time}' if ev.end_time else ''))
            if ev.location:
                parts.append(f'Location: {ev.location}')
            if ev.meeting_link:
                parts.append(f'Meeting Link: {ev.meeting_link}')
            context_parts.append(' | '.join(parts))

        # ── 2. Tasks ──
        tasks = _entity_matches(
            Task,
            ['title', 'description'],
            {'owner': user},
        ).exclude(status='completed').order_by('due_date', 'due_time')[:3]
        for t in tasks:
            parts = [f'Task: {t.title}']
            if t.due_date:
                parts.append(f'Due: {t.due_date}')
            if t.due_time:
                parts.append(f'Due Time: {t.due_time}')
            context_parts.append(' | '.join(parts))

        # ── 3. Campaigns ──
        campaigns = _entity_matches(
            Campaign,
            ['name', 'subject', 'body'],
            {'owner': user},
        ).order_by('-created_at')[:3]
        for c in campaigns:
            parts = [f'Campaign: {c.name}']
            if c.subject:
                parts.append(f'Subject: {c.subject}')
            context_parts.append(' | '.join(parts))

        # ── 4. Contacts ──
        contact_fields = ['full_name', 'email', 'company', 'job_title', 'notes']
        if keywords:
            q = Q()
            for kw in keywords:
                for f in contact_fields:
                    q |= _match_q(f, kw)
            contacts = user.contacts.filter(q).distinct()[:3]
            for c in contacts:
                parts = [f'Contact: {c.full_name}']
                if c.email:
                    parts.append(f'Email: {c.email}')
                if c.company:
                    parts.append(f'Company: {c.company}')
                if c.job_title:
                    parts.append(f'Position: {c.job_title}')
                if c.notes:
                    parts.append(f'Notes: {c.notes[:200]}')
                context_parts.append(' | '.join(parts))

        # ── 5. Leads ──
        leads = _entity_matches(
            Lead,
            ['lead_name', 'contact_person', 'email', 'notes'],
            {'owner': user},
        ).order_by('-created_at')[:3]
        for l in leads:
            parts = [f'Lead: {l.lead_name}']
            if l.email:
                parts.append(f'Email: {l.email}')
            if l.get_status_display():
                parts.append(f'Status: {l.get_status_display()}')
            if l.expected_revenue:
                parts.append(f'Expected Revenue: ${l.expected_revenue:,.2f}')
            context_parts.append(' | '.join(parts))

        # ── 6. Workflows ──
        workflows = _entity_matches(
            Workflow,
            ['name', 'description'],
            {'owner': user},
        ).order_by('-created_at')[:3]
        for w in workflows:
            parts = [f'Workflow: {w.name}']
            parts.append(f'Active: {"Yes" if w.is_active else "No"}')
            if w.description:
                parts.append(f'Description: {w.description[:200]}')
            context_parts.append(' | '.join(parts))

        # ── 7. Notifications ──
        notifications = _entity_matches(
            Notification,
            ['title', 'message'],
            {'owner': user},
        ).order_by('-created_at')[:3]
        for n in notifications:
            parts = [f'Notification: {n.title}']
            parts.append(f'Read: {"Yes" if n.is_read else "No"}')
            if n.message:
                parts.append(f'Message: {n.message[:200]}')
            context_parts.append(' | '.join(parts))

        return '\n'.join(context_parts) if context_parts else ''

    def execute(self, text, user):
        if not user or not user.is_authenticated:
            return None

        params = self._parse(text)
        recipient_name = params.get('recipient', '').strip()
        email_addr = params.get('email_address', '').strip() or None
        purpose = params.get('purpose', '').strip()
        company_name = None
        position_name = None
        crm_source = ''
        bulk_all = params.get('bulk_all', False)

        # ── Bulk email to everyone ──
        if bulk_all:
            all_people = self._find_all_people(user)
            if not all_people:
                return (
                    'I could not find any contacts or leads with email '
                    'addresses to send to.'
                )
            recipient_list = [p for p in all_people if p['email']]
            if not recipient_list:
                return (
                    'No recipients with email addresses were found in '
                    'your contacts or leads.'
                )

            sender_name = user.get_full_name() or user.username
            sender_email = user.email or ''
            sender_company = (getattr(user, 'company', None) or '').strip() or None
            total = len(recipient_list)
            purpose_text = purpose or 'a general update'

            to_summary = '\n'.join(
                f'  {p["name"]} <{p["email"]}>' for p in recipient_list[:20]
            )
            if total > 20:
                to_summary += f'\n  ... and {total - 20} more'

            crm_context = self._gather_crm_context(
                'All Recipients', None, user,
            )
            prompt = self._build_prompt(
                'All Recipients', None, None, None, purpose_text,
                sender_name, sender_email, sender_company, crm_context,
                'Bulk',
            )
            from assistant.services.ai_service import AIService
            from assistant.services.ai_crm_service import MockMessage
            ai = AIService()
            raw = ai.generate_response([MockMessage('user', prompt)])
            raw = self._strip_reasoning(raw)
            result = self._parse_json_response(raw)
            if result:
                subject = result.get('subject', '').strip()
                body = result.get('body', '').strip()
            else:
                subject = ''
                body = raw.strip()
            body = self._post_process_body(body, sender_name, sender_email)
            body = self._clean_placeholders(body, sender_company)
            if not subject:
                subject = 'Follow-up' if purpose else 'Hello'
            body = re.sub(
                r'(?:\n\s*)?'
                r'(?:Best\s+regards|Kind\s+regards|Regards|Sincerely|'
                r'Thanks|Thank\s+you|Warmly|Best|Cheers|'
                r'Yours\s+truly|Yours\s+sincerely|Respectfully)'
                r'[,!.]?[ \t]*'
                r'(?:\n.*)*$',
                '', body, flags=re.IGNORECASE | re.DOTALL,
            ).rstrip('\n')
            sig_lines = ['Best regards,', '', sender_name]
            if sender_email:
                sig_lines.extend(['', sender_email])
            body = body + '\n\n' + '\n'.join(sig_lines) if body else '\n'.join(sig_lines)

            lines = [
                f'Email drafted for **{total} recipient(s)**',
                '',
                'To:',
                to_summary,
                '',
                f'From:\n{sender_name} <{sender_email}>',
                '',
                f'Subject:\n{subject}',
                '',
                'Body:',
                '',
                body,
            ]
            return '\n'.join(lines)

        if email_addr:
            # Direct email provided — skip CRM module search
            if not recipient_name:
                recipient_name = email_addr.split('@')[0]
            company_name = None
        else:
            if not recipient_name:
                return (
                    'Who would you like to send the email to? '
                    'Please provide a recipient name or email address.'
                )

            # Search across CRM modules
            people = self._find_matching_people(recipient_name, user)

            if len(people) > 1:
                lines = [
                    f'I found multiple people matching **"{recipient_name}"**:',
                    '',
                ]
                for idx, p in enumerate(people, 1):
                    lines.append(f'{idx}.')
                    lines.append(p['name'])
                    lines.append(f'({p["source"]})')
                    if p['email']:
                        lines.append(p['email'])
                    lines.append('')
                lines.append('Which one would you like to email? Please specify.')
                return '\n'.join(lines)

            if people:
                person = people[0]
                recipient_name = person['name']
                email_addr = person['email'] or None
                company_name = person['company'] or None
                position_name = person.get('position') or None
                crm_source = person.get('source', '')
            # else: recipient not found in CRM — continue with parsed name

        # Retrieve logged-in sender info
        sender_name = user.get_full_name() or user.username
        sender_email = user.email or ''
        sender_company = (getattr(user, 'company', None) or '').strip() or None

        # Gather CRM context for personalised email content
        crm_context = self._gather_crm_context(recipient_name, email_addr, user)

        # Build prompt and call LLM
        prompt = self._build_prompt(
            recipient_name, email_addr, company_name, position_name, purpose,
            sender_name, sender_email, sender_company, crm_context, crm_source,
        )

        from assistant.services.ai_service import AIService
        from assistant.services.ai_crm_service import MockMessage
        ai = AIService()
        raw = ai.generate_response([MockMessage('user', prompt)])

        # Strip any internal reasoning / chain-of-thought before display
        raw = self._strip_reasoning(raw)

        # Parse JSON or fall back to raw text
        result = self._parse_json_response(raw)
        if result:
            subject = result.get('subject', '').strip()
            body = result.get('body', '').strip()
        else:
            subject = ''
            body = raw.strip()

        # Replace any hallucinated placeholders with real sender info
        body = self._post_process_body(body, sender_name, sender_email)

        # Final cleanup — remove every remaining placeholder token the LLM
        # may have included despite the prompt instructions.
        body = self._clean_placeholders(body, sender_company)

        # Ensure subject always exists
        if not subject:
            subject = 'Follow-up' if purpose else 'Hello'

        # Strip any trailing LLM-generated closing from body, then append the
        # complete canonical signature (Best regards, + name + email).
        body = re.sub(
            r'(?:\n\s*)?'
            r'(?:Best\s+regards|Kind\s+regards|Regards|Sincerely|'
            r'Thanks|Thank\s+you|Warmly|Best|Cheers|'
            r'Yours\s+truly|Yours\s+sincerely|Respectfully)'
            r'[,!.]?[ \t]*'
            r'(?:\n.*)*$',
            '', body, flags=re.IGNORECASE | re.DOTALL,
        ).rstrip('\n')
        sig_lines = ['Best regards,', '', sender_name]
        if sender_email:
            sig_lines.extend(['', sender_email])
        body = body + '\n\n' + '\n'.join(sig_lines) if body else '\n'.join(sig_lines)

        to_lines = ['To:']
        if email_addr:
            to_lines.append(f'{recipient_name} <{email_addr}>')
        else:
            to_lines.append(recipient_name)

        from_lines = ['From:']
        from_lines.append(f'{sender_name} <{sender_email}>')

        lines = [
            '\n'.join(to_lines),
            '',
            '\n'.join(from_lines),
            '',
            f'Subject:\n{subject}',
            '',
            'Body:',
            '',
            body,
        ]

        result = '\n'.join(lines)

        # Validate all required sections are present
        required = ['To:', 'From:', 'Subject:', 'Body:', 'Best regards,']
        missing = [s for s in required if s not in result]
        if missing:
            logger.error(
                'Email output missing sections: %s. Regenerating...', missing,
            )
            result = (
                f'To:\n{recipient_name}\n\n'
                f'From:\n{sender_name}\n\n'
                f'Subject:\n{subject}\n\n'
                f'Body:\n\nDear {recipient_name.split()[0] if recipient_name else "there"},\n\n'
                f'{purpose or "I hope this message finds you well."}\n\n'
                f'{'\n'.join(sig_lines)}'
            )

        if email_addr:
            from emails.models import SMTPConfig, EmailMessage
            from emails.services import send_email_message
            smtp = SMTPConfig.objects.filter(owner=user).first()
            if smtp:
                final_body = body + '\n\n' + '\n'.join(sig_lines) if body else '\n'.join(sig_lines)
                try:
                    email_msg = EmailMessage.objects.create(
                        owner=user,
                        smtp_config=smtp,
                        to_emails=email_addr,
                        subject=subject,
                        body_plain=final_body,
                        body_html=f'<pre style="font-family:inherit;white-space:pre-wrap">{final_body}</pre>',
                        status='queued',
                    )
                    success, error = send_email_message(email_msg)
                    from django.urls import reverse
                    view_url = reverse('emails:detail', args=[email_msg.pk])
                    if hasattr(self, '_current_request') and self._current_request is not None:
                        view_url = self._current_request.build_absolute_uri(view_url)
                    if success:
                        result += f'\n\n---\n**Email sent successfully** to {recipient_name} <{email_addr}>\n[View in Sent Mail]({view_url})'
                    else:
                        result += f'\n\n---\nDraft saved. Failed to send: {error}\n[View details]({view_url})'
                except Exception as e:
                    result += f'\n\n---\nDraft saved. Send error: {e}'
            else:
                result += '\n\n---\n*Draft saved. Configure [SMTP Settings](/emails/smtp/config/) to send emails automatically.*'

        return result


@register
class SendEmailAction(BaseAction):
    """Detect real email-sending intents, compose via LLM, and dispatch
    the message through the user's SMTP configuration.

    Uses the same recipient-resolution and AI-generation pipeline as
    ComposeEmailAction, but then **actually sends** the email and
    returns a delivery result with the logged-in user's real identity.
    """

    action_type = 'send_email'
    keywords = frozenset({
        'send', 'email', 'mail', 'dispatch', 'transmit',
    })
    patterns = [
        re.compile(r'send\s+(?:an?\s+)?(?:email|mail)\b'),
        re.compile(r'send\s+(?:this|that|the)\s+(?:email|mail)\b'),
        re.compile(r'send\s+(?:a\s+)?\w+\s+(?:email|mail)\b'),
        re.compile(r'send\s+(?:a\s+)?follow[-\s]up'),
        re.compile(r'dispatch\s+(?:an?\s+)?(?:email|mail)\b'),
        re.compile(r'email\s+\w+'),
    ]

    _BARE_EMAIL_RE = re.compile(r'^[\w.+-]+@[\w-]+(?:\.[\w-]+)+$')

    def detect(self, text):
        text_stripped = text.strip()
        if self._BARE_EMAIL_RE.match(text_stripped):
            return True
        return super().detect(text_stripped)

    @staticmethod
    def _resolve_sender(user):
        """Return (sender_name, sender_email) from the logged-in user,
        falling back to SMTPConfig sender fields if profile is missing."""
        from emails.models import SMTPConfig
        name = user.get_full_name() or user.username or ''
        email = user.email or ''
        if not name or not email:
            smtp = SMTPConfig.objects.filter(owner=user).first()
            if smtp:
                name = name or smtp.sender_name or smtp.username
                email = email or smtp.effective_sender_email
        return name, email

    def execute(self, text, user):
        if not user or not user.is_authenticated:
            return None

        from emails.models import SMTPConfig, EmailMessage
        from emails.services import send_email_message

        print(f'[SEND-EMAIL] Intent detected: {text!r}', flush=True)
        print(f'[SEND-EMAIL] user_id={user.id} email={user.email!r} username={user.username!r}', flush=True)

        smtp = SMTPConfig.objects.filter(owner=user).first()
        if not smtp:
            print('[SEND-EMAIL] ABORT: No SMTP config found for user', flush=True)
            from django.urls import reverse
            smtp_url = reverse('emails:smtp_config')
            if hasattr(self, '_current_request') and self._current_request is not None:
                smtp_url = self._current_request.build_absolute_uri(smtp_url)
            return (
                'SMTP is not configured.\n\n'
                'Please configure SMTP Settings before sending emails.\n\n'
                f'[SMTP Settings]({smtp_url})'
            )

        print(f'[SEND-EMAIL] SMTP config loaded: id={smtp.id} host={smtp.host} port={smtp.port} use_tls={smtp.use_tls} username={smtp.username}', flush=True)

        sender_name, sender_email = self._resolve_sender(user)
        print(f'[SEND-EMAIL] Sender resolved: name={sender_name!r} email={sender_email!r}', flush=True)

        print('[SEND-EMAIL] Calling ComposeEmailAction to generate email content...', flush=True)
        composer = ComposeEmailAction()

        # ── Multi-recipient: split "X and Y" into separate sends ──
        multi_m = re.search(r'(?:to|for)\s+(.+?)$', text, re.IGNORECASE)
        if multi_m:
            raw_names = re.split(r'\s*,\s*(?:and\s+)?|\s+and\s+', multi_m.group(1))
            raw_names = [n.strip() for n in raw_names if n.strip()]
            # Only treat as multi-recipient if every segment looks like
            # a short name (no prepositions, no long phrases).
            _PREPOSITIONS = re.compile(
                r'\b(?:about|on|at|in|for|from|regarding|concerning|'
                r'write|draft|compose|create|send|subject|body)\b',
                re.IGNORECASE,
            )
            if len(raw_names) > 1 and all(
                len(n) < 40 and not _PREPOSITIONS.search(n)
                for n in raw_names
            ):
                print(f'[SEND-EMAIL] Multi-recipient detected: {raw_names}', flush=True)
                prefix = text[:multi_m.start()]
                results = []
                for name in raw_names:
                    single = f'{prefix}to {name}'
                    print(f'[SEND-EMAIL] Processing: {single!r}', flush=True)
                    r = self._send_one(composer, single, user, smtp, sender_name, sender_email)
                    if r:
                        results.append(r)
                return '\n'.join(results) if results else 'Could not send emails to any recipients.'

        print('[SEND-EMAIL] Single recipient path...', flush=True)
        return self._send_one(composer, text, user, smtp, sender_name, sender_email)

    def _generate_default_email(self, text, user, sender_name, sender_email):
        """Generate email content directly via LLM without asking questions.

        Used as a fallback when ComposeEmailAction returns a clarifying
        question instead of composed email content.
        """
        from assistant.services.ai_service import AIService
        from assistant.services.ai_crm_service import MockMessage

        composer = ComposeEmailAction()
        params = composer._parse(text)
        recipient_name = params.get('recipient', '').strip()
        email_addr = params.get('email_address', '').strip() or None
        purpose = params.get('purpose', '').strip()

        if not email_addr and not recipient_name:
            return None

        if not recipient_name and email_addr:
            recipient_name = email_addr.split('@')[0]

        company_name = None
        position_name = None
        crm_source = ''
        if not email_addr:
            people = composer._find_matching_people(recipient_name, user)
            if people:
                person = people[0]
                recipient_name = person['name']
                email_addr = person['email'] or None
                company_name = person['company'] or None
                position_name = person.get('position') or None
                crm_source = person.get('source', '')

        prompt = (
            'You are a professional email writer. Generate a brief, '
            'professional email. Do NOT ask questions. Do NOT list options. '
            'Just generate the email directly.\n\n'
            f'RECIPIENT: {recipient_name or "Unknown"}\n'
            f'EMAIL: {email_addr or "Unknown"}\n'
        )
        if company_name:
            prompt += f'COMPANY: {company_name}\n'
        if purpose:
            prompt += f'PURPOSE: {purpose}\n'
        prompt += (
            f'\nSENDER: {sender_name} <{sender_email}>\n\n'
            'Generate a detailed, professional email (8-15 sentences minimum). '
            'Include a proper greeting, context, key details, and a clear call to action. '
            'Make it feel personal and thoughtful. '
            'Respond ONLY with JSON:\n'
            '{"subject": "<subject>", "body": "<body>"}\n'
        )

        ai = AIService()
        raw = ai.generate_response([MockMessage('user', prompt)])
        raw = composer._strip_reasoning(raw)
        result = composer._parse_json_response(raw)
        if result:
            subject = result.get('subject', '').strip() or 'Hello'
            body = result.get('body', '').strip()
        else:
            subject = 'Hello'
            body = raw.strip() if raw else ''

        if not body:
            return None

        to_line = f'{recipient_name} <{email_addr}>' if email_addr else recipient_name
        return (
            f'To:\n{to_line}\n\n'
            f'From:\n{sender_name} <{sender_email}>\n\n'
            f'Subject:\n{subject}\n\n'
            f'Body:\n\n{body}'
        )

    def _send_one(self, composer, text, user, smtp, sender_name, sender_email):
        from emails.models import EmailMessage
        from emails.services import send_email_message

        raw_result = composer.execute(text, user)
        if raw_result is None:
            print('[SEND-EMAIL] ABORT: ComposeEmailAction returned None', flush=True)
            return None

        print(f'[SEND-EMAIL] ComposeEmail result ({len(raw_result)} chars): {raw_result[:300]!r}', flush=True)

        # If ComposeEmailAction returned a clarifying question instead of
        # email content, generate a default email directly via LLM.
        _is_question = (
            'I found multiple people matching' in raw_result
            or ('?' in raw_result and len(raw_result) < 500)
            or raw_result.strip().startswith(('Which', 'Who', 'What'))
            or 'Who would you like' in raw_result
        )
        if _is_question:
            print('[SEND-EMAIL] ComposeEmail returned a question — generating default email', flush=True)
            raw_result = self._generate_default_email(text, user, sender_name, sender_email)
            if not raw_result:
                return 'Could not generate an email for this request.'

        if 'I found multiple people matching' in raw_result:
            print('[SEND-EMAIL] Multi-match disambiguation needed, passing through', flush=True)
            return raw_result

        to_match = re.search(r'^To:\s*\n\s*(.+?)(?:\n|$)', raw_result, re.MULTILINE)
        subject_match = re.search(r'^Subject:\s*\n\s*(.+?)(?:\n|$)', raw_result, re.MULTILINE)
        body_match = re.search(r'^Body:\s*\n(.*)', raw_result, re.DOTALL | re.MULTILINE)

        if not to_match:
            to_match = re.search(r'^To:\s*(.+?)(?:\n|$)', raw_result, re.MULTILINE)
        if not subject_match:
            subject_match = re.search(r'^Subject:\s*(.+?)(?:\n|$)', raw_result, re.MULTILINE)

        to_line = to_match.group(1).strip() if to_match else ''
        subject = subject_match.group(1).strip() if subject_match else 'No Subject'

        body = body_match.group(1).strip() if body_match else ''
        body = re.sub(
            r'(?:\n\s*)?'
            r'(?:Best\s+regards|Kind\s+regards|Regards|Sincerely|'
            r'Thanks|Thank\s+you|Warmly|Best|Cheers|'
            r'Yours\s+truly|Yours\s+sincerely|Respectfully)'
            r'[,!.]?[ \t]*'
            r'(?:\n.*)*$',
            '', body, flags=re.IGNORECASE | re.DOTALL,
        ).rstrip('\n')

        print(f'[SEND-EMAIL] Parsed: to_line={to_line!r} subject={subject!r}', flush=True)

        email_addr = ''
        name_part = to_line
        addr_match = re.search(r'<([^>]+)>', to_line)
        if addr_match:
            email_addr = addr_match.group(1).strip()
            name_part = to_line.split('<')[0].strip()
        elif '@' in to_line:
            email_addr = to_line.strip()

        print(f'[SEND-EMAIL] Resolved: name_part={name_part!r} email_addr={email_addr!r}', flush=True)

        if not email_addr:
            print(f'[SEND-EMAIL] ABORT: Could not extract email address from to_line={to_line!r}', flush=True)
            return (
                f'I found **{name_part}** but could not determine their email '
                f'address. Please specify an email address.'
            )

        sig = f'Best regards,\n\n{sender_name}'
        if sender_email:
            sig += f'\n{sender_email}'
        final_body = body + '\n\n' + sig if body else sig

        print(f'[SEND-EMAIL] Creating EmailMessage object...', flush=True)
        email = EmailMessage.objects.create(
            owner=user,
            smtp_config=smtp,
            to_emails=email_addr,
            subject=subject,
            body_plain=final_body,
            body_html=f'<pre style="font-family:inherit;white-space:pre-wrap">{final_body}</pre>',
            status='queued',
        )
        print(f'[SEND-EMAIL] EmailMessage created: id={email.id} status={email.status}', flush=True)

        print(f'[SEND-EMAIL] Calling send_email_message(email_id={email.id})...', flush=True)
        success, error = send_email_message(email)
        print(f'[SEND-EMAIL] send_email_message returned: success={success} error={error!r}', flush=True)

        email.refresh_from_db()
        print(f'[SEND-EMAIL] EmailMessage final status: id={email.id} status={email.status}', flush=True)

        from django.urls import reverse
        view_url = reverse('emails:detail', args=[email.pk])
        if hasattr(self, '_current_request') and self._current_request is not None:
            view_url = self._current_request.build_absolute_uri(view_url)

        if success:
            print(f'[SEND-EMAIL] SUCCESS: Email sent to {email_addr}', flush=True)
            return (
                '**Email Sent Successfully**\n\n'
                f'**To:**\n{name_part} <{email_addr}>\n\n'
                f'**From:**\n{sender_name} <{sender_email}>\n\n'
                f'**Subject:**\n{subject}\n\n'
                f'**Status:**\nSent\n\n'
                f'[View in Sent Mail]({view_url})'
            )
        else:
            print(f'[SEND-EMAIL] FAILED: {error}', flush=True)
            return (
                f'❌ Failed to send email to **{name_part}**: {error}\n\n'
                f'[View details]({view_url})'
            )


# ═══════════════════════════════════════════════════════════════════════════
#  analytics — AI CRM Analytics
# ═══════════════════════════════════════════════════════════════════════════

@register
class AnalyticsAction(BaseAction):
    """Answer analytical questions by querying the CRM database directly,
    without using the LLM.

    Supports counting contacts, leads, tasks, events, campaigns, workflows,
    and notifications with filters for status, priority, and time ranges.
    """

    action_type = 'analytics'
    keywords = frozenset({
        'how many', 'count', 'total', 'number of', 'how much',
        'expected revenue', 'revenue',
        'busy', 'packed', 'schedule', 'free', 'look',
        'statistics', 'stats', 'show', 'view',
        'pipeline', 'forecast', 'opportunities', 'top deals',
        'won deals', 'lost deals', 'deal analytics', 'sales summary',
        'pipeline summary', 'revenue forecast', 'expected revenue',
        'largest opportunities', 'win rate',
    })
    patterns = [
        re.compile(r'(?:how many|count|total|number of|how much)'),
        re.compile(r'expected revenue'),
        re.compile(r'(?:show|view|get)\s+(?:crm\s+)?(?:statistics?|stats?|summary|report|dashboard|analytics)\b'),
        re.compile(r'\b(?:how\s+(?:busy|packed)|am\s+i\s+(?:busy|free)|what\s+(?:does\s+)?(?:my\s+)?(?:week|schedule|day)\s+look)\b'),
        re.compile(r'(?:show|view|get|what(?:\'s|\s+is)?)\s+(?:my\s+)?pipeline\b'),
        re.compile(r'(?:sales|pipeline)\s+(?:summary|overview|analytics)\b'),
        re.compile(r'(?:revenue|sales)\s+forecast\b'),
        re.compile(r'expected\s+revenue\s+(?:this\s+)?month\b'),
        re.compile(r'(?:top|largest|biggest)\s+(?:deals?|opportunities?)\b'),
        re.compile(r'(?:won|closed\s+won)\s+deals?\b'),
        re.compile(r'(?:lost|closed\s+lost)\s+deals?\b'),
        re.compile(r'deal\s+analytics\b'),
        re.compile(r'win\s+rate\b'),
    ]

    # ── Entity detection (ordered: more specific first) ──
    _ENTITY_MAP = [
        (re.compile(r'\bcontacts?\b', re.IGNORECASE), 'contact'),
        (re.compile(r'\bleads?\b|\bprospects?\b|\bdeals?\b', re.IGNORECASE), 'lead'),
        (re.compile(r'\btasks?\b|\bto-?dos?\b', re.IGNORECASE), 'task'),
        (re.compile(r'\bmeetings?\b|\bappointments?\b', re.IGNORECASE), 'meeting'),
        (re.compile(r'\b(?:calendar\s+)?events?\b', re.IGNORECASE), 'event'),
        (re.compile(r'\bcampaigns?\b', re.IGNORECASE), 'campaign'),
        (re.compile(r'\bworkflows?\b|\bautomations?\b', re.IGNORECASE), 'workflow'),
        (re.compile(r'\bnotifications?\b|\balerts?\b', re.IGNORECASE), 'notification'),
    ]

    # ── Filter patterns ──
    _STATUS_MAP = {
        'active': re.compile(r'\bactive\b', re.IGNORECASE),
        'inactive': re.compile(r'\binactive\b', re.IGNORECASE),
        'completed': re.compile(r'\bcompleted\b|\bdone\b|\bfinished\b', re.IGNORECASE),
        'pending': re.compile(r'\bpending\b', re.IGNORECASE),
        'scheduled': re.compile(r'\bscheduled\b|\bupcoming\b', re.IGNORECASE),
        'cancelled': re.compile(r'\bcancelled?\b', re.IGNORECASE),
        'draft': re.compile(r'\bdraft\b', re.IGNORECASE),
        'sent': re.compile(r'\bsent\b', re.IGNORECASE),
        'qualified': re.compile(r'\bqualified\b', re.IGNORECASE),
        'contacted': re.compile(r'\bcontacted\b', re.IGNORECASE),
        'negotiation': re.compile(r'\bnegotiation\b', re.IGNORECASE),
        'proposal_sent': re.compile(r'\bproposal\s+sent\b', re.IGNORECASE),
        'won': re.compile(r'\bwon\b', re.IGNORECASE),
        'lost': re.compile(r'\blost\b', re.IGNORECASE),
        'unread': re.compile(r'\bunread\b', re.IGNORECASE),
        'overdue': re.compile(r'\boverdue\b|\bover\s*due\b', re.IGNORECASE),
    }
    _PRIORITY_MAP = {
        'high': re.compile(r'\bhigh[-\s]priority\b|\bhigh\b', re.IGNORECASE),
        'medium': re.compile(r'\bmedium[-\s]priority\b|\bmedium\b', re.IGNORECASE),
        'low': re.compile(r'\blow[-\s]priority\b|\blow\b', re.IGNORECASE),
    }
    _TIME_MAP = {
        'today': re.compile(r'\btoday\b', re.IGNORECASE),
        'tomorrow': re.compile(r'\btomorrow\b', re.IGNORECASE),
        'this_week': re.compile(r'\bthis\s+week\b', re.IGNORECASE),
        'next_week': re.compile(r'\bnext\s+week\b', re.IGNORECASE),
        'this_month': re.compile(r'\bthis\s+month\b', re.IGNORECASE),
        'last_month': re.compile(r'\blast\s+month\b', re.IGNORECASE),
        'this_year': re.compile(r'\bthis\s+year\b', re.IGNORECASE),
    }
    _TIME_LABELS = {
        'today': 'for today',
        'tomorrow': 'for tomorrow',
        'this_week': 'this week',
        'next_week': 'next week',
        'this_month': 'this month',
        'last_month': 'last month',
        'this_year': 'this year',
    }

    _BUSY_RE = re.compile(
        r'\b(?:busy|packed|schedule|am\s+i\s+(?:busy|free)|'
        r'how\s+(?:busy|packed)|'
        r'what\s+(?:does\s+)?(?:my\s+)?(?:week|schedule|day)\s+look)\b',
        re.IGNORECASE,
    )

    # ------------------------------------------------------------------
    # Detection helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _detect_entity(text):
        for pattern, entity in AnalyticsAction._ENTITY_MAP:
            if pattern.search(text):
                return entity
        return None

    @staticmethod
    def _detect_filter(text, filter_map):
        for key, pattern in filter_map.items():
            if pattern.search(text):
                return key
        return None

    # ------------------------------------------------------------------
    # Date-range logic
    # ------------------------------------------------------------------

    @staticmethod
    def _get_date_range(time_key):
        today = date.today()
        if time_key == 'today':
            return today, today
        elif time_key == 'tomorrow':
            d = today + timedelta(days=1)
            return d, d
        elif time_key == 'this_week':
            start = today - timedelta(days=today.weekday())
            return start, start + timedelta(days=6)
        elif time_key == 'next_week':
            start = today - timedelta(days=today.weekday()) + timedelta(days=7)
            return start, start + timedelta(days=6)
        elif time_key == 'this_month':
            start = today.replace(day=1)
            if start.month == 12:
                end = start.replace(year=start.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end = start.replace(month=start.month + 1, day=1) - timedelta(days=1)
            return start, end
        elif time_key == 'last_month':
            first_this = today.replace(day=1)
            end = first_this - timedelta(days=1)
            return end.replace(day=1), end
        elif time_key == 'this_year':
            start = today.replace(month=1, day=1)
            return start, today.replace(month=12, day=31)
        return None, None

    @staticmethod
    def _apply_time_filter(qs, time_key, field_name):
        if not time_key:
            return qs
        start, end = AnalyticsAction._get_date_range(time_key)
        if not start:
            return qs
        return qs.filter(**{
            f'{field_name}__gte': start,
            f'{field_name}__lt': end + timedelta(days=1),
        })

    # ------------------------------------------------------------------
    # Entity-specific filter helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _count_leads(user, status, priority, time_key):
        from django.db.models import Q
        qs = user.leads.all()
        if status:
            mapping = {
                'active': ['New', 'Contacted', 'Qualified', 'Proposal Sent', 'Negotiation'],
                'inactive': ['Won', 'Lost'],
                'qualified': ['Qualified'],
                'contacted': ['Contacted'],
                'negotiation': ['Negotiation'],
                'proposal_sent': ['Proposal Sent'],
                'won': ['Won'],
                'lost': ['Lost'],
                'pending': ['New', 'Contacted'],
                'completed': ['Won'],
            }
            if status in mapping:
                qs = qs.filter(status__in=mapping[status])
        if priority:
            qs = qs.filter(priority__iexact=priority.capitalize())
        qs = AnalyticsAction._apply_time_filter(qs, time_key, 'created_at')
        return qs.count()

    @staticmethod
    def _count_tasks(user, status, priority, time_key):
        from django.db.models import Q
        qs = user.tasks.all()
        if status == 'overdue':
            qs = qs.filter(due_date__lt=date.today()).exclude(status='completed')
        elif status:
            mapping = {
                'completed': ['completed'],
                'pending': ['pending'],
                'active': ['pending', 'in_progress'],
                'in_progress': ['in_progress'],
            }
            if status in mapping:
                qs = qs.filter(status__in=mapping[status])
        if priority:
            qs = qs.filter(priority__iexact=priority)
        qs = AnalyticsAction._apply_time_filter(qs, time_key, 'created_at')
        return qs.count()

    @staticmethod
    def _count_events(user, entity, status, time_key):
        qs = user.events.all()
        if entity == 'meeting':
            qs = qs.filter(event_type='meeting')
        if status:
            mapping = {
                'scheduled': ['scheduled'],
                'completed': ['completed'],
                'cancelled': ['cancelled'],
                'upcoming': ['scheduled'],
            }
            if status in mapping:
                qs = qs.filter(status__in=mapping[status])
        qs = AnalyticsAction._apply_time_filter(qs, time_key, 'start_date')
        return qs.count()

    @staticmethod
    def _count_campaigns(user, status, time_key):
        qs = user.campaigns.all()
        if status:
            mapping = {
                'draft': ['Draft'],
                'scheduled': ['Scheduled'],
                'sent': ['Sent'],
                'active': ['Scheduled', 'Sent'],
            }
            if status in mapping:
                qs = qs.filter(status__in=mapping[status])
        qs = AnalyticsAction._apply_time_filter(qs, time_key, 'created_at')
        return qs.count()

    @staticmethod
    def _count_workflows(user, status, time_key):
        qs = user.workflows.all()
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)
        qs = AnalyticsAction._apply_time_filter(qs, time_key, 'created_at')
        return qs.count()

    @staticmethod
    def _count_notifications(user, status, time_key):
        qs = user.notifications.all()
        if status == 'unread':
            qs = qs.filter(is_read=False)
        qs = AnalyticsAction._apply_time_filter(qs, time_key, 'created_at')
        return qs.count()

    @staticmethod
    def _query_revenue(user, time_key):
        from django.db.models import Sum
        qs = user.leads.all()
        qs = AnalyticsAction._apply_time_filter(qs, time_key, 'created_at')
        total = qs.aggregate(total=Sum('expected_revenue'))['total'] or 0
        if total == 0:
            return 'You have no expected revenue recorded yet.'
        time_str = AnalyticsAction._TIME_LABELS.get(time_key, '')
        if time_str:
            return f'Your total expected revenue {time_str} is ${total:,.2f}.'
        return f'Your total expected revenue is ${total:,.2f}.'

    # ------------------------------------------------------------------
    # Busyness / schedule summary
    # ------------------------------------------------------------------

    @staticmethod
    def _count_busyness(user, time_key):
        from django.db.models import Q
        from datetime import date

        start, end = AnalyticsAction._get_date_range(time_key or 'today')
        if not start:
            start = end = date.today()

        tasks = user.tasks.filter(
            ~Q(status='completed'),
            due_date__gte=start,
            due_date__lte=end,
        )
        task_count = tasks.count()

        events = user.events.filter(
            status='scheduled',
            start_date__gte=start,
            start_date__lte=end,
        )
        event_count = events.count()

        return task_count, event_count, start, end

    @staticmethod
    def _format_busyness(user, time_key):
        task_count, event_count, start, end = AnalyticsAction._count_busyness(user, time_key)
        time_label = AnalyticsAction._TIME_LABELS.get(time_key, '')

        total = task_count + event_count

        logger.info(
            'Busyness for %s (time_key=%s): tasks=%d, events=%d',
            user, time_key, task_count, event_count,
        )

        if total == 0:
            if time_label:
                return f'You have nothing scheduled {time_label}. You\'re free!'
            return 'You have nothing scheduled. You\'re free!'

        parts = []
        if task_count:
            parts.append(f'{task_count} task{"s" if task_count != 1 else ""}')
        if event_count:
            names = 'meetings' if event_count > 1 else 'meeting'
            parts.append(f'{event_count} {names}')

        item_str = ' and '.join(parts)

        if total <= 2:
            busyness = 'not very busy'
        elif total <= 5:
            busyness = 'moderately busy'
        else:
            busyness = 'very busy'

        if time_label:
            cap = time_label[0].upper() + time_label[1:] if time_label else ''
            return f'{cap}, you have {item_str}. You\'re {busyness}!'
        return f'You have {item_str}. You\'re {busyness}!'

    # ------------------------------------------------------------------
    # Response formatting
    # ------------------------------------------------------------------

    @staticmethod
    def _crm_summary(user):
        if not user or not user.is_authenticated:
            return None
        c = user.contacts.count()
        l = user.leads.count()
        t = user.tasks.count()
        e = user.events.count()
        ca = user.campaigns.count()
        w = user.workflows.count()
        n = user.notifications.count()
        lines = [
            '**CRM Statistics Overview**',
            '',
            f'Contacts:      {c}',
            f'Leads:         {l}',
            f'Tasks:         {t}',
            f'Events:        {e}',
            f'Campaigns:     {ca}',
            f'Workflows:     {w}',
            f'Notifications: {n}',
            '',
            'Tip: Ask "how many X are Y" for filtered counts.',
        ]
        return '\n'.join(lines)

    @staticmethod
    def _format_response(entity, count, status, priority, time_key):
        display_names = {
            'contact': 'contacts',
            'lead': 'leads',
            'task': 'tasks',
            'meeting': 'meetings',
            'event': 'events',
            'campaign': 'campaigns',
            'workflow': 'workflows',
            'notification': 'notifications',
        }
        name = display_names.get(entity, entity + 's')

        time_str = AnalyticsAction._TIME_LABELS.get(time_key, '')

        if time_str:
            if status:
                return f'You have {count} {status} {name} {time_str}.'
            if priority:
                return f'You have {count} {priority}-priority {name} {time_str}.'
            return f'You have {count} {name} {time_str}.'

        if priority:
            return f'You currently have {count} {priority}-priority {name}.'
        if status:
            return f'You currently have {count} {status} {name}.'
        if count == 1:
            return f'You currently have 1 {entity}.'
        return f'You currently have {count} {name}.'

    @staticmethod
    def _help_message():
        return (
            'I can help you analyse your CRM data. '
            'Try asking:\n'
            '- How many contacts do I have?\n'
            '- How many high priority leads?\n'
            '- How many tasks are overdue?\n'
            '- How many meetings this week?\n'
            '- How many workflows are active?\n'
            '- What is my expected revenue?\n'
            '- Show pipeline summary\n'
            '- Revenue forecast\n'
            '- Top deals\n'
            '- Won deals\n'
            '- Deal analytics'
        )

    # ------------------------------------------------------------------
    # Pipeline-specific query helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _is_pipeline_query(text_lower):
        pipeline_keywords = [
            'pipeline', 'forecast', 'top deals', 'top opportunity',
            'won deals', 'lost deals', 'deal analytics', 'win rate',
            'sales summary', 'pipeline summary', 'revenue forecast',
            'largest opportunity',
        ]
        return any(kw in text_lower for kw in pipeline_keywords)

    @staticmethod
    def _pipeline_summary(user):
        from django.db.models import Sum
        from deals.models import Deal
        from decimal import Decimal

        qs = Deal.objects.filter(owner=user)
        total = qs.count()
        won = qs.filter(stage='Won')
        lost = qs.filter(stage='Lost')
        open_deals = qs.exclude(stage__in=['Won', 'Lost'])

        won_count = won.count()
        lost_count = lost.count()
        closed_count = won_count + lost_count
        win_rate = (won_count / closed_count * 100) if closed_count else 0

        won_value = won.aggregate(v=Sum('value'))['v'] or Decimal('0')
        pipeline_value = open_deals.aggregate(v=Sum('value'))['v'] or Decimal('0')
        avg_value = (qs.aggregate(v=Sum('value'))['v'] or Decimal('0')) / total if total else Decimal('0')

        stages = []
        stage_order = ['New', 'Qualified', 'Proposal Sent', 'Negotiation', 'Contract Review', 'Won', 'Lost']
        for s in stage_order:
            count = qs.filter(stage=s).count()
            if count:
                stages.append(f'  {s}: {count}')

        lines = [
            '**Sales Pipeline Summary**',
            '',
            f'Total Deals:     {total}',
            f'Open Deals:      {open_deals.count()}',
            f'Won Deals:       {won_count}',
            f'Lost Deals:      {lost_count}',
            f'Pipeline Value:  ${pipeline_value:,.2f}',
            f'Won Revenue:     ${won_value:,.2f}',
            f'Avg Deal Value:  ${avg_value:,.2f}',
            f'Win Rate:        {win_rate:.1f}%',
            '',
        ]
        if stages:
            lines.append('Stage Breakdown:')
            lines.extend(stages)

        return '\n'.join(lines)

    @staticmethod
    def _pipeline_stages(user):
        from deals.models import Deal

        qs = Deal.objects.filter(owner=user)
        stage_order = ['New', 'Qualified', 'Proposal Sent', 'Negotiation', 'Contract Review', 'Won', 'Lost']

        lines = ['**Pipeline by Stage**', '']
        for s in stage_order:
            count = qs.filter(stage=s).count()
            if count:
                lines.append(f'  {s}: {count} deal{"s" if count != 1 else ""}')

        total = qs.count()
        if total:
            lines.append(f'\nTotal: {total} deals')
        else:
            lines.append('No deals in pipeline.')

        return '\n'.join(lines)

    @staticmethod
    def _revenue_forecast(user):
        from deals.models import Deal
        from decimal import Decimal
        from datetime import date, timedelta

        today = date.today()
        month_start = today.replace(day=1)
        if month_start.month == 12:
            next_month_start = month_start.replace(year=month_start.year + 1, month=1, day=1)
        else:
            next_month_start = month_start.replace(month=month_start.month + 1, day=1)

        if next_month_start.month == 12:
            next_month_end = next_month_start.replace(year=next_month_start.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            next_month_end = next_month_start.replace(month=next_month_start.month + 1, day=1) - timedelta(days=1)

        open_deals = Deal.objects.filter(owner=user).exclude(stage__in=['Won', 'Lost'])

        this_month_deals = open_deals.filter(
            expected_close_date__gte=month_start,
            expected_close_date__lte=(month_start.replace(month=month_start.month + 1, day=1) - timedelta(days=1)),
        )
        next_month_deals = open_deals.filter(
            expected_close_date__gte=next_month_start,
            expected_close_date__lte=next_month_end,
        )

        this_month_est = sum(
            (d.value or Decimal('0')) * (d.probability or 0) / 100
            for d in this_month_deals
        )
        next_month_est = sum(
            (d.value or Decimal('0')) * (d.probability or 0) / 100
            for d in next_month_deals
        )
        projected = sum(
            (d.value or Decimal('0')) * (d.probability or 0) / 100
            for d in open_deals
        )

        return (
            '**Revenue Forecast**\n'
            f'Estimated This Month:  ${this_month_est:,.2f}\n'
            f'Estimated Next Month:  ${next_month_est:,.2f}\n'
            f'Projected Pipeline:    ${projected:,.2f}\n'
            f'\nBased on {open_deals.count()} open deals with expected close dates.'
        )

    @staticmethod
    def _top_deals(user, limit=5):
        from deals.models import Deal

        deals = (
            Deal.objects.filter(owner=user)
            .exclude(stage__in=['Won', 'Lost'])
            .order_by('-value')[:limit]
        )

        lines = ['**Top Active Opportunities**', '']
        for i, d in enumerate(deals, 1):
            val = f'${d.value:,.2f}' if d.value else 'No value'
            lines.append(f'{i}. {d.deal_name} — {val} ({d.stage}, {d.probability}% probable)')

        if not deals:
            lines.append('No open deals found.')

        return '\n'.join(lines)

    @staticmethod
    def _won_deals(user, limit=5):
        from deals.models import Deal

        deals = (
            Deal.objects.filter(owner=user, stage='Won')
            .order_by('-updated_at')[:limit]
        )

        lines = ['**Recent Won Deals**', '']
        for d in deals:
            val = f'${d.value:,.2f}' if d.value else 'No value'
            lines.append(f'  {d.deal_name} — {val} (closed {d.updated_at.strftime("%b %d, %Y")})')

        if not deals:
            lines.append('No won deals yet.')

        return '\n'.join(lines)

    @staticmethod
    def _lost_deals(user, limit=5):
        from deals.models import Deal

        deals = (
            Deal.objects.filter(owner=user, stage='Lost')
            .order_by('-updated_at')[:limit]
        )

        lines = ['**Recent Lost Deals**', '']
        for d in deals:
            val = f'${d.value:,.2f}' if d.value else 'No value'
            lines.append(f'  {d.deal_name} — {val} (lost {d.updated_at.strftime("%b %d, %Y")})')

        if not deals:
            lines.append('No lost deals.')

        return '\n'.join(lines)

    # ------------------------------------------------------------------
    # Entry point
    # ------------------------------------------------------------------

    def execute(self, text, user):
        if not user or not user.is_authenticated:
            return None

        text_lower = text.lower()

        # Pipeline-specific queries (handled first before entity detection)
        if self._is_pipeline_query(text_lower):
            if 'forecast' in text_lower or 'expected revenue' in text_lower:
                return self._revenue_forecast(user)
            if 'top deals' in text_lower or 'top opportunity' in text_lower or 'largest opportunity' in text_lower:
                return self._top_deals(user)
            if 'won deals' in text_lower:
                return self._won_deals(user)
            if 'lost deals' in text_lower:
                return self._lost_deals(user)
            if 'stage' in text_lower:
                return self._pipeline_stages(user)
            if 'win rate' in text_lower:
                return self._pipeline_summary(user)
            return self._pipeline_summary(user)

        # Detect time filter early (shared by all paths)
        time_key = self._detect_filter(text_lower, self._TIME_MAP)

        # Special case: revenue query (SUM, not COUNT)
        if 'revenue' in text_lower:
            return self._query_revenue(user, time_key)

        # Special case: overview / statistics / summary (show all counts)
        if re.search(
            r'\b(?:statistics?|stats?|dashboard|overview|summary)\b',
            text_lower, re.IGNORECASE,
        ) and not self._detect_entity(text_lower):
            return self._crm_summary(user)

        # Detect entity
        entity = self._detect_entity(text_lower)
        if not entity:
            if self._BUSY_RE.search(text_lower):
                if not time_key:
                    if re.search(r'\bweek\b', text_lower):
                        time_key = 'this_week'
                    elif re.search(r'\bmonth\b', text_lower):
                        time_key = 'this_month'
                    elif re.search(r'\byear\b', text_lower):
                        time_key = 'this_year'
                    else:
                        time_key = 'today'
                return self._format_busyness(user, time_key)
            return self._help_message()

        status = self._detect_filter(text_lower, self._STATUS_MAP)
        priority = self._detect_filter(text_lower, self._PRIORITY_MAP)

        # Route to entity-specific counter
        if entity == 'contact':
            qs = self._apply_time_filter(user.contacts.all(), time_key, 'created_at')
            count = qs.count()
        elif entity == 'lead':
            count = self._count_leads(user, status, priority, time_key)
        elif entity == 'task':
            count = self._count_tasks(user, status, priority, time_key)
        elif entity in ('meeting', 'event'):
            count = self._count_events(user, entity, status, time_key)
        elif entity == 'campaign':
            count = self._count_campaigns(user, status, time_key)
        elif entity == 'workflow':
            count = self._count_workflows(user, status, time_key)
        elif entity == 'notification':
            count = self._count_notifications(user, status, time_key)
        else:
            return self._help_message()

        return self._format_response(entity, count, status, priority, time_key)


# ═══════════════════════════════════════════════════════════════════════════
#  recommendations — Enterprise AI Smart Recommendations
# ═══════════════════════════════════════════════════════════════════════════

@register
class RecommendationsAction(BaseAction):
    """Answer recommendation / priority questions using live CRM data.

    Collects every category of action item from the database directly
    and returns a prioritised list of actionable recommendations.
    Never uses the LLM — always reads real rows from the database.
    """

    action_type = 'recommendations'
    keywords = frozenset({
        'what should', 'recommend', 'priorit', 'attention', 'overdue',
        'follow-up', 'follow up', 'coming up', 'action items',
        'focus', 'today', 'any', 'which', 'need',
    })
    patterns = [
        re.compile(r'what\s+should\s+(?:i\s+)?do\s+today'),
        re.compile(r"(?:today'?s?\s+)?recommendations?"),
        re.compile(r'show\s+my\s+priorit'),
        re.compile(r'what\s+needs\s+my\s+attention'),
        re.compile(r'any\s+overdue'),
        re.compile(r'which\s+leads?\s+require\s+follow'),
        re.compile(r'what\s+meetings?\s+are\s+coming'),
        re.compile(r'which\s+campaigns?\s+need'),
        re.compile(r"(?:today'?s?\s+)?action\s+items?"),
        re.compile(r'(?:give\s+me\s+)?(?:my\s+)?business\s+recommendations?'),
        re.compile(r'what\s+should\s+i\s+focus\s+on\s+today'),
        re.compile(r'show\s+(?:my\s+)?(?:today\'?s?\s+)?(?:priorit|action)'),
    ]

    @staticmethod
    def _collect_data(user):
        from datetime import date, timedelta
        today = date.today()

        data = {}

        # High priority tasks (pending or in_progress)
        high_priority_tasks = user.tasks.filter(
            priority='high',
        ).exclude(status='completed').order_by('due_date')
        data['high_priority_tasks'] = list(high_priority_tasks.values('title', 'due_date', 'due_time', 'status')[:10])
        data['high_priority_task_count'] = high_priority_tasks.count()

        # Overdue tasks
        overdue_tasks = user.tasks.filter(
            due_date__lt=today,
        ).exclude(status='completed').order_by('due_date')
        data['overdue_tasks'] = list(overdue_tasks.values('title', 'due_date', 'due_time')[:10])
        data['overdue_task_count'] = overdue_tasks.count()

        # Pending tasks (not overdue, not completed)
        pending_tasks = user.tasks.filter(
            status='pending',
            due_date__gte=today,
        ).order_by('due_date')
        data['pending_tasks'] = list(pending_tasks.values('title', 'due_date', 'due_time')[:10])
        data['pending_task_count'] = pending_tasks.count()

        # Meetings today
        meetings_today = user.events.filter(
            start_date=today, event_type='meeting', status='scheduled',
        ).order_by('start_time')
        data['meetings_today'] = list(meetings_today.values('title', 'start_time', 'location', 'contact__full_name', 'lead__lead_name')[:10])
        data['meetings_today_count'] = meetings_today.count()

        # Meetings tomorrow
        meetings_tomorrow = user.events.filter(
            start_date=today + timedelta(days=1), event_type='meeting', status='scheduled',
        ).order_by('start_time')
        data['meetings_tomorrow'] = list(meetings_tomorrow.values('title', 'start_time', 'location')[:10])
        data['meetings_tomorrow_count'] = meetings_tomorrow.count()

        # Upcoming meetings (beyond tomorrow)
        upcoming_meetings = user.events.filter(
            start_date__gt=today + timedelta(days=1), event_type='meeting', status='scheduled',
        ).order_by('start_date', 'start_time')
        data['upcoming_meetings'] = list(upcoming_meetings.values('title', 'start_date', 'start_time', 'location')[:10])
        data['upcoming_meetings_count'] = upcoming_meetings.count()

        # High priority leads (High or Urgent)
        high_priority_leads = user.leads.filter(
            priority__in=['High', 'Urgent'],
        ).order_by('-created_at')
        data['high_priority_leads'] = list(high_priority_leads.values('lead_name', 'email', 'status', 'created_at')[:10])
        data['high_priority_lead_count'] = high_priority_leads.count()

        # Leads with no recent follow-up (created > 7 days ago, still active)
        stale_leads = user.leads.filter(
            created_at__lt=today - timedelta(days=7),
            status__in=['New', 'Contacted', 'Qualified'],
        ).order_by('created_at')
        data['stale_leads'] = list(stale_leads.values('lead_name', 'email', 'status', 'created_at')[:10])
        data['stale_lead_count'] = stale_leads.count()

        # Recently created leads (last 7 days)
        recent_leads = user.leads.filter(
            created_at__gte=today - timedelta(days=7),
        ).order_by('-created_at')
        data['recent_leads'] = list(recent_leads.values('lead_name', 'email', 'status')[:10])
        data['recent_lead_count'] = recent_leads.count()

        # Active campaigns (Scheduled or Sent)
        active_campaigns = user.campaigns.filter(
            status__in=['Scheduled', 'Sent'],
        ).order_by('-scheduled_at')
        data['active_campaigns'] = list(active_campaigns.values('name', 'status', 'scheduled_at')[:10])
        data['active_campaign_count'] = active_campaigns.count()

        # Campaigns ending soon (Scheduled with past or today scheduled_at)
        campaigns_ending = user.campaigns.filter(
            status='Scheduled',
            scheduled_at__lte=today,
        ).order_by('scheduled_at')
        data['campaigns_ending'] = list(campaigns_ending.values('name', 'scheduled_at')[:10])
        data['campaigns_ending_count'] = campaigns_ending.count()

        # Unread notifications
        unread = user.notifications.filter(is_read=False).order_by('-created_at')
        data['unread_notifications'] = list(unread.values('title', 'message')[:10])
        data['unread_notification_count'] = unread.count()

        # Inactive workflows
        inactive_workflows = user.workflows.filter(is_active=False).order_by('-created_at')
        data['inactive_workflows'] = list(inactive_workflows.values('name', 'description')[:10])
        data['inactive_workflow_count'] = inactive_workflows.count()

        return data

    @staticmethod
    def _format_recommendations(data):
        from datetime import date
        items = []

        # 1. High Priority Tasks
        for t in data['high_priority_tasks']:
            due = ''
            if t.get('due_date'):
                due = f" (due {t['due_date']})"
                if t.get('due_time'):
                    due = f" (due {t['due_date']} at {t['due_time']})"
            items.append(f'Complete **{t["title"]}**{due}')

        # 2. Overdue Tasks
        for t in data['overdue_tasks']:
            due = ''
            if t.get('due_date'):
                due = f" (was due {t['due_date']})"
            items.append(f'Catch up on **{t["title"]}**{due}')

        # 3. Today's Meetings
        for m in data['meetings_today']:
            time_str = f' at {m["start_time"]}' if m.get('start_time') else ''
            loc_str = f' — {m["location"]}' if m.get('location') else ''
            items.append(f'Attend **{m["title"]}**{time_str}{loc_str}')
        if data['meetings_today_count'] > len(data['meetings_today']):
            items.append(f'Attend {data["meetings_today_count"]} meeting(s) scheduled today')

        # 4. High Priority Leads
        for l in data['high_priority_leads']:
            items.append(f'Follow up with **{l["lead_name"]}** (high priority lead)')

        # 5. Stale leads (no recent follow-up)
        for l in data['stale_leads']:
            created = l.get('created_at')
            if created:
                if hasattr(created, 'date'):
                    d = created.date()
                else:
                    d = created if isinstance(created, date) else date.today()
                days_ago = (date.today() - d).days
            else:
                days_ago = ''
            items.append(f'Contact **{l["lead_name"]}** — last follow-up {days_ago} day(s) ago')

        # 6. Unread notifications
        if data['unread_notification_count']:
            items.append(f'Read **{data["unread_notification_count"]}** unread notification(s)')

        # 7. Campaign deadlines
        for c in data['campaigns_ending']:
            items.append(f'Review campaign **{c["name"]}** — action needed')

        # 8. Upcoming meetings
        for m in data['upcoming_meetings']:
            date_str = f' on {m["start_date"]}' if m.get('start_date') else ''
            items.append(f'Prepare for **{m["title"]}**{date_str}')

        # 9. Inactive workflows
        for w in data['inactive_workflows']:
            items.append(f'Check **{w["name"]}** workflow (currently inactive)')

        if not items:
            return (
                'Everything looks great.\n\n'
                'You currently have no overdue tasks or urgent items.\n'
                'Enjoy your productive day.'
            )

        lines = ['**Today\'s Recommendations**', '']
        for item in items[:15]:
            lines.append(f'\u2022 {item}')

        return '\n'.join(lines)

    def execute(self, text, user):
        if not user or not user.is_authenticated:
            return None
        data = self._collect_data(user)
        return self._format_recommendations(data)


# ═══════════════════════════════════════════════════════════════════════════
#  dashboard — Enterprise AI Dashboard Summary
# ═══════════════════════════════════════════════════════════════════════════

@register
class DashboardAction(BaseAction):
    """Answer dashboard summary questions using live CRM database data.

    Aggregates information from all existing modules without using the LLM
    for data collection — always queries the database directly.
    """

    action_type = 'dashboard'
    keywords = frozenset({
        'dashboard', 'summary', 'overview', 'how is', 'crm',
        'show', 'view', 'get', 'give',
    })
    patterns = [
        re.compile(r'\b(?:crm\s+)?(?:summary|overview|dashboard)\b'),
        re.compile(r'(?:show|view|get)\s+(?:my\s+)?(?:crm\s+)?(?:dashboard|summary|overview)\b'),
        re.compile(r'(?:give|show|get)\s+me\s+(?:my\s+)?(?:crm\s+)?(?:dashboard|summary|overview)\b'),
        re.compile(r'how\s+is\s+(?:my\s+)?crm\b'),
    ]

    @staticmethod
    def _collect_dashboard_data(user):
        from datetime import date, timedelta
        today = date.today()
        data = {}

        data['contacts_total'] = user.contacts.count()

        leads = user.leads.all()
        data['leads_total'] = leads.count()
        data['leads_qualified'] = leads.filter(status='Qualified').count()
        data['leads_new'] = leads.filter(status='New').count()
        data['leads_converted'] = leads.filter(status='Won').count()

        tasks = user.tasks.all()
        data['tasks_total'] = tasks.count()
        data['tasks_pending'] = tasks.filter(status='pending').count()
        data['tasks_completed'] = tasks.filter(status='completed').count()
        data['tasks_overdue'] = tasks.filter(
            due_date__lt=today,
        ).exclude(status='completed').count()

        events = user.events.all()
        data['meetings_today'] = events.filter(
            start_date=today, event_type='meeting', status='scheduled',
        ).count()
        data['meetings_tomorrow'] = events.filter(
            start_date=today + timedelta(days=1), event_type='meeting', status='scheduled',
        ).count()
        data['upcoming_meetings'] = events.filter(
            start_date__gte=today, event_type='meeting', status='scheduled',
        ).count()

        campaigns = user.campaigns.all()
        data['campaigns_total'] = campaigns.count()
        data['campaigns_active'] = campaigns.filter(status__in=['Scheduled', 'Sent']).count()
        data['campaigns_completed'] = campaigns.filter(status='Sent').count()

        workflows = user.workflows.all()
        data['workflows_total'] = workflows.count()
        data['workflows_active'] = workflows.filter(is_active=True).count()

        data['notifications_total'] = user.notifications.count()
        data['notifications_unread'] = user.notifications.filter(is_read=False).count()

        return data

    @staticmethod
    def _format_dashboard(data):
        lines = ['**CRM Summary**', '']
        lines.append(f'Contacts')
        lines.append(f'{data["contacts_total"]}')
        lines.append('')
        lines.append(f'Leads')
        lines.append(f'{data["leads_total"]}')
        lines.append(f'Qualified Leads')
        lines.append(f'{data["leads_qualified"]}')
        lines.append('')
        lines.append(f'Meetings Today')
        lines.append(f'{data["meetings_today"]}')
        lines.append('')
        lines.append(f'Pending Tasks')
        lines.append(f'{data["tasks_pending"]}')
        lines.append('')
        active_display = data['campaigns_active']
        total_camp = data['campaigns_total']
        lines.append(f'Campaigns')
        lines.append(f'{active_display} Active')
        lines.append('')
        lines.append(f'Unread Notifications')
        lines.append(f'{data["notifications_unread"]}')
        return '\n'.join(lines)

    def execute(self, text, user):
        if not user or not user.is_authenticated:
            return None
        data = self._collect_dashboard_data(user)
        return self._format_dashboard(data)


# ═══════════════════════════════════════════════════════════════════════════
#  crm_insights — CRM Insights & Recommendations
# ═══════════════════════════════════════════════════════════════════════════

@register
class CrmInsightsAction(BaseAction):
    """Analyse CRM data and provide intelligent business insights,
    today's priorities, CRM summary, and actionable recommendations.

    Uses Django ORM for all data collection; only uses the LLM to
    generate human-friendly recommendation text from real CRM data.
    """

    action_type = 'crm_insights'
    keywords = frozenset({
        'summary', 'insights', 'priorities', 'recommendations',
        'focus', 'urgent', 'overdue', 'pending', 'follow up',
        'what should', 'how is my', 'what to do',
        'business insights', 'summarize',
    })
    patterns = [
        re.compile(r'(?:summarize|summary|overview)\s+(?:my\s+)?crm'),
        re.compile(r'(?:insight|how\s+(?:is|are)\s+(?:my|the)|business)'),
        re.compile(r'(?:priorit|recommend|focus|what\s+should)'),
        re.compile(r'(?:urgent|overdue|pending|follow\s+up)'),
        re.compile(r'(?:today\'?s?\s+(?:priorit|agenda|plan|task))'),
        re.compile(r'(?:any\s+(?:urgent|pending|follow))'),
    ]

    # ── Intent detection ──

    _SUMMARY_RE = re.compile(r'\b(summarize|summary|overview)\b', re.IGNORECASE)
    _TODAY_RE = re.compile(
        r'\b(today|priority|priorities|agenda|urgent|overdue|pending|follow\s+up|what\s+should\s+i\s+do)\b',
        re.IGNORECASE,
    )
    _RECOMMEND_RE = re.compile(
        r'\b(recommend|recommendations?|focus\s+on|what\s+to\s+do|what\s+should\s+i\s+focus)\b',
        re.IGNORECASE,
    )

    # ── CRM data collection ──

    @staticmethod
    def _collect_crm_data(user):
        """Gather counts and items from every CRM module via ORM."""
        from datetime import date, timedelta
        today = date.today()
        data = {}

        # ── Contacts ──
        data['contacts_total'] = user.contacts.count()

        # ── Leads ──
        leads = user.leads.all()
        data['leads_total'] = leads.count()
        data['leads_high_priority'] = leads.filter(priority__in=['High', 'Urgent']).count()
        data['leads_new'] = leads.filter(status='New').count()
        data['leads_inactive'] = leads.filter(status__in=['Won', 'Lost']).count()
        data['leads_high_priority_list'] = list(
            leads.filter(priority__in=['High', 'Urgent']).values('lead_name', 'email')[:5]
        )

        # ── Tasks ──
        tasks = user.tasks.all()
        data['tasks_total'] = tasks.count()
        data['tasks_overdue'] = tasks.filter(
            due_date__lt=today,
        ).exclude(status='completed').count()
        data['tasks_due_today'] = tasks.filter(
            due_date=today,
        ).exclude(status='completed').count()
        data['tasks_pending'] = tasks.filter(status='pending').count()
        data['tasks_completed'] = tasks.filter(status='completed').count()
        data['tasks_overdue_list'] = list(
            tasks.filter(due_date__lt=today).exclude(status='completed').values('title', 'due_date')[:5]
        )
        data['tasks_due_today_list'] = list(
            tasks.filter(due_date=today).exclude(status='completed').values('title')[:5]
        )

        # ── Events / Meetings ──
        events = user.events.all()
        data['meetings_today'] = events.filter(
            start_date=today, event_type='meeting', status='scheduled',
        ).count()
        data['meetings_tomorrow'] = events.filter(
            start_date=today + timedelta(days=1), event_type='meeting', status='scheduled',
        ).count()
        data['events_upcoming'] = events.filter(
            start_date__gte=today, status='scheduled',
        ).count()
        data['meetings_today_list'] = list(
            events.filter(start_date=today, event_type='meeting', status='scheduled')
            .values('title', 'start_time', 'location')[:5]
        )
        data['meetings_tomorrow_list'] = list(
            events.filter(start_date=today + timedelta(days=1), event_type='meeting', status='scheduled')
            .values('title', 'start_time', 'location')[:5]
        )
        data['upcoming_events_list'] = list(
            events.filter(start_date__gte=today, status='scheduled')
            .order_by('start_date', 'start_time')
            .values('title', 'start_date', 'start_time')[:5]
        )

        # ── Campaigns ──
        campaigns = user.campaigns.all()
        data['campaigns_total'] = campaigns.count()
        data['campaigns_draft'] = campaigns.filter(status='Draft').count()
        data['campaigns_scheduled'] = campaigns.filter(status='Scheduled').count()
        data['campaigns_sent'] = campaigns.filter(status='Sent').count()
        data['campaigns_draft_list'] = list(
            campaigns.filter(status='Draft').values('name')[:5]
        )
        data['campaigns_scheduled_list'] = list(
            campaigns.filter(status='Scheduled').values('name', 'scheduled_at')[:5]
        )

        # ── Workflows ──
        workflows = user.workflows.all()
        data['workflows_total'] = workflows.count()
        data['workflows_active'] = workflows.filter(is_active=True).count()
        data['workflows_inactive'] = workflows.filter(is_active=False).count()

        # ── Notifications ──
        data['notifications_total'] = user.notifications.count()
        data['notifications_unread'] = user.notifications.filter(is_read=False).count()

        return data

    # ── Response builders ──

    @staticmethod
    def _build_summary(data):
        lines = ['# CRM Summary', '']

        lines.append('## Contacts')
        lines.append(f'- Total: **{data["contacts_total"]}**')
        lines.append('')

        lines.append('## Leads')
        lines.append(f'- Total: **{data["leads_total"]}**')
        lines.append(f'- High Priority: **{data["leads_high_priority"]}**')
        lines.append('')

        lines.append('## Tasks')
        lines.append(f'- Total: **{data["tasks_total"]}**')
        lines.append(f'- Overdue: **{data["tasks_overdue"]}**')
        lines.append(f'- Due Today: **{data["tasks_due_today"]}**')
        lines.append('')

        lines.append('## Meetings')
        lines.append(f'- Today: **{data["meetings_today"]}**')
        lines.append(f'- Tomorrow: **{data["meetings_tomorrow"]}**')
        lines.append(f'- Upcoming: **{data["events_upcoming"]}**')
        lines.append('')

        lines.append('## Campaigns')
        lines.append(f'- Total: **{data["campaigns_total"]}**')
        lines.append(f'- Draft: **{data["campaigns_draft"]}**')
        lines.append(f'- Scheduled: **{data["campaigns_scheduled"]}**')
        lines.append('')

        lines.append('## Workflows')
        lines.append(f'- Active: **{data["workflows_active"]}**')
        lines.append(f'- Inactive: **{data["workflows_inactive"]}**')
        lines.append('')

        lines.append('## Notifications')
        lines.append(f'- Unread: **{data["notifications_unread"]}**')

        return '\n'.join(lines)

    @staticmethod
    def _build_today(data):
        lines = ['# Today\'s Priorities', '']
        found = False

        if data['tasks_overdue']:
            found = True
            lines.append('## Overdue Tasks')
            for t in data['tasks_overdue_list']:
                lines.append(f'- {t["title"]} (overdue)')
            lines.append('')

        if data['tasks_due_today']:
            found = True
            lines.append('## Due Today')
            for t in data['tasks_due_today_list']:
                lines.append(f'- {t["title"]}')
            lines.append('')

        if data['meetings_today']:
            found = True
            lines.append('## Meetings Today')
            for m in data['meetings_today_list']:
                time_str = f' at {m["start_time"]}' if m['start_time'] else ''
                loc_str = f' — {m["location"]}' if m['location'] else ''
                lines.append(f'- {m["title"]}{time_str}{loc_str}')
            lines.append('')

        if data['leads_high_priority']:
            found = True
            lines.append('## High-Priority Leads')
            for l in data['leads_high_priority_list']:
                lines.append(f'- {l["lead_name"]} ({l["email"] or "no email"})')
            lines.append('')

        if data['campaigns_draft']:
            found = True
            lines.append('## Draft Campaigns')
            for c in data['campaigns_draft_list']:
                lines.append(f'- {c["name"]} — ready to review')
            lines.append('')

        if not found:
            lines.append('Nothing urgent needs your attention right now.')
            lines.append('')

        # Add quick recommendations based on ORM data
        lines.append('## Quick Recommendations')
        recommendations = []

        if data['tasks_overdue']:
            recommendations.append(f'Complete {data["tasks_overdue"]} overdue task(s) first.')
        if data['leads_high_priority']:
            recommendations.append(f'Follow up with {data["leads_high_priority"]} high-priority lead(s).')
        if data['meetings_tomorrow']:
            recommendations.append(f'Prepare for {data["meetings_tomorrow"]} meeting(s) tomorrow.')
        if data['campaigns_draft']:
            recommendations.append(f'Review and launch {data["campaigns_draft"]} draft campaign(s).')
        if data['notifications_unread']:
            recommendations.append(f'Check {data["notifications_unread"]} unread notification(s).')

        if recommendations:
            for i, r in enumerate(recommendations, 1):
                lines.append(f'{i}. {r}')
        else:
            lines.append('Everything looks up to date. Keep up the good work!')

        return '\n'.join(lines)

    @staticmethod
    def _build_insights(data):
        lines = ['# CRM Insights', '']

        # Overall health
        total_items = (
            data['contacts_total'] + data['leads_total'] + data['tasks_total']
            + data['campaigns_total'] + data['workflows_total']
        )
        issues = []

        if data['tasks_overdue']:
            issues.append(f'{data["tasks_overdue"]} overdue tasks')
        if data['leads_high_priority'] and data['leads_high_priority'] > 2:
            issues.append(f'{data["leads_high_priority"]} high-priority leads requiring attention')
        if data['campaigns_draft']:
            issues.append(f'{data["campaigns_draft"]} draft campaigns not yet launched')
        if data['notifications_unread']:
            issues.append(f'{data["notifications_unread"]} unread notifications')

        if issues:
            lines.append('## Areas Needing Attention')
            for issue in issues:
                lines.append(f'- {issue.capitalize()}')
            lines.append('')

        lines.append('## At a Glance')
        lines.append(f'- **{data["contacts_total"]}** contacts in your CRM')
        lines.append(f'- **{data["leads_total"]}** leads ({data["leads_high_priority"]} high priority)')
        lines.append(f'- **{data["tasks_total"]}** tasks ({data["tasks_pending"]} pending, {data["tasks_completed"]} completed)')
        lines.append(f'- **{data["events_upcoming"]}** upcoming events')
        lines.append(f'- **{data["campaigns_total"]}** campaigns ({data["campaigns_draft"]} draft, {data["campaigns_scheduled"]} scheduled)')
        lines.append(f'- **{data["workflows_active"]}** active workflows')
        lines.append(f'- **{data["notifications_unread"]}** unread notifications')
        lines.append('')

        # Quick wins
        wins = []
        if data['meetings_today']:
            wins.append(f'You have {data["meetings_today"]} meeting(s) today — be prepared.')
        if data['tasks_due_today']:
            wins.append(f'{data["tasks_due_today"]} task(s) due today — tackle them first.')
        if data['campaigns_scheduled']:
            wins.append(f'{data["campaigns_scheduled"]} campaign(s) scheduled — ready to go.')
        if data['workflows_active']:
            wins.append(f'{data["workflows_active"]} workflow(s) running — your automation is on track.')

        if wins:
            lines.append('## Quick Wins')
            for w in wins:
                lines.append(f'- {w}')

        return '\n'.join(lines)

    @staticmethod
    def _build_recommendations(data, user):
        """Use ORM data to build a prompt, then call the LLM for human-friendly
        recommendations."""
        prompt_lines = [
            'You are an experienced Sales Manager and CRM Consultant.',
            'Based on the CRM data below, provide 3-5 specific, actionable',
            'recommendations. Be concise and direct. Use bullet points.',
            '',
            '--- CRM DATA ---',
            f'Contacts: {data["contacts_total"]}',
            f'Leads: {data["leads_total"]} ({data["leads_high_priority"]} high priority, {data["leads_new"]} new)',
            f'Tasks: {data["tasks_total"]} ({data["tasks_overdue"]} overdue, {data["tasks_due_today"]} due today, {data["tasks_pending"]} pending)',
            f'Meetings Today: {data["meetings_today"]}',
            f'Meetings Tomorrow: {data["meetings_tomorrow"]}',
            f'Upcoming Events: {data["events_upcoming"]}',
            f'Campaigns: {data["campaigns_total"]} ({data["campaigns_draft"]} draft, {data["campaigns_scheduled"]} scheduled)',
            f'Workflows: {data["workflows_total"]} ({data["workflows_active"]} active)',
            f'Notifications: {data["notifications_total"]} ({data["notifications_unread"]} unread)',
            '',
            'Recommendations:',
        ]

        if data['tasks_overdue_list']:
            prompt_lines.append('Overdue tasks:')
            for t in data['tasks_overdue_list']:
                prompt_lines.append(f'  - {t["title"]}')
            prompt_lines.append('')

        if data['tasks_due_today_list']:
            prompt_lines.append('Tasks due today:')
            for t in data['tasks_due_today_list']:
                prompt_lines.append(f'  - {t["title"]}')
            prompt_lines.append('')

        if data['meetings_today_list']:
            prompt_lines.append('Meetings today:')
            for m in data['meetings_today_list']:
                prompt_lines.append(f'  - {m["title"]}')
            prompt_lines.append('')

        prompt = '\n'.join(prompt_lines)

        from assistant.services.ai_service import AIService
        from assistant.services.ai_crm_service import MockMessage
        ai = AIService()
        raw = ai.generate_response([MockMessage('user', prompt)])
        return raw.strip() or 'No recommendations could be generated.'

    # ── Intent routing ──

    @staticmethod
    def _detect_intent(text_lower):
        if CrmInsightsAction._SUMMARY_RE.search(text_lower):
            return 'summary'
        if CrmInsightsAction._RECOMMEND_RE.search(text_lower):
            return 'recommendations'
        if CrmInsightsAction._TODAY_RE.search(text_lower):
            return 'today'
        return 'insights'

    # ── Entry point ──

    def execute(self, text, user):
        if not user or not user.is_authenticated:
            return None

        text_lower = text.lower()
        data = self._collect_crm_data(user)
        intent = self._detect_intent(text_lower)

        if intent == 'summary':
            return self._build_summary(data)
        elif intent == 'today':
            return self._build_today(data)
        elif intent == 'recommendations':
            return self._build_recommendations(data, user)
        else:
            return self._build_insights(data)


# ═══════════════════════════════════════════════════════════════════════════
#  smart_actions — AI Automation & Smart Actions
# ═══════════════════════════════════════════════════════════════════════════

@register
class SmartActions(BaseAction):
    """Execute multiple CRM actions from a single natural language request.

    Detects multi-action intents (meeting with someone, lead + follow-up,
    meeting prep, workflow creation, campaign launch), parses entities,
    and runs a pipeline of existing CRUD actions via synthetic text.
    Returns a piped success summary with smart suggestions.
    """

    action_type = 'smart_actions'
    keywords = frozenset({
        'meeting', 'meet', 'follow up', 'follow-up', 'followup',
        'prepare', 'launch', 'launches', 'notify', 'when', 'remind',
        'and', 'then', 'also',
    })
    patterns = [
        re.compile(r'(?:meeting|meet)\s+with\s+\w+'),
        re.compile(r'(?:create|add|new)\s+.*\b(lead|contact)\b.*\b(?:and|then)\b'),
        re.compile(r'(?:prepare|ready)\s+for\s+(?:tomorrow|today|the|this)'),
        re.compile(r'(?:when|if)\s+.*\b(lead|contact|task)\b.*\b(?:created|updated|completed|high)\b'),
        re.compile(r'(?:launch|create)\s+.*\bcampaign\b.*\b(?:and|notify)\b'),
        re.compile(r'\band\s+(?:create|schedule|notify|send|then)'),
        re.compile(r'\bremind\s+me\b.*\b(?:after|when|in)\b'),
    ]

    # ── Pipeline definitions ──
    # Each pipeline is a list of dicts:
    #   label       — short label like "Calendar Event"
    #   action_cls  — class to instantiate
    #   build_text  — callable(original_text) -> synthetic text
    #   check       — optional callable(results_so_far) -> bool (skip if False)

    @staticmethod
    def _pipeline_meeting(text, user):
        """Meeting with someone → Event + Notification + suggest Task."""
        steps = []

        # Extract person name
        m = re.search(r'(?:meeting|meet)\s+with\s+(\w+(?:\s+\w+)?)', text, re.IGNORECASE)
        person = m.group(1).strip() if m else 'the contact'

        # Strip "I have a / I've got a / I've a / etc" from the original to
        # produce clean synthetic text for each sub-action.
        clean = re.sub(
            r'^(?:i\s+(?:have|got|have\s+got)\s+a\s+|i\'?v?e?\s+got\s+a\s+)',
            '', text, flags=re.IGNORECASE,
        ).strip()

        event_text = clean
        # If "meeting with X" already present, ensure the target is present
        # for CreateEventAction's parser
        if 'schedule' not in event_text.lower() and 'create' not in event_text.lower():
            event_text = 'schedule ' + event_text

        steps.append({
            'label': 'Calendar Event',
            'action_cls': CreateEventAction,
            'build_text': lambda orig, t=event_text: t,
        })

        # Build notification text from extracted info
        notif_title = f'Reminder: Meeting with {person}'
        notif_text = (
            f'create a notification titled "{notif_title}" '
            f'message "Don\'t forget your meeting with {person}."'
        )
        steps.append({
            'label': 'Reminder Notification',
            'action_cls': CreateNotificationAction,
            'build_text': lambda orig, n=notif_text: n,
        })

        suggestion = f"💡 Would you like me to schedule a follow-up task regarding **{person}**?"
        return steps, suggestion

    @staticmethod
    def _pipeline_lead_followup(text, user):
        """Create lead + follow-up task."""
        # Extract lead name: everything after "lead" prefix and before "and"
        body = re.sub(
            r'.*?\b(?:lead|prospect|deal)\s+',
            '', text, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'\s+and\s+.*$', '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'^(?:for|named|called|about)\s+', '', body, flags=re.IGNORECASE,
        ).strip()
        lead_name = _extract_title(body) or 'New Lead'

        # Clean original text for lead creation: strip everything after "and"
        lead_text = re.sub(
            r'\s+and\s+.*$', '', text, flags=re.IGNORECASE,
        ).strip()
        if not re.search(r'\b(?:lead|prospect|deal)\b', lead_text, re.IGNORECASE):
            lead_text = lead_text + ' lead'
        # Ensure "create" prefix present
        if 'create' not in lead_text.lower() and 'add' not in lead_text.lower() and 'new' not in lead_text.lower():
            lead_text = 'create a ' + lead_text

        steps = [
            {
                'label': 'Lead',
                'action_cls': CreateLeadAction,
                'build_text': lambda orig, lt=lead_text: lt,
            },
            {
                'label': 'Follow-up Task',
                'action_cls': CreateTaskAction,
                'build_text': lambda orig: (
                    f'create a task: follow up with {lead_name} '
                    f'priority high'
                ),
            },
        ]

        suggestion = f"💡 Lead **{lead_name}** created. Would you like me to set a reminder to follow up in 3 days?"
        return steps, suggestion

    @staticmethod
    def _pipeline_meeting_prep(text, user):
        """Prepare for tomorrow's meeting → find meeting + create agenda + remind."""
        notif_text = (
            'create a notification titled "Prepare for tomorrow\'s meeting" '
            'message "Get ready for your meeting tomorrow. Review agenda and prepare notes."'
        )

        steps = [
            {
                'label': 'Find Meeting',
                'action_cls': None,  # handled inline
                'build_text': None,
                '_inline_result': None,
            },
            {
                'label': 'Reminder Notification',
                'action_cls': CreateNotificationAction,
                'build_text': lambda orig, n=notif_text: n,
            },
        ]

        # Inline: find tomorrow's meetings
        from datetime import date, timedelta
        tomorrow = date.today() + timedelta(days=1)
        meetings = list(user.events.filter(
            start_date=tomorrow, event_type='meeting', status='scheduled',
        ).values('title', 'start_time', 'location')[:3])

        suggestion = ''
        if not meetings:
            meeting_info = "I didn't find any meetings scheduled for tomorrow."
        else:
            lines = []
            for m in meetings:
                time_str = f' at {m["start_time"]}' if m['start_time'] else ''
                loc_str = f' — {m["location"]}' if m['location'] else ''
                lines.append(f'- **{m["title"]}**{time_str}{loc_str}')
            meeting_info = 'Found tomorrow\'s meetings:\n' + '\n'.join(lines)
            # Add agenda generation as extra result
            agenda_items = []
            for m in meetings:
                agenda_items.append(f'- {m["title"]}: Discuss agenda, review progress, action items')
            meeting_info += '\n\nSuggested Agenda:\n' + '\n'.join(agenda_items[:3])
            suggestion = '💡 Would you like me to send this agenda to attendees?'

        steps[0]['_inline_result'] = meeting_info

        return steps, suggestion

    @staticmethod
    def _pipeline_workflow(text, user):
        """When X happens, remind me → create entity + create workflow."""
        text_lower = text.lower()

        # Detect entity type
        entity = 'lead'  # default
        if re.search(r'\bcontact\b', text_lower):
            entity = 'contact'
        elif re.search(r'\btask\b', text_lower):
            entity = 'task'
        elif re.search(r'\bcampaign\b', text_lower):
            entity = 'campaign'

        trigger_map = {
            'lead': 'lead_created',
            'contact': 'contact_created',
            'task': 'task_created',
            'campaign': 'campaign_created',
        }

        # Extract entity name from text
        body = re.sub(
            r'(?:when|if)\s+(?:a\s+|an\s+)?(?:new\s+)?(?:high\s+priority\s+)?'
            r'(?:lead|contact|task|campaign).*?'
            r'(?:is\s+)?(?:created|updated|completed)\s*,?\s*',
            '', text, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'\s+remind\s+me\s+.*$', '', body, flags=re.IGNORECASE).strip()
        entity_name = _extract_title(body) or f'New {entity.title()}'

        steps = []
        trigger_slug = trigger_map[entity]

        # Create the entity
        if entity == 'lead':
            lead_text = f'create a lead named {entity_name} priority high'
            steps.append({
                'label': 'Lead',
                'action_cls': CreateLeadAction,
                'build_text': lambda orig, lt=lead_text: lt,
            })
        elif entity == 'contact':
            contact_text = f'create a contact named {entity_name}'
            steps.append({
                'label': 'Contact',
                'action_cls': CreateContactAction,
                'build_text': lambda orig, ct=contact_text: ct,
            })
        elif entity == 'task':
            task_text = f'create a task {entity_name} priority high'
            steps.append({
                'label': 'Task',
                'action_cls': CreateTaskAction,
                'build_text': lambda orig, tt=task_text: tt,
            })
        elif entity == 'campaign':
            campaign_text = f'create a campaign named {entity_name}'
            steps.append({
                'label': 'Campaign',
                'action_cls': CreateCampaignAction,
                'build_text': lambda orig, ct=campaign_text: ct,
            })

        # Create Workflow
        wf_name = trigger_slug.replace('_', ' ').title() + ' Reminder'
        wf_text = (
            f'create a workflow named "{wf_name}" '
            f'trigger: {trigger_slug} '
            f'actions: create_notification'
        )
        steps.append({
            'label': 'Workflow',
            'action_cls': CreateWorkflowAction,
            'build_text': lambda orig, wt=wf_text: wt,
        })

        # Create notification
        notif_text = (
            f'create a notification titled "Workflow Set Up for {entity_name}" '
            f'message "Automation created: you will be notified when this {entity} is updated."'
        )
        steps.append({
            'label': 'Notification',
            'action_cls': CreateNotificationAction,
            'build_text': lambda orig, nt=notif_text: nt,
        })

        suggestion = f'The workflow will automatically notify you when this {entity} is updated or created.'
        return steps, suggestion

    @staticmethod
    def _pipeline_campaign(text, user):
        """Launch campaign + notify → create campaign + workflow + notification."""
        body = re.sub(
            r'.*?\b(?:campaign)\s+',
            '', text, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'\s+and\s+notify\s+me\s+.*$', '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(
            r'^(?:for|named|called|about)\s+', '', body, flags=re.IGNORECASE,
        ).strip()
        campaign_name = _extract_title(body) or 'New Campaign'

        # Clean campaign text: strip "and notify me..." part
        campaign_text = re.sub(
            r'\s+and\s+notify\s+me\s+.*$', '', text, flags=re.IGNORECASE,
        ).strip()
        if 'create' not in campaign_text.lower() and 'launch' not in campaign_text.lower():
            campaign_text = 'create ' + campaign_text
        if 'campaign' not in campaign_text.lower():
            campaign_text = campaign_text + ' campaign'

        wf_text = (
            'create a workflow named "Campaign Completion Notifier" '
            'trigger: campaign_completed '
            'actions: create_notification'
        )
        notif_text = (
            f'create a notification titled "Campaign Launched: {campaign_name}" '
            f'message "Your campaign has been launched. You will be notified when it completes."'
        )

        steps = [
            {
                'label': 'Campaign',
                'action_cls': CreateCampaignAction,
                'build_text': lambda orig, ct=campaign_text: ct,
            },
            {
                'label': 'Workflow',
                'action_cls': CreateWorkflowAction,
                'build_text': lambda orig, wt=wf_text: wt,
            },
            {
                'label': 'Notification',
                'action_cls': CreateNotificationAction,
                'build_text': lambda orig, nt=notif_text: nt,
            },
        ]

        suggestion = f'Campaign **{campaign_name}** is live! I will notify you when it finishes.'
        return steps, suggestion

    # ── Workflow detection ──

    _WORKFLOW_DETECTORS = [
        ('meeting_pipeline', lambda t: bool(re.search(r'(?:meeting|meet)\s+with\s+\w+', t, re.IGNORECASE))),
        ('meeting_prep', lambda t: bool(re.search(r'(?:prepare|ready)\s+for\s+(?:tomorrow|today|\bthe\b|\bthis\b)', t, re.IGNORECASE))),
        ('lead_followup', lambda t: (
            bool(re.search(r'(?:create|add|new).*\b(lead|prospect|deal)\b', t, re.IGNORECASE))
            and bool(re.search(r'\b(?:and|then)\b.*\b(?:follow|schedule|task)\b', t, re.IGNORECASE))
        )),
        ('workflow', lambda t: bool(re.search(r'(?:when|if)\s+.*\b(?:lead|contact|task|campaign)\b.*\b(?:created|updated|completed|high)\b', t, re.IGNORECASE))),
        ('campaign', lambda t: bool(re.search(r'(?:launch|create)\s+.*\bcampaign\b.*\b(?:and|notify)\b', t, re.IGNORECASE))),
    ]

    @staticmethod
    def _detect_workflow(text_lower):
        for name, detector in SmartActions._WORKFLOW_DETECTORS:
            if detector(text_lower):
                return name
        return None

    # ── Pipeline runner ──

    @staticmethod
    def _run_pipeline(steps, text, user):
        """Execute a list of steps. Each step calls the action class's
        execute method with the synthetic text built from *text*."""
        results = []
        for step in steps:
            action_cls = step['action_cls']
            build_text = step['build_text']
            inline_result = step.get('_inline_result')

            if action_cls is None and inline_result:
                results.append((step['label'], True, inline_result))
                continue

            try:
                action = action_cls()
                synthetic = build_text(text)
                result = action.execute(synthetic, user)
                results.append((step['label'], True, result))
            except Exception as e:
                logger.exception('SmartActions step %s failed', step['label'])
                results.append((step['label'], False, str(e)))

        return results

    # ── Response formatter ──

    @staticmethod
    def _format_response(results, suggestion):
        lines = []
        for label, success, msg in results:
            icon = '✅' if success else '❌'
            lines.append(f'{icon} {label}')
            if success:
                # Extract just the first line of the result for brevity
                first_line = msg.split('\n')[0][:120]
                lines.append(f'   {first_line}')
            else:
                lines.append(f'   Error: {msg[:120]}')
        lines.append('')

        all_ok = all(success for _, success, _ in results)
        if all_ok:
            lines.append('**Everything completed successfully.**')
        else:
            lines.append('**Some steps failed. See details above.**')

        if suggestion:
            lines.append('')
            lines.append(f'💡 {suggestion}')

        return '\n'.join(lines)

    # ── Entry point ──

    def execute(self, text, user):
        if not user or not user.is_authenticated:
            return None

        text_lower = text.lower()
        wf = self._detect_workflow(text_lower)
        if wf is None:
            return None

        pipeline_map = {
            'meeting_pipeline': self._pipeline_meeting,
            'lead_followup': self._pipeline_lead_followup,
            'meeting_prep': self._pipeline_meeting_prep,
            'workflow': self._pipeline_workflow,
            'campaign': self._pipeline_campaign,
        }

        builder = pipeline_map.get(wf)
        if builder is None:
            return None

        steps, suggestion = builder(text, user)
        results = self._run_pipeline(steps, text, user)
        return self._format_response(results, suggestion)


# ═══════════════════════════════════════════════════════════════════════════
#  crm_reports — AI CRM Reports (CSV / Excel / PDF)
# ═══════════════════════════════════════════════════════════════════════════

@register
class CrmReportsAction(BaseAction):
    """Generate professional CRM reports from natural language.

    Supported formats:
      - CSV   — lightweight, universally openable
      - XLSX  — formatted Excel workbook
      - PDF   — printable report via reportlab

    Entity types: contacts, leads, tasks, events, campaigns,
    workflows, notifications, summary, full/all.
    """

    action_type = 'crm_reports'
    keywords = frozenset({
        'export', 'report', 'download', 'generate', 'print',
        'summary', 'sales', 'csv', 'pdf', 'excel', 'xlsx',
    })
    patterns = [
        re.compile(r'\b(?:export|generate|download|print)\s+(?:all\s+)?(?:contacts|leads|tasks?|events?|meetings?|campaigns?|workflows?|notifications?)'),
        re.compile(r'\b(?:export|generate|download|print)\s+(?:(?:complete|full|all)\s+)?crm\b'),
        re.compile(r'\b(?:generate|create)\s+(?:a\s+)?(?:crm\s+)?(?:summary|sales)\s+report\b'),
        re.compile(r'\b(?:export|save|download)\s+(?:as|to|in)\s+(?:pdf|csv|excel|xlsx)\b'),
        re.compile(r'\b(?:today\'?s?\s+)?(?:tasks?|events?|meetings?)\s+(?:as|to)?\s*(?:pdf|csv|excel|xlsx)'),
        re.compile(r'\bacitve\s+campaigns?\b'),
    ]

    _ENTITY_ALIASES = {
        'contacts': 'contacts', 'contact': 'contacts',
        'leads': 'leads', 'lead': 'leads',
        'tasks': 'tasks', 'task': 'tasks',
        'events': 'events', 'event': 'events',
        'meetings': 'events', 'meeting': 'events',
        'campaigns': 'campaigns', 'campaign': 'campaigns',
        'workflows': 'workflows', 'workflow': 'workflows',
        'notifications': 'notifications', 'notification': 'notifications',
        'summary': 'summary', 'sales': 'summary',
        'complete': 'full', 'full': 'full', 'all': 'full',
    }

    _FORMAT_ALIASES = {
        'csv': 'csv', 'xlsx': 'xlsx', 'excel': 'xlsx',
        'pdf': 'pdf',
    }

    # ── Detection helpers ──

    @staticmethod
    def _detect_entity(text_lower):
        """Return entity slug or None."""
        # Check for summary / sales report first
        if re.search(r'\b(summary|sales)\s+report\b', text_lower):
            return 'summary'
        if re.search(r'\b(?:complete|full|all)\s+crm\b', text_lower) or text_lower.strip() == 'export crm':
            return 'full'

        # Check for active campaigns
        if re.search(r'\bactive\s+campaigns?\b', text_lower):
            return 'campaigns'

        # Check for today's tasks / events
        if re.search(r'\btoday\'?s?\s+(?:tasks?|events?|meetings?)\b', text_lower):
            entity_match = re.search(r'\b(tasks?|events?|meetings?)\b', text_lower)
            word = entity_match.group(1)
            return CrmReportsAction._ENTITY_ALIASES.get(word, 'tasks')

        # General entity detection
        for word, slug in sorted(
            CrmReportsAction._ENTITY_ALIASES.items(),
            key=lambda x: -len(x[0]),
        ):
            if (
                word in text_lower
                and slug not in ('summary', 'full')
            ):
                # Prefer more specific match — don't match if the word
                # is inside "complete crm" or "summary report"
                if re.search(r'\b' + re.escape(word) + r'\b', text_lower):
                    return slug

        # If only "export" / "generate" / "report" is found, default to summary
        if re.search(r'\b(?:export|report|generate|print)\b', text_lower):
            return 'summary'

        return None

    @staticmethod
    def _detect_format(text_lower):
        """Return format slug or default to pdf."""
        for fmt in ('pdf', 'xlsx', 'excel', 'csv'):
            if fmt in text_lower:
                return CrmReportsAction._FORMAT_ALIASES.get(fmt, 'pdf')
        return 'pdf'

    # ── Data queries ──

    @staticmethod
    def _query_contacts(user, extra_filters=None):
        qs = user.contacts.all().order_by('full_name')
        if extra_filters:
            qs = qs.filter(**extra_filters)
        return list(qs.values(
            'full_name', 'email', 'phone', 'company',
            'job_title', 'tags', 'created_at',
        ))

    @staticmethod
    def _query_leads(user, extra_filters=None):
        qs = user.leads.all().order_by('lead_name')
        if extra_filters:
            qs = qs.filter(**extra_filters)
        return list(qs.values(
            'lead_name', 'email', 'phone', 'priority',
            'status', 'expected_revenue', 'created_at',
        ))

    @staticmethod
    def _query_tasks(user, extra_filters=None):
        qs = user.tasks.all().order_by('-created_at')
        if extra_filters:
            qs = qs.filter(**extra_filters)
        results = []
        for t in qs.select_related('contact'):
            results.append({
                'title': t.title,
                'priority': t.priority,
                'status': t.status,
                'due_date': t.due_date,
                'contact_name': t.contact.full_name if t.contact else '',
            })
        return results

    @staticmethod
    def _query_events(user, extra_filters=None):
        qs = user.events.all().order_by('-start_date')
        if extra_filters:
            qs = qs.filter(**extra_filters)
        return list(qs.values(
            'title', 'start_date', 'start_time', 'location',
            'status', 'event_type',
        ))

    @staticmethod
    def _query_campaigns(user, extra_filters=None):
        qs = user.campaigns.all().order_by('-created_at')
        if extra_filters:
            qs = qs.filter(**extra_filters)
        return list(qs.values(
            'name', 'status', 'scheduled_at', 'created_at',
        ))

    @staticmethod
    def _query_workflows(user, extra_filters=None):
        qs = user.workflows.all().order_by('-created_at')
        if extra_filters:
            qs = qs.filter(**extra_filters)
        results = []
        for w in qs.prefetch_related('actions'):
            action_names = [a.get_action_type_display() for a in w.actions.all()]
            results.append({
                'name': w.name,
                'trigger_type': w.get_trigger_type_display(),
                'is_active': w.is_active,
                'actions': ', '.join(action_names) if action_names else '—',
            })
        return results

    @staticmethod
    def _query_notifications(user, extra_filters=None):
        qs = user.notifications.all().order_by('-created_at')
        if extra_filters:
            qs = qs.filter(**extra_filters)
        return list(qs.values(
            'title', 'message', 'is_read', 'created_at',
        ))

    # ── File generation ──

    @staticmethod
    def _save_file(filename, content_bytes, user=None):
        """Write *content_bytes* to ``media/reports/<user_id>/`` and return
        the relative URL path. Filenames include a unique short-UUID so
        previous reports are never overwritten."""
        import os
        import uuid
        from django.conf import settings

        # Namespace by user so reports are never exposed across users
        subdir = f'reports/user_{user.pk}' if user else 'reports'
        reports_dir = os.path.join(settings.MEDIA_ROOT, subdir)
        os.makedirs(reports_dir, exist_ok=True)

        # Insert a short unique id before the extension
        stem, ext = os.path.splitext(filename)
        unique_stem = f'{stem}_{uuid.uuid4().hex[:8]}'
        safe_filename = f'{unique_stem}{ext}'

        filepath = os.path.join(reports_dir, safe_filename)
        with open(filepath, 'wb') as f:
            f.write(content_bytes)

        return f'{settings.MEDIA_URL}{subdir}/{safe_filename}'

    def _build_download_url(self, path):
        """Convert a media-relative path to an absolute download URL that
        forces the browser to download the file instead of opening it
        inline."""
        from django.conf import settings
        from django.urls import reverse
        # Strip the MEDIA_URL prefix to get the path under reports/
        relative_path = path
        if relative_path.startswith(settings.MEDIA_URL):
            relative_path = relative_path[len(settings.MEDIA_URL):]
        # Strip the leading 'reports/' segment
        if relative_path.startswith('reports/'):
            relative_path = relative_path[len('reports/'):]
        # Build the download view URL
        dl_view_path = reverse('assistant:report_download') + '?path=' + quote(relative_path)
        if hasattr(self, '_current_request') and self._current_request is not None:
            return self._current_request.build_absolute_uri(dl_view_path)
        return dl_view_path

    @staticmethod
    def _rows_to_csv(rows, field_names, field_labels):
        import csv, io
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(field_labels)
        for row in rows:
            writer.writerow(str(row.get(f, '') or '') for f in field_names)
        return buf.getvalue().encode('utf-8-sig')

    @staticmethod
    def _rows_to_xlsx(rows, field_names, field_labels, sheet_name):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')

        for col_idx, label in enumerate(field_labels, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, row in enumerate(rows, 2):
            for col_idx, field in enumerate(field_names, 1):
                val = row.get(field, '')
                if val is None:
                    val = ''
                ws.cell(row=row_idx, column=col_idx, value=str(val))

        for col_idx in range(1, len(field_labels) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 22

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    @staticmethod
    def _rows_to_pdf(rows, field_names, field_labels, title):
        import io
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet

        import io
        buf = io.BytesIO()
        # Use landscape if many columns
        page_size = landscape(A4) if len(field_labels) > 6 else A4
        doc = SimpleDocTemplate(buf, pagesize=page_size, topMargin=20*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()

        elements = []
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 10*mm))

        header_row = field_labels
        data_rows = []
        for row in rows:
            data_rows.append([str(row.get(f, '') or '') for f in field_names])

        table_data = [header_row] + data_rows
        col_count = len(field_labels)
        available_width = page_size[0] - 40*mm
        col_width = available_width / max(col_count, 1)

        table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ]))

        # Handle page-break for many rows
        max_rows_per_page = 40
        if len(table_data) > max_rows_per_page + 1:
            elements.append(table)
        else:
            elements.append(table)

        doc.build(elements)
        buf.seek(0)
        return buf.getvalue()

    # ── Summary data ──

    @staticmethod
    def _build_summary_data(user):
        from datetime import date
        from django.db import models
        today = date.today()
        contacts_count = user.contacts.count()
        leads_count = user.leads.count()
        leads_high = user.leads.filter(priority__in=['High', 'Urgent']).count()
        total_revenue = (
            user.leads.aggregate(total=models.Sum('expected_revenue'))['total'] or 0
        )
        tasks_total = user.tasks.count()
        tasks_completed = user.tasks.filter(status='completed').count()
        tasks_pending = user.tasks.filter(status='pending').count()
        events_total = user.events.filter(start_date__gte=today).count()
        campaigns_total = user.campaigns.count()
        workflows_active = user.workflows.filter(is_active=True).count()
        notifications_unread = user.notifications.filter(is_read=False).count()
        return {
            'contacts_count': contacts_count,
            'leads_count': leads_count,
            'leads_high': leads_high,
            'total_revenue': total_revenue,
            'tasks_total': tasks_total,
            'tasks_completed': tasks_completed,
            'tasks_pending': tasks_pending,
            'events_total': events_total,
            'campaigns_total': campaigns_total,
            'workflows_active': workflows_active,
            'notifications_unread': notifications_unread,
        }

    @staticmethod
    def _generate_summary_report(user, fmt):
        from datetime import datetime, date
        data = CrmReportsAction._build_summary_data(user)
        today_str = datetime.now().strftime('%Y%m%d_%H%M%S')

        if fmt == 'csv':
            rows = [
                {'metric': 'Contacts', 'value': data['contacts_count']},
                {'metric': 'Leads', 'value': data['leads_count']},
                {'metric': 'High Priority Leads', 'value': data['leads_high']},
                {'metric': 'Expected Revenue', 'value': f'${data["total_revenue"]:,.2f}'},
                {'metric': 'Tasks', 'value': data['tasks_total']},
                {'metric': 'Tasks Completed', 'value': data['tasks_completed']},
                {'metric': 'Tasks Pending', 'value': data['tasks_pending']},
                {'metric': 'Upcoming Events', 'value': data['events_total']},
                {'metric': 'Campaigns', 'value': data['campaigns_total']},
                {'metric': 'Active Workflows', 'value': data['workflows_active']},
                {'metric': 'Unread Notifications', 'value': data['notifications_unread']},
            ]
            field_names = ['metric', 'value']
            field_labels = ['Metric', 'Value']
            content = CrmReportsAction._rows_to_csv(rows, field_names, field_labels)
            filename = f'crm_summary_{today_str}.csv'
            url = CrmReportsAction._save_file(filename, content, user)
            return url, f'**CRM Summary Report** exported as CSV.'

        elif fmt == 'xlsx':
            rows = [
                {'metric': 'Contacts', 'value': data['contacts_count']},
                {'metric': 'Leads', 'value': data['leads_count']},
                {'metric': 'High Priority Leads', 'value': data['leads_high']},
                {'metric': 'Expected Revenue', 'value': f'${data["total_revenue"]:,.2f}'},
                {'metric': 'Tasks', 'value': data['tasks_total']},
                {'metric': 'Tasks Completed', 'value': data['tasks_completed']},
                {'metric': 'Tasks Pending', 'value': data['tasks_pending']},
                {'metric': 'Upcoming Events', 'value': data['events_total']},
                {'metric': 'Campaigns', 'value': data['campaigns_total']},
                {'metric': 'Active Workflows', 'value': data['workflows_active']},
                {'metric': 'Unread Notifications', 'value': data['notifications_unread']},
            ]
            field_names = ['metric', 'value']
            field_labels = ['Metric', 'Value']
            content = CrmReportsAction._rows_to_xlsx(rows, field_names, field_labels, 'Summary')
            filename = f'crm_summary_{today_str}.xlsx'
            url = CrmReportsAction._save_file(filename, content, user)
            return url, f'**CRM Summary Report** exported as Excel.'

        else:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
            )
            from reportlab.lib.styles import getSampleStyleSheet

            import io
            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm)
            styles = getSampleStyleSheet()

            elements = []
            elements.append(Paragraph('CRM Summary Report', styles['Title']))
            elements.append(Paragraph(f'Generated: {date.today().strftime("%B %d, %Y")}', styles['Normal']))
            elements.append(Spacer(1, 8*mm))

            summary_items = [
                ['Metric', 'Value'],
                ['Contacts', str(data['contacts_count'])],
                ['Leads', str(data['leads_count'])],
                ['High Priority Leads', str(data['leads_high'])],
                ['Expected Revenue', f'${data["total_revenue"]:,.2f}'],
                ['Tasks', str(data['tasks_total'])],
                ['Tasks Completed', str(data['tasks_completed'])],
                ['Tasks Pending', str(data['tasks_pending'])],
                ['Upcoming Events', str(data['events_total'])],
                ['Campaigns', str(data['campaigns_total'])],
                ['Active Workflows', str(data['workflows_active'])],
                ['Unread Notifications', str(data['notifications_unread'])],
            ]
            t = Table(summary_items, colWidths=[120*mm, 80*mm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ]))
            elements.append(t)

            doc.build(elements)
            buf.seek(0)
            content = buf.getvalue()
            filename = f'crm_summary_{today_str}.pdf'
            url = CrmReportsAction._save_file(filename, content, user)
            return url, f'**CRM Summary Report** exported as PDF.'

    # ── Per-entity report generators ──

    _ENTITY_REPORT = {
        'contacts': {
            'label': 'Contacts',
            'query_method': '_query_contacts',
            'fields': ['full_name', 'email', 'phone', 'company', 'job_title', 'tags', 'created_at'],
            'headers': ['Name', 'Email', 'Phone', 'Company', 'Position', 'Tags', 'Created Date'],
        },
        'leads': {
            'label': 'Leads',
            'query_method': '_query_leads',
            'fields': ['lead_name', 'email', 'phone', 'priority', 'status', 'expected_revenue', 'created_at'],
            'headers': ['Name', 'Email', 'Phone', 'Priority', 'Status', 'Expected Revenue', 'Created Date'],
        },
        'tasks': {
            'label': 'Tasks',
            'query_method': '_query_tasks',
            'fields': ['title', 'priority', 'status', 'due_date', 'contact_name'],
            'headers': ['Title', 'Priority', 'Status', 'Due Date', 'Contact'],
        },
        'events': {
            'label': 'Events',
            'query_method': '_query_events',
            'fields': ['title', 'start_date', 'start_time', 'location', 'status', 'event_type'],
            'headers': ['Title', 'Date', 'Time', 'Location', 'Status', 'Type'],
        },
        'campaigns': {
            'label': 'Campaigns',
            'query_method': '_query_campaigns',
            'fields': ['name', 'status', 'scheduled_at', 'created_at'],
            'headers': ['Name', 'Status', 'Scheduled At', 'Created Date'],
        },
        'workflows': {
            'label': 'Workflows',
            'query_method': '_query_workflows',
            'fields': ['name', 'trigger_type', 'is_active', 'actions'],
            'headers': ['Name', 'Trigger', 'Active', 'Actions'],
        },
        'notifications': {
            'label': 'Notifications',
            'query_method': '_query_notifications',
            'fields': ['title', 'message', 'is_read', 'created_at'],
            'headers': ['Title', 'Message', 'Read', 'Created Date'],
        },
    }

    @staticmethod
    def _generate_entity_report(entity, user, fmt, extra_filters=None):
        from datetime import datetime
        meta = CrmReportsAction._ENTITY_REPORT.get(entity)
        if not meta:
            return None, f'Unknown entity: {entity}'

        query_method_name = meta['query_method']
        query_method = getattr(CrmReportsAction, query_method_name)
        rows = query_method(user, extra_filters)
        field_names = meta['fields']
        field_labels = meta['headers']
        label = meta['label']
        today_str = datetime.now().strftime('%Y%m%d_%H%M%S')

        if not rows:
            return None, f'No **{label.lower()}** found to export.'

        if fmt == 'csv':
            content = CrmReportsAction._rows_to_csv(rows, field_names, field_labels)
            filename = f'{entity}_{today_str}.csv'
        elif fmt == 'xlsx':
            content = CrmReportsAction._rows_to_xlsx(rows, field_names, field_labels, label)
            filename = f'{entity}_{today_str}.xlsx'
        else:
            content = CrmReportsAction._rows_to_pdf(rows, field_names, field_labels, f'{label} Report')
            filename = f'{entity}_{today_str}.pdf'

        url = CrmReportsAction._save_file(filename, content, user)
        fmt_display = fmt.upper()
        return url, f'**{label}** report exported as {fmt_display}.'

    @staticmethod
    def _generate_full_report(user, fmt):
        from datetime import datetime
        today_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        entities = ['contacts', 'leads', 'tasks', 'events', 'campaigns', 'workflows', 'notifications']

        if fmt == 'xlsx':
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook()
            wb.remove(wb.active)

            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center')

            for entity in entities:
                meta = CrmReportsAction._ENTITY_REPORT[entity]
                query_method = getattr(CrmReportsAction, meta['query_method'])
                rows = query_method(user)
                field_names = meta['fields']
                field_labels = meta['headers']
                label = meta['label']

                ws = wb.create_sheet(title=label[:31])
                for col_idx, hl in enumerate(field_labels, 1):
                    cell = ws.cell(row=1, column=col_idx, value=hl)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align

                for row_idx, row in enumerate(rows, 2):
                    for col_idx, fn in enumerate(field_names, 1):
                        val = row.get(fn, '')
                        if val is None:
                            val = ''
                        ws.cell(row=row_idx, column=col_idx, value=str(val))

                for col_idx in range(1, len(field_labels) + 1):
                    ws.column_dimensions[chr(64 + col_idx)].width = 22

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            content = buf.getvalue()
            filename = f'full_crm_{today_str}.xlsx'
            url = CrmReportsAction._save_file(filename, content, user)
            return url, '**Complete CRM** exported as Excel with 7 sheets.'

        # For PDF and CSV: generate one file per entity and return the last
        if fmt == 'csv':
            for entity in entities:
                meta = CrmReportsAction._ENTITY_REPORT[entity]
                query_method = getattr(CrmReportsAction, meta['query_method'])
                rows = query_method(user)
                field_names = meta['fields']
                field_labels = meta['headers']
                content = CrmReportsAction._rows_to_csv(rows, field_names, field_labels)
                filename = f'{entity}_{today_str}.csv'
                CrmReportsAction._save_file(filename, content, user)
            msg = '**Complete CRM** exported as 7 CSV files in the reports folder.'
            return None, msg

        if fmt == 'pdf':
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
            )
            from reportlab.lib.styles import getSampleStyleSheet

            import io
            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm)
            styles = getSampleStyleSheet()
            elements = []

            for entity in entities:
                meta = CrmReportsAction._ENTITY_REPORT[entity]
                query_method = getattr(CrmReportsAction, meta['query_method'])
                rows = query_method(user)
                field_names = meta['fields']
                field_labels = meta['headers']
                label = meta['label']
                if elements:
                    elements.append(PageBreak())
                elements.append(Paragraph(f'{label} Report', styles['Heading1']))
                elements.append(Spacer(1, 5*mm))

                header_row = field_labels
                data_rows = []
                for row in rows:
                    data_rows.append([str(row.get(f, '') or '') for f in field_names])

                table_data = [header_row] + data_rows
                col_count = len(field_labels)
                available_width = A4[0] - 40*mm
                col_width = max(available_width / max(col_count, 1), 25*mm)

                if col_count > 5:
                    # Need landscape
                    doc.pagesize = landscape(A4)
                    available_width = landscape(A4)[0] - 40*mm
                    col_width = available_width / col_count

                t = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
                ]))
                elements.append(t)

            doc.build(elements)
            buf.seek(0)
            content = buf.getvalue()
            filename = f'full_crm_{today_str}.pdf'
            url = CrmReportsAction._save_file(filename, content, user)
            return url, '**Complete CRM** exported as PDF with 7 sections.'

    # ── Response formatter ──

    @staticmethod
    def _parse_filename_from_url(url):
        """Extract the filename from a relative media URL."""
        if not url or '/' not in url:
            return 'report'
        return url.rsplit('/')[-1]

    @staticmethod
    def _fmt_label(entity):
        return entity.replace('_', ' ').title()

    def _format_download_response(self, url, msg, entity, fmt):
        """Return a rich download response with absolute URL, filename and
        metadata."""
        if not url:
            return f'❌ Failed to generate report.\n\n{msg}' if msg else '❌ Failed to generate report.'

        absolute_url = self._build_download_url(url)
        filename = self._parse_filename_from_url(url)

        fmt_icons = {'pdf': '📄', 'xlsx': '📊', 'csv': '📁'}
        icon = fmt_icons.get(fmt, '📄')
        fmt_upper = fmt.upper()
        label = self._fmt_label(entity)

        from datetime import datetime
        now_str = datetime.now().strftime('%B %d, %Y at %I:%M %p').lstrip('0')

        lines = [
            f'✅ {label} Report Generated Successfully',
            '',
            f'**File Name:**',
            f'{filename}',
            '',
            f'**Format:**',
            f'{icon} {fmt_upper} (.{fmt})',
            '',
            f'**Generated:**',
            f'{now_str}',
            '',
            f'**Download:**',
            f'{icon} [⬇ Download {label} Report]({absolute_url})',
            '',
            f'---',
            f'*Secure link — only you can access this report.*',
        ]
        return '\n'.join(lines)

    # ── Entry point ──

    def execute(self, text, user):
        if not user or not user.is_authenticated:
            return None

        text_lower = text.lower()
        entity = self._detect_entity(text_lower)
        if entity is None:
            return None

        fmt = self._detect_format(text_lower)

        from django.db import models as django_models
        models = django_models  # for .Sum in summary

        try:
            if entity == 'summary':
                url, msg = self._generate_summary_report(user, fmt)
                return self._format_download_response(url, msg, entity, fmt)

            if entity == 'full':
                url, msg = self._generate_full_report(user, fmt)
                return self._format_download_response(url, msg, entity, fmt)

            # Apply time-based filters
            extra_filters = {}
            if re.search(r'\btoday\'?s?\b', text_lower):
                from datetime import date
                today = date.today()
                if entity == 'tasks':
                    extra_filters['due_date'] = today
                elif entity == 'events':
                    extra_filters['start_date'] = today
            elif re.search(r'\bthis\s+month\'?s?\b', text_lower):
                from datetime import date
                today = date.today()
                if entity == 'events':
                    extra_filters['start_date__year'] = today.year
                    extra_filters['start_date__month'] = today.month
                elif entity == 'tasks':
                    extra_filters['due_date__year'] = today.year
                    extra_filters['due_date__month'] = today.month
            elif re.search(r'\bactive\b', text_lower) and entity == 'campaigns':
                extra_filters['status'] = 'Scheduled'

            url, msg = self._generate_entity_report(entity, user, fmt, extra_filters)
            return self._format_download_response(url, msg, entity, fmt)

        except Exception as e:
            logger.exception('CrmReportsAction failed for user=%s', user)
            return f'❌ Failed to generate report.\n\n**Reason:** {e}'


class CreateCompanyAction(BaseAction):
    action_type = 'create_company'
    keywords = frozenset({'create', 'add', 'new', 'company', 'companies', 'organization', 'org'})
    patterns = [re.compile(r'(create|add|new).*(compan|organi)')]

    def execute(self, text, user):
        from companies.models import Company

        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:create|add|new)\s+(?:a\s+|an\s+|the\s+)?(?:company|organization|org)\s+(?:called\s+|named\s+)?',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'\s*company\s*$', '', body, flags=re.IGNORECASE).strip()
        name = body.strip().rstrip('.,!;')

        if not name or len(name) < 2:
            return 'Please specify a company name, e.g. "Create a company called Microsoft".'

        if Company.objects.filter(name__iexact=name, owner=user).exists():
            return f'A company named "{name}" already exists.'

        company = Company.objects.create(owner=user, name=name)
        from workflows.services.engine import fire_trigger
        fire_trigger('company_created', company)
        return (
            f'✅ **Company created:** {name}\n\n'
            f'You can now add contacts, leads, and deals to this company.'
        )


class UpdateCompanyAction(BaseAction):
    action_type = 'update_company'
    keywords = frozenset({
        'update', 'edit', 'change', 'modify', 'set', 'rename',
        'company', 'companies', 'organization',
    })
    patterns = [
        re.compile(r'\b(update|edit|change|modify|set|rename)\b.*(compan|organi)'),
        re.compile(r'(compan|organi).*(?:name|website|email|phone|industry|status)\s+(?:to|as)'),
    ]

    _FIELD_MAP = {
        'name': 'name', 'company': 'name',
        'website': 'website', 'site': 'website', 'url': 'website',
        'email': 'email', 'mail': 'email',
        'phone': 'phone', 'telephone': 'phone', 'mobile': 'phone',
        'industry': 'industry',
        'status': 'status',
        'description': 'description',
        'city': 'city', 'state': 'state', 'country': 'country',
        'employees': 'employees',
    }

    def _extract_updates(self, text):
        updates = {}
        text_lower = text.lower()

        patterns = [
            (r'(?:name|rename)\s+(?:to\s+)?["\']?(.+?)["\']?(?:\s+company|\s*$)', 'name'),
            (r'website\s+(?:to\s+|:?\s*)(.+)', 'website'),
            (r'email\s+(?:to\s+|:?\s*)(.+)', 'email'),
            (r'phone\s+(?:to\s+|:?\s*)(.+)', 'phone'),
            (r'industry\s+(?:to\s+|:?\s*)(.+)', 'industry'),
            (r'status\s+(?:to\s+|:?\s*)(.+)', 'status'),
            (r'city\s+(?:to\s+|:?\s*)(.+)', 'city'),
            (r'state\s+(?:to\s+|:?\s*)(.+)', 'state'),
            (r'country\s+(?:to\s+|:?\s*)(.+)', 'country'),
        ]

        for pat, field in patterns:
            m = re.search(pat, text_lower)
            if m:
                val = m.group(1).strip().strip('"\'')
                if val:
                    updates[field] = val

        return updates

    def execute(self, text, user):
        from companies.models import Company

        text_lower = text.lower()

        name_match = re.search(
            r'(?:compan|organi).*?[:"]?\s*(.+?)(?:\s+(?:website|email|phone|industry|status|name)\s+(?:to|as)|\s*$|\.\s*)',
            text_lower,
        )
        company_name = None
        if name_match:
            cand = name_match.group(1).strip().strip('"\'')
            if cand and not any(w in cand for w in ['update', 'edit', 'change', 'modify', 'set', 'rename', 'website', 'email', 'phone', 'industry', 'status']):
                company_name = cand

        if not company_name:
            parts = re.split(r'\s+(?:website|email|phone|industry|status)\s+(?:to|as)\s+', text_lower, maxsplit=1)
            company_name = parts[0].strip()
            company_name = re.sub(
                r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
                r'(?:update|edit|change|modify|set|rename)\s+(?:the\s+)?',
                '', company_name, flags=re.IGNORECASE,
            ).strip()
            company_name = re.sub(r'\s+(?:compan|organi).*$', '', company_name, flags=re.IGNORECASE).strip()

        if not company_name or len(company_name) < 2:
            return 'Please specify which company to update.'

        company = Company.objects.filter(owner=user, name__icontains=company_name).first()
        if not company:
            return f'Company "{company_name}" not found.'

        updates = self._extract_updates(text)
        if not updates:
            return f'What would you like to update for "{company.name}"? Try "Update {company.name} website to https://..."'

        changed = []
        for field, value in updates.items():
            setattr(company, field, value)
            changed.append(f'{field} → {value}')
        company.save()

        from workflows.services.engine import fire_trigger
        fire_trigger('company_updated', company)

        return (
            f'✅ **{company.name}** updated:\n'
            + '\n'.join(f'• {c}' for c in changed)
        )


class DeleteCompanyAction(BaseAction):
    action_type = 'delete_company'
    keywords = frozenset({'delete', 'remove', 'erase', 'cancel', 'company', 'companies', 'organization'})
    patterns = [
        re.compile(r'(delete|remove|erase|cancel)\b.*(compan|organi)'),
    ]

    def execute(self, text, user):
        from companies.models import Company

        body = text
        body = re.sub(
            r'^(?:please\s+)?(?:i\s+(?:want\s+(?:to\s+)?)?)?'
            r'(?:delete|remove|erase|cancel)\s+(?:the\s+|this\s+)?(?:company|organization|org)\s+',
            '', body, flags=re.IGNORECASE,
        ).strip()
        body = re.sub(r'\s*company\s*$', '', body, flags=re.IGNORECASE).strip()
        name = body.strip().rstrip('.,!;')

        if not name or len(name) < 2:
            return 'Please specify which company to delete.'

        company = Company.objects.filter(owner=user, name__icontains=name).first()
        if not company:
            return f'Company "{name}" not found.'

        company_name = company.name
        company.delete()

        from workflows.services.engine import fire_trigger
        fire_trigger('company_deleted', company)

        return f'🗑️ **Company deleted:** {company_name}'


class ViewCompanyAction(BaseAction):
    action_type = 'view_company'
    keywords = frozenset({
        'company', 'companies', 'organization', 'organizations', 'org', 'orgs',
        'show', 'view', 'search', 'find', 'list', 'all', 'which', 'what',
    })
    patterns = [
        re.compile(r'(show|view|list|search|find).*(compan|organi)'),
        re.compile(r'(compan|organi).*(details|info|information|about|list)'),
        re.compile(r'(companies|organizations)\s+(in|with|from|that|which|having)'),
        re.compile(r'(which|what)\s+compan'),
        re.compile(r'search\s+(?:for\s+)?\w+'),
    ]

    def execute(self, text, user):
        from companies.models import Company
        from django.db.models import Count

        text_lower = text.lower()

        # "which company has the most contacts"
        if re.search(r'(?:which|what)\s+compan.*most\s+contacts', text_lower):
            companies = Company.objects.filter(owner=user).annotate(
                contact_count=Count('contacts')
            ).order_by('-contact_count')
            if not companies:
                return 'No companies found.'
            top = companies.first()
            if top.contact_count == 0:
                return 'No companies have any contacts yet.'
            return f'**{top.name}** has the most contacts ({top.contact_count} contacts).'

        # "show companies in X country"
        country_match = re.search(r'in\s+(\w+(?:\s+\w+)?)', text_lower)
        if country_match and ('companies' in text_lower or 'organizations' in text_lower):
            country = country_match.group(1).strip().title()
            qs = Company.objects.filter(owner=user, country__icontains=country).order_by('-created_at')[:10]
            if not qs:
                return f'No companies found in {country}.'
            lines = [f'**Companies in {country}**', '']
            for c in qs:
                lines.append(f'\u2022 {c.name} — {c.industry or "—"} ({c.status})')
            return '\n'.join(lines)

        # "show companies with more than X employees"
        emp_match = re.search(r'more\s+than\s+(\d[\d,]*)\s+employees', text_lower)
        if emp_match:
            threshold = int(emp_match.group(1).replace(',', ''))
            qs = Company.objects.filter(owner=user, employees__gte=threshold).order_by('-employees')[:10]
            if not qs:
                return f'No companies with more than {threshold:,} employees found.'
            lines = [f'**Companies with {threshold:,}+ employees**', '']
            for c in qs:
                lines.append(f'\u2022 {c.name} — {c.employees:,} employees ({c.industry or "—"})')
            return '\n'.join(lines)

        # "show companies in X industry / IT industry"
        industry_match = re.search(r'in\s+(.+?)\s+(?:industry|sector)', text_lower)
        if industry_match:
            ind_query = industry_match.group(1).strip().lower()
            ind_map = {
                'technology': 'Technology', 'tech': 'Technology', 'it': 'Technology',
                'healthcare': 'Healthcare', 'health': 'Healthcare',
                'finance': 'Finance', 'banking': 'Finance',
                'education': 'Education', 'edtech': 'Education',
                'manufacturing': 'Manufacturing',
                'retail': 'Retail', 'ecommerce': 'Retail',
                'real estate': 'Real Estate',
                'consulting': 'Consulting',
                'media': 'Media',
            }
            industry_val = ind_map.get(ind_query, ind_query.title())
            qs = Company.objects.filter(owner=user, industry__iexact=industry_val).order_by('-created_at')[:10]
            if not qs:
                return f'No companies found in the {industry_val} industry.'
            lines = [f'**Companies in {industry_val} industry**', '']
            for c in qs:
                lines.append(f'\u2022 {c.name} — {c.city or "—"} ({c.status})')
            return '\n'.join(lines)

        # "show all companies" / "list companies"
        if re.search(r'(?:show|view|list)\s+(?:all\s+)?(?:companies|organizations)', text_lower):
            qs = Company.objects.filter(owner=user).order_by('-created_at')[:20]
            if not qs:
                return 'No companies found. Create your first company!'
            lines = ['**All Companies**', '']
            for c in qs:
                lines.append(f'\u2022 {c.name} — {c.industry or "—"} ({c.status})')
            if qs.count() == 20:
                lines.append('*(showing first 20)*')
            return '\n'.join(lines)

        # "search for X" — find a company by name
        search_match = re.search(r'(?:search|find)\s+(?:for\s+)?(.+)', text_lower)
        if search_match:
            query = search_match.group(1).strip()
        else:
            query = text_lower.strip()
            query = re.sub(r'^(?:show|view|open|tell\s+me\s+about)\s+(?:the\s+)?(?:company|organization|org)\s+', '', query).strip()

        if query and len(query) >= 2:
            company = Company.objects.filter(owner=user, name__icontains=query).first()
            if not company:
                company = Company.objects.filter(owner=user, name__iexact=query).first()
            if company:
                contacts_count = company.contacts.count()
                leads_count = company.leads.count()
                deals_count = company.deals.count()
                return (
                    f'**{company.name}**\n'
                    f'Industry: {company.industry or "—"}\n'
                    f'Website: {company.website or "—"}\n'
                    f'Email: {company.email or "—"}\n'
                    f'Phone: {company.phone or "—"}\n'
                    f'Location: {company.city or "—"}, {company.country or "—"}\n'
                    f'Employees: {company.employees or "—"}\n'
                    f'Revenue: ${company.annual_revenue:,.2f}' if company.annual_revenue else 'Revenue: —\n'
                    f'Status: {company.status}\n'
                    f'Contacts: {contacts_count} | Leads: {leads_count} | Deals: {deals_count}'
                )

        return 'Please specify a company name or try "Show all companies".'
