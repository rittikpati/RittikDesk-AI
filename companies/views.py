import csv
import io
import json

from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView, FormView, View
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.db.models import Q, Count, Sum
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect
from core.mixins import OwnerFilterMixin
from workflows.services.engine import fire_trigger
from activities.services import log_activity
from .models import Company
from .forms import CompanyForm, CompanyImportForm

SORT_MAP = {
    'newest': '-created_at',
    'oldest': 'created_at',
    'name_asc': 'name',
    'name_desc': '-name',
    'industry_asc': 'industry',
    'industry_desc': '-industry',
    'city_asc': 'city',
    'city_desc': '-city',
    'country_asc': 'country',
    'country_desc': '-country',
    'employees_asc': 'employees',
    'employees_desc': '-employees',
    'revenue_asc': 'annual_revenue',
    'revenue_desc': '-annual_revenue',
    'status_asc': 'status',
    'status_desc': '-status',
}

INDUSTRY_CHOICES = [
    'Technology', 'Healthcare', 'Finance', 'Education', 'Manufacturing',
    'Retail', 'Real Estate', 'Consulting', 'Media', 'Telecommunications',
    'Transportation', 'Energy', 'Hospitality', 'Agriculture', 'Other',
]
STATUS_CHOICES = ['Active', 'Inactive', 'Lead', 'Prospect', 'Customer', 'Partner', 'Former']


def apply_filters(qs, search, industry_filter, status_filter, country_filter, sort):
    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(website__icontains=search) |
            Q(email__icontains=search) |
            Q(phone__icontains=search) |
            Q(city__icontains=search) |
            Q(state__icontains=search) |
            Q(country__icontains=search) |
            Q(description__icontains=search) |
            Q(industry__icontains=search)
        )
    if industry_filter and industry_filter in INDUSTRY_CHOICES:
        qs = qs.filter(industry=industry_filter)
    if status_filter and status_filter in STATUS_CHOICES:
        qs = qs.filter(status=status_filter)
    if country_filter:
        qs = qs.filter(country__icontains=country_filter)
    ordering = SORT_MAP.get(sort, '-created_at')
    qs = qs.order_by(ordering)
    return qs


class CompanyListView(OwnerFilterMixin, ListView):
    model = Company
    template_name = 'companies/company_list.html'
    context_object_name = 'companies'
    paginate_by = 10

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.GET.get('search', '').strip()
        industry_filter = self.request.GET.get('industry', '').strip()
        status_filter = self.request.GET.get('status', '').strip()
        country_filter = self.request.GET.get('country', '').strip()
        sort = self.request.GET.get('sort', 'newest')
        return apply_filters(qs, search, industry_filter, status_filter, country_filter, sort)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['current_industry'] = self.request.GET.get('industry', '')
        context['current_status'] = self.request.GET.get('status', '')
        context['current_country'] = self.request.GET.get('country', '')
        context['current_sort'] = self.request.GET.get('sort', 'newest')
        context['industries'] = INDUSTRY_CHOICES
        context['statuses'] = STATUS_CHOICES
        context['countries'] = (
            Company.objects.filter(owner=self.request.user)
            .exclude(country='')
            .values_list('country', flat=True)
            .distinct().order_by('country')
        )
        return context


class CompanySearchJsonView(OwnerFilterMixin, View):
    model = Company

    def get(self, request):
        search = request.GET.get('search', '').strip()
        industry_filter = request.GET.get('industry', '').strip()
        status_filter = request.GET.get('status', '').strip()
        country_filter = request.GET.get('country', '').strip()
        sort = request.GET.get('sort', 'newest')
        qs = Company.objects.filter(owner=request.user)
        qs = apply_filters(qs, search, industry_filter, status_filter, country_filter, sort)

        data = []
        for c in qs:
            data.append({
                'id': c.pk,
                'name': c.name,
                'industry': c.industry or '',
                'website': c.website or '',
                'email': c.email or '',
                'phone': c.phone or '',
                'city': c.city or '',
                'state': c.state or '',
                'country': c.country or '',
                'employees': c.employees,
                'annual_revenue': str(c.annual_revenue) if c.annual_revenue else '',
                'status': c.status,
                'detail_url': reverse('companies:detail', args=[c.pk]),
                'update_url': reverse('companies:update', args=[c.pk]),
                'delete_url': reverse('companies:delete', args=[c.pk]),
            })

        return JsonResponse({
            'companies': data,
            'count': len(data),
        })


class CompanyCreateView(OwnerFilterMixin, CreateView):
    model = Company
    form_class = CompanyForm
    template_name = 'companies/company_form.html'
    success_url = reverse_lazy('companies:list')

    def form_valid(self, form):
        form.instance.owner = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, f'Company "{form.instance.name}" created successfully.')
        fire_trigger('company_created', form.instance)
        log_activity(self.request.user, 'company_created',
                     name=form.instance.name,
                     object_id=form.instance.pk, object_repr=form.instance.name,
                     detail_url=f'/companies/{form.instance.pk}/',
                     description=f'New company created: {form.instance.name}')
        return response


class CompanyUpdateView(OwnerFilterMixin, UpdateView):
    model = Company
    form_class = CompanyForm
    template_name = 'companies/company_form.html'
    success_url = reverse_lazy('companies:list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Company "{form.instance.name}" updated successfully.')
        fire_trigger('company_updated', form.instance)
        log_activity(self.request.user, 'company_updated',
                     name=form.instance.name,
                     object_id=form.instance.pk, object_repr=form.instance.name,
                     detail_url=f'/companies/{form.instance.pk}/',
                     description=f'Company "{form.instance.name}" updated')
        return response


class CompanyDetailView(OwnerFilterMixin, DetailView):
    model = Company
    template_name = 'companies/company_detail.html'
    context_object_name = 'company'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.object

        context['contacts'] = company.contacts.filter(owner=self.request.user).order_by('-created_at')[:10]
        context['leads'] = company.leads.filter(owner=self.request.user).order_by('-created_at')[:10]
        context['deals'] = company.deals.filter(owner=self.request.user).order_by('-created_at')[:10]
        context['tasks'] = company.tasks.filter(owner=self.request.user).order_by('-created_at')[:10]
        context['events'] = company.events.filter(owner=self.request.user).order_by('-start_date')[:10]

        context['contacts_count'] = company.contacts.filter(owner=self.request.user).count()
        context['leads_count'] = company.leads.filter(owner=self.request.user).count()
        context['deals_count'] = company.deals.filter(owner=self.request.user).count()
        context['tasks_count'] = company.tasks.filter(owner=self.request.user).count()
        context['events_count'] = company.events.filter(owner=self.request.user).count()

        recent = []
        for c in company.contacts.filter(owner=self.request.user).order_by('-created_at')[:5]:
            recent.append({'type': 'contact', 'text': f'Contact <strong>{c.full_name}</strong> added', 'time': c.created_at})
        for l in company.leads.filter(owner=self.request.user).order_by('-created_at')[:5]:
            recent.append({'type': 'lead', 'text': f'Lead <strong>{l.lead_name}</strong> created — {l.status}', 'time': l.created_at})
        for d in company.deals.filter(owner=self.request.user).order_by('-created_at')[:5]:
            recent.append({'type': 'deal', 'text': f'Deal <strong>{d.deal_name}</strong> created — {d.stage}', 'time': d.created_at})
        for t in company.tasks.filter(owner=self.request.user).order_by('-created_at')[:5]:
            recent.append({'type': 'task', 'text': f'Task <strong>{t.title}</strong> created', 'time': t.created_at})
        for e in company.events.filter(owner=self.request.user).order_by('-created_at')[:5]:
            recent.append({'type': 'event', 'text': f'Event <strong>{e.title}</strong> scheduled', 'time': e.created_at})
        recent.sort(key=lambda x: x['time'], reverse=True)
        context['recent_activities'] = recent[:10]

        return context


class CompanyDeleteView(OwnerFilterMixin, DeleteView):
    model = Company
    template_name = 'companies/company_confirm_delete.html'
    success_url = reverse_lazy('companies:list')
    context_object_name = 'company'

    def form_valid(self, form):
        name = self.request.POST.get('confirm_name', '').strip()
        if name.lower() != self.object.name.lower():
            messages.error(self.request, 'Please type the company name exactly to confirm deletion.')
            return self.form_invalid(form)
        messages.success(self.request, f'Company "{self.object.name}" deleted.')
        fire_trigger('company_deleted', self.object)
        log_activity(self.request.user, 'company_deleted',
                     name=self.object.name,
                     object_id=self.object.pk, object_repr=self.object.name,
                     description=f'Company "{self.object.name}" deleted')
        return super().form_valid(form)


class CompanyBulkDeleteView(OwnerFilterMixin, View):
    model = Company

    def post(self, request):
        ids = request.POST.getlist('ids')
        if not ids:
            return JsonResponse({'error': 'No companies selected.'}, status=400)
        qs = self.get_queryset().filter(pk__in=ids)
        count = qs.count()
        qs.delete()
        return JsonResponse({'success': True, 'deleted': count})


class CompanyBulkUpdateView(OwnerFilterMixin, View):
    model = Company

    def post(self, request):
        ids = request.POST.getlist('ids')
        new_status = request.POST.get('status', '').strip()
        if not ids:
            return JsonResponse({'error': 'No companies selected.'}, status=400)
        qs = self.get_queryset().filter(pk__in=ids)
        update_data = {}
        if new_status and new_status in STATUS_CHOICES:
            update_data['status'] = new_status
        if update_data:
            qs.update(**update_data)
        return JsonResponse({'success': True, 'updated': qs.count()})


class CompanyExportCSVView(OwnerFilterMixin, View):
    model = Company

    def get(self, request):
        ids = request.GET.get('ids', '')
        qs = self.get_queryset()
        if ids:
            id_list = [int(i) for i in ids.split(',') if i.isdigit()]
            qs = qs.filter(pk__in=id_list)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="companies.csv"'
        writer = csv.writer(response)
        writer.writerow([
            'Name', 'Industry', 'Website', 'Email', 'Phone',
            'Address', 'City', 'State', 'Country', 'Postal Code',
            'Employees', 'Annual Revenue', 'Company Size', 'LinkedIn URL',
            'Description', 'Status',
        ])
        for c in qs:
            writer.writerow([
                c.name, c.industry, c.website, c.email, c.phone,
                c.address, c.city, c.state, c.country, c.postal_code,
                c.employees, str(c.annual_revenue) if c.annual_revenue else '',
                c.company_size, c.linkedin_url, c.description, c.status,
            ])
        return response


class CompanyExportXLSXView(OwnerFilterMixin, View):
    model = Company

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        ids = request.GET.get('ids', '')
        qs = self.get_queryset()
        if ids:
            id_list = [int(i) for i in ids.split(',') if i.isdigit()]
            qs = qs.filter(pk__in=id_list)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Companies'

        headers = [
            'Name', 'Industry', 'Website', 'Email', 'Phone',
            'Address', 'City', 'State', 'Country', 'Postal Code',
            'Employees', 'Annual Revenue', 'Company Size', 'LinkedIn URL',
            'Description', 'Status',
        ]
        header_fill = PatternFill(start_color='6C63FF', end_color='6C63FF', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, c in enumerate(qs, 2):
            ws.cell(row=row_idx, column=1, value=c.name)
            ws.cell(row=row_idx, column=2, value=c.industry)
            ws.cell(row=row_idx, column=3, value=c.website)
            ws.cell(row=row_idx, column=4, value=c.email)
            ws.cell(row=row_idx, column=5, value=c.phone)
            ws.cell(row=row_idx, column=6, value=c.address)
            ws.cell(row=row_idx, column=7, value=c.city)
            ws.cell(row=row_idx, column=8, value=c.state)
            ws.cell(row=row_idx, column=9, value=c.country)
            ws.cell(row=row_idx, column=10, value=c.postal_code)
            ws.cell(row=row_idx, column=11, value=c.employees)
            ws.cell(row=row_idx, column=12, value=float(c.annual_revenue) if c.annual_revenue else '')
            ws.cell(row=row_idx, column=13, value=c.company_size)
            ws.cell(row=row_idx, column=14, value=c.linkedin_url)
            ws.cell(row=row_idx, column=15, value=c.description)
            ws.cell(row=row_idx, column=16, value=c.status)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="companies.xlsx"'
        wb.save(response)
        return response


class CompanyExportPDFView(OwnerFilterMixin, View):
    model = Company

    def get(self, request):
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        ids = request.GET.get('ids', '')
        qs = self.get_queryset()
        if ids:
            id_list = [int(i) for i in ids.split(',') if i.isdigit()]
            qs = qs.filter(pk__in=id_list)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=16, spaceAfter=12)
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7, leading=9)

        elements = [Paragraph('Companies Export', title_style), Spacer(1, 8)]

        headers = ['Name', 'Industry', 'City', 'Country', 'Employees', 'Revenue', 'Status']
        table_data = [headers]

        for c in qs:
            table_data.append([
                Paragraph(str(c.name)[:30], cell_style),
                c.industry or '-',
                c.city or '-',
                c.country or '-',
                str(c.employees) if c.employees else '-',
                f'${float(c.annual_revenue):,.2f}' if c.annual_revenue else '-',
                c.status,
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C63FF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (4, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5ff')]),
        ]))

        elements.append(table)
        doc.build(elements)
        buf.seek(0)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="companies.pdf"'
        response.write(buf.read())
        return response


class CompanyImportView(LoginRequiredMixin, FormView):
    template_name = 'companies/company_import.html'
    form_class = CompanyImportForm
    success_url = reverse_lazy('companies:list')

    def form_valid(self, form):
        csv_file = form.cleaned_data['csv_file']
        decoded = csv_file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(decoded))
        created = 0
        errors = []
        for row_num, row in enumerate(reader, 2):
            name = row.get('name', '').strip()
            if not name:
                errors.append(f'Row {row_num}: name is required.')
                continue
            if Company.objects.filter(name__iexact=name, owner=self.request.user).exists():
                errors.append(f'Row {row_num}: "{name}" already exists.')
                continue
            try:
                Company.objects.create(
                    owner=self.request.user,
                    name=name,
                    industry=row.get('industry', '').strip(),
                    website=row.get('website', '').strip(),
                    email=row.get('email', '').strip(),
                    phone=row.get('phone', '').strip(),
                    city=row.get('city', '').strip(),
                    state=row.get('state', '').strip(),
                    country=row.get('country', '').strip(),
                    postal_code=row.get('postal_code', '').strip(),
                    employees=row.get('employees', '').strip() or None,
                    annual_revenue=row.get('annual_revenue', '').strip() or None,
                    description=row.get('description', '').strip(),
                    status=row.get('status', '').strip() or 'Active',
                )
                created += 1
            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')

        if created:
            messages.success(self.request, f'{created} company(ies) imported successfully.')
        if errors:
            for err in errors[:10]:
                messages.warning(self.request, err)
            if len(errors) > 10:
                messages.warning(self.request, f'... and {len(errors) - 10} more errors.')
        return super().form_valid(form)
