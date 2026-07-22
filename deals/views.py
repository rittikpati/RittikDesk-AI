import csv
import io
import json

from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView, TemplateView, View
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from core.mixins import OwnerFilterMixin
from workflows.services.engine import fire_trigger
from .models import Deal
from .forms import DealForm

SORT_MAP = {
    'newest': '-created_at',
    'oldest': 'created_at',
    'name_asc': 'deal_name',
    'name_desc': '-deal_name',
    'company_asc': 'company',
    'company_desc': '-company',
    'value_asc': 'value',
    'value_desc': '-value',
    'close_date_asc': 'expected_close_date',
    'close_date_desc': '-expected_close_date',
    'probability_asc': 'probability',
    'probability_desc': '-probability',
}

STAGE_CHOICES = ['New', 'Qualified', 'Proposal Sent', 'Negotiation', 'Contract Review', 'Won', 'Lost']
PRIORITY_CHOICES = ['Low', 'Medium', 'High', 'Urgent']
SOURCE_CHOICES = ['Website', 'Referral', 'LinkedIn', 'Facebook', 'Instagram', 'Cold Email', 'Event', 'Other']


def apply_filters(qs, search, stage_filter, priority_filter, source_filter, sort):
    if search:
        qs = qs.filter(
            Q(deal_name__icontains=search) |
            Q(company__icontains=search) |
            Q(description__icontains=search) |
            Q(notes__icontains=search) |
            Q(owner__username__icontains=search) |
            Q(owner__email__icontains=search) |
            Q(owner__first_name__icontains=search) |
            Q(owner__last_name__icontains=search)
        )
    if stage_filter and stage_filter in STAGE_CHOICES:
        qs = qs.filter(stage=stage_filter)
    if priority_filter and priority_filter in PRIORITY_CHOICES:
        qs = qs.filter(priority=priority_filter)
    if source_filter and source_filter in SOURCE_CHOICES:
        qs = qs.filter(source=source_filter)
    ordering = SORT_MAP.get(sort, '-created_at')
    qs = qs.order_by(ordering)
    return qs


class DealListView(OwnerFilterMixin, ListView):
    model = Deal
    template_name = 'deals/deal_list.html'
    context_object_name = 'deals'
    paginate_by = 10

    def get_queryset(self):
        qs = super().get_queryset().select_related('contact', 'lead')
        search = self.request.GET.get('search', '').strip()
        stage_filter = self.request.GET.get('stage', '').strip()
        priority_filter = self.request.GET.get('priority', '').strip()
        source_filter = self.request.GET.get('source', '').strip()
        sort = self.request.GET.get('sort', 'newest')
        return apply_filters(qs, search, stage_filter, priority_filter, source_filter, sort)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_deals'] = self.object_list.count()
        context['search_query'] = self.request.GET.get('search', '')
        context['active_stage'] = self.request.GET.get('stage', '')
        context['active_priority'] = self.request.GET.get('priority', '')
        context['active_source'] = self.request.GET.get('source', '')
        context['sort_by'] = self.request.GET.get('sort', 'newest')
        context['stage_choices'] = STAGE_CHOICES
        context['priority_choices'] = PRIORITY_CHOICES
        context['source_choices'] = SOURCE_CHOICES
        context['sort_choices'] = [
            ('newest', 'Newest First'),
            ('oldest', 'Oldest First'),
            ('name_asc', 'Name (A–Z)'),
            ('name_desc', 'Name (Z–A)'),
            ('company_asc', 'Company (A–Z)'),
            ('company_desc', 'Company (Z–A)'),
            ('value_asc', 'Value (Low)'),
            ('value_desc', 'Value (High)'),
            ('close_date_asc', 'Close Date (Earliest)'),
            ('close_date_desc', 'Close Date (Latest)'),
            ('probability_asc', 'Probability (Low)'),
            ('probability_desc', 'Probability (High)'),
        ]
        return context


class DealCreateView(LoginRequiredMixin, CreateView):
    model = Deal
    form_class = DealForm
    template_name = 'deals/deal_form.html'
    success_url = reverse_lazy('deals:list')

    def form_valid(self, form):
        form.instance.owner = self.request.user
        form.instance.update_status_from_stage()
        response = super().form_valid(form)
        fire_trigger('deal_created', form.instance)
        messages.success(self.request, f'Deal "{form.instance.deal_name}" created successfully.')
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add Deal'
        return context


class DealDetailView(OwnerFilterMixin, DetailView):
    model = Deal
    template_name = 'deals/deal_detail.html'
    context_object_name = 'deal'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        deal = self.object
        if deal.value and deal.stage in ('New', 'Qualified', 'Proposal Sent', 'Negotiation', 'Contract Review'):
            from decimal import Decimal
            context['weighted_value'] = deal.value * (Decimal(str(deal.probability)) / Decimal('100'))
        else:
            context['weighted_value'] = None
        return context


class DealUpdateView(OwnerFilterMixin, UpdateView):
    model = Deal
    form_class = DealForm
    template_name = 'deals/deal_form.html'
    success_url = reverse_lazy('deals:list')

    def form_valid(self, form):
        old_stage = self.get_object().stage
        form.instance.update_status_from_stage()
        response = super().form_valid(form)
        fire_trigger('deal_updated', form.instance)
        new_stage = form.instance.stage
        if old_stage != new_stage:
            fire_trigger('deal_stage_changed', form.instance)
            if new_stage == 'Won':
                fire_trigger('deal_won', form.instance)
            elif new_stage == 'Lost':
                fire_trigger('deal_lost', form.instance)
        messages.success(self.request, f'Deal "{form.instance.deal_name}" updated successfully.')
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit Deal'
        context['is_update'] = True
        return context


class DealDeleteView(OwnerFilterMixin, DeleteView):
    model = Deal
    template_name = 'deals/deal_confirm_delete.html'
    success_url = reverse_lazy('deals:list')
    context_object_name = 'deal'

    def form_valid(self, form):
        messages.success(self.request, f'Deal "{self.object.deal_name}" deleted successfully.')
        return super().form_valid(form)


class DealSearchJsonView(LoginRequiredMixin, View):
    def get(self, request):
        search = request.GET.get('search', '').strip()
        stage_filter = request.GET.get('stage', '').strip()
        priority_filter = request.GET.get('priority', '').strip()
        source_filter = request.GET.get('source', '').strip()
        sort = request.GET.get('sort', 'newest')
        qs = Deal.objects.filter(owner=request.user).select_related('contact', 'lead')
        qs = apply_filters(qs, search, stage_filter, priority_filter, source_filter, sort)

        data = []
        for d in qs:
            data.append({
                'id': d.pk,
                'deal_name': d.deal_name,
                'company': d.company or '',
                'contact_name': d.contact.full_name if d.contact else '',
                'lead_name': d.lead.lead_name if d.lead else '',
                'deal_owner': d.owner.get_full_name() or d.owner.username,
                'value': str(d.value) if d.value else '',
                'currency': d.currency,
                'stage': d.stage,
                'probability': d.probability,
                'priority': d.priority,
                'expected_close_date': str(d.expected_close_date) if d.expected_close_date else '',
                'source': d.source,
                'status': d.status,
                'detail_url': reverse('deals:detail', args=[d.pk]),
                'update_url': reverse('deals:update', args=[d.pk]),
                'delete_url': reverse('deals:delete', args=[d.pk]),
            })

        return JsonResponse({
            'deals': data,
            'count': len(data),
            'search': search,
            'stage': stage_filter,
            'priority': priority_filter,
            'source': source_filter,
            'sort': sort,
        })


class DealBulkDeleteView(OwnerFilterMixin, View):
    def post(self, request):
        ids = request.POST.getlist('ids')
        deals = Deal.objects.filter(owner=request.user, pk__in=ids)
        count = deals.count()
        deals.delete()
        messages.success(request, f'{count} deal(s) deleted successfully.')
        return JsonResponse({'deleted': count})


class DealBulkUpdateView(OwnerFilterMixin, View):
    def post(self, request):
        ids = request.POST.getlist('ids')
        stage = request.POST.get('stage', '').strip()
        priority = request.POST.get('priority', '').strip()
        deals = Deal.objects.filter(owner=request.user, pk__in=ids)
        updated = 0
        for deal in deals:
            changed = False
            if stage and stage in STAGE_CHOICES and deal.stage != stage:
                deal.stage = stage
                deal.update_status_from_stage()
                changed = True
            if priority and priority in PRIORITY_CHOICES and deal.priority != priority:
                deal.priority = priority
                changed = True
            if changed:
                deal.save()
                fire_trigger('deal_updated', deal)
                updated += 1
        messages.success(request, f'{updated} deal(s) updated successfully.')
        return JsonResponse({'updated': updated})


class DealExportCSVView(OwnerFilterMixin, View):
    model = Deal

    def get(self, request):
        ids = request.GET.get('ids', '')
        qs = self.get_queryset().select_related('contact', 'lead')
        if ids:
            id_list = [int(i) for i in ids.split(',') if i.isdigit()]
            qs = qs.filter(pk__in=id_list)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="deals.csv"'
        writer = csv.writer(response)
        writer.writerow([
            'Deal Name', 'Company', 'Contact', 'Lead', 'Deal Owner',
            'Value', 'Currency', 'Stage', 'Probability', 'Expected Close Date',
            'Source', 'Priority', 'Status', 'Description', 'Notes',
        ])
        for d in qs:
            writer.writerow([
                d.deal_name, d.company,
                d.contact.full_name if d.contact else '',
                d.lead.lead_name if d.lead else '',
                d.owner.get_full_name() or d.owner.username,
                str(d.value) if d.value else '',
                d.currency, d.stage, d.probability,
                str(d.expected_close_date) if d.expected_close_date else '',
                d.source, d.priority, d.status, d.description, d.notes,
            ])
        return response


class DealExportXLSXView(OwnerFilterMixin, View):
    model = Deal

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        ids = request.GET.get('ids', '')
        qs = self.get_queryset().select_related('contact', 'lead')
        if ids:
            id_list = [int(i) for i in ids.split(',') if i.isdigit()]
            qs = qs.filter(pk__in=id_list)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Deals'

        headers = [
            'Deal Name', 'Company', 'Contact', 'Lead', 'Deal Owner',
            'Value', 'Currency', 'Stage', 'Probability', 'Expected Close Date',
            'Source', 'Priority', 'Status', 'Description', 'Notes',
        ]
        header_fill = PatternFill(start_color='6C63FF', end_color='6C63FF', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, d in enumerate(qs, 2):
            ws.cell(row=row_idx, column=1, value=d.deal_name)
            ws.cell(row=row_idx, column=2, value=d.company)
            ws.cell(row=row_idx, column=3, value=d.contact.full_name if d.contact else '')
            ws.cell(row=row_idx, column=4, value=d.lead.lead_name if d.lead else '')
            ws.cell(row=row_idx, column=5, value=d.owner.get_full_name() or d.owner.username)
            ws.cell(row=row_idx, column=6, value=float(d.value) if d.value else '')
            ws.cell(row=row_idx, column=7, value=d.currency)
            ws.cell(row=row_idx, column=8, value=d.stage)
            ws.cell(row=row_idx, column=9, value=d.probability)
            ws.cell(row=row_idx, column=10, value=str(d.expected_close_date) if d.expected_close_date else '')
            ws.cell(row=row_idx, column=11, value=d.source)
            ws.cell(row=row_idx, column=12, value=d.priority)
            ws.cell(row=row_idx, column=13, value=d.status)
            ws.cell(row=row_idx, column=14, value=d.description)
            ws.cell(row=row_idx, column=15, value=d.notes)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="deals.xlsx"'
        wb.save(response)
        return response


class DealExportPDFView(OwnerFilterMixin, View):
    model = Deal

    def get(self, request):
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        ids = request.GET.get('ids', '')
        qs = self.get_queryset().select_related('contact', 'lead')
        if ids:
            id_list = [int(i) for i in ids.split(',') if i.isdigit()]
            qs = qs.filter(pk__in=id_list)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=16, spaceAfter=12)
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7, leading=9)

        elements = [Paragraph('Deals Export', title_style), Spacer(1, 8)]

        headers = ['Name', 'Company', 'Value', 'Stage', 'Prob.', 'Close Date', 'Priority', 'Status']
        table_data = [headers]

        for d in qs:
            table_data.append([
                Paragraph(str(d.deal_name)[:30], cell_style),
                Paragraph(str(d.company)[:20], cell_style),
                f'{d.currency} {float(d.value):,.2f}' if d.value else '-',
                d.stage,
                f'{d.probability}%',
                str(d.expected_close_date) if d.expected_close_date else '-',
                d.priority,
                d.status,
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C63FF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5ff')]),
        ]))

        elements.append(table)
        doc.build(elements)
        buf.seek(0)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="deals.pdf"'
        response.write(buf.read())
        return response


class DealKanbanView(OwnerFilterMixin, ListView):
    model = Deal
    template_name = 'deals/kanban.html'
    context_object_name = 'deals'

    def get_queryset(self):
        return super().get_queryset().select_related('contact', 'lead').order_by('-value', '-probability')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        stages = ['New', 'Qualified', 'Proposal Sent', 'Negotiation', 'Contract Review', 'Won', 'Lost']
        stage_data = {}
        for stage in stages:
            stage_data[stage] = [d for d in context['deals'] if d.stage == stage]

        pipeline_stats = Deal.objects.filter(owner=self.request.user)
        total = pipeline_stats.count()
        won = pipeline_stats.filter(stage='Won').count()
        lost = pipeline_stats.filter(stage='Lost').count()
        open_deals = pipeline_stats.exclude(stage__in=['Won', 'Lost']).count()
        revenue = pipeline_stats.filter(stage='Won').aggregate(Sum('value'))['value__sum'] or 0
        potential = pipeline_stats.exclude(stage__in=['Won', 'Lost']).aggregate(Sum('value'))['value__sum'] or 0
        avg_size = revenue / won if won else 0
        win_rate = round((won / total * 100), 1) if total else 0
        loss_rate = round((lost / total * 100), 1) if total else 0
        conv_rate = round((won / (won + lost) * 100), 1) if (won + lost) else 0

        context['stage_data'] = stage_data
        context['stages'] = stages
        context['total_deals'] = total
        context['won_deals'] = won
        context['lost_deals'] = lost
        context['open_deals'] = open_deals
        context['revenue'] = revenue
        context['potential_revenue'] = potential
        context['avg_deal_size'] = avg_size
        context['win_rate'] = win_rate
        context['loss_rate'] = loss_rate
        context['conversion_rate'] = conv_rate
        return context


class DealKanbanUpdateView(LoginRequiredMixin, View):
    def post(self, request):
        deal_id = request.POST.get('deal_id')
        new_stage = request.POST.get('stage')
        if not deal_id or not new_stage or new_stage not in STAGE_CHOICES:
            return JsonResponse({'error': 'Invalid parameters'}, status=400)
        deal = get_object_or_404(Deal, pk=deal_id, owner=request.user)
        old_stage = deal.stage
        if old_stage != new_stage:
            deal.stage = new_stage
            deal.update_status_from_stage()
            deal.save()
            fire_trigger('deal_stage_changed', deal)
            if new_stage == 'Won':
                fire_trigger('deal_won', deal)
            elif new_stage == 'Lost':
                fire_trigger('deal_lost', deal)
        return JsonResponse({'success': True, 'deal_id': deal_id, 'stage': new_stage})
